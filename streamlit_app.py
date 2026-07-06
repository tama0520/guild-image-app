#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
パチスロ・パチンコ データ画像生成 Web アプリ
Streamlit + PIL で実装

起動方法:
    streamlit run streamlit_app.py

依存ライブラリ:
    pip install streamlit pandas openpyxl pillow
"""

import copy
import datetime
import math
import io
import json
import os
import platform
import re
import subprocess
import sys
import tempfile
import traceback
import unicodedata
import zipfile

from openpyxl import load_workbook

# フォントディレクトリ（このスクリプトと同階層の fonts/ フォルダ）
_FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")

# .env をモジュール起動時に一度だけ読み込む（ローカル用）
# Cloud では st.secrets が使われるため .env がなくてもエラーにしない
try:
    from pathlib import Path as _Path
    from dotenv import load_dotenv as _load_dotenv
    for _ep in [_Path(__file__).parent / ".env", _Path(".env")]:
        if _ep.exists():
            _load_dotenv(_ep, override=True)
            break
except Exception:
    pass

import logging as _logging
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from PIL import Image, ImageDraw, ImageFont, ImageFilter

# components.html(height=0) を使う不可視JSインジェクション用に警告を抑制
# st.iframe は height=0 を受け付けないためレイアウトが崩れる（2026-06-01削除予定）
_logging.getLogger("streamlit.deprecation_util").setLevel(_logging.ERROR)

# =============================================================================
# ■ ①設定データ
#   ─ ここを編集するだけで店舗・画像種類を追加できます ─
# =============================================================================

# 店舗 → 利用できる画像種類リスト
_IMAGE_TYPES = [
    "全台データ画像",
    "高配分データ画像",
    "並び画像",
    "末尾画像",
    "その他の優秀台画像",
]

# 店舗は cats.jpg の並び順（4列×3行）で定義する
STORES: dict[str, list[str]] = {
    "新宿歌舞伎町": _IMAGE_TYPES,
    "西武新宿":     _IMAGE_TYPES,
    "新大久保":     _IMAGE_TYPES,
    "高田馬場":     _IMAGE_TYPES,
    "上野本館":     _IMAGE_TYPES,
    "上野新館":     _IMAGE_TYPES,
    "渋谷新館":     _IMAGE_TYPES,
    "赤坂見附":     _IMAGE_TYPES,
    "新小岩":       _IMAGE_TYPES,
    "溝の口本館":   _IMAGE_TYPES,
    "溝の口新館":   _IMAGE_TYPES,
    "稲毛":         _IMAGE_TYPES,
    "秋葉原":       _IMAGE_TYPES,
}

# オススメ機種ピックアップ・素材メモ・詳細結果テキスト機能を持つ店舗
# 新店舗に同機能を追加するにはここに店舗名を追記するだけでよい
EXTENDED_FEATURE_STORES: frozenset[str] = frozenset({"新宿歌舞伎町", "西武新宿", "新大久保", "高田馬場", "上野本館", "上野新館", "渋谷新館", "赤坂見附", "新小岩", "溝の口本館", "溝の口新館", "稲毛"})

# 店舗ごとの結果テキスト絵文字設定 {store: (e1, e2)}
# e1=冒頭アイコン, e2=セクション見出しアイコン
# 未登録店舗は ("💫", "👑") がデフォルトで使われる
STORE_EMOJI_CONFIG: dict[str, tuple[str, str]] = {
    "新大久保":   ("🐝", "🍯"),
    "新小岩":     ("✨", "🍀"),
    "溝の口新館": ("📢", "🎁"),
    "西武新宿":   ("✨", "👑"),
    "赤坂見附":   ("🔥", "👑"),
}

# 店舗ごとの結果テキスト後処理置換リスト {store: [(旧, 新), ...]}
# generate_report_text 後にこの順で str.replace を適用する
STORE_RESULT_TRANSFORMS: dict[str, list[tuple[str, str]]] = {
    "赤坂見附": [
        ("👑高配分機種",    "👑高配分以上機種👑"),
        ("👑並び仕掛け",    "👑並び👑"),
        ("👑全台系濃厚機種", "👑全台系濃厚機種👑"),
        ("👑その他の優秀台", "👑その他の優秀台👑"),
        ("→", "➡"),
    ],
}

# ローテ結果テキストの絵文字設定 {store: (機種名囲み絵文字, 枚数ティア絵文字)}
# 未登録店舗は ("🌌", "🔥") がデフォルト
ROTE_EMOJI_CONFIG: dict[str, tuple[str, str]] = {
    "西武新宿":   ("👑", "💥"),
    "上野本館":   ("🗼", "👑"),
    "新大久保":   ("🍯", "🐝"),
    "溝の口新館": ("🌠", "🚨"),
}

# ローテ結果テキスト・台番プレフィックス絵文字 {store: emoji}
# 未登録店舗は "💫" がデフォルト
ROTE_BAN_EMOJI_CONFIG: dict[str, str] = {
    "新大久保": "✨",
}

# ローテ画像・台番列の背景色 {store: hex}
# 未登録店舗は "#00FFCC" がデフォルト
ROTE_BAN_COLOR_CONFIG: dict[str, str] = {
    "高田馬場":     "#EA5A96",
    "上野本館":     "#EA5A96",
    "渋谷新館":     "#F7EBCB",
    "新大久保":     "#AED6F1",
    "新宿歌舞伎町": "#FFF2CD",
}

# 拡張機能店舗のオススメブロック絵文字設定
# section_emoji: 「{emoji}{title}の優秀台」ヘッダー絵文字
# block_emojis:  各ブロック内の機種名文頭絵文字（4ブロック分）
# item_emoji:    その他の優秀台の台番プレフィックス絵文字
STORE_REC_CONFIG: dict[str, dict] = {
    "西武新宿": {
        "section_emoji": "🏆",
        "block_emojis":  ["🍀", "⚡️", "⭐", "🎯"],
        "item_emoji":    "📍",
    },
    "新小岩": {
        "section_emoji": "🍀",
        "block_emojis":  ["🎖️", "💥", "🤡", "🌺", "🎖️", "🎖️"],
        "item_emoji":    "🚩",
        # ブロックインデックス→カッコ内に表示する短縮名リスト（指定なしは「の優秀台」形式）
        "block_header_names": {
            0: ["スマスロ北斗", "北斗転生2", "東京喰種", "ヴヴヴ2", "かぐや様"],
            1: ["カバネリ海門", "モンキーV", "炎炎2", "真打吉宗"],
        },
    },
}

# 画像種類ごとの条件項目定義
# type: "text" | "int" | "float"  で入力ウィジェットが自動切り替わります
IMAGE_CONDITIONS: dict[str, list[dict]] = {
    "並び画像": [
        {
            "name":    "ranges",
            "label":   "台番範囲（例: 409-413, 315-317）",
            "type":    "text",
            "default": "",
        },
    ],
    "末尾画像": [
        {
            "name":    "tail",
            "label":   "末尾（例: 5、ゾロ目は「ゾロ目」と入力）",
            "type":    "text",
            "default": "",
        },
    ],
    "全台データ画像": [
        {
            "name":    "keyword",
            "label":   "機種名キーワード（部分一致・空白で全台）",
            "type":    "text",
            "default": "",
        },
        {
            "name":    "title",
            "label":   "画像タイトル",
            "type":    "text",
            "default": "全台データ",
        },
    ],
    "高配分データ画像": [
        {
            "name":    "keyword",
            "label":   "機種名キーワード（部分一致・空白で全台）",
            "type":    "text",
            "default": "",
        },
        {
            "name":    "title",
            "label":   "画像タイトル",
            "type":    "text",
            "default": "高配分データ",
        },
    ],
    "優秀台ピックアップ": [
        {
            "name":    "diff_min",
            "label":   "差枚 下限（枚）",
            "type":    "int",
            "default": 3000,
        },
        {
            "name":    "games_min",
            "label":   "G数 下限（G）",
            "type":    "int",
            "default": 5000,
        },
        {
            "name":    "title",
            "label":   "画像タイトル",
            "type":    "text",
            "default": "優秀台ピックアップ",
        },
    ],
    "機種別まとめ": [
        {
            "name":    "min_count",
            "label":   "最低台数（台）",
            "type":    "int",
            "default": 1,
        },
        {
            "name":    "title",
            "label":   "画像タイトル",
            "type":    "text",
            "default": "機種別まとめ",
        },
    ],
    "その他の優秀台画像": [
        {
            "name":    "bans",
            "label":   "台番（カンマ区切り）",
            "type":    "text",
            "default": "",
        },
        {
            "name":    "title",
            "label":   "画像タイトル",
            "type":    "text",
            "default": "その他の優秀台",
        },
    ],
}

# Excel 列名の正規化マップ（左が正規名、右が候補列名リスト・先に見つかったものを採用）
# ※ 正規名は既存スクリプト（convert_稲毛_*.py）の表示列名に合わせる
COLUMN_ALIASES: dict[str, list[str]] = {
    "台番":    ["台番", "台No", "台no", "台NO", "号機", "番台", "台番号"],
    "機種名":  [
        "機種名", "機種名（正式名）", "機種名（データサイト表記）",
        "機種名（表記）", "機種",
    ],
    "差枚":    ["差枚", "差枚数", "差玉", "差枚(枚)", "差"],
    "BB":      ["BB", "bb", "BIG", "big", "BIG回数", "BB回数"],
    "RB":      ["RB", "rb", "REG", "reg", "REG回数", "RB回数"],
    "AT":      ["ART", "art", "AT", "at", "ART回数", "AT回数"],   # 表示は AT
    "ゲーム数": ["G数", "G数(G)", "ゲーム数", "総ゲーム数", "G", "回転数", "スピン数"],  # 表示はゲーム数
}

# 機種名変換テーブルのパス
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
NAME_MAP_PATH = os.path.join(BASE_DIR, "機種名変換.xlsx")

# 出力先デスクトップ（どのPCでも自動検出）
_DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
# Streamlit Cloud 判定（Linux = Cloud、Windows = ローカル）
_IS_CLOUD = platform.system() != "Windows"

# 店舗設定（オススメ機種など）の永続化先
STORE_SETTINGS_DIR = os.path.join(BASE_DIR, "store_settings")
# ローテ機種名の永続化先
_ROTE_SAVE_FILE = os.path.join(BASE_DIR, "rote_machines.json")
# 週間オススメ表・項目の永続化先
_WEEKLY_SAVE_FILE = os.path.join(BASE_DIR, "weekly_items.json")
# 週間結果テキスト機種名の永続化先
_WRT_SAVE_FILE = os.path.join(BASE_DIR, "wrt_machines.json")

# 機種画像（パネル/液晶）関連。Cloudでも読めるようリポジトリ内に配置する。
# デスクトップの「パネル」フォルダの中身を assets/machine_images/ にコピーして管理する。
_MACHINE_IMAGES_DIR = os.path.join(BASE_DIR, "assets", "machine_images")
_MACHINE_IMAGE_MASTER_PATH = os.path.join(BASE_DIR, "masters", "machine_image_master.xlsx")
_MACHINE_IMG_EXTS = (".png", ".jpg", ".jpeg", ".webp")

# =============================================================================
# ■ 店舗別設定
#   新店舗を追加するときは STORES（上部）と STORE_CONFIG の両方に追記する。
#   設定を省略した項目は DEFAULT_STORE_CONFIG の値が使われる。
# =============================================================================

DEFAULT_STORE_CONFIG: dict = {
    # ジャグラーシリーズ: (機種名, 合算確率閾値, 差枚ボーナス閾値)
    "juggler_jobs": [
        ("マイジャグV",         140, 1000),
        ("ファンキー2",         145, 1000),
        ("ゴージャグ3",         135, 1000),
        ("アイム",              146, 1000),
        ("ネオアイム",          146, 1000),
        ("ハピジャグV",         142, 1000),
        ("ウルトラミラジャグ",   143, 1000),
        ("ジャグラーガールズ",   137, 1000),
        ("ミスジャグ",          139, 1000),
    ],
    # 全台系個別画像を作らず +1000枚台のみその他へ回す機種
    "manual_exclude": set(),
    # G数2000G以上 かつ RB閾値以上 かつ 差枚>=0 で優秀台に入れる機種
    "rb_threshold_machines": {
        "スマスロ北斗の拳", "東京リベンジャーズ", "炎炎ノ消防隊2", "北斗転生2",
        "モンキーターンV", "モンハンライズ", "防振り",
        "東京喰種", "カバネリ海門決戦",
    },
    "rb_min":        15,     # RB閾値
    "diff_bonus":    1000,   # その他優秀台の差枚下限
    "juggler_g_min": 2000,   # ジャグラーの最低G数
    # 合算確率基準でフィルターする非ジャグラー機種: (機種名, 合算確率閾値, 差枚ボーナス閾値)
    # 条件: (合算確率_num <= 閾値 AND diff >= 0) OR diff >= 差枚ボーナス
    # ※ この機種は全台プラスでも全台系（Step1）には入れない
    "prob_jobs": [
        ("Lハナビ",          152, 1000),
        ("新ハナビ",         150, 1000),
        ("スマートキンハナV", 165, 1000),
        ("アレックスB",      166, 1000),
        ("LBサンダーV",    152, 1000),
        ("LBシェイク",     160, 1000),
        ("クレアの秘宝伝",  152, 1000),
        ("ヱヴァ約束",     172, 1000),
        ("ディスクアップUR", 200, 1000),
        ("うみねこ2",      190, 1000),
        ("バーサスリヴァイズ", 150, 1000),
        ("ホウオウ天翔30", 160, 1000),
        ("キンハナ30",     157, 1000),
        ("キンハナV30",    165, 1000),
        ("ディスクアップ2", 170, 1000),
        ("かぐや様",        150, 1000),
    ],
    # 総台数9台以下: 勝率50%以上 かつ +1000枚台>=3 で高配分個別画像を生成し統合画像から除外
    "small_machine_rule": {"max_total": 9, "min_1k": 3},
}

# 店舗別オーバーライド（空 dict = デフォルト設定をそのまま使用）
# キーは STORES の店舗名と一致させること
STORE_CONFIG: dict[str, dict] = {
    "新宿歌舞伎町": {},
    "西武新宿":     {},
    "新大久保":     {},
    "高田馬場":     {},
    "上野本館":     {},
    "上野新館":     {},
    "渋谷新館":     {"poster_extra_exclude": {"北斗転生2", "カバネリ海門決戦"},
                   "min7_machines": {"スマスロ北斗の拳", "東京喰種", "北斗転生2", "カバネリ海門決戦"}},
    "赤坂見附":     {},
    "新小岩":       {},
    "溝の口本館":   {},
    "溝の口新館":   {},
    "稲毛":         {},
    "秋葉原":       {},
    # ── 新店舗を追加する場合は STORES とここに同時追記 ───────────────
    # "○○店": {
    #     "juggler_jobs": [...],   # ジャグラー機種が違う場合
    #     "manual_exclude": {...}, # 手動除外機種が違う場合
    # },
}


def get_store_config(store: str) -> dict:
    """店舗名からデフォルト設定に上書きを適用した設定 dict を返す。"""
    cfg = copy.deepcopy(DEFAULT_STORE_CONFIG)
    cfg.update(STORE_CONFIG.get(store, {}))
    cfg["juggler_series"] = {j[0] for j in cfg["juggler_jobs"]}
    cfg["prob_jobs_map"]  = {j[0]: (j[1], j[2]) for j in cfg["prob_jobs"]}
    return cfg


# 自動処理: 店舗 → 並びスクリプト（並び画像オプション用・subprocess）
_NARABI_GENERIC = os.path.join(BASE_DIR, "convert_narabi_pil.py")
STORE_NARABI_SCRIPT: dict[str, str] = {
    "新宿歌舞伎町": _NARABI_GENERIC,
    "西武新宿":     _NARABI_GENERIC,
    "新大久保":     _NARABI_GENERIC,
    "高田馬場":     _NARABI_GENERIC,
    "上野本館":     _NARABI_GENERIC,
    "上野新館":     _NARABI_GENERIC,
    "渋谷新館":     _NARABI_GENERIC,
    "赤坂見附":     _NARABI_GENERIC,
    "新小岩":       _NARABI_GENERIC,
    "溝の口本館":   _NARABI_GENERIC,
    "溝の口新館":   _NARABI_GENERIC,
    "稲毛":   _NARABI_GENERIC,
    "秋葉原": _NARABI_GENERIC,
}

# 列ごとの最低表示幅（px）― 旧スクリプトの CSS 幅と同値。scale=150/96 で乗算して実使用
MIN_COL_WIDTHS: dict[str, int] = {
    # CSS content幅 + padding(8px×2=16px) の合計 → ×scale で実ピクセルに
    # 明示幅なし列（台番・ゲーム数）は 0 にしてテキスト幅で自動決定
    "台番":    0,    # CSS 幅未指定 → 自動
    "機種名":  186,  # 170 + 16
    "ゲーム数": 0,   # CSS 幅未指定 → 自動
    "合算確率": 96,  # 80 + 16
    "差枚":    106,  # 90 + 16
    "差枚数":  106,  # 90 + 16
    "BB":      56,   # 40 + 16
    "RB":      56,
    "BIG":     56,
    "REG":     56,
    "AT":      56,
}
# 内部列名 → 旧スクリプト表示名への変換（draw_table_image に渡す直前に適用）
_DISPLAY_RENAME: dict[str, str] = {"BB": "BIG", "RB": "REG", "差枚": "差枚数"}

# =============================================================================
# ■ ②デザイン定数
# =============================================================================

C_TITLE_BG        = "#2F559E"   # タイトルバー背景（青）
C_TITLE_FG        = "#FFFFFF"   # タイトルバー文字
C_REDLINE         = "#CC0000"   # タイトル下の赤アクセントライン
C_HEADER_BG       = "#4472C4"   # 全台系ヘッダー背景（青）
C_HEADER_FG       = "#FFFFFF"   # 全台系ヘッダー文字
C_MACH_HEADER_BG  = "#f3e6c8"   # 機種別ヘッダー背景（クリーム）
C_MACH_HEADER_FG  = "#4B0082"   # 機種別ヘッダー文字（紫）
C_ROW_BG          = "#FFFFFF"   # データ行背景（全行統一）
C_BORDER          = "#AAAAAA"   # 罫線
C_PLUS            = "#0000CC"   # 差枚プラス色
C_MINUS           = "#CC0000"   # 差枚マイナス色
C_ZERO            = "#000000"   # 差枚ゼロ・通常文字
C_SUMMARY_BG_RGBA = (255, 182, 193, 255)  # ピンクバー背景（RGBA）

IMG_FONT_SZ     = 14
TITLE_FONT_SZ   = 34    # タイトルバーの文字サイズ（既存スクリプトの 36pt に近い大きさ）
SUMMARY_FONT_SZ = 14    # ピンクバーの文字サイズ
CELL_PAD        = 8     # セル内水平余白(px)
ROW_H           = 28    # font=14px + padding=4px×2 → ×1.5625 ≈ 44px
HEADER_H        = 28    # 同上
TITLE_H         = 75    # タイトルバーの高さ（文字サイズに合わせて拡大）
REDLINE_H       = 5     # タイトル下の赤ラインの高さ(px)
GAP_SUM         = -8    # ピンクバー内 % と（の間のカーニング調整(px)

# =============================================================================
# ■ ③フォントユーティリティ
# =============================================================================

_font_cache: dict[int, ImageFont.ImageFont] = {}

def load_font(size: int) -> ImageFont.ImageFont:
    """
    日本語フォントを読み込む。
    優先順: fonts/MochiyPopOne-Regular.ttf → fonts/NotoSansJP-Regular.ttf
            → Windows フォント（ローカル実行時フォールバック）
    fonts/ に何もなければ st.error でどのファイルが不足しているか表示する。
    """
    if size in _font_cache:
        return _font_cache[size]

    candidates = [
        # ① プロジェクト同梱フォント（Cloud・ローカル共通）
        (os.path.join(_FONTS_DIR, "MochiyPopOne-Regular.ttf"), None),
        (os.path.join(_FONTS_DIR, "NotoSansJP-Regular.ttf"),   None),
        # ② Windows ローカル実行時フォールバック
        (r"C:\Users\23-3\AppData\Local\Microsoft\Windows\Fonts\MochiyPopOne-Regular.ttf", None),
        (r"C:\Windows\Fonts\meiryo.ttc",  0),
        (r"C:\Windows\Fonts\msgothic.ttc", 1),
        (r"C:\Windows\Fonts\YuGoth-M.ttc", 0),
    ]
    for path, idx in candidates:
        if not os.path.exists(path):
            continue
        try:
            font = (ImageFont.truetype(path, size)
                    if idx is None
                    else ImageFont.truetype(path, size, index=idx))
            _font_cache[size] = font
            return font
        except Exception:
            continue

    # フォントが一切見つからない場合はエラーを表示して停止
    mochiy_path = os.path.join(_FONTS_DIR, "MochiyPopOne-Regular.ttf")
    noto_path   = os.path.join(_FONTS_DIR, "NotoSansJP-Regular.ttf")
    st.error(
        "【フォントエラー】日本語フォントが見つかりません。\n"
        f"以下のいずれかを `fonts/` フォルダに配置してください:\n"
        f"  ・{mochiy_path}\n"
        f"  ・{noto_path}"
    )
    st.stop()
    return ImageFont.load_default()  # unreachable but satisfies type checker

# =============================================================================
# ■ ④データユーティリティ
# =============================================================================

# ── 稲毛 差枚補正 ──────────────────────────────────────────────────────────────
_PIPELINE_KEEP_VALS: set[int] = {
    18800, 18500, 18400, 18300, 18000, 17800, 17500, 17400, 17700, 17000,
    16800, 16500, 16400, 16700, 16000, 15800, 15600, 15500, 15400, 15000,
    14800, 14500, 14400, 14300, 14000, 13800, 13500, 13400, 13200, 13000,
    12900, 11000, 10800, 10500, 10400, 10600, 10300, 10000,  9900,  9800,
     9500,  9400,  9300,  9000,  8800,  8600,  8500,  8400,  8200,  8000,
     7800,  7700,  7500,  7400,  7100,  7000,  6900,  6800,  6500,  6400,
     6200,  6000,  5800,  5500,  5400,  5700,  5000,  4800,  4500,  4400,
     4600,  4000,  3800,  3600,  3500,  3400,  3300,  3000,  2800,  2600,
     2500,  2200,  2000,  1800,  1500,  1400,  1200,  1100,
    -6500, -5500, -4500, -3500, -2500, -1500,  -500,
    -6000, -5800, -5000, -4600, -4000, -3400, -3100, -3000,
    -2300, -2200, -2000, -1600, -1200, -1000,  -800,  -300,
}


def _pipeline_calc_d(d_val) -> int:
    """差枚補正。keep_vals に含まれる値はそのまま、それ以外は範囲別加算後に50の倍数へ丸め。"""
    if pd.isna(d_val) or d_val == 0:
        return 0
    d = int(d_val)
    if d in _PIPELINE_KEEP_VALS:
        res = d
    elif d == 2400:
        res = d + 30
    else:
        if   101  <= d <=   999: add = 20
        elif 1000 <= d <=  1599: add = 30
        elif 1600 <= d <=  3500: add = 20
        elif 3501 <= d <=  4499: add = 30
        elif 5001 <= d <= 20000: add = 30
        elif -1999 <= d <=  -101: add = 30
        elif -3499 <= d <= -2000: add = 30
        elif -20000 <= d <= -3500: add = 25
        else: add = 20
        res = d + add
    sign = 1 if d > 0 else -1
    multiple = 50 * sign
    result = round(res / multiple) * multiple
    if result > 19000:
        result = 19000
    return result


def normalize_df(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    DataFrame の列名を COLUMN_ALIASES の正規名に変換する。
    戻り値: (変換済み DataFrame, 見つからなかった正規名リスト)
    """
    rename_map: dict[str, str] = {}
    for std_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in df.columns:
                rename_map[alias] = std_name
                break

    df = df.rename(columns=rename_map)

    # 数値列を強制変換（変換できない値は 0 にする）
    for col in ["台番", "差枚", "BB", "RB", "AT", "ゲーム数"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    missing = [c for c in COLUMN_ALIASES if c not in df.columns]
    return df, missing


def _read_csv_raw(data: bytes) -> pd.DataFrame:
    """CSV バイト列を読み込んで生DataFrameを返す（集計行除去・列名NFKC正規化）。"""
    df = pd.read_csv(io.BytesIO(data), encoding="cp932")
    # 台番号が数値でない集計行（総平均・総合計）を除去
    df = df[pd.to_numeric(df.iloc[:, 0], errors="coerce").notna()].copy()
    # ＢＢ回数→BB回数、ＲＢ回数→RB回数 など全角英数を半角に正規化
    df.columns = [unicodedata.normalize("NFKC", str(c)) for c in df.columns]
    # CSV の「差」列はアウト-セーフ（店側視点）なので符号を反転してプレイヤー視点に変換
    if "差" in df.columns:
        df["差"] = pd.to_numeric(df["差"], errors="coerce").fillna(0) * -1
    # 機種名列の半角カナを全角に正規化（juggler_jobs 等との比較を正しく行うため）
    kisha_col = next((c for c in df.columns if c in ("機種名", "機種名（正式名）", "機種名（データサイト表記）", "機種名（表記）", "機種")), None)
    if kisha_col:
        df[kisha_col] = df[kisha_col].astype(str).apply(lambda s: unicodedata.normalize("NFKC", s))
    # BB/RB回数列がない場合（確率列のみのCSV）は確率とゲーム数から逆算して合成
    if "BB回数" not in df.columns and "BB確率" in df.columns and "ゲーム数" in df.columns:
        _g   = pd.to_numeric(df["ゲーム数"], errors="coerce").fillna(0)
        _bbp = pd.to_numeric(df["BB確率"],  errors="coerce").replace(0, float("nan"))
        _rbp = pd.to_numeric(df["RB確率"],  errors="coerce").replace(0, float("nan"))
        df["BB回数"] = (_g / _bbp).round().fillna(0).astype(int)
        df["RB回数"] = (_g / _rbp).round().fillna(0).astype(int)
    return df


def _read_uploaded_df(uploaded) -> pd.DataFrame:
    """UploadedFile (xlsx / xls / csv) を読み込んで生DataFrameを返す。"""
    data = uploaded.getvalue()
    uploaded.seek(0)
    if uploaded.name.lower().endswith(".csv"):
        return _read_csv_raw(data)
    return pd.read_excel(io.BytesIO(data))


def _normalize_key(s: str) -> str:
    """機種名マッチング用の正規化（全角スペース・半角スペース・記号統一）"""
    s = unicodedata.normalize("NFKC", str(s))   # 全角英数→半角、全角スペース→半角スペース
    s = re.sub(r"[\s\u3000]+", "", s)            # 空白を全除去
    s = s.strip()
    return s

def build_name_map_from_df(conv_df: pd.DataFrame) -> tuple[dict, dict]:
    """変換マスタDataFrameから(完全一致マップ, 正規化マップ)を構築する"""
    name_map: dict[str, str] = {}
    for _, row in conv_df.iterrows():
        k = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        v = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
        if k and v and k != "nan" and v != "nan":
            name_map[k] = v
    name_map_norm = {_normalize_key(k): v for k, v in name_map.items()}
    return name_map, name_map_norm


@st.cache_data
def load_name_map() -> tuple[dict, dict]:
    """
    機種名変換テーブルを読み込んでキャッシュする。
    戻り値: (完全一致マップ, 正規化マップ)
    """
    if not os.path.exists(NAME_MAP_PATH):
        return {}, {}
    try:
        conv = pd.read_excel(NAME_MAP_PATH, header=1)
        return build_name_map_from_df(conv)
    except Exception:
        return {}, {}


def _apply_map(df: pd.DataFrame, name_map: dict, name_map_norm: dict) -> tuple[pd.DataFrame, int]:
    """
    name_map / name_map_norm を使って「機種名」列を変換する。
    戻り値: (変換済み DataFrame, 変換件数)
    """
    count = 0

    def convert(raw: str) -> str:
        nonlocal count
        raw = str(raw).strip()
        if raw in name_map:
            count += 1
            return name_map[raw]
        key = _normalize_key(raw)
        if key in name_map_norm:
            count += 1
            return name_map_norm[key]
        return raw

    df = df.copy()
    df["機種名"] = df["機種名"].apply(convert)
    return df, count


def apply_name_conversion(df: pd.DataFrame) -> pd.DataFrame:
    """
    機種名変換テーブルを使って機種名を正式名に変換する。
    完全一致 → 正規化一致 の順で検索し、どちらにも一致しなければ元の名前を保持する。
    """
    name_map, name_map_norm = load_name_map()
    if not name_map:
        return df
    df, _ = _apply_map(df, name_map, name_map_norm)
    return df


# =============================================================================
# ■ 機種画像（パネル/液晶）の紐づけ
#   簡略名 → 画像グループID（masters/machine_image_master.xlsx）
#   画像グループID → assets/machine_images/{id}_panel.png / {id}_01.png ...
# =============================================================================

@st.cache_data
def load_machine_image_master() -> dict:
    """
    簡略名→画像グループIDのマスタを読み込む。
    列: 「簡略名」「画像グループID」（列名優先・無ければ位置0,1でフォールバック）。
    戻り値: {簡略名: 画像グループID}
    """
    if not os.path.exists(_MACHINE_IMAGE_MASTER_PATH):
        return {}
    try:
        df = pd.read_excel(_MACHINE_IMAGE_MASTER_PATH)
    except Exception:
        return {}
    cols = list(df.columns)
    has_named = "簡略名" in cols and "画像グループID" in cols
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        if has_named:
            sn_raw, gid_raw = row["簡略名"], row["画像グループID"]
        else:
            sn_raw = row.iloc[0] if len(row) > 0 else None
            gid_raw = row.iloc[1] if len(row) > 1 else None
        sn = str(sn_raw).strip() if pd.notna(sn_raw) else ""
        gid = str(gid_raw).strip() if pd.notna(gid_raw) else ""
        if sn and gid and sn != "nan" and gid != "nan":
            mapping[sn] = gid
    return mapping


def _rel_machine_path(abs_path: str) -> str:
    """BASE_DIR からの相対パスを / 区切りで返す（辞書の値表記を統一）。"""
    return os.path.relpath(abs_path, BASE_DIR).replace(os.sep, "/")


def _find_panel_image(image_id: str) -> str | None:
    """{image_id}_panel.* を探して相対パスを返す。無ければ None。"""
    for ext in _MACHINE_IMG_EXTS:
        p = os.path.join(_MACHINE_IMAGES_DIR, f"{image_id}_panel{ext}")
        if os.path.exists(p):
            return _rel_machine_path(p)
    return None


def _find_screen_images(image_id: str) -> list[str]:
    """{image_id}_01.* / _02.* ... の液晶画像を番号順に返す。"""
    if not os.path.isdir(_MACHINE_IMAGES_DIR):
        return []
    pat = re.compile(rf"^{re.escape(image_id)}_(\d+)$")
    found: list[tuple[int, str]] = []
    for fn in os.listdir(_MACHINE_IMAGES_DIR):
        stem, ext = os.path.splitext(fn)
        if ext.lower() not in _MACHINE_IMG_EXTS:
            continue
        m = pat.match(stem)
        if m:
            found.append((int(m.group(1)), fn))
    found.sort(key=lambda x: x[0])
    return [_rel_machine_path(os.path.join(_MACHINE_IMAGES_DIR, fn)) for _, fn in found]


def get_machine_images(short_name: str) -> dict | None:
    """
    簡略名から画像グループIDを引き、パネル/液晶画像のパスを返す。
    戻り値: {"short_name", "image_id", "panel"(str|None), "screens"(list[str])}
            紐づけが無ければ None（エラーで止めない）。
    """
    master = load_machine_image_master()
    sn = (short_name or "").strip()
    if not sn:
        return None
    image_id = master.get(sn)
    if not image_id:
        # 正規化一致（スペース・全角除去）でも探す
        norm = {_normalize_key(k): v for k, v in master.items()}
        image_id = norm.get(_normalize_key(sn))
    if not image_id:
        return None
    return {
        "short_name": sn,
        "image_id": image_id,
        "panel": _find_panel_image(image_id),
        "screens": _find_screen_images(image_id),
    }


# ── 画像グループIDの自動予測（簡略名 → image_id の候補作成） ────────────────
# 比較前に除去する接頭ワード（画像グループIDはローマ字なので日本語のみ）
_MATCH_STRIP_WORDS = ("スマスロ", "パチスロ", "ぱちスロ", "ぱちんこ", "パチンコ", "新台")

_kakasi_instance = None


def _to_romaji(s: str) -> str:
    """
    日本語（漢字・かな）をヘボン式ローマ字に変換する。
    pykakasi が無ければ元文字列を返す（英字簡略名はそのまま比較される）。
    """
    global _kakasi_instance
    if _kakasi_instance is None:
        try:
            import pykakasi
            _kakasi_instance = pykakasi.kakasi()
        except Exception:
            _kakasi_instance = False
    if not _kakasi_instance:
        return str(s)
    try:
        return "".join(item["hepburn"] for item in _kakasi_instance.convert(str(s)))
    except Exception:
        return str(s)


def _normalize_for_match(s: str) -> str:
    """
    予測マッチ用の正規化。接頭ワード除去 → ローマ字化 → 小文字化 →
    英数以外を除去 → 先頭のL/P/Sマーカーを除去する。
    画像グループID（ローマ字）と日本語簡略名を同じ土俵で比較できるようにする。
    """
    import unicodedata
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    for w in _MATCH_STRIP_WORDS:
        t = t.replace(w, "")
    t = _to_romaji(t).lower()
    # 英数以外（記号・空白・変換漏れの記号）を除去
    t = re.sub(r"[^0-9a-z]", "", t)
    # 先頭の l / p / s マーカー（機種区分）を1文字だけ除外
    t = re.sub(r"^[lps]", "", t)
    return t


def list_image_group_ids() -> list[str]:
    """assets/machine_images/ 内の *_panel.* から画像グループID一覧を返す。"""
    if not os.path.isdir(_MACHINE_IMAGES_DIR):
        return []
    ids: set[str] = set()
    for fn in os.listdir(_MACHINE_IMAGES_DIR):
        stem, ext = os.path.splitext(fn)
        if ext.lower() not in _MACHINE_IMG_EXTS:
            continue
        if stem.endswith("_panel"):
            ids.add(stem[: -len("_panel")])
    return sorted(ids)


def predict_image_id(short_name: str, group_ids: list[str]) -> tuple[str | None, float, str]:
    """
    簡略名に対する画像グループIDの最有力候補を返す。
    戻り値: (画像グループID or None, 類似度スコア, 判定理由)
      1. 完全一致（正規化後） → score 1.0
      2. 部分一致（一方が他方を含む） → score 0.9
      3. 類似一致（difflib） → SequenceMatcher の ratio
    """
    from difflib import SequenceMatcher
    sn = _normalize_for_match(short_name)
    if not sn:
        return None, 0.0, ""
    best_id, best_score, best_reason = None, 0.0, ""
    for gid in group_ids:
        gn = _normalize_for_match(gid)
        if not gn:
            continue
        if sn == gn:
            score, reason = 1.0, "完全一致"
        elif sn in gn or gn in sn:
            # 短すぎる部分一致の暴発を抑える（1文字包含は類似度に回す）
            if min(len(sn), len(gn)) >= 2:
                score, reason = 0.9, "部分一致"
            else:
                score, reason = SequenceMatcher(None, sn, gn).ratio(), "類似"
        else:
            score, reason = SequenceMatcher(None, sn, gn).ratio(), "類似"
        if score > best_score:
            best_id, best_score, best_reason = gid, score, reason
    return best_id, round(best_score, 3), best_reason


def append_machine_image_master(rows: list[tuple[str, str]]) -> int:
    """
    (簡略名, 画像グループID) を machine_image_master.xlsx に追記する。
    既存の簡略名は重複登録しない。戻り値: 実際に追加した件数。
    """
    os.makedirs(os.path.dirname(_MACHINE_IMAGE_MASTER_PATH), exist_ok=True)
    if os.path.exists(_MACHINE_IMAGE_MASTER_PATH):
        try:
            df = pd.read_excel(_MACHINE_IMAGE_MASTER_PATH)
        except Exception:
            df = pd.DataFrame(columns=["簡略名", "画像グループID"])
    else:
        df = pd.DataFrame(columns=["簡略名", "画像グループID"])
    if "簡略名" not in df.columns or "画像グループID" not in df.columns:
        df = df.rename(columns={df.columns[0]: "簡略名", df.columns[1]: "画像グループID"})
    existing = {str(v).strip() for v in df["簡略名"].tolist()}
    new_rows = []
    for sn, gid in rows:
        sn, gid = str(sn).strip(), str(gid).strip()
        if sn and gid and sn not in existing:
            new_rows.append({"簡略名": sn, "画像グループID": gid})
            existing.add(sn)
    if not new_rows:
        return 0
    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    df.to_excel(_MACHINE_IMAGE_MASTER_PATH, index=False)
    load_machine_image_master.clear()  # キャッシュ破棄
    return len(new_rows)


def save_machine_image_mapping(short_name: str, gid: str, overwrite: bool = False) -> tuple[str, str]:
    """
    簡略名→画像グループIDを1件 追加/更新する。
    戻り値: (action, gid)
      action = "added"（新規追加） / "updated"（上書き更新） / "exists"（既存・overwrite=False）
    """
    os.makedirs(os.path.dirname(_MACHINE_IMAGE_MASTER_PATH), exist_ok=True)
    if os.path.exists(_MACHINE_IMAGE_MASTER_PATH):
        try:
            df = pd.read_excel(_MACHINE_IMAGE_MASTER_PATH)
        except Exception:
            df = pd.DataFrame(columns=["簡略名", "画像グループID"])
    else:
        df = pd.DataFrame(columns=["簡略名", "画像グループID"])
    if "簡略名" not in df.columns or "画像グループID" not in df.columns:
        df = df.rename(columns={df.columns[0]: "簡略名", df.columns[1]: "画像グループID"})
    sn, gid = str(short_name).strip(), str(gid).strip()
    mask = df["簡略名"].astype(str).str.strip() == sn
    if mask.any():
        if not overwrite:
            return "exists", str(df.loc[mask, "画像グループID"].iloc[0]).strip()
        df.loc[mask, "画像グループID"] = gid
        action = "updated"
    else:
        df = pd.concat([df, pd.DataFrame([{"簡略名": sn, "画像グループID": gid}])], ignore_index=True)
        action = "added"
    df.to_excel(_MACHINE_IMAGE_MASTER_PATH, index=False)
    load_machine_image_master.clear()  # キャッシュ破棄
    return action, gid


def _sync_machine_image_master() -> tuple[bool, str]:
    """
    machine_image_master.xlsx をGitHubへ反映する。
    ローカル: git add→commit→pull→push。Cloud: GitHub API でバイナリを直接PUT。
    """
    repo_path = "masters/machine_image_master.xlsx"
    if _IS_CLOUD:
        import urllib.request, urllib.error, base64 as _b64
        token = get_secret_value("GITHUB_TOKEN", "")
        if not token:
            return False, "GITHUB_TOKEN未設定（Cloudで同期するにはSecrets設定が必要）"
        repo = "tama0520/guild-image-app"
        url = f"https://api.github.com/repos/{repo}/contents/{repo_path}"
        hdrs = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
        }
        try:
            with open(_MACHINE_IMAGE_MASTER_PATH, "rb") as _f:
                data = _f.read()
        except Exception as e:
            return False, f"ファイル読込失敗: {e}"
        for _attempt in range(3):
            sha = None
            try:
                req = urllib.request.Request(url, headers=hdrs)
                with urllib.request.urlopen(req, timeout=10) as r:
                    sha = json.loads(r.read())["sha"]
            except Exception:
                sha = None  # 未存在なら新規作成
            payload = {
                "message": "update machine_image_master",
                "content": _b64.b64encode(data).decode("ascii"),
                "branch": "main",
            }
            if sha:
                payload["sha"] = sha
            body = json.dumps(payload).encode("utf-8")
            try:
                req2 = urllib.request.Request(url, data=body, headers=hdrs, method="PUT")
                with urllib.request.urlopen(req2, timeout=15) as r:
                    r.read()
                return True, "GitHubに同期しました"
            except urllib.error.HTTPError as e:
                if e.code == 409:
                    continue
                return False, f"同期失敗: {e}"
            except Exception as e:
                return False, f"同期失敗: {e}"
        return False, "同期失敗: SHA競合が解消されませんでした"
    else:
        import subprocess
        try:
            subprocess.run(["git", "add", repo_path], cwd=BASE_DIR, capture_output=True, check=True)
            diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=BASE_DIR, capture_output=True)
            if diff.returncode == 0:
                return True, "変更なし（push不要）"
            subprocess.run(["git", "commit", "-m", "update machine_image_master"],
                           cwd=BASE_DIR, capture_output=True, check=True)
            subprocess.run(["git", "pull", "--rebase", "origin", "main"],
                           cwd=BASE_DIR, capture_output=True, check=True)
            subprocess.run(["git", "push", "origin", "main"],
                           cwd=BASE_DIR, capture_output=True, check=True)
            return True, "GitHubにpushしました"
        except subprocess.CalledProcessError as e:
            return False, f"push失敗: {e.stderr.decode('utf-8', errors='replace').strip()}"


def round_games(v) -> int:
    """ゲーム数を 50G / 100G 単位に丸める（既存スクリプトと同じロジック）"""
    n = int(v); r = n % 100
    if r <= 24:   return (n // 100) * 100
    elif r <= 74: return (n // 100) * 100 + 50
    else:         return (n // 100 + 1) * 100


def fmt_diff(v) -> str:
    """差枚を '+X,XXX枚' 形式でフォーマット（マイナスはそのまま、ゼロは ±0枚）"""
    try:
        n = int(v)
        if n > 0:   return f"+{n:,}枚"
        elif n < 0: return f"{n:,}枚"
        else:       return "±0枚"
    except Exception:
        return str(v)


def fmt_games(v) -> str:
    """ゲーム数を 'X,XXXG' 形式でフォーマット"""
    try:
        return f"{int(v):,}G"
    except Exception:
        return str(v)


def fmt_prob(bb, rb, g) -> str:
    """
    合算確率を '1/XXX.X' 形式でフォーマット。
    BB+RB=0 または G数=0 の場合は '――' を返す。
    """
    try:
        total = int(bb) + int(rb)
        g_val = int(g)
        if total == 0 or g_val == 0:
            return "――"
        return f"1/{g_val / total:.1f}"
    except Exception:
        return "――"

# =============================================================================
# ■ ⑤テーブル画像描画（共通）
# =============================================================================

def _text_w(draw: ImageDraw.ImageDraw, text: str, font) -> int:
    """テキストの描画幅を返す"""
    bb = draw.textbbox((0, 0), text, font=font)
    return bb[2] - bb[0]


def _text_h(draw: ImageDraw.ImageDraw, text: str, font) -> int:
    """テキストの描画高さを返す"""
    bb = draw.textbbox((0, 0), text, font=font)
    return bb[3] - bb[1]


def _add_margin(img: Image.Image, pad: int = 20, color: str = "#CFEEEE") -> Image.Image:
    """表画像の外側四方に均等な余白を追加して返す。表自体は変更しない。"""
    new_w = img.width  + pad * 2
    new_h = img.height + pad * 2
    canvas = Image.new("RGB", (new_w, new_h), color)
    canvas.paste(img, (pad, pad))
    return canvas


def draw_table_image(
    headers: list[str],
    rows: list[list],
    diff_col_idx: int | None = None,
    title: str | None = None,
    summary_stat: dict | None = None,
    header_bg: str = C_HEADER_BG,
    header_fg: str = C_HEADER_FG,
    diff_cell_bg: bool = False,
    scale: float = 1.0,
) -> Image.Image:
    """
    PIL でテーブル画像を描画して返す。
    既存スクリプト（convert_稲毛_*.py 等）と同じ視覚デザインを再現する。

    Parameters
    ----------
    headers      : 列名リスト
    rows         : データ行リスト（各行は文字列リスト）
    diff_col_idx : 差枚列のインデックス（プラス/マイナス色分け用）
    title        : タイトルバー文字列（None でタイトルなし）
    summary_stat : ピンクバー用の集計値 dict
                   {"total_diff": int, "avg_diff": int,
                    "win_count": int, "total_count": int}
                   None の場合はピンクバーなし
    header_bg    : ヘッダー背景色（全台系=C_HEADER_BG / 機種別=C_MACH_HEADER_BG）
    header_fg    : ヘッダー文字色
    scale        : 150/96 を渡すと playwright 150DPI 相当の寸法になる
    """
    _font_sz    = round(IMG_FONT_SZ    * scale)
    _sum_sz     = round(SUMMARY_FONT_SZ * scale)
    _row_h      = round(ROW_H          * scale)
    _header_h   = round(HEADER_H       * scale)
    _cell_pad   = round(CELL_PAD       * scale)
    _min_col_w  = {k: round(v * scale) for k, v in MIN_COL_WIDTHS.items()}

    fn_data    = load_font(_font_sz)
    fn_header  = load_font(_font_sz)
    fn_title   = load_font(TITLE_FONT_SZ)
    fn_summary = load_font(_sum_sz)

    if not rows:
        raise ValueError("出力するデータが 0 件です。")

    # ── 列幅計算 ────────────────────────────────────────────────────
    dummy = Image.new("RGB", (1, 1))
    d0    = ImageDraw.Draw(dummy)

    col_w: list[int] = []
    for ci, h in enumerate(headers):
        w = _text_w(d0, str(h), fn_header) + _cell_pad * 2
        for row in rows:
            if ci < len(row):
                w = max(w, _text_w(d0, str(row[ci]), fn_data) + _cell_pad * 2)
        min_w = _min_col_w.get(str(h), round(30 * scale))
        col_w.append(max(w, min_w))

    # BIG と REG を同幅に揃える（旧スクリプトと同じ見た目）
    big_idx = headers.index("BIG") if "BIG" in headers else None
    reg_idx = headers.index("REG") if "REG" in headers else None
    if big_idx is not None and reg_idx is not None:
        shared = max(col_w[big_idx], col_w[reg_idx])
        col_w[big_idx] = col_w[reg_idx] = shared

    total_w = sum(col_w)

    # ── 各パーツの高さを計算 ─────────────────────────────────────────
    title_h   = TITLE_H + REDLINE_H if title else 0
    table_h   = _header_h + len(rows) * _row_h
    pink_h    = int(_row_h * 1.3) if summary_stat else 0
    total_h   = title_h + table_h + pink_h

    img  = Image.new("RGBA", (total_w, total_h), (255, 255, 255, 255))
    draw = ImageDraw.Draw(img)

    # ── タイトルバー ─────────────────────────────────────────────────
    y = 0
    if title:
        # 青タイトルバー
        draw.rectangle([(0, 0), (total_w - 1, TITLE_H - 1)], fill=C_TITLE_BG)
        tb = draw.textbbox((0, 0), title, font=fn_title)
        tx = (total_w - (tb[2] - tb[0])) // 2 - tb[0]
        ty = (TITLE_H  - (tb[3] - tb[1])) // 2 - tb[1]
        draw.text((tx, ty), title, fill=C_TITLE_FG, font=fn_title)
        y += TITLE_H
        # 赤アクセントライン（既存スクリプトの red_line と同じ）
        draw.rectangle([(0, y), (total_w - 1, y + REDLINE_H - 1)], fill=C_REDLINE)
        y += REDLINE_H

    # ── ヘッダー行 ───────────────────────────────────────────────────
    x = 0
    for ci, h in enumerate(headers):
        draw.rectangle(
            [(x, y), (x + col_w[ci] - 1, y + _header_h - 1)],
            fill=header_bg, outline=C_BORDER,
        )
        tb = draw.textbbox((0, 0), str(h), font=fn_header)
        tx = x + (col_w[ci] - (tb[2] - tb[0])) // 2 - tb[0]
        ty = y + (_header_h  - (tb[3] - tb[1])) // 2 - tb[1]
        draw.text((tx, ty), str(h), fill=header_fg, font=fn_header)
        x += col_w[ci]
    y += _header_h

    # ── データ行 ─────────────────────────────────────────────────────
    for ri, row in enumerate(rows):
        bg = C_ROW_BG
        x  = 0
        for ci in range(len(headers)):
            cell = str(row[ci]) if ci < len(row) else ""
            draw.rectangle(
                [(x, y), (x + col_w[ci] - 1, y + _row_h - 1)],
                fill=bg, outline=C_BORDER,
            )
            tb     = draw.textbbox((0, 0), cell, font=fn_data)
            tw, th = tb[2] - tb[0], tb[3] - tb[1]
            ty_c   = y + (_row_h - th) // 2 - tb[1]

            if diff_col_idx is not None and ci == diff_col_idx:
                # 差枚列のみ右寄せ・プラス/マイナスで色分け
                raw = (cell.replace("+", "").replace("枚", "")
                            .replace(",", "").replace("±0", "0"))
                try:
                    v = int(raw)
                    color = C_PLUS if v > 0 else (C_MINUS if v < 0 else C_ZERO)
                    if diff_cell_bg and v != 0:
                        cell_fill = "#DBEAFE" if v > 0 else "#FCE4EC"
                        draw.rectangle(
                            [(x + 1, y + 1), (x + col_w[ci] - 2, y + _row_h - 2)],
                            fill=cell_fill,
                        )
                except Exception:
                    color = C_ZERO
                tx = x + col_w[ci] - tw - _cell_pad - tb[0]
                draw.text((tx, ty_c), cell, fill=color, font=fn_data)

            else:
                # 差枚以外はすべてセンタリング
                tx = x + (col_w[ci] - tw) // 2 - tb[0]
                draw.text((tx, ty_c), cell, fill=C_ZERO, font=fn_data)

            x += col_w[ci]
        y += _row_h

    # ── ピンクサマリーバー（GAP_SUM=-8 で %と（を詰める）────────────
    if summary_stat:
        s         = summary_stat
        win_rate  = s["win_count"] / s["total_count"] * 100 if s["total_count"] > 0 else 0.0
        # %まで含む前半と （X/X台） の後半に分けて描画（GAP_SUM で字間調整）
        part1 = (f"総差枚：{fmt_diff(s['total_diff'])}"
                 f"　平均：{fmt_diff(s['avg_diff'])}"
                 f"　勝率：{win_rate:.1f}%")
        part2 = f"（{s['win_count']}/{s['total_count']}台）"

        draw.rectangle(
            [(0, y), (total_w - 1, y + pink_h - 1)],
            fill=C_SUMMARY_BG_RGBA[:3], outline=C_BORDER,
        )
        bb1  = draw.textbbox((0, 0), part1, font=fn_summary)
        th_s = bb1[3] - bb1[1]
        ty_s = y + (pink_h - th_s) // 2 - bb1[1]
        draw.text((_cell_pad - bb1[0], ty_s), part1, fill=C_ZERO, font=fn_summary)

        # GAP_SUM=-8 で % と（ の間を詰める
        x2   = _cell_pad + (bb1[2] - bb1[0]) + GAP_SUM
        bb2  = draw.textbbox((0, 0), part2, font=fn_summary)
        draw.text((x2 - bb2[0], ty_s), part2, fill=C_ZERO, font=fn_summary)

    return img.convert("RGB")

# =============================================================================
# ■ ⑥画像生成ハンドラー
#   各関数の引数: (df: 正規化済み DataFrame, conditions: 条件 dict)
#   戻り値: PIL Image
# =============================================================================

def _active_sub_cols(df: pd.DataFrame) -> list[str]:
    """BB / RB / AT のうち全行が 0 でない列だけ返す"""
    return [c for c in ["BB", "RB", "AT"]
            if c in df.columns and not (df[c] == 0).all()]


def _add_prob_col(df: pd.DataFrame) -> pd.DataFrame:
    """BB・RB・ゲーム数 が存在する場合に合算確率列を追加したコピーを返す"""
    if all(c in df.columns for c in ["BB", "RB", "ゲーム数"]):
        df = df.copy()
        df["合算確率"] = df.apply(
            lambda r: fmt_prob(r["BB"], r["RB"], r["ゲーム数"]), axis=1
        )
    return df


def _format_display_cols(df: pd.DataFrame) -> pd.DataFrame:
    """ゲーム数（丸め処理あり）・差枚を表示用文字列に変換したコピーを返す"""
    df = df.copy()
    if "ゲーム数" in df.columns:
        df["ゲーム数"] = df["ゲーム数"].apply(lambda v: fmt_games(round_games(v)))
    if "差枚" in df.columns:
        df["差枚"] = df["差枚"].apply(fmt_diff)
    return df


def generate_全台データ画像(df: pd.DataFrame, conditions: dict) -> Image.Image:
    """
    【全台データ画像】
    キーワードで機種名を部分一致検索し、一致した台を一覧表示する。
    機種別JPGと同デザイン（クリームヘッダー＋青タイトルバー＋ピンクサマリーバー）。
    """
    keyword = conditions.get("keyword", "").strip()
    title   = conditions.get("title", "").strip()

    # キーワードフィルター（空なら全台）
    if keyword:
        mask = df["機種名"].astype(str).str.contains(keyword, na=False)
        df   = df[mask].copy()
        if df.empty:
            raise ValueError(f"キーワード「{keyword}」に一致する台がありません。")

    suffix    = conditions.get("_title_suffix", "")
    img_title = (f"{keyword}{suffix}" if keyword else (title if title else "全台データ"))

    # ピンクバー用の集計（フォーマット前の数値で計算）
    stat = {
        "total_diff":  int(df["差枚"].sum()),
        "avg_diff":    int(round(df["差枚"].mean())),
        "win_count":   int((df["差枚"] > 0).sum()),
        "total_count": len(df),
    }

    return _build_machine_img(df, img_title, stat)



def generate_高配分データ画像(df: pd.DataFrame, conditions: dict) -> Image.Image:
    """【高配分データ画像】プラス台のみ・ピンクバーなし・タイトルに(優秀台)を付加。"""
    keyword = conditions.get("keyword", "").strip()
    title   = conditions.get("title", "").strip()

    if keyword:
        mask = df["機種名"].astype(str).str.contains(keyword, na=False)
        df   = df[mask].copy()
        if df.empty:
            raise ValueError(f"キーワード「{keyword}」に一致する台がありません。")

    df = df[df["差枚"] > 0].copy()
    if df.empty:
        raise ValueError("プラスの台がありません。")

    img_title = (f"{keyword}(優秀台)" if keyword else (title if title else "高配分データ"))
    return _build_machine_img(df, img_title, summary_stat=None)


def generate_優秀台ピックアップ(df: pd.DataFrame, conditions: dict) -> Image.Image:
    """
    【優秀台ピックアップ】
    差枚下限 + G数下限 の両方を満たす台を抽出して表示する。
    タイトルは「（優秀台）」形式・ピンクバーなし（convert_稲毛_*.py の優秀台画像スタイルに対応）。
    """
    diff_min  = int(conditions.get("diff_min", 3000))
    games_min = int(conditions.get("games_min", 5000))
    title     = conditions.get("title", "優秀台ピックアップ").strip() or "優秀台ピックアップ"

    result = df[(df["差枚"] >= diff_min) & (df["ゲーム数"] >= games_min)].copy()
    if result.empty:
        raise ValueError(
            f"差枚 ≥ {diff_min:,}枚 かつ ゲーム数 ≥ {games_min:,}G に一致する台がありません。"
        )

    result   = _add_prob_col(result)
    active   = _active_sub_cols(result)
    disp     = ["台番", "機種名", "ゲーム数"] + active
    if "合算確率" in result.columns:
        disp.append("合算確率")
    disp.append("差枚")
    disp     = [c for c in disp if c in result.columns]
    result   = _format_display_cols(result[disp])
    headers  = list(result.columns)
    rows     = result.values.tolist()
    diff_idx = headers.index("差枚") if "差枚" in headers else None

    # 優秀台画像: ピンクバーなし・機種別ヘッダー色（convert_稲毛_*.py の all_plus=False スタイル）
    return draw_table_image(
        headers, rows,
        diff_col_idx=diff_idx,
        title=title,
        summary_stat=None,
        header_bg=C_MACH_HEADER_BG,
        header_fg=C_MACH_HEADER_FG,
        scale=150/96,
    )


def generate_機種別まとめ(df: pd.DataFrame, conditions: dict) -> Image.Image:
    """
    【機種別まとめ】
    機種ごとに台数・総差枚・平均差枚・勝率・合算確率を集計して表示する。
    """
    min_count = int(conditions.get("min_count", 1))
    title     = conditions.get("title", "機種別まとめ").strip() or "機種別まとめ"

    records: list[list] = []
    for machine, grp in df.groupby("機種名", sort=False):
        if len(grp) < min_count:
            continue
        total      = len(grp)
        total_diff = int(grp["差枚"].sum())
        avg_diff   = int(round(grp["差枚"].mean()))
        win_cnt    = int((grp["差枚"] > 0).sum())
        win_rate   = f"{win_cnt / total * 100:.1f}%"
        prob       = fmt_prob(grp["BB"].sum(), grp["RB"].sum(), grp["ゲーム数"].sum()) \
                     if all(c in grp.columns for c in ["BB", "RB", "ゲーム数"]) else "――"
        records.append([
            str(machine),
            f"{total}台",
            fmt_diff(total_diff),
            fmt_diff(avg_diff),
            win_rate,
            f"{win_cnt}/{total}",
            prob,
        ])

    if not records:
        raise ValueError("集計できる機種がありません。")

    headers = ["機種名", "台数", "総差枚", "平均差枚", "勝率", "勝/全", "合算確率"]
    # 機種別まとめ: 全台系ヘッダー色・ピンクバーなし
    return draw_table_image(
        headers, records,
        diff_col_idx=None,
        title=title,
        summary_stat=None,
        header_bg=C_HEADER_BG,
        header_fg=C_HEADER_FG,
        scale=150/96,
    )


def generate_その他の優秀台(df: pd.DataFrame, conditions: dict) -> Image.Image:
    """
    【その他の優秀台】
    指定した台番の台を抽出して表示する。
    """
    bans_str = conditions.get("bans", "").strip()
    title    = conditions.get("title", "その他の優秀台").strip() or "その他の優秀台"

    if not bans_str:
        raise ValueError("台番を入力してください。")

    bans = set()
    for s in bans_str.split(","):
        s = s.strip()
        if s.isdigit():
            bans.add(int(s))

    if not bans:
        raise ValueError("有効な台番がありません。")

    result = df[df["台番"].astype(int).isin(bans)].copy()
    if result.empty:
        raise ValueError("指定した台番が見つかりません。")

    result   = _add_prob_col(result)
    active   = _active_sub_cols(result)
    disp     = ["台番", "機種名", "ゲーム数"] + active
    if "合算確率" in result.columns:
        disp.append("合算確率")
    disp.append("差枚")
    disp     = [c for c in disp if c in result.columns]
    result   = _format_display_cols(result[disp])
    headers  = list(result.columns)
    rows     = result.values.tolist()
    diff_idx = headers.index("差枚") if "差枚" in headers else None

    # その他の優秀台: 機種別ヘッダー色・ピンクバーなし（convert_1000plus_test.py と同スタイル）
    return draw_table_image(
        headers, rows,
        diff_col_idx=diff_idx,
        title=title,
        summary_stat=None,
        header_bg=C_MACH_HEADER_BG,
        header_fg=C_MACH_HEADER_FG,
        scale=150/96,
    )


# =============================================================================
# ■ ⑥-R ローテ画像生成
# =============================================================================

def _rote_diff_color(v: int) -> tuple[str, str]:
    """差枚値に対応する (背景色, 文字色) を返す。"""
    if v >= 10000:
        return ("rainbow", "white")
    if v >= 5000:
        return ("#FF4343", "black")   # aaa.jpg 実測値
    if v >= 3000:
        return ("#FFC000", "black")   # aaa.jpg 実測値
    return ("#FFFF00", "black")       # aaa.jpg 実測値


def generate_rote_image(df: pd.DataFrame, machine_names: list[str], date_label: str = "", store: str = "") -> Image.Image:
    """ローテ用機種別差枚一覧画像を生成する（aaa.jpg 参照）。"""
    import datetime, io as _io

    # ── 定数 ───────────────────────────────────────────────────────────────
    SC = 2
    ROW_H    = 26 * SC   # データ行高
    MAC_H    = 32 * SC   # 機種ヘッダー行高
    LEG_H    = 26 * SC   # 凡例行高（データ行と統一）
    COL_HDR_H = 34 * SC  # 台番・日付列ヘッダー行高（少し広め）
    GAP_H    = 10 * SC   # 凡例↔列ヘッダー間の余白
    COL_BAN  = 80 * SC
    COL_DIFF = 130 * SC  # 枚数列を少し狭く（-13%）
    W = COL_BAN + COL_DIFF

    C_BAN_BG  = ROTE_BAN_COLOR_CONFIG.get(store, "#00FFCC")  # 台番列
    C_HDR_BG  = "#606060"   # 列ヘッダー背景 濃いグレー
    C_HDR_FG  = "#FFFFFF"   # 列ヘッダー文字 白
    C_BORDER  = "#000000"
    C_EMPTY   = "#FFFFFF"
    C_MAC_BG  = "#000000"
    C_MAC_FG  = "#FFFFFF"

    FONT_SZ = int(15 * SC)  # 約115%（一段階大きく）
    font = load_font(FONT_SZ)

    # ── 日付ラベル ─────────────────────────────────────────────────────────
    if not date_label:
        today  = datetime.date.today()
        dow    = ["月", "火", "水", "木", "金", "土", "日"][today.weekday()]
        date_label = f"{today.month}/{today.day}({dow})"

    # ── 機種データ収集 ─────────────────────────────────────────────────────
    name_col = "機種名" if "機種名" in df.columns else (
                "機種名（正式名）" if "機種名（正式名）" in df.columns else None)
    machines_data: list[tuple[str, "pd.DataFrame"]] = []
    for raw in machine_names:
        kw = (raw or "").strip()
        if not kw or name_col is None:
            continue
        mask = df[name_col].astype(str).str.contains(kw, na=False)
        sub  = df[mask].copy().sort_values("台番") if mask.any() else None
        if sub is not None and not sub.empty:
            machines_data.append((kw, sub))

    # ── 高さ計算 ───────────────────────────────────────────────────────────
    total_h = LEG_H * 4 + GAP_H + COL_HDR_H   # 凡例4行 + 余白 + 列ヘッダー
    for _, sub in machines_data:
        total_h += MAC_H + ROW_H * len(sub)

    img = Image.new("RGB", (W, total_h), "white")
    d   = ImageDraw.Draw(img)

    # ── ヘルパー ───────────────────────────────────────────────────────────
    def cell(x, y, w, h, bg, text="", fg="black", align="center"):
        d.rectangle([x, y, x + w - 1, y + h - 1], fill=bg, outline=C_BORDER)
        if text:
            bb = d.textbbox((0, 0), text, font=font)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
            ty = y + (h - th) // 2 - bb[1]  # bb[1] 補正で真の上下センター
            if align == "center":
                tx = x + (w - tw) // 2
            elif align == "right":
                tx = x + w - tw - int(8 * SC)
            else:
                tx = x + int(8 * SC)
            d.text((tx, ty), text, fill=fg, font=font)

    def rainbow_cell(x, y, w, h, text=""):
        # aaa.jpg 実測に合わせた虹色（緑:#82F78C、赤:#FF4343 基準）
        stops = [
            (255, 67, 67),   # 赤
            (255, 192, 0),   # オレンジ
            (255, 255, 0),   # 黄
            (130, 247, 140), # 緑（#82F78C）
            (0,  100, 255),  # 青
            (180,  0, 255),  # 紫
        ]
        n = len(stops) - 1
        for px in range(w):
            t  = px / max(w - 1, 1) * n
            i  = min(int(t), n - 1)
            f  = t - i
            c1, c2 = stops[i], stops[i + 1]
            rgb = tuple(int(c1[k] + f * (c2[k] - c1[k])) for k in range(3))
            d.line([(x + px, y + 1), (x + px, y + h - 2)], fill=rgb)
        d.rectangle([x, y, x + w - 1, y + h - 1], outline=C_BORDER)
        if text:
            bb = d.textbbox((0, 0), text, font=font)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
            tx = x + (w - tw) // 2
            ty = y + (h - th) // 2 - bb[1]  # bb[1] 補正で真の上下センター
            d.text((tx, ty), text, fill="black", font=font)

    # ── 凡例 ───────────────────────────────────────────────────────────────
    cy = 0
    legends = [
        ("#FFFF00", "black",  "1,000枚～"),
        ("#FFC000", "black",  "3,000枚～"),
        ("#FF4343", "black",  "5,000枚～"),
        (None,      "black",  "万枚オーバー"),
    ]
    for (bg, fg, label) in legends:
        if bg is None:
            rainbow_cell(0, cy, COL_BAN, LEG_H)
        else:
            cell(0, cy, COL_BAN, LEG_H, bg)
        cell(COL_BAN, cy, COL_DIFF, LEG_H, "#FFFFFF", label, "black", align="center")
        cy += LEG_H

    # 凡例↔列ヘッダー間の余白（外枠マージンと同色）
    d.rectangle([0, cy, W - 1, cy + GAP_H - 1], fill="#CFEEEE")
    cy += GAP_H

    # ── 列ヘッダー ─────────────────────────────────────────────────────────
    cell(0,       cy, COL_BAN,  COL_HDR_H, C_HDR_BG, "台番",     C_HDR_FG)
    cell(COL_BAN, cy, COL_DIFF, COL_HDR_H, C_HDR_BG, date_label, C_HDR_FG)
    cy += COL_HDR_H

    # ── 機種別データ行 ─────────────────────────────────────────────────────
    for (mac_name, sub) in machines_data:
        # 機種名ヘッダー（黒背景・白文字・全幅）
        cell(0, cy, W, MAC_H, C_MAC_BG, mac_name, C_MAC_FG)
        cy += MAC_H

        for _, row in sub.iterrows():
            ban_str = str(int(row["台番"]))
            diff_v  = row.get("差枚", None)

            cell(0, cy, COL_BAN, ROW_H, C_BAN_BG, ban_str, "black")

            if pd.isna(diff_v) or int(diff_v) < 1000:
                cell(COL_BAN, cy, COL_DIFF, ROW_H, C_EMPTY)
            else:
                v        = int(diff_v)
                diff_str = f"+{v:,}"
                bg, fg   = _rote_diff_color(v)
                if bg == "rainbow":
                    rainbow_cell(COL_BAN, cy, COL_DIFF, ROW_H, diff_str)
                else:
                    cell(COL_BAN, cy, COL_DIFF, ROW_H, bg, diff_str, fg, align="center")

            cy += ROW_H

    return img


def generate_ranking_image(df: pd.DataFrame, machine_inputs: list[str], date_label: str = "", store: str = ""):
    """ランキング画像を生成する。全機種の差枚+1000枚以上を降順で一覧表示。"""
    import datetime as _dt

    SC = 2
    ROW_H     = 26 * SC
    COL_HDR_H = 34 * SC
    COL_BAN   = 80 * SC
    COL_DIFF  = 130 * SC
    W = COL_BAN + COL_DIFF

    C_HDR_BG = "#333333"
    C_HDR_FG = "#FFFFFF"
    C_BORDER = "#000000"
    C_BAN_BG = ROTE_BAN_COLOR_CONFIG.get(store, "#00FFCC")

    FONT_SZ = int(15 * SC)
    font = load_font(FONT_SZ)

    if not date_label:
        today = _dt.date.today()
        dow = ["月", "火", "水", "木", "金", "土", "日"][today.weekday()]
        date_label = f"{today.month}/{today.day}({dow})"

    name_col = "機種名" if "機種名" in df.columns else (
               "機種名（正式名）" if "機種名（正式名）" in df.columns else None)
    rows = []
    seen_bans: set[int] = set()
    for raw in machine_inputs:
        kw = (raw or "").strip()
        if not kw or name_col is None:
            continue
        mask = df[name_col].astype(str).str.contains(kw, na=False)
        sub = df[mask].copy()
        if sub.empty:
            continue
        for _, row in sub.iterrows():
            diff_v = pd.to_numeric(row.get("差枚", None), errors="coerce")
            if pd.isna(diff_v) or int(diff_v) < 1000:
                continue
            ban = int(row["台番"])
            if ban in seen_bans:
                continue
            seen_bans.add(ban)
            rows.append({"台番": ban, "差枚": int(diff_v)})

    if not rows:
        return None

    rows.sort(key=lambda r: (-r["差枚"], r["台番"]))

    total_h = COL_HDR_H + ROW_H * len(rows)
    img = Image.new("RGB", (W, total_h), "white")
    d   = ImageDraw.Draw(img)

    def cell(x, y, w, h, bg, text="", fg="black", align="center"):
        d.rectangle([x, y, x + w - 1, y + h - 1], fill=bg, outline=C_BORDER)
        if text:
            bb = d.textbbox((0, 0), text, font=font)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
            ty = y + (h - th) // 2 - bb[1]
            if align == "center":
                tx = x + (w - tw) // 2
            elif align == "right":
                tx = x + w - tw - int(8 * SC)
            else:
                tx = x + int(8 * SC)
            d.text((tx, ty), text, fill=fg, font=font)

    stops = [
        (255, 67, 67), (255, 192, 0), (255, 255, 0),
        (130, 247, 140), (0, 100, 255), (180, 0, 255),
    ]

    def rainbow_cell(x, y, w, h, text=""):
        n = len(stops) - 1
        for px in range(w):
            t  = px / max(w - 1, 1) * n
            i  = min(int(t), n - 1)
            f  = t - i
            c1, c2 = stops[i], stops[i + 1]
            rgb = tuple(int(c1[k] + f * (c2[k] - c1[k])) for k in range(3))
            d.line([(x + px, y + 1), (x + px, y + h - 2)], fill=rgb)
        d.rectangle([x, y, x + w - 1, y + h - 1], outline=C_BORDER)
        if text:
            bb = d.textbbox((0, 0), text, font=font)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
            tx = x + (w - tw) // 2
            ty = y + (h - th) // 2 - bb[1]
            d.text((tx, ty), text, fill="black", font=font)

    cy = 0
    cell(0, cy, COL_BAN, COL_HDR_H, C_HDR_BG, "台番", C_HDR_FG)
    cell(COL_BAN, cy, COL_DIFF, COL_HDR_H, C_HDR_BG, date_label, C_HDR_FG)
    cy += COL_HDR_H

    for r in rows:
        ban_str  = str(r["台番"])
        v        = r["差枚"]
        diff_str = f"+{v:,}"
        bg, fg   = _rote_diff_color(v)
        cell(0, cy, COL_BAN, ROW_H, C_BAN_BG, ban_str, "black")
        if bg == "rainbow":
            rainbow_cell(COL_BAN, cy, COL_DIFF, ROW_H, diff_str)
        else:
            cell(COL_BAN, cy, COL_DIFF, ROW_H, bg, diff_str, fg, align="center")
        cy += ROW_H

    # 外周を内側と同じ太さ（2px）で上書き
    d.rectangle([0, 0, W - 1, total_h - 1], outline=C_BORDER, width=2)

    return img


# =============================================================================
# ■ ⑦画像種類 → ハンドラー関数マッピング
#   新しい画像種類を追加するときはここにも追加してください
# =============================================================================

IMAGE_HANDLERS: dict[str, callable] = {
    "全台データ画像":    generate_全台データ画像,
    "高配分データ画像":  generate_高配分データ画像,
    "優秀台ピックアップ": generate_優秀台ピックアップ,
    "機種別まとめ":      generate_機種別まとめ,
    "その他の優秀台画像":    generate_その他の優秀台,
}

# =============================================================================
# ■ ⑧Streamlit ページ
# =============================================================================

def _navigate(page: str, store: str | None = None, itype: str | None = None) -> None:
    """ページ遷移: session_state と URL query_params を同時に更新してブラウザ履歴に積む"""
    st.session_state.page = page
    if store is not None:
        st.session_state.selected_store = store
    if itype is not None:
        st.session_state.selected_image_type = itype

    params: dict[str, str] = {"page": page}
    s = st.session_state.get("selected_store", "")
    t = st.session_state.get("selected_image_type", "")
    if page in ("image_type", "work", "auto", "auto_slump", "rote") and s:
        params["store"] = s
    if page == "work" and t:
        params["type"] = t
    st.query_params.from_dict(params)
    st.rerun()


def _sync_from_query_params() -> None:
    """ブラウザの戻る/進むボタン対応: URL の query_params → session_state を同期する"""
    qp = st.query_params
    qp_page = qp.get("page", "")
    if qp_page and qp_page != st.session_state.get("page"):
        st.session_state.page = qp_page
        qp_store = qp.get("store", "")
        if qp_store:
            st.session_state.selected_store = qp_store
        qp_type = qp.get("type", "")
        if qp_type:
            st.session_state.selected_image_type = qp_type

def img_to_bytes(img: Image.Image) -> bytes:
    """PIL Image を PNG バイト列に変換する（ダウンロード用）"""
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def show_store_page() -> None:
    """画面1: 店舗選択"""
    st.markdown("---")

    # 4列固定で並べる
    store_list = list(STORES.keys())
    cols       = st.columns(4)

    for i, store in enumerate(store_list):
        with cols[i % len(cols)]:
            if st.button(store, key=f"store_{store}", use_container_width=True):
                _navigate("image_type", store=store)


def show_image_type_page() -> None:
    """画面2: 画像種類選択"""
    store = st.session_state.selected_store

    st.markdown(f"## 【{store}】")
    st.markdown("---")

    # ── 一括自動処理ボタン群 ─────────────────────────────────────────
    # 高田馬場は4ボタン全てを同じ列幅に統一するため専用ブロックで描画
    if store == "高田馬場":
        st.markdown(
            """<style>
            .st-key-rote_mode_btn button {
                background-color: #1976D2 !important;
                border-color: #1565C0 !important;
                color: white !important;
            }
            .st-key-rote_mode_btn button:hover {
                background-color: #1565C0 !important;
                border-color: #1565C0 !important;
            }
            .st-key-weekly_result_text_btn button {
                background-color: #388E3C !important;
                border-color: #2E7D32 !important;
                color: white !important;
            }
            .st-key-weekly_result_text_btn button:hover {
                background-color: #2E7D32 !important;
                border-color: #2E7D32 !important;
            }
            .st-key-auto_article_btn button {
                background-color: #7B1FA2 !important;
                border-color: #6A1B9A !important;
                color: white !important;
            }
            .st-key-auto_article_btn button:hover {
                background-color: #6A1B9A !important;
                border-color: #6A1B9A !important;
            }
            </style>""",
            unsafe_allow_html=True,
        )
        _top_l, _top_r = st.columns(2)
        with _top_l:
            if st.button(
                "⚡ 結果ポスト用",
                key="auto_mode_btn",
                type="primary",
                use_container_width=True,
            ):
                _navigate("auto")
        with _top_r:
            if st.button(
                "📰 記事用",
                key="auto_article_btn",
                use_container_width=True,
            ):
                _navigate("auto_article")
        _bot_l, _bot_r = st.columns(2)
        with _bot_l:
            if st.button(
                "📋 ローテ用",
                key="rote_mode_btn",
                use_container_width=True,
            ):
                _navigate("rote")
        with _bot_r:
            if st.button(
                "📅 1週間分の結果テキスト",
                key="weekly_result_text_btn",
                use_container_width=True,
            ):
                _navigate("weekly_result_text")
    else:
        # 高田馬場以外
        if store == "上野本館":
            # 上野本館：結果ポスト用 ＋ スランプ付き結果ポスト ＋ ローテ用
            st.markdown(
                """<style>
                .st-key-auto_slump_btn button {
                    background-color: #00ACC1 !important;
                    border-color: #00838F !important;
                    color: white !important;
                }
                .st-key-auto_slump_btn button:hover {
                    background-color: #00838F !important;
                    border-color: #00838F !important;
                }
                .st-key-rote_mode_btn button {
                    background-color: #1976D2 !important;
                    border-color: #1565C0 !important;
                    color: white !important;
                }
                .st-key-rote_mode_btn button:hover {
                    background-color: #1565C0 !important;
                    border-color: #1565C0 !important;
                }
                </style>""",
                unsafe_allow_html=True,
            )
            _col_l, _col_r = st.columns(2)
            with _col_l:
                if st.button(
                    "⚡ 結果ポスト用",
                    key="auto_mode_btn",
                    type="primary",
                    use_container_width=True,
                ):
                    _navigate("auto")
            with _col_r:
                if st.button(
                    "📊 スランプ付き結果ポスト",
                    key="auto_slump_btn",
                    use_container_width=True,
                ):
                    _navigate("auto_slump")
            if st.button(
                "📋 ローテ用",
                key="rote_mode_btn",
                use_container_width=True,
            ):
                _navigate("rote")
        elif store in ("溝の口本館", "溝の口新館", "西武新宿", "渋谷新館", "新宿歌舞伎町", "新大久保"):
            # ローテあり：2列横並び
            st.markdown(
                """<style>
                .st-key-rote_mode_btn button {
                    background-color: #1976D2 !important;
                    border-color: #1565C0 !important;
                    color: white !important;
                }
                .st-key-rote_mode_btn button:hover {
                    background-color: #1565C0 !important;
                    border-color: #1565C0 !important;
                }
                </style>""",
                unsafe_allow_html=True,
            )
            _col_l, _col_r = st.columns(2)
            with _col_l:
                if st.button(
                    "⚡ 結果ポスト用",
                    key="auto_mode_btn",
                    type="primary",
                    use_container_width=True,
                ):
                    _navigate("auto")
            with _col_r:
                if st.button(
                    "📋 ローテ用",
                    key="rote_mode_btn",
                    use_container_width=True,
                ):
                    _navigate("rote")
        elif store in ("稲毛", "上野新館", "新小岩"):
            # 稲毛・上野新館・新小岩：結果ポスト用 ＋ スランプ付き結果ポスト
            st.markdown(
                """<style>
                .st-key-auto_slump_btn button {
                    background-color: #00ACC1 !important;
                    border-color: #00838F !important;
                    color: white !important;
                }
                .st-key-auto_slump_btn button:hover {
                    background-color: #00838F !important;
                    border-color: #00838F !important;
                }
                </style>""",
                unsafe_allow_html=True,
            )
            _col_l, _col_r = st.columns(2)
            with _col_l:
                if st.button(
                    "⚡ 結果ポスト用",
                    key="auto_mode_btn",
                    type="primary",
                    use_container_width=True,
                ):
                    _navigate("auto")
            with _col_r:
                if st.button(
                    "📊 スランプ付き結果ポスト",
                    key="auto_slump_btn",
                    use_container_width=True,
                ):
                    _navigate("auto_slump")
        elif store == "秋葉原":
            # 秋葉原：スランプ付き結果ポスト（左）＋ 記事用（右）
            st.markdown(
                """<style>
                .st-key-auto_slump_btn button {
                    background-color: #00ACC1 !important;
                    border-color: #00838F !important;
                    color: white !important;
                }
                .st-key-auto_slump_btn button:hover {
                    background-color: #00838F !important;
                    border-color: #00838F !important;
                }
                .st-key-auto_article_btn button {
                    background-color: #7B1FA2 !important;
                    border-color: #6A1B9A !important;
                    color: white !important;
                }
                .st-key-auto_article_btn button:hover {
                    background-color: #6A1B9A !important;
                    border-color: #6A1B9A !important;
                }
                </style>""",
                unsafe_allow_html=True,
            )
            _col_l, _col_r = st.columns(2)
            with _col_l:
                if st.button(
                    "📊 スランプ付き結果ポスト",
                    key="auto_slump_btn",
                    use_container_width=True,
                ):
                    _navigate("auto_slump")
            with _col_r:
                if st.button(
                    "📰 記事用",
                    key="auto_article_btn",
                    use_container_width=True,
                ):
                    _navigate("auto_article")
        else:
            # ローテなし：結果ポスト用のみ
            if st.button(
                "⚡ 結果ポスト用",
                key="auto_mode_btn",
                type="primary",
                use_container_width=True,
            ):
                _navigate("auto")

    st.markdown("---")
    st.markdown("**個別に生成する場合は以下から選択：**")

    image_types = STORES.get(store, [])
    if not image_types:
        st.warning("この店舗に対応する画像種類が設定されていません。")
    else:
        for itype in image_types:
            if st.button(itype, key=f"itype_{itype}", use_container_width=True):
                _navigate("work", itype=itype)

    st.markdown("---")
    if st.button("← 店舗選択に戻る", key="back_to_store"):
        _navigate("store")


def show_work_page() -> None:
    """画面3: Excel アップロード・条件設定・画像生成"""
    store      = st.session_state.selected_store
    image_type = st.session_state.selected_image_type

    st.markdown(f"## 【{store}】　{image_type}")
    st.markdown("---")

    # ── Excel アップロード ──────────────────────────────────────────
    st.markdown("### ① Excel ファイルをアップロード")
    uploaded = st.file_uploader(
        "xlsx または xls を選択してください",
        type=["xlsx", "xls"],
        key="excel_upload",
    )

    df = None
    if uploaded:
        try:
            df_raw    = pd.read_excel(uploaded)
            df, missing = normalize_df(df_raw)

            # 必須列チェック
            required = ["台番", "機種名", "差枚"]
            lack     = [c for c in required if c not in df.columns]
            if lack:
                st.error(
                    f"❌ 必須列が見つかりません: {lack}\n\n"
                    f"実際の列名: {list(df_raw.columns)}"
                )
                df = None
            else:
                # 機種名変換テーブルを適用
                name_map_tmp, name_map_norm_tmp = load_name_map()
                df, nc_count = _apply_map(df, name_map_tmp, name_map_norm_tmp)
                st.success(
                    f"✅ {len(df):,} 行を読み込みました"
                    + (f"　（機種名 {nc_count} 件を変換済み）" if nc_count else "")
                )
                if missing:
                    st.warning(f"⚠️ 以下の列は見つかりませんでした: {missing}")
                with st.expander("📋 データプレビュー（先頭 5 行）"):
                    st.dataframe(df.head(), use_container_width=True)

        except Exception as e:
            st.error(f"❌ Excel の読み込みに失敗しました: {e}")
            df = None

    # ── 条件設定 ──────────────────────────────────────────────────
    st.markdown("### ② 条件を設定")
    cond_defs  = IMAGE_CONDITIONS.get(image_type, [])
    conditions: dict = {}

    for cdef in cond_defs:
        key_id = f"cond_{cdef['name']}"   # session_state キー

        if cdef["type"] == "text":
            if cdef["name"] == "keyword":
                try:
                    _kw_candidates = load_machine_candidates()
                except Exception:
                    _kw_candidates = []
                render_machine_autocomplete_input(cdef["label"], key_id, _kw_candidates)
                conditions[cdef["name"]] = st.session_state.get(key_id, "")
            else:
                conditions[cdef["name"]] = st.text_input(
                    cdef["label"],
                    value=str(cdef["default"]),
                    key=key_id,
                )

        elif cdef["type"] == "int":
            conditions[cdef["name"]] = st.number_input(
                cdef["label"],
                value=int(cdef["default"]),
                min_value=0,
                step=100,
                key=key_id,
            )

        elif cdef["type"] == "float":
            conditions[cdef["name"]] = st.number_input(
                cdef["label"],
                value=float(cdef["default"]),
                min_value=0.0,
                step=0.1,
                format="%.2f",
                key=key_id,
            )

    # ── 実行ボタン＋戻るボタン ────────────────────────────────────
    st.markdown("### ③ 実行")

    run_clicked = run_zentai = run_yushu = False

    if image_type == "末尾画像":
        _c1, _c2, _c3 = st.columns([2, 2, 1])
        with _c1:
            run_zentai = st.button("全台", disabled=(df is None), type="primary",
                                   use_container_width=True, key="run_zentai_btn")
        with _c2:
            run_yushu = st.button("優秀台", disabled=(df is None), type="primary",
                                  use_container_width=True, key="run_yushu_btn")
        with _c3:
            if st.button("← 戻る", use_container_width=True, key="back_to_itype"):
                _navigate("image_type")
    else:
        col_run, col_back = st.columns([4, 1])
        with col_run:
            run_clicked = st.button(
                "▶ 画像を生成する",
                disabled=(df is None),
                type="primary",
                use_container_width=True,
                key="run_btn",
            )
        with col_back:
            if st.button("← 戻る", use_container_width=True, key="back_to_itype"):
                _navigate("image_type")

    if df is None and uploaded is None:
        st.info("⬆️ まず Excel ファイルをアップロードしてください。")

    # ── 画像生成 ──────────────────────────────────────────────────
    if (run_clicked or run_zentai or run_yushu) and df is not None:

        # ── 末尾画像（全台 / 優秀台 の2モード） ──────────────────────
        if image_type == "末尾画像":
            tail = str(conditions.get("tail", "")).strip()
            _ZORORME = {"00","11","22","33","44","55","66","77","88","99"}
            if tail == "ゾロ目":
                filtered = df[df["台番"].apply(
                    lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                )].copy()
                base_label = "末尾ゾロ目の台"
            elif tail.isdigit() and len(tail) in (1, 2):
                filtered = df[df["台番"].astype(str).str[-len(tail):] == tail].copy()
                base_label = f"末尾{tail}番台"
            else:
                st.error("❌ 末尾を入力してください（例: 5、22、または「ゾロ目」）。")
                filtered = None
                base_label = ""

            if filtered is not None:
                try:
                    if filtered.empty:
                        st.error(f"❌ {base_label} の台が見つかりません。")
                    else:
                        if run_yushu:
                            filtered = filtered[filtered["差枚"] > 0].copy()
                            if filtered.empty:
                                st.error(f"❌ {base_label}でプラスの台がありません。")
                                return
                            img_title = f"{base_label}の優秀台"
                            stat = None
                        else:
                            img_title = base_label
                            stat = {
                                "total_diff":  int(filtered["差枚"].sum()),
                                "avg_diff":    int(round(filtered["差枚"].mean())),
                                "win_count":   int((filtered["差枚"] > 0).sum()),
                                "total_count": len(filtered),
                            }
                        with st.spinner("画像を生成中..."):
                            img = _build_machine_img(filtered, img_title, stat)
                            st.success("✅ 画像を生成しました！")
                            st.image(img, caption=img_title, use_container_width=True)
                            if not _IS_CLOUD:
                                stem       = os.path.splitext(uploaded.name)[0]
                                dir_stem   = stem.replace("_20S", "")
                                output_dir = os.path.join(_DESKTOP, dir_stem)
                                os.makedirs(output_dir, exist_ok=True)
                                safe = _make_safe_fn(img_title)
                                save_path = os.path.join(output_dir, f"{safe}.jpg")
                                _save_jpeg(img, save_path)
                                st.info(f"💾 `{save_path}` に保存しました")
                except Exception as e:
                    st.error(f"❌ 予期しないエラーが発生しました: {e}")
                    with st.expander("詳細（開発者向け）"):
                        st.code(traceback.format_exc())

        # ── 並び画像（特殊フロー：範囲ごとに別ファイル保存） ──────────
        elif image_type == "並び画像":
            ranges_text = str(conditions.get("ranges", "")).strip()
            if not ranges_text:
                st.error("❌ 台番範囲を入力してください。")
            else:
                narabi_ranges = parse_ranges(ranges_text)
                if not narabi_ranges:
                    st.error("❌ 有効な範囲が見つかりません。例: 409-413, 315-317")
                else:
                    stem       = os.path.splitext(uploaded.name)[0]
                    dir_stem   = stem.replace("_20S", "")
                    if not _IS_CLOUD:
                        output_dir = os.path.join(_DESKTOP, dir_stem)
                        narabi_dir = os.path.join(output_dir, "並び画像")
                        os.makedirs(narabi_dir, exist_ok=True)
                    else:
                        narabi_dir = tempfile.mkdtemp()

                    ban_to_idx = {int(row["台番"]): i for i, row in df.iterrows()}

                    with st.spinner("並び画像を生成中..."):
                        saved_paths = []
                        preview_imgs = []
                        try:
                            for ban_list in narabi_ranges:
                                indices = []
                                for ban in ban_list:
                                    if ban in ban_to_idx:
                                        indices.append(ban_to_idx[ban])
                                if not indices:
                                    st.warning(f"⚠️ 台番 {ban_list[0]}-{ban_list[-1]} がExcelに見つかりません")
                                    continue

                                grp      = df.loc[indices].copy().reset_index(drop=True)
                                diff_raw_s = grp["差枚"].copy()

                                machines = list(dict.fromkeys(str(m) for m in grp["機種名"]))
                                n        = len(grp)
                                if len(machines) == 1:
                                    title_n = f"{machines[0]}({n}台並び)"
                                elif len(machines) == 2:
                                    title_n = f"{machines[0]}+{machines[1]}({n}台並び)"
                                else:
                                    title_n = f"{machines[0]}～{machines[-1]}({n}台並び)"

                                stat = {
                                    "total_diff":  int(diff_raw_s.sum()),
                                    "avg_diff":    int(round(diff_raw_s.mean())),
                                    "win_count":   int((diff_raw_s > 0).sum()),
                                    "total_count": n,
                                }
                                img_n = _build_machine_img(grp, title_n, stat)
                                preview_imgs.append((title_n, img_n))

                                safe = _make_safe_fn(title_n)
                                path = os.path.join(narabi_dir, f"{safe}.jpg")
                                _save_jpeg(img_n, path)
                                saved_paths.append(path)

                            if preview_imgs:
                                st.success(f"✅ {len(preview_imgs)}件の並び画像を生成しました！")
                                for title_n, img_n in preview_imgs:
                                    st.image(img_n, caption=title_n, use_container_width=True)
                                if not _IS_CLOUD:
                                    for p in saved_paths:
                                        st.info(f"💾 `{p}` に保存しました")
                                else:
                                    _narabi_zip = _make_zip_bytes(narabi_dir)
                                    st.download_button("📥 並び画像をZIPでダウンロード",
                                                       _narabi_zip, f"{dir_stem}_narabi.zip",
                                                       "application/zip", key="narabi_zip_dl")

                        except Exception as e:
                            st.error(f"❌ 予期しないエラーが発生しました: {e}")
                            with st.expander("詳細（開発者向け）"):
                                st.code(traceback.format_exc())

        else:
            # ── 通常フロー ──────────────────────────────────────────
            handler = IMAGE_HANDLERS.get(image_type)

            if handler is None:
                st.error(f"❌ 「{image_type}」のハンドラーが未実装です。IMAGE_HANDLERS に追加してください。")
                return

            with st.spinner("画像を生成中..."):
                try:
                    img       = handler(df, conditions)
                    img_bytes = img_to_bytes(img)

                    st.success("✅ 画像を生成しました！")
                    st.image(img, caption=f"{store} — {image_type}", use_container_width=True)

                    title     = str(conditions.get("title", image_type))
                    file_name = f"{store}_{title}.png"
                    st.download_button(
                        label="💾 PNG をダウンロード",
                        data=img_bytes,
                        file_name=file_name,
                        mime="image/png",
                        type="primary",
                        key="download_btn",
                    )

                    # 出力フォルダへ自動保存（ローカルのみ）
                    if not _IS_CLOUD:
                        stem       = os.path.splitext(uploaded.name)[0]
                        dir_stem   = stem.replace("_20S", "")
                        output_dir = os.path.join(_DESKTOP, dir_stem)
                        os.makedirs(output_dir, exist_ok=True)
                        keyword_save = str(conditions.get("keyword", "")).strip()
                        save_name  = keyword_save if keyword_save else title
                        safe_title = _make_safe_fn(save_name)
                        save_path  = os.path.join(output_dir, f"{safe_title}.jpg")
                        _save_jpeg(img, save_path)
                        st.info(f"💾 `{save_path}` に保存しました")

                except ValueError as e:
                    st.error(f"❌ {e}")

                except Exception as e:
                    st.error(f"❌ 予期しないエラーが発生しました: {e}")
                    with st.expander("詳細（開発者向け）"):
                        st.code(traceback.format_exc())

# =============================================================================
# ■ ⑩ 自動処理パイプライン（PIL ベース・全店舗共通）
# =============================================================================

def _save_jpeg(img: Image.Image, path: str, target_kb: int = 250) -> None:
    """JPEG品質をバイナリサーチで target_kb に近づけて保存する。"""
    TARGET = target_kb * 1024
    lo, hi, best_q, best_d = 1, 95, 85, float("inf")
    for _ in range(15):
        mid = (lo + hi) // 2
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=mid, subsampling=0)
        size = buf.tell()
        d = abs(size - TARGET)
        if d < best_d:
            best_d, best_q = d, mid
        if d / TARGET <= 0.02:
            break
        if size < TARGET:
            lo = mid + 1
        else:
            hi = mid - 1
    img.save(path, format="JPEG", quality=best_q, subsampling=0)


def _make_zip_bytes(dir_path: str) -> bytes:
    """dir_path 以下のファイルをすべてZIP化してbytesで返す（サブフォルダ含む）。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _dirs, files in os.walk(dir_path):
            for fname in sorted(files):
                fpath = os.path.join(root, fname)
                arcname = os.path.relpath(fpath, dir_path)
                zf.write(fpath, arcname)
    return buf.getvalue()


_CIRCLE_TO_ASCII = {
    "⓪": "0", "①": "1", "②": "2", "③": "3", "④": "4", "⑤": "5",
    "⑥": "6", "⑦": "7", "⑧": "8", "⑨": "9", "⑩": "10", "⑪": "11",
    "⑫": "12", "⑬": "13", "⑭": "14", "⑮": "15", "⑯": "16", "⑰": "17",
    "⑱": "18", "⑲": "19", "⑳": "20",
}


def _make_safe_fn(name: str) -> str:
    # 丸囲み数字（⓪①…）は一部環境でファイルが見つからない扱いになるため通常数字へ
    s = str(name)
    for _c, _d in _CIRCLE_TO_ASCII.items():
        s = s.replace(_c, _d)
    return re.sub(r'[\\/:*?"<>|]', '_', s)


def _check_github_token() -> tuple[bool, str]:
    """GITHUB_TOKEN が secrets/.env に設定されているか確認する。"""
    token = get_secret_value("GITHUB_TOKEN", "")
    if not token:
        return False, "GITHUB_TOKEN が未設定です"
    return True, f"GITHUB_TOKEN 設定済み（先頭8文字: {str(token)[:8]}…）"


def _github_push_file(content_str: str, repo_path: str = "weekly_items.json") -> tuple[bool, str]:
    """Cloud専用: GitHub APIでファイルを直接更新する。
    st.secrets["GITHUB_TOKEN"] が必要。
    """
    import urllib.request, urllib.error, base64 as _b64
    token = get_secret_value("GITHUB_TOKEN", "")
    if not token:
        return False, "GITHUB_TOKEN未設定"
    repo = "tama0520/guild-image-app"
    url = f"https://api.github.com/repos/{repo}/contents/{repo_path}"
    hdrs = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
    }
    # SHAを取得してPUT（409 Conflict時は最大3回リトライ）
    for _attempt in range(3):
        try:
            req = urllib.request.Request(url, headers=hdrs)
            with urllib.request.urlopen(req, timeout=10) as r:
                sha = json.loads(r.read())["sha"]
        except Exception as e:
            return False, f"SHA取得失敗: {e}"
        body = json.dumps({
            "message": "auto: Cloud上のチェック状態を保存",
            "content": _b64.b64encode(content_str.encode("utf-8")).decode("ascii"),
            "sha": sha,
            "branch": "main",
        }).encode("utf-8")
        try:
            req2 = urllib.request.Request(url, data=body, headers=hdrs, method="PUT")
            with urllib.request.urlopen(req2, timeout=15) as r:
                r.read()
            return True, "GitHubに同期しました"
        except urllib.error.HTTPError as e:
            if e.code == 409:
                continue  # SHA競合 → 最新SHAで再試行
            return False, f"同期失敗: {e}"
        except Exception as e:
            return False, f"同期失敗: {e}"
    return False, "同期失敗: SHA競合が解消されませんでした"


def _git_auto_pull() -> tuple[bool, str]:
    """ローカル環境専用: 起動時にgit pull --rebaseでリモートの変更を取り込む。"""
    import subprocess
    try:
        # 未コミット変更があっても pull できるよう stash で退避
        stash_r = subprocess.run(
            ["git", "stash"],
            cwd=BASE_DIR, capture_output=True,
        )
        stashed = b"No local changes" not in stash_r.stdout and stash_r.returncode == 0
        r = subprocess.run(
            ["git", "pull", "--rebase", "origin", "main"],
            cwd=BASE_DIR, capture_output=True, check=True,
        )
        if stashed:
            subprocess.run(["git", "stash", "pop"], cwd=BASE_DIR, capture_output=True)
        msg = r.stdout.decode("utf-8", errors="replace").strip()
        return True, msg or "Already up to date."
    except subprocess.CalledProcessError as e:
        return False, e.stderr.decode("utf-8", errors="replace").strip()


def _git_auto_push(label: str = "auto") -> tuple[bool, str]:
    """ローカル環境専用: 設定ファイルをgit add→commit→pushする。
    戻り値: (成功, メッセージ)
    """
    import subprocess
    targets = [
        "weekly_items.json",
        "auto_page_inputs.json",
        "rote_machines.json",
        "store_settings",
    ]
    try:
        # 先にステージング（pull前に行うことでunstaged changesによるpull失敗を防ぐ）
        subprocess.run(
            ["git", "add"] + targets,
            cwd=BASE_DIR, capture_output=True, check=True,
        )
        # 差分がなければpullのみしてスキップ
        diff = subprocess.run(
            ["git", "diff", "--cached", "--quiet"],
            cwd=BASE_DIR, capture_output=True,
        )
        if diff.returncode == 0:
            subprocess.run(
                ["git", "pull", "--rebase", "origin", "main"],
                cwd=BASE_DIR, capture_output=True, check=True,
            )
            return True, "変更なし（push不要）"
        # コミット → リモートの変更を取り込む → push
        subprocess.run(
            ["git", "commit", "-m", f"auto: {label}後の設定を保存"],
            cwd=BASE_DIR, capture_output=True, check=True,
        )
        subprocess.run(
            ["git", "pull", "--rebase", "origin", "main"],
            cwd=BASE_DIR, capture_output=True, check=True,
        )
        subprocess.run(
            ["git", "push", "origin", "main"],
            cwd=BASE_DIR, capture_output=True, check=True,
        )
        return True, "GitHubにpushしました"
    except subprocess.CalledProcessError as e:
        return False, f"push失敗: {e.stderr.decode('utf-8', errors='replace').strip()}"


def _load_pipeline_df(excel_path: str) -> tuple[pd.DataFrame, pd.Series]:
    """Excel/CSV を読み込み、列正規化・機種名変換・補助列追加を行って返す。
    戻り値: (df, diff_raw)  diff_raw は生の差枚 int Series。"""
    if excel_path.lower().endswith(".csv"):
        with open(excel_path, "rb") as _f:
            df_raw = _read_csv_raw(_f.read())
    else:
        df_raw = pd.read_excel(excel_path)
    df, _ = normalize_df(df_raw)
    nm, nm_norm = load_name_map()
    if nm:
        df, _ = _apply_map(df, nm, nm_norm)
    df["ゲーム数_rounded"] = df["ゲーム数"].apply(round_games)
    df["合算確率_num"] = df.apply(
        lambda r: (r["ゲーム数"] / (r["BB"] + r["RB"])
                   if (r["BB"] + r["RB"]) > 0 else float("inf")),
        axis=1,
    )
    diff_raw = df["差枚"].copy()
    return df, diff_raw


def _stat_from_diff(diff_raw_s: pd.Series) -> dict:
    return {
        "total_diff":  int(diff_raw_s.sum()),
        "avg_diff":    int(round(diff_raw_s.mean())),
        "win_count":   int((diff_raw_s > 0).sum()),
        "total_count": len(diff_raw_s),
    }


def _build_machine_img(
    df_m: pd.DataFrame,
    title: str,
    summary_stat: dict | None = None,
) -> Image.Image:
    """パイプライン用機種別画像。旧スクリプトと同じ寸法で生成する。
    title が「（優秀台）」で終わる場合は2パーツ描画（gap=-22）。"""

    # ── テーブル部分（タイトル・ピンクバーなし）─────────────────────
    df_d = _add_prob_col(df_m)
    active = _active_sub_cols(df_d)
    disp = ["台番", "機種名", "ゲーム数"] + active
    if "合算確率" in df_d.columns:
        disp.append("合算確率")
    disp.append("差枚")
    disp = [c for c in disp if c in df_d.columns]
    df_d     = _format_display_cols(df_d[disp])
    headers  = [_DISPLAY_RENAME.get(h, h) for h in df_d.columns]
    rows     = df_d.values.tolist()
    diff_idx = headers.index("差枚数") if "差枚数" in headers else None
    table_img = draw_table_image(
        headers, rows,
        diff_col_idx=diff_idx,
        title=None,
        summary_stat=None,
        header_bg=C_MACH_HEADER_BG,
        header_fg=C_MACH_HEADER_FG,
        scale=150/96,
    )
    w = table_img.width

    # ── タイトルバー（横幅に比例してBAR_Hを自動計算・視覚的に統一）──────
    LINE_H  = 6
    BAR_H   = round(w * 73 / 950)   # 標準幅950pxのとき73px
    FONT_SZ = round(BAR_H * 40 / 73)

    bar  = Image.new("RGBA", (w, BAR_H),  (38, 76, 161, 255))
    line = Image.new("RGBA", (w, LINE_H), (204, 0, 0, 255))
    bd   = ImageDraw.Draw(bar)
    font = load_font(FONT_SZ)

    SUB = "（優秀台）"
    if title.endswith(SUB):
        main_text = title[:-len(SUB)].replace('\uff65', '\u30fb')
        sub_text  = SUB
        GAP_TITLE = -22
        b1 = bd.textbbox((0, 0), main_text, font=font)
        b2 = bd.textbbox((0, 0), sub_text,  font=font)
        w1, w2  = b1[2]-b1[0], b2[2]-b2[0]
        total_w = w1 + GAP_TITLE + w2
        x1 = (w - total_w) // 2 - b1[0]
        x2 = x1 + w1 + GAP_TITLE - b2[0]
        ty = (BAR_H - (b1[3]-b1[1])) // 2 - b1[1]
        bd.text((x1, ty), main_text, fill=(255, 255, 255, 255), font=font)
        bd.text((x2, ty), sub_text,  fill=(255, 255, 255, 255), font=font)
    else:
        disp_title = title.replace('\uff65', '\u30fb')
        bb = bd.textbbox((0, 0), disp_title, font=font)
        tx = (w - (bb[2]-bb[0])) // 2 - bb[0]
        ty = (BAR_H - (bb[3]-bb[1])) // 2 - bb[1]
        bd.text((tx, ty), disp_title, fill=(255, 255, 255, 255), font=font)

    # ── ピンクサマリーバー（全機種固定高さ）─────────────────────────
    if summary_stat:
        pink_h = int(round(ROW_H * 150 / 96) * 1.2)  # 44px × 1.2 = 52px 固定
        font_sum   = load_font(24)

        s        = summary_stat
        win_rate = s["win_count"] / s["total_count"] * 100 if s["total_count"] > 0 else 0.0
        part1 = (f"総差枚：{fmt_diff(s['total_diff'])}"
                 f"　平均：{fmt_diff(s['avg_diff'])}"
                 f"　勝率：{win_rate:.1f}%")
        part2 = f"（{s['win_count']}/{s['total_count']}台）"

        pink = Image.new("RGBA", (w, pink_h), (255, 182, 193, 255))
        pd_  = ImageDraw.Draw(pink)
        bb1  = pd_.textbbox((0, 0), part1, font=font_sum)
        ty_p = (pink_h - (bb1[3]-bb1[1])) // 2 - bb1[1]
        pd_.text((8, ty_p), part1, fill=(0, 0, 0, 255), font=font_sum)
        x2   = 8 + (bb1[2]-bb1[0]) + GAP_SUM
        bb2  = pd_.textbbox((0, 0), part2, font=font_sum)
        pd_.text((x2 - bb2[0], ty_p), part2, fill=(0, 0, 0, 255), font=font_sum)
        pd_.rectangle([0, 0, w-1, pink_h-1], outline=(170, 170, 170, 255))

        total_h = BAR_H + LINE_H + table_img.height + pink_h
        final   = Image.new("RGBA", (w, total_h), (255, 255, 255, 255))
        final.paste(bar,       (0, 0))
        final.paste(line,      (0, BAR_H))
        final.paste(table_img, (0, BAR_H + LINE_H))
        final.paste(pink,      (0, BAR_H + LINE_H + table_img.height))
    else:
        ImageDraw.Draw(table_img).line(
            [(0, table_img.height-1), (w-1, table_img.height-1)],
            fill=(170, 170, 170, 255), width=1,
        )
        total_h = BAR_H + LINE_H + table_img.height
        final   = Image.new("RGBA", (w, total_h), (255, 255, 255, 255))
        final.paste(bar,       (0, 0))
        final.paste(line,      (0, BAR_H))
        final.paste(table_img, (0, BAR_H + LINE_H))

    return final.convert("RGB")


def _build_machine_img_no_bar(df_m: pd.DataFrame) -> Image.Image:
    """タイトルバー・ピンクバーなしの機種別テーブル画像（記事用）。"""
    df_d   = _add_prob_col(df_m)
    active = _active_sub_cols(df_d)
    disp   = ["台番", "機種名", "ゲーム数"] + active
    if "合算確率" in df_d.columns:
        disp.append("合算確率")
    disp.append("差枚")
    disp     = [c for c in disp if c in df_d.columns]
    df_d     = _format_display_cols(df_d[disp])
    headers  = [_DISPLAY_RENAME.get(h, h) for h in df_d.columns]
    rows     = df_d.values.tolist()
    diff_idx = headers.index("差枚数") if "差枚数" in headers else None
    return draw_table_image(
        headers, rows,
        diff_col_idx=diff_idx,
        title=None,
        summary_stat=None,
        header_bg=C_MACH_HEADER_BG,
        header_fg=C_MACH_HEADER_FG,
        scale=150 / 96,
    ).convert("RGB")


def _build_article_machine_img(
    df_m: pd.DataFrame,
    machine_name: str,
    summary_stat: dict | None = None,
) -> Image.Image:
    """記事用機種別画像。タイトルバー・ピンクバーなし。表＋白サマリーエリア（クラウン＋機種名＋統計）。"""

    # ── テーブル部分（diff_cell_bg=True で差枚セルに背景色）─────────
    df_d   = _add_prob_col(df_m)
    active = _active_sub_cols(df_d)
    disp   = ["台番", "機種名", "ゲーム数"] + active
    if "合算確率" in df_d.columns:
        disp.append("合算確率")
    disp.append("差枚")
    disp     = [c for c in disp if c in df_d.columns]
    df_d     = _format_display_cols(df_d[disp])
    headers  = [_DISPLAY_RENAME.get(h, h) for h in df_d.columns]
    rows     = df_d.values.tolist()
    diff_idx = headers.index("差枚数") if "差枚数" in headers else None
    table_img = draw_table_image(
        headers, rows,
        diff_col_idx=diff_idx,
        title=None,
        summary_stat=None,
        header_bg=C_MACH_HEADER_BG,
        header_fg=C_MACH_HEADER_FG,
        diff_cell_bg=False,
        scale=150 / 96,
    )
    w = table_img.width
    ImageDraw.Draw(table_img).line(
        [(0, table_img.height - 1), (w - 1, table_img.height - 1)],
        fill=(170, 170, 170, 255), width=1,
    )

    # ── 白サマリーエリア ─────────────────────────────────────────────
    PAD      = 10
    CROWN_H  = 72
    FN_MACH  = load_font(34)
    FN_STAT  = load_font(26)
    LINE_GAP = 8

    # クラウン画像（JPGなのでアルファなし）
    crown_path = os.path.join(BASE_DIR, "crown.jpg")
    crown_img  = None
    crown_w    = 0
    if os.path.exists(crown_path):
        _ci     = Image.open(crown_path).convert("RGB")
        ow, oh  = _ci.size
        crown_w = round(ow * CROWN_H / oh)
        crown_img = _ci.resize((crown_w, CROWN_H), Image.LANCZOS)

    # テキストサイズ測定
    dummy   = Image.new("RGB", (1, 1))
    d0      = ImageDraw.Draw(dummy)
    bb_mach = d0.textbbox((0, 0), machine_name, font=FN_MACH)
    mach_h  = bb_mach[3] - bb_mach[1]
    row1_h  = max(CROWN_H, mach_h)

    stat_lines: list[str] = []
    if summary_stat:
        s  = summary_stat
        wr = s["win_count"] / s["total_count"] * 100 if s["total_count"] > 0 else 0.0
        stat_lines = [
            f"勝率　：{wr:.1f}% ({s['win_count']}/{s['total_count']}台)",
            f"総差枚：{fmt_diff(s['total_diff']).replace('枚', ' 枚')}",
            f"平均　：{fmt_diff(s['avg_diff']).replace('枚', ' 枚')}",
        ]

    bb_stat0      = d0.textbbox((0, 0), stat_lines[0] if stat_lines else "勝率", font=FN_STAT)
    stat_h        = bb_stat0[3] - bb_stat0[1]
    stats_total_h = (stat_h + LINE_GAP) * len(stat_lines)

    summary_h = PAD + row1_h + PAD + stats_total_h + PAD
    summary   = Image.new("RGB", (w, summary_h), (255, 255, 255))
    sd        = ImageDraw.Draw(summary)

    # クラウン貼り付け
    if crown_img:
        cy = PAD + (row1_h - CROWN_H) // 2
        summary.paste(crown_img, (PAD, cy))

    # 機種名（クラウン右隣、垂直中央）
    text_x = PAD + crown_w + PAD
    mach_y = PAD + (row1_h - mach_h) // 2 - bb_mach[1]
    sd.text((text_x - bb_mach[0], mach_y), machine_name, fill=(0, 0, 0), font=FN_MACH)

    # 統計行
    sy = PAD + row1_h + PAD
    for line in stat_lines:
        bb = d0.textbbox((0, 0), line, font=FN_STAT)
        sd.text((PAD - bb[0], sy - bb[1]), line, fill=(0, 0, 0), font=FN_STAT)
        sy += stat_h + LINE_GAP

    # テーブル + サマリーを結合
    total_h = table_img.height + summary_h
    final   = Image.new("RGB", (w, total_h), (255, 255, 255))
    final.paste(table_img.convert("RGB"), (0, 0))
    final.paste(summary, (0, table_img.height))
    return final


def run_step1_main(
    df: pd.DataFrame,
    diff_raw: pd.Series,
    output_dir: str,
    stem: str,
    cfg: dict,
    log,
    article_mode: bool = False,
) -> tuple[list[str], list[dict]]:
    """Step 1: 全台系PNG + 全台プラス機種別JPG を生成する。
    戻り値: (generated, zen_dai_list)"""
    manual_exclude   = cfg["manual_exclude"]
    juggler_series   = cfg["juggler_series"]
    juggler_g_min    = cfg["juggler_g_min"]
    prob_jobs_map    = cfg["prob_jobs_map"]
    juggler_prob_map = {m: thr for m, thr, _ in cfg.get("juggler_jobs", [])}
    generated: list[str] = []
    zen_dai_list: list[dict] = []

    # ── 全台系 PNG（タイトルなし・青ヘッダー）──────────────────────
    df_d = _add_prob_col(df)
    active = _active_sub_cols(df_d)
    disp = ["台番", "機種名", "ゲーム数"] + active
    if "合算確率" in df_d.columns:
        disp.append("合算確率")
    disp.append("差枚")
    disp = [c for c in disp if c in df_d.columns]
    df_disp  = _format_display_cols(df_d[disp])
    headers  = [_DISPLAY_RENAME.get(h, h) for h in df_disp.columns]
    rows     = df_disp.values.tolist()
    diff_idx = headers.index("差枚数") if "差枚数" in headers else None
    img_all  = draw_table_image(
        headers, rows,
        diff_col_idx=diff_idx,
        title=None,
        summary_stat=None,
        header_bg=C_HEADER_BG,
        header_fg=C_HEADER_FG,
        scale=150/96,
    )
    # 全台系PNG は全店舗で生成しない（不使用のため）

    # ── 機種別 JPG（全台プラスのみ）──────────────────────────────
    for machine, grp in df.groupby("機種名", sort=False):
        if machine in manual_exclude:
            continue
        if machine in prob_jobs_map:
            prob_thr, _ = prob_jobs_map[machine]
            _dr_check = diff_raw.loc[grp.index]
            _all_g = bool((grp["ゲーム数_rounded"] > 2000).all())
            if "合算確率_num" in grp.columns:
                _all_q = bool(((grp["合算確率_num"] <= prob_thr) & (_dr_check > 0)).all())
            else:
                _all_q = False
            _all_1k = bool((_dr_check >= 1000).all())
            if (not (_all_g and _all_q) and not _all_1k) or len(grp) <= 1:
                continue  # 全台条件未達 → Step3で処理
            # 全台が確率条件を満たす OR 全台+1000枚以上 → 全台系として処理（fall-through）
        total_raw = len(grp)
        dr_all = diff_raw.loc[grp.index]  # G数フィルター前の全台差枚
        _g_all = grp["ゲーム数_rounded"] if "ゲーム数_rounded" in grp.columns else None
        _grp_orig = grp  # 合算確率チェック用（フィルター前）
        if machine in juggler_series:
            grp = grp[grp["ゲーム数_rounded"] >= juggler_g_min]
        if grp.empty:
            continue
        dr_m = diff_raw.loc[grp.index]
        # 各台が「+1000枚以上」または「G数>=2000かつプラスかつ合算確率達成」なら全台系
        # ジャグラーは合算確率条件も必須
        _jug_prob_thr = juggler_prob_map.get(machine) if machine in juggler_series else None
        if _g_all is not None:
            if _jug_prob_thr is not None and "合算確率_num" in _grp_orig.columns:
                all_plus = bool(((dr_all >= 1000) | ((dr_all >= 0) & (_g_all >= 2000) & (_grp_orig["合算確率_num"] <= _jug_prob_thr))).all())
            else:
                all_plus = bool(((dr_all >= 1000) | ((dr_all >= 0) & (_g_all >= 2000))).all())
        else:
            all_plus = bool((dr_all >= 0).all())
        if not all_plus or len(dr_m) <= 1:
            continue  # 条件未達・1台以下 はStep3で処理
        title = machine.replace('\uff65', '\u30fb')
        if article_mode:
            img = _build_article_machine_img(grp, title, _stat_from_diff(dr_m))
        else:
            img = _build_machine_img(grp, title, _stat_from_diff(dr_m))
        out   = os.path.join(output_dir, f"{_make_safe_fn(machine)}.jpg")
        _save_jpeg(img, out)
        generated.append(out)
        log(f"  {machine}（{len(dr_m)}台）")
        zen_dai_list.append({
            "name":         machine,
            "count":        int((dr_m > 0).sum()),
            "total":        total_raw,
            "diffs":        sorted([int(d) for d in dr_m.tolist() if int(d) >= 1000], reverse=True),
            "all_avg_diff": int(round(dr_m.mean())),
            "bans":         [int(b) for b in grp["台番"].tolist()],
        })

    return generated, zen_dai_list


def run_step2_juggler(
    df: pd.DataFrame,
    diff_raw: pd.Series,
    output_dir: str,
    cfg: dict,
    narabi_bans: set[int],
    log,
    recommended_machines: set[str] = set(),
    suebangai_bans: set[int] = set(),
    zen_dai_juggler_machines: set[str] = set(),
    article_mode: bool = False,
    sonota_exclude: set[str] = frozenset(),
) -> tuple[list[str], pd.DataFrame | None, pd.Series | None, list[dict], list[dict]]:
    """Step 2: ジャグラーシリーズ優秀台フィルター。
    少数機種は統合画像へ。5台以下なら overflow として Step 3 へ渡す。
    戻り値: (generated, overflow_df, overflow_diff, high_ratio_list, jug_excellent_list)"""
    juggler_jobs  = cfg["juggler_jobs"]
    juggler_g_min = cfg["juggler_g_min"]
    generated:         list[str]         = []
    pool_dfs:          list[pd.DataFrame] = []
    pool_diffs:        list[pd.Series]    = []
    high_ratio_list:   list[dict]         = []
    jug_excellent_list: list[dict]        = []

    _jug_all_bans = narabi_bans | suebangai_bans
    for machine, prob_threshold, diff_bonus in juggler_jobs:
        all_for_m_orig = df[df["機種名"] == machine].copy()
        all_for_m = all_for_m_orig.copy()
        if _jug_all_bans:
            all_for_m = all_for_m[~all_for_m["台番"].isin(_jug_all_bans)]
        # 高配分判定・画像生成は並び含む全台ベース（Step3と同じ方針）
        total_all  = len(all_for_m_orig)
        total_orig = total_all
        count_orig = int((diff_raw.loc[all_for_m_orig.index] > 0).sum())
        mdf = all_for_m_orig[all_for_m_orig["ゲーム数_rounded"] >= juggler_g_min]
        if mdf.empty:
            continue
        dr_m = diff_raw.loc[mdf.index]
        # 全台プラスチェックはG数フィルター前の全台で行う（G数未達マイナス台が隠れるバグ防止）
        _dr_all_orig = diff_raw.loc[all_for_m_orig.index]
        _all_plus = bool(((_dr_all_orig >= 1000) | ((_dr_all_orig >= 0) & (all_for_m_orig["ゲーム数_rounded"] >= juggler_g_min) & (all_for_m_orig["合算確率_num"] <= prob_threshold))).all())
        if _all_plus:
            log(f"  {machine} 全台プラス→スキップ")
            continue
        mask     = ((mdf["合算確率_num"] <= prob_threshold) & (dr_m >= 0)) | (dr_m >= diff_bonus)
        filtered = mdf[mask].copy().reset_index(drop=True)
        dr_f     = dr_m[mask].reset_index(drop=True)
        if filtered.empty:
            continue
        # pool・excellent_list 用は並び台・末尾台除外
        mdf_ex      = all_for_m[all_for_m["ゲーム数_rounded"] >= juggler_g_min]
        dr_m_ex     = diff_raw.loc[mdf_ex.index]
        mask_ex     = ((mdf_ex["合算確率_num"] <= prob_threshold) & (dr_m_ex >= 0)) | (dr_m_ex >= diff_bonus)
        filtered_ex = mdf_ex[mask_ex].copy().reset_index(drop=True)
        dr_f_ex     = dr_m_ex[mask_ex].reset_index(drop=True)
        # +1000枚以上の台を excellent_list 用に収集（並び除外・個別・オススメ機種は除外）
        if machine not in recommended_machines:
            for _i in range(len(filtered_ex)):
                _d = int(dr_f_ex.iloc[_i])
                if _d >= 1000:
                    jug_excellent_list.append({
                        "name": machine, "diff": _d,
                        "ban": int(filtered_ex.iloc[_i]["台番"]),
                    })
        _small_rule = cfg.get("small_machine_rule")
        _meets_small_jug = False
        if _small_rule and total_all <= _small_rule["max_total"]:
            _dr_all = diff_raw.loc[all_for_m_orig.index]
            _plus_all = int((_dr_all > 0).sum())
            _cnt_1k_all = int((_dr_all >= 1000).sum())
            if _plus_all >= math.ceil(total_all / 2) and _cnt_1k_all >= _small_rule["min_1k"]:
                _meets_small_jug = True
        if machine not in recommended_machines and (_meets_small_jug or (len(filtered) >= 2 and (len(filtered) >= total_all / 2 or len(filtered) >= 10))):
            if article_mode:
                img = _build_machine_img_no_bar(filtered)
            else:
                img = _build_machine_img(filtered, machine.replace('･', '・') + "（優秀台）", None)
            out   = os.path.join(output_dir, f"{_make_safe_fn(machine)}_高配分.jpg")
            _save_jpeg(img, out)
            generated.append(out)
            log(f"  {machine} 高配分: {len(filtered)}台")
            high_ratio_list.append({
                "name":       machine,
                "count":      count_orig,
                "total":      total_orig,
                "diffs":      sorted([int(d) for d in dr_f.tolist() if int(d) >= 1000], reverse=True),
                "all_avg_diff": int(round(diff_raw.loc[all_for_m_orig.index].mean())),
                "has_image":  True,
                "bans":       [int(b) for b in filtered["台番"].tolist()],
            })
        else:
            if machine in recommended_machines:
                log(f"  {machine} →個別画像指定のためジャグラー高配分・統合から除外")
            else:
                pool_dfs.append(filtered_ex)
                pool_diffs.append(dr_f_ex)
                log(f"  {machine} →統合({len(filtered_ex)}/{len(all_for_m)}台)")
                # 統合画像に入った機種も勝率50%以上ならテキストのみ high_ratio_list に追加
                if count_orig >= math.ceil(total_orig / 2):
                    high_ratio_list.append({
                        "name":         machine,
                        "count":        count_orig,
                        "total":        total_orig,
                        "diffs":        sorted([int(d) for d in dr_f.tolist() if int(d) >= 1000], reverse=True),
                        "all_avg_diff": int(round(_dr_all_orig.mean())),
                        "has_image":    False,
                    })

    if not pool_dfs:
        return generated, None, None, high_ratio_list, jug_excellent_list, None

    pool_dfs   = [d for d in pool_dfs   if not d.empty]
    pool_diffs = [d for d in pool_diffs if not d.empty]
    if not pool_dfs:
        return generated, None, None, high_ratio_list, jug_excellent_list, None
    combined    = pd.concat(pool_dfs,   ignore_index=True)
    dr_combined = pd.concat(pool_diffs, ignore_index=True)
    if combined.empty:
        return generated, None, None, high_ratio_list, jug_excellent_list, None
    order       = combined["台番"].argsort()
    combined    = combined.iloc[order].reset_index(drop=True)
    dr_combined = dr_combined.iloc[order].reset_index(drop=True)

    if len(combined) <= 5:
        log(f"  ジャグラー統合 {len(combined)}台 → overflow")
        return generated, combined, dr_combined, high_ratio_list, jug_excellent_list, None

    # オススメ機種にジャグラーが含まれる場合は統合画像を作らずoverflowへ
    _juggler_names = {m for m, _, _ in juggler_jobs}
    if sonota_exclude & _juggler_names:
        log(f"  ジャグラーシリーズ優秀台: オススメ機種に含まれるため統合画像スキップ → overflow")
        return generated, combined, dr_combined, high_ratio_list, jug_excellent_list, None

    juggler_series_set = {m for m, _, _ in juggler_jobs}
    has_narabi_jug = bool(narabi_bans) and not df[
        df["台番"].isin(narabi_bans) & df["機種名"].isin(juggler_series_set)
    ].empty
    juggler_recommended = {m for m in recommended_machines if m in juggler_series_set}
    has_other_jug_img = bool(high_ratio_list) or bool(zen_dai_juggler_machines) or has_narabi_jug or bool(juggler_recommended)
    if article_mode:
        img = _build_machine_img_no_bar(combined)
    else:
        title_jug = "その他のジャグラーシリーズの優秀台" if has_other_jug_img else "ジャグラーシリーズの優秀台"
        img = _build_machine_img(combined, title_jug, None)
    out = os.path.join(output_dir, "ジャグラーシリーズ優秀台.jpg")
    _save_jpeg(img, out, target_kb=800)
    generated.append(out)
    log(f"  ジャグラーシリーズ優秀台: {len(combined)}台")
    return generated, None, None, high_ratio_list, jug_excellent_list, combined


def run_step3_other(
    df: pd.DataFrame,
    diff_raw: pd.Series,
    output_dir: str,
    cfg: dict,
    narabi_bans: set[int],
    overflow_df:   pd.DataFrame | None,
    overflow_diff: pd.Series    | None,
    log,
    recommended_machines: set[str] = set(),
    suebangai_bans: set[int] = set(),
    article_mode: bool = False,
    sonota_exclude: set[str] = frozenset(),
) -> tuple[list[str], list[dict], list[dict]]:
    """Step 3: 非ジャグラー機種の優秀台 + その他の優秀台ピックアップ統合画像。
    戻り値: (generated, high_ratio_list, excellent_list)"""
    juggler_series     = cfg["juggler_series"]
    manual_exclude     = cfg["manual_exclude"]
    rb_thresh_machines = cfg["rb_threshold_machines"]
    rb_min             = cfg["rb_min"]
    diff_bonus         = cfg["diff_bonus"]
    prob_jobs_map      = cfg["prob_jobs_map"]
    generated:       list[str]         = []
    other_dfs:       list[pd.DataFrame] = []
    other_diffs:     list[pd.Series]    = []
    high_ratio_list: list[dict]         = []
    excellent_list:  list[dict]         = []

    for machine, grp in df.groupby("機種名", sort=False):
        if machine in juggler_series:
            continue
        if grp.empty:
            continue
        dr_m     = diff_raw.loc[grp.index]
        total    = len(grp)
        if "ゲーム数_rounded" in grp.columns:
            all_plus = bool(((dr_m >= 1000) | ((dr_m >= 0) & (grp["ゲーム数_rounded"] >= 2000))).all())
        else:
            all_plus = bool((dr_m >= 0).all())
        # excellent_list・その他の優秀台用は並び台・末尾台除外
        _ex_bans = narabi_bans | suebangai_bans
        if _ex_bans:
            grp_ex = grp[~grp["台番"].isin(_ex_bans)].copy()
            dr_ex  = diff_raw.loc[grp_ex.index]
        else:
            grp_ex, dr_ex = grp, dr_m

        if all_plus:
            if machine in manual_exclude:
                mask_1k = dr_ex >= 1000
                if mask_1k.any():
                    if machine not in recommended_machines and machine not in sonota_exclude:
                        other_dfs.append(grp_ex[mask_1k].copy().reset_index(drop=True))
                        other_diffs.append(dr_ex[mask_1k].reset_index(drop=True))
                    if machine not in recommended_machines:
                        for idx in dr_ex[mask_1k].index:
                            excellent_list.append({"name": machine, "diff": int(diff_raw.loc[idx]), "ban": int(grp_ex.loc[idx, "台番"])})
                continue
            elif total == 1 and int(dr_m.iloc[0]) >= 1000:
                if not grp_ex.empty:
                    if machine not in recommended_machines and machine not in sonota_exclude:
                        other_dfs.append(grp_ex.copy().reset_index(drop=True))
                        other_diffs.append(dr_ex.reset_index(drop=True))
                        excellent_list.append({"name": machine, "diff": int(dr_ex.iloc[0]), "ban": int(grp_ex.iloc[0]["台番"])})
                continue
            elif bool((grp["ゲーム数_rounded"] > 2000).all()):
                if machine in prob_jobs_map:
                    # prob_jobs機種はStep1で確率条件チェックがある → 実際に生成されたか確認
                    _p_thr, _ = prob_jobs_map[machine]
                    if "合算確率_num" in grp.columns:
                        _step1_ok = bool(((grp["合算確率_num"] <= _p_thr) & (dr_m > 0)).all())
                    else:
                        _step1_ok = False
                    _all_1k_s3 = bool((dr_m >= 1000).all())
                    if _step1_ok or _all_1k_s3:
                        continue  # Step 1 で生成済み（確率条件 or 全台+1000枚）
                    # Step 1 を確率条件未達でスキップ → 高配分フィルターへ fall-through
                else:
                    continue  # 全台G数>2000 → Step 1 で生成済み
            elif total > 1 and (machine not in prob_jobs_map or bool((dr_m >= 1000).all())):
                continue  # all_plus かつ 2台以上（非prob_jobs or 全台+1000枚）→ Step 1 で生成済み
            # prob_jobs で確率条件未達またはG数条件未達 → 高配分フィルターへ fall-through

        if machine in manual_exclude:
            mask_1k = dr_ex >= 1000
            if mask_1k.any():
                if machine not in recommended_machines and machine not in sonota_exclude:
                    other_dfs.append(grp_ex[mask_1k].copy().reset_index(drop=True))
                    other_diffs.append(dr_ex[mask_1k].reset_index(drop=True))
                for idx in dr_ex[mask_1k].index:
                    excellent_list.append({"name": machine, "diff": int(diff_raw.loc[idx]), "ban": int(grp_ex.loc[idx, "台番"])})
            continue

        # 高配分: 全台（並び含む）でカウント・生成
        _use_rb_mask = False
        if machine in prob_jobs_map:
            prob_thr, diff_bon = prob_jobs_map[machine]
            if "合算確率_num" in grp.columns and "ゲーム数_rounded" in grp.columns:
                mask = (grp["ゲーム数_rounded"] >= 2000) & (((grp["合算確率_num"] <= prob_thr) & (dr_m >= 0)) | (dr_m >= diff_bon))
            elif "ゲーム数_rounded" in grp.columns:
                mask = (grp["ゲーム数_rounded"] >= 2000) & (dr_m >= diff_bon)
            else:
                mask = dr_m >= diff_bon
        elif machine in rb_thresh_machines:
            # AT機: +1000枚は無条件で含む・RBパスはG数>=2000を必須
            _std_count = int((dr_m >= 1000).sum())
            _rb_col = next((c for c in ["RB", "REG"] if c in grp.columns), None)
            if _rb_col and "ゲーム数_rounded" in grp.columns:
                mask = (dr_m >= 1000) | ((grp["ゲーム数_rounded"] >= 2000) & (grp[_rb_col] >= rb_min) & (dr_m >= 0))
            elif "ゲーム数_rounded" in grp.columns:
                mask = dr_m >= 1000
            else:
                mask = dr_m >= 1000
            # 画像生成閾値: 7台以上かつ通常条件が少ない場合は35%
            if total >= 7 and _std_count < max(2, math.ceil(total / 2)):
                _use_rb_mask = True
        else:
            if "ゲーム数_rounded" in grp.columns:
                _cnt_1k = int((dr_m >= 1000).sum())
                if _cnt_1k >= 2:
                    # +1000枚台が2台以上ある場合: ±0以上かつG数>=2000の台も含める
                    mask = (dr_m >= 1000) | ((dr_m >= 0) & (grp["ゲーム数_rounded"] >= 2000))
                else:
                    mask = (dr_m >= 1000) | ((dr_m > 0) & (grp["ゲーム数_rounded"] >= 5000))
            else:
                mask = dr_m >= 1000
        filtered = grp[mask].copy().reset_index(drop=True)
        dr_f     = dr_m[mask].reset_index(drop=True)
        count_f  = len(filtered)

        if count_f == 0:
            continue
        if _use_rb_mask:
            _meets_thr = count_f >= max(2, math.ceil(total * 0.35))
        else:
            _meets_thr = count_f >= max(2, math.ceil(total / 2))
        _small_rule = cfg.get("small_machine_rule")
        _meets_small = False
        if _small_rule and total <= _small_rule["max_total"]:
            _plus_cnt = int((dr_m > 0).sum())
            _cnt_1k = int((dr_m >= 1000).sum())
            if _plus_cnt >= math.ceil(total / 2) and _cnt_1k >= _small_rule["min_1k"]:
                _meets_small = True
        _min7_machines = cfg.get("min7_machines", set())
        _meets_min7 = machine in _min7_machines and count_f >= 7
        if _meets_thr or _meets_small or _meets_min7:
            if machine not in recommended_machines:
                if article_mode:
                    img = _build_machine_img_no_bar(filtered)
                else:
                    img = _build_machine_img(filtered, machine.replace('･', '・') + "（優秀台）", None)
                out   = os.path.join(output_dir, f"{_make_safe_fn(machine)}_高配分.jpg")
                _save_jpeg(img, out)
                generated.append(out)
                log(f"  {machine}: {count_f}/{total}台")
                high_ratio_list.append({
                    "name":         machine,
                    "count":        int((dr_m > 0).sum()),
                    "total":        total,
                    "diffs":        sorted([int(d) for d in dr_f.tolist() if int(d) >= 1000], reverse=True),
                    "all_avg_diff": int(round(dr_m.mean())),
                    "has_image":    True,
                    "bans":         [int(b) for b in filtered["台番"].tolist()],
                })
        else:
            # 勝率50%以上 → テキストのみ high_ratio_list に追加（画像なし）
            # 表示差枚は+1,000枚以上のみ（稲毛指示と同様）
            plus_count = int((dr_m > 0).sum())
            if total >= 2 and plus_count >= math.ceil(total / 2) and machine not in recommended_machines:
                high_ratio_list.append({
                    "name":         machine,
                    "count":        plus_count,
                    "total":        total,
                    "diffs":        sorted([int(d) for d in dr_m.tolist() if int(d) >= 1000], reverse=True),
                    "all_avg_diff": int(round(dr_m.mean())),
                    "has_image":    False,
                })
            # excellent pool: 並び台除外版を使用
            mask_ex = dr_ex >= 1000
            filt_ex = grp_ex[mask_ex].copy().reset_index(drop=True)
            dr_f_ex = dr_ex[mask_ex].reset_index(drop=True)
            if machine not in recommended_machines and machine not in sonota_exclude and not filt_ex.empty:
                other_dfs.append(filt_ex)
                other_diffs.append(dr_f_ex)
                for _i in range(len(filt_ex)):
                    excellent_list.append({"name": machine, "diff": int(dr_f_ex.iloc[_i]), "ban": int(filt_ex.iloc[_i]["台番"])})

    if overflow_df is not None and overflow_diff is not None and not overflow_df.empty:
        # ジャグラー overflow を excellent pool へ（sonota_exclude機種は除外）
        if sonota_exclude and "機種名" in overflow_df.columns:
            _ov_keep      = ~overflow_df["機種名"].isin(sonota_exclude)
            overflow_df   = overflow_df[_ov_keep].reset_index(drop=True)
            overflow_diff = overflow_diff[_ov_keep.values].reset_index(drop=True)
        if not overflow_df.empty:
            for i, row in overflow_df.iterrows():
                excellent_list.append({"name": str(row["機種名"]), "diff": int(overflow_diff.iloc[i]), "ban": int(row["台番"])})
            other_dfs.append(overflow_df)
            other_diffs.append(overflow_diff)

    if not other_dfs:
        log("  その他の優秀台: 該当台なし")
        return generated, high_ratio_list, excellent_list

    combined    = pd.concat(other_dfs,   ignore_index=True)
    dr_combined = pd.concat(other_diffs, ignore_index=True)
    order       = combined["台番"].argsort()
    combined    = combined.iloc[order].reset_index(drop=True)
    dr_combined = dr_combined.iloc[order].reset_index(drop=True)

    img = _build_machine_img_no_bar(combined) if article_mode else _build_machine_img(combined, "その他の優秀台ピックアップ", None)
    out = os.path.join(output_dir, "その他の優秀台ピックアップ.jpg")
    _save_jpeg(img, out, target_kb=800)
    generated.append(out)
    log(f"  その他の優秀台ピックアップ: {len(combined)}台")
    return generated, high_ratio_list, excellent_list



_CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩"


def expand_machine_numbers(text: str) -> list[int]:
    """"2027-2029　2045-2048" や "2024.2029.2037..." から台番リストを返す。"""
    text = re.sub(r'[\.。…]+$', '', text.strip())
    tokens = re.findall(r'\d{3,4}(?:[-－]\d{3,4})?', text)
    result = []
    for token in tokens:
        m = re.match(r'(\d{3,4})[-－](\d{3,4})', token)
        if m:
            result.extend(range(int(m.group(1)), int(m.group(2)) + 1))
        else:
            result.append(int(token))
    return result


def _parse_section_line(text: str) -> dict:
    """"①塊（ブドウ図柄沢山）" → {label, comment, numbers}"""
    comment_m = re.search(r'[（(](.+?)[）)]', text)
    comment = comment_m.group(1) if comment_m else ""
    label = re.sub(r'[（(].+?[）)]', '', text).strip()
    return {"label": label, "comment": comment, "numbers": []}


def parse_result_memo(raw_text: str) -> list[dict]:
    """素材メモを機種ブロックのリストに変換する。
    「・機種名」行はオプション。なければ machine="" の単一ブロックとして扱う。
    Returns: [{"machine": str, "sections": [{"label": str, "comment": str, "numbers": [int]}]}]"""
    blocks: list[dict] = []
    current_machine: str = ""
    current_sections: list[dict] = []
    current_section: dict | None = None
    has_machine_line = any(
        line.strip() and line.strip()[0] in ('・', '•', '●', '･')
        for line in raw_text.splitlines()
    )

    def _flush_machine():
        nonlocal current_section
        secs = list(current_sections)
        if current_section is not None:
            secs.append(current_section)
            current_section = None
        if secs:
            blocks.append({"machine": current_machine, "sections": secs})

    for raw_line in raw_text.strip().splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if line[0] in ('・', '•', '●', '･'):
            _flush_machine()
            current_sections.clear()
            current_section = None
            rest = line[1:].strip()
            sec_pos = next((i for i, c in enumerate(rest) if c in _CIRCLED), -1)
            if sec_pos != -1:
                current_machine = rest[:sec_pos].strip()
                current_section = _parse_section_line(rest[sec_pos:])
            else:
                current_machine = rest

        elif line[0] in _CIRCLED:
            if current_section is not None:
                current_sections.append(current_section)
            current_section = _parse_section_line(line)

        elif re.search(r'\d{3,4}', line):
            nums = expand_machine_numbers(line)
            if nums:
                if current_section is None:
                    current_section = {"label": "", "comment": "", "numbers": []}
                current_section["numbers"].extend(nums)

    _flush_machine()
    return blocks


def _ban_to_diff(df, diff_raw, num: int):
    """台番 → (差枚int | None)"""
    row = df[df["台番"].apply(lambda x: int(x) if pd.notna(x) else -1) == num]
    if row.empty:
        return None
    return int(diff_raw.loc[row.index[0]])


def format_result_memo_sections(
    parsed_blocks: list[dict], df, diff_raw
) -> tuple[str, list[int]]:
    """パース済みブロックに差枚データを付与してテキスト化する。
    machine="" の場合は台番から機種名を自動検出して 🚨ヘッダーを付ける。
    Returns: (formatted_text, missing_ban_list)"""
    name_map, _ = load_name_map()
    lines: list[str] = []
    missing: list[int] = []

    for block in parsed_blocks:
        machine = block["machine"]
        if machine:
            short = name_map.get(machine, machine)
            lines.append(f"🚨{short}")
        else:
            # 最初の台番から機種名を自動検出
            first_nums = [n for sec in block["sections"] for n in sec["numbers"]]
            if first_nums:
                row = df[df["台番"].apply(lambda x: int(x) if pd.notna(x) else -1) == first_nums[0]]
                if not row.empty:
                    detected = str(row.iloc[0]["機種名"])
                    short = name_map.get(detected, detected)
                    lines.append(f"🚨{short}")

        for sec in block["sections"]:
            label = sec["label"]
            comment = sec["comment"]
            if comment:
                lines.append(f"{label}({comment})")
            elif label:
                lines.append(label)
            for num in sec["numbers"]:
                val = _ban_to_diff(df, diff_raw, num)
                if val is None:
                    missing.append(num)
                    lines.append(f"【{num}番台】（データなし）")
                else:
                    sign = "+" if val >= 0 else ""
                    lines.append(f"【{num}番台】{sign}{val:,}枚")
            lines.append("")  # セクション間の空行

    return "\n".join(lines).rstrip(), missing


def insert_formatted_result_before_other_picks(
    result_text: str, formatted_text: str, store_name: str = ""
) -> str:
    """formatted_text を「{e2}その他の優秀台」直前に挿入して返す。
    新小岩は「{e2}全台系濃厚機種」の前に挿入する。"""
    if not formatted_text.strip():
        return result_text
    e2 = STORE_EMOJI_CONFIG.get(store_name, ("💫", "👑"))[1]
    if store_name == "新小岩":
        marker = f"{e2}全台系濃厚機種"
    else:
        marker = f"{e2}その他の優秀台"
    idx = result_text.find(marker)
    if idx == -1:
        return result_text + "\n\n" + formatted_text
    return result_text[:idx] + formatted_text + "\n\n" + result_text[idx:]


_REC_BLOCK_EMOJIS = ["🌺", "✨", "⭐", "🎯"]


def generate_recommended_result_text(
    recommended_blocks: list[dict], df, diff_raw,
    exclude_machines: set | None = None,
    store_name: str = "",
) -> str:
    """オススメ機種ブロックから +1,000枚以上の台番をピックアップしたテキストを生成する。
    各ブロックが「{section_emoji}{title}の優秀台」セクションになる。"""
    name_map, _ = load_name_map()
    sections: list[str] = []
    _exclude = exclude_machines or set()

    _rec_cfg      = STORE_REC_CONFIG.get(store_name, {})
    _sec_emoji    = _rec_cfg.get("section_emoji", "🍯")
    _blk_emojis   = _rec_cfg.get("block_emojis", _REC_BLOCK_EMOJIS)

    for i, block in enumerate(recommended_blocks):
        title    = block.get("title", "").strip()
        machines = [m.strip() for m in block.get("machines", []) if m.strip()]
        if not title or not machines:
            continue

        emoji = _blk_emojis[i] if i < len(_blk_emojis) else "🎯"
        machine_parts: list[str] = []

        for machine in machines:
            if machine in _exclude:
                continue
            grp = df[df["機種名"] == machine]
            if grp.empty:
                continue
            dr_m = diff_raw.loc[grp.index]
            good = grp[dr_m >= 1000].copy()
            if good.empty:
                continue
            short = name_map.get(machine, machine)
            lines = [f"{emoji}{short}"]
            for idx in good.sort_values("台番").index:
                ban = int(good.loc[idx, "台番"])
                val = int(diff_raw.loc[idx])
                sign = "+" if val >= 0 else ""
                lines.append(f"【{ban}番台】{sign}{val:,}枚")
            machine_parts.append("\n".join(lines))

        if not machine_parts:
            continue
        clean_title = title[:-len("の優秀台")] if title.endswith("の優秀台") else title
        _block_header_names = _rec_cfg.get("block_header_names", {})
        if i in _block_header_names:
            header_line = f"{_sec_emoji}{clean_title}({'・'.join(_block_header_names[i])})"
        else:
            header_line = f"{_sec_emoji}{clean_title}の優秀台"
        sections.append(header_line + "\n" + "\n\n".join(machine_parts))

    return "\n\n".join(sections)


def _fmt_diff(n: int) -> str:
    if n == 0:
        return "±0枚"
    sign = "+" if n > 0 else ""
    return f"{sign}{n:,}枚"


def _format_diffs(diffs: list[int], wrap: int = 3) -> str:
    """差枚リストを重複×N統合・wrap台ごと折り返しフォーマット"""
    if not diffs:
        return ""
    tokens: list[str] = []
    i = 0
    while i < len(diffs):
        n = 1
        while i + n < len(diffs) and diffs[i + n] == diffs[i]:
            n += 1
        tokens.append(f"{_fmt_diff(diffs[i])}×{n}" if n > 1 else _fmt_diff(diffs[i]))
        i += n
    lines = []
    for j in range(0, len(tokens), wrap):
        chunk = tokens[j:j + wrap]
        suffix = "、" if j + wrap < len(tokens) else ""
        lines.append("、".join(chunk) + suffix)
    return "\n".join(lines)


def generate_report_text(
    store_name: str,
    date,
    zen_dai_list: list[dict],
    high_ratio_list: list[dict],
    nami_list: list[dict],
    excellent_list: list[dict],
    diff_raw=None,
    df=None,
    suebangai_data: list[dict] | None = None,
    jug_sue_data: list[dict] | None = None,
) -> str:
    """画像生成で使ったデータをそのまま文章化して返す"""
    weekday_jp = ["月", "火", "水", "木", "金", "土", "日"]
    if date is not None:
        date_str = f"{date.month}/{date.day}({weekday_jp[date.weekday()]})"
    else:
        date_str = ""

    def _diff_emoji(diffs: list[int]) -> str:
        return "🌋" if diffs and max(diffs) >= 4000 else "💎"

    # 西武新宿のみ: 高配分の(1/2台)機種は高配分に載せず、+2,000枚以上の台をその他の優秀台へ回す
    _demoted_names: set[str] = set()
    if store_name == "西武新宿":
        _demoted_names = {
            item["name"] for item in high_ratio_list
            if item.get("count") == 1 and item.get("total") == 2
        }

    def zen_dai_section() -> str:
        if not zen_dai_list:
            return "（なし）"
        lines = []
        sorted_list = sorted(
            zen_dai_list,
            key=lambda x: x.get("all_avg_diff", 0),
            reverse=True,
        )
        for item in sorted_list:
            avg = item.get("all_avg_diff", 0)
            avg_str = f"→平均{_fmt_diff(avg)}" if avg > 800 else ""
            lines.append(f"🎖️{item['name']}({item['count']}/{item['total']}台){avg_str}")
            if item["diffs"]:
                emoji = _diff_emoji(item["diffs"])
                lines.append(f"{emoji}{_format_diffs(item['diffs'])}")
        return "\n".join(lines)

    def high_ratio_section() -> str:
        _hr_list = [it for it in high_ratio_list if it["name"] not in _demoted_names]
        if not _hr_list:
            return "（なし）"
        lines = []
        sorted_list = sorted(
            _hr_list,
            key=lambda x: x["all_avg_diff"] if "all_avg_diff" in x
                          else (int(round(sum(x["diffs"]) / len(x["diffs"]))) if x["diffs"] else 0),
            reverse=True,
        )
        for item in sorted_list:
            if "all_avg_diff" in item:
                avg = item["all_avg_diff"]
            else:
                avg = int(round(sum(item["diffs"]) / len(item["diffs"]))) if item["diffs"] else 0
            avg_str = f"→平均{_fmt_diff(avg)}" if avg > 800 else ""
            lines.append(f"🎖️{item['name']}({item['count']}/{item['total']}台){avg_str}")
            if item["diffs"]:
                emoji = _diff_emoji(item["diffs"])
                lines.append(f"{emoji}{_format_diffs(item['diffs'])}")
        return "\n".join(lines)

    def nami_section() -> str:
        if not nami_list:
            return "（なし）"
        if any("machine" in item and "ban_range" in item for item in nami_list):
            machine_order = []
            grouped: dict[str, list] = {}
            for item in nami_list:
                m = item.get("machine") or item["title"].split("(")[0]
                if m not in grouped:
                    machine_order.append(m)
                    grouped[m] = []
                grouped[m].append(item)
            lines = []
            for m in machine_order:
                lines.append(f"🍡{m}")
                for item in grouped[m]:
                    br = item.get("ban_range", "")
                    n = item["count"]
                    avg = _fmt_diff(item["avg_diff"])
                    lines.append(f"{br}番台({n}台並び)→平均{avg}" if br else f"({n}台並び)→平均{avg}")
            return "\n".join(lines)
        lines = [f"🍡{item['title']}→平均{_fmt_diff(item['avg_diff'])}" for item in nami_list]
        return "\n".join(lines)

    def suebangai_section() -> str:
        _circle_map = {"0":"⓪","1":"①","2":"②","3":"③","4":"④",
                       "5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
        lines = []
        if suebangai_data:
            lines.append("👑優秀末尾")
            for _item in suebangai_data:
                _t = _item["tail"]
                _label = "末尾ゾロ目の台" if _t == "ゾロ目" else f"末尾{_circle_map.get(_t, _t)}番台"
                lines.append(f"🎁{_label}({_item['win_count']}/{_item['total']}台)→平均{_fmt_diff(_item['avg_diff'])}")
        if jug_sue_data:
            lines.append("👑ジャグラーの優秀末尾")
            for _item in jug_sue_data:
                _t = _item["tail"]
                _label = "末尾ゾロ目番台" if _t == "ゾロ目" else f"末尾{_circle_map.get(_t, _t)}番台"
                lines.append(f"🎁{_label}({_item['win_count']}/{_item['total']}台)→平均{_fmt_diff(_item['avg_diff'])}")
        return "\n".join(lines)

    def excellent_section() -> str:
        _item_emoji = STORE_REC_CONFIG.get(store_name, {}).get("item_emoji", "🚩")
        # 降格した(1/2台)機種は除外対象から外し、その+2,000枚以上の台をその他の優秀台に含める
        high_ratio_names = {item["name"] for item in high_ratio_list} - _demoted_names
        _poster_ex = get_store_config(store_name).get("poster_extra_exclude", set())
        _ex_src = list(excellent_list)
        if _demoted_names and df is not None and diff_raw is not None:
            _seen = {(x["name"], x.get("ban")) for x in _ex_src}
            for _nm in _demoted_names:
                _sub = df[df["機種名"] == _nm]
                _dr = diff_raw.loc[_sub.index]
                for _i in _sub.index:
                    _d = int(_dr.loc[_i])
                    _ban = int(_sub.loc[_i, "台番"])
                    if _d >= 2000 and (_nm, _ban) not in _seen:
                        _ex_src.append({"name": _nm, "ban": _ban, "diff": _d})
                        _seen.add((_nm, _ban))
        filtered = [x for x in _ex_src if x["diff"] >= 2000 and x["name"] not in high_ratio_names and x["name"] not in _poster_ex]
        if not filtered:
            return "（なし）"
        sorted_items = sorted(filtered, key=lambda x: x["diff"], reverse=True)
        lines = [
            f"{_item_emoji}【{item['ban']}番台】{item['name']}→{_fmt_diff(item['diff'])}"
            if item.get("ban") is not None
            else f"{_item_emoji}{item['name']}→{_fmt_diff(item['diff'])}"
            for item in sorted_items
        ]
        return "\n".join(lines)

    def shibuyashinkan_poster_section() -> str:
        cfg_s = get_store_config(store_name)
        jug_names = cfg_s.get("juggler_series", set())

        def find_high(name):
            return next((x for x in high_ratio_list if x["name"] == name), None)

        def nami_for(name):
            return [x for x in nami_list if x.get("machine") == name]

        def nami_detail_lines(name):
            res = []
            for item in nami_for(name):
                br = item.get("ban_range", "")
                n = item["count"]
                avg = _fmt_diff(item["avg_diff"])
                res.append(f"{br}番台({n}台並び)→平均{avg}" if br else f"({n}台並び)→平均{avg}")
            return res

        def extract_bans_1k(machine_name):
            if df is None or diff_raw is None:
                return []
            sub = df[df["機種名"] == machine_name]
            if sub.empty:
                return []
            dr_sub = diff_raw.loc[sub.index]
            mask = dr_sub >= 1000
            sub_f = sub[mask]
            dr_f = dr_sub[mask]
            if sub_f.empty:
                return []
            order = dr_f.sort_values(ascending=False).index
            return [(int(sub_f.loc[i, "台番"]), int(dr_f.loc[i])) for i in order]

        def _checked_pins(table_num: int) -> list[str]:
            """同日の⑤表入力でチェック/選択が入った項目のみ返す。日付/start_date不明・未設定時は全件返す。
            t3: multiselect UIのためcell_machinesで活性判定（選択あり=活性）
            t1/t2: checkbox UIのためchecksで活性判定
            """
            all_items = _load_weekly_items(store_name, table_num)

            def _all_nonempty():
                return [it.strip() for it in all_items if (it or "").strip()]

            if date is None:
                return _all_nonempty()
            start_str = _load_weekly_start_date(store_name, table_num)
            if not start_str:
                return _all_nonempty()
            try:
                start = datetime.date.fromisoformat(start_str)
                _d = date.date() if hasattr(date, "date") else date
                col = (_d - start).days
            except Exception:
                return _all_nonempty()
            if not (0 <= col <= 6):
                return _all_nonempty()

            tdata = _weekly_table_data(store_name, table_num)
            if table_num == 3:
                # t3: multiselect UI → cell_machinesで判定
                cm = tdata.get("cell_machines", {})
                if not cm:
                    return _all_nonempty()
                result = []
                for i, it in enumerate(all_items):
                    if not (it or "").strip():
                        continue
                    if cm.get(f"{i},{col}", []):
                        result.append(it.strip())
                return result
            else:
                # t1/t2: checkbox UI → checksで判定
                raw_checks = tdata.get("checks", [])
                if not raw_checks:
                    return _all_nonempty()
                all_checks = _load_weekly_checks(store_name, table_num)
                result = []
                for i, it in enumerate(all_items):
                    if not (it or "").strip():
                        continue
                    ck = all_checks[i][col] if i < len(all_checks) and col < len(all_checks[i]) else False
                    if ck:
                        result.append(it.strip())
                return result

        L = ["📈オススメポスター機種の仕掛け📈"]

        # ── スマスロ北斗の拳 ──
        _kita_pins = _checked_pins(1)
        KITA = "スマスロ北斗の拳"
        L.append(f"【{KITA}】")
        h_kita = find_high(KITA)
        for pin in _kita_pins:
            L.append(f"📌{pin}")
        if h_kita:
            L.append(f"🎖️{KITA}({h_kita['count']}/{h_kita['total']}台)")
        if nami_for(KITA):
            L.append(f"🍡{KITA}")
            L.extend(nami_detail_lines(KITA))
        L.append("")

        # ── 北斗転生2 ──
        KITA2 = "北斗転生2"
        L.append(f"【{KITA2}】")
        for ban, dv in extract_bans_1k(KITA2):
            L.append(f"🚩【{ban}番台】+{dv:,}枚")
        L.append("")

        # ── ジャグラーシリーズ ──
        _jug_pins = _checked_pins(3)
        L.append("【ジャグラーシリーズ】")
        for pin in _jug_pins:
            L.append(f"📌{pin}")
        for z in [x for x in zen_dai_list if x["name"] in jug_names]:
            L.append(f"🎖️{z['name']}({z['count']}/{z['total']}台)→平均{_fmt_diff(z.get('all_avg_diff', 0))}")
        for h in [x for x in high_ratio_list if x["name"] in jug_names]:
            L.append(f"🎖️{h['name']}({h['count']}/{h['total']}台)→平均{_fmt_diff(h.get('all_avg_diff', 0))}")
        jug_m_order = list(dict.fromkeys(x.get("machine", "") for x in nami_list if x.get("machine", "") in jug_names))
        for m in jug_m_order:
            L.append(f"🍡{m}")
            L.extend(nami_detail_lines(m))
        L.append("")

        # ── 東京喰種 ──
        _ghoul_pins = _checked_pins(2)
        GHOUL = "東京喰種"
        L.append(f"【{GHOUL}】")
        for pin in _ghoul_pins:
            L.append(f"📌{pin}")
        h_ghoul = find_high(GHOUL)
        if h_ghoul:
            L.append(f"🎖️{GHOUL}({h_ghoul['count']}/{h_ghoul['total']}台)")
            n1k = len(h_ghoul.get("diffs", []))
            L.append(f"🎖️{GHOUL}→{n1k}台が+1,000枚以上")
        L.extend(nami_detail_lines(GHOUL))
        L.append("")

        # ── カバネリ海門決戦 ──
        KABA = "カバネリ海門決戦"
        L.append(f"【{KABA}】")
        for ban, dv in extract_bans_1k(KABA):
            L.append(f"🚩【{ban}番台】+{dv:,}枚")
        L.append("")

        # ── ヴァルヴレイヴ2 ──
        VALV = "ヴァルヴレイヴ2"
        L.append(f"【{VALV}】")
        L.append("📌リーゼロッテ(全台系!?)")
        L.append("📌ショーコとサキ/二人(2台並び!?)")
        L.append("📌VVV並び(3台並び!?)")
        L.append("📌革命分岐(高配分◎!?)")
        valv_2 = [x for x in nami_list if x.get("machine") == VALV and x["count"] == 2]
        for item in valv_2:
            br = item.get("ban_range", ""); n = item["count"]; avg = _fmt_diff(item["avg_diff"])
            L.append(f"{br}番台({n}台並び)→平均{avg}" if br else f"({n}台並び)→平均{avg}")
        valv_3 = [x for x in nami_list if x.get("machine") == VALV and x["count"] >= 3]
        for item in valv_3:
            br = item.get("ban_range", ""); n = item["count"]; avg = _fmt_diff(item["avg_diff"])
            L.append(f"{br}番台({n}台並び)→平均{avg}" if br else f"({n}台並び)→平均{avg}")
        L.append("")

        # ── 3F週間オススメポスター ──
        L.append("【3F週間オススメポスター】")
        L.append("")

        # ── BT機 ──
        L.append("【BT機】")

        return "\n".join(L)

    def summary_section() -> str:
        if diff_raw is None:
            return ""
        diffs = [int(d) for d in diff_raw.dropna().tolist()]
        if not diffs:
            return ""
        total = sum(diffs)
        avg = int(round(total / len(diffs)))

        def _s(n: int) -> str:
            sign = "+" if n >= 0 else "-"
            return f"{sign}{abs(n):,}枚"

        c10k = sum(1 for d in diffs if d >= 10000)
        c5k  = sum(1 for d in diffs if d >= 5000)
        c1k  = sum(1 for d in diffs if d >= 1000)
        date_label = f"{date.month}/{date.day}" if date else ""
        lines = [
            f"📈{date_label}の結果📈",
            f"🏆総差枚：{_s(total)}",
            f"🏆平均差枚：{_s(avg)}",
            "",
            f"🌋万枚オーバーが{c10k}台！",
            f"💥+5,000枚オーバーが{c5k}台！",
            f"💎{c1k}台が+1,000枚オーバー！",
        ]
        return "\n".join(lines)

    def juggler_summary_section() -> str:
        if df is None:
            return ""
        try:
            jug_series = set(get_store_config(store_name)["juggler_series"])
        except Exception:
            return ""
        jug = df[df["機種名"].isin(jug_series)]
        if jug.empty:
            return ""
        total_diff = int(jug["差枚"].sum())
        avg_diff   = int(round(jug["差枚"].mean()))
        over_1k    = int((jug["差枚"] >= 1000).sum())
        def _s2(n: int) -> str:
            sign = "+" if n >= 0 else "-"
            return f"{sign}{abs(n):,}枚"
        return "\n".join([
            "🤡本日のジャグラー全体の結果は",
            f"総差枚{_s2(total_diff)}、平均差枚{_s2(avg_diff)}！",
            f"💎{over_1k}台が+1,000枚オーバー！",
        ])

    header = f"{date_str}📝分析結果📝" if date_str else "📝分析結果📝"
    e1, e2 = STORE_EMOJI_CONFIG.get(store_name, ("💫", "👑"))
    _STORE_DISPLAY_NAMES: dict[str, str] = {
        "西武新宿": "エスパス 西武 新宿",
    }
    store_display = _STORE_DISPLAY_NAMES.get(store_name, f"エスパス{store_name}")
    parts = [
        header,
        store_display,
        "",
        f"{e1}本日は注目ポイント多数！",
        f"{e1}全台系、並び、優秀台を確認！",
        "",
    ]
    if store_name == "渋谷新館":
        parts += [
            "📮ななこポスト考察📮",
            "📍",
            "📍",
            "📍",
            "📍",
            "📍",
            "📍",
            "📍",
            "✅渋谷ななこのポストからは連日ヒントを確認🧐",
            "",
        ]
    parts += [
        f"{e2}全台系濃厚機種",
        zen_dai_section(),
        "",
        f"{e2}高配分機種",
        high_ratio_section(),
        "",
        f"{e2}並び仕掛け",
        nami_section(),
        "",
    ]
    _sue_sec = suebangai_section()
    if _sue_sec:
        parts += [_sue_sec, ""]
    if store_name == "渋谷新館":
        parts.append(shibuyashinkan_poster_section())
        parts.append("")
    parts += [
        f"{e2}その他の優秀台",
        excellent_section(),
    ]
    _summary = summary_section()
    if _summary:
        parts += ["", _summary]
    _jug_sum = juggler_summary_section()
    if _jug_sum:
        parts += ["", _jug_sum]
    return "\n".join(parts)


def run_auto_pipeline(
    excel_path: str,
    output_dir: str,
    store: str,
    narabi_bans: set[int],
    log,
    narabi_ranges: list | None = None,
    recommended_machines: set[str] = set(),
    suebangai_tails: list[str] = [],
    article_mode: bool = False,
    sonota_exclude: set[str] = frozenset(),
    jug_suebangai_tails: list[str] = [],
    variety_bans: set[int] = set(),
) -> dict:
    """3ステップパイプラインを実行する。
    戻り値: {"ok": bool, "files": list[str], "error": str | None,
             "zen_dai_list", "high_ratio_list", "nami_list", "excellent_list", "date"}"""
    try:
        cfg  = get_store_config(store)
        stem = os.path.splitext(os.path.basename(excel_path))[0]
        os.makedirs(output_dir, exist_ok=True)

        log("データ読み込み中…")
        df, diff_raw = _load_pipeline_df(excel_path)

        # 差枚補正（補正前の生値を nami_list 用に保持）
        diff_raw_original = diff_raw.copy()
        if "差枚" in df.columns:
            df["差枚"] = df["差枚"].apply(_pipeline_calc_d)
            diff_raw   = df["差枚"].copy()

        # 末尾台番セット（その他の優秀��ピックアップから除外）
        suebangai_bans: set[int] = set()
        for _sue_t in suebangai_tails:
            _st = _sue_t.strip()
            if _st == "ゾロ目":
                suebangai_bans |= {
                    int(b) for b in df["台番"]
                    if (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                }
            elif _st.isdigit() and len(_st) in (1, 2):
                suebangai_bans |= {
                    int(b) for b in df["台番"]
                    if str(int(b))[-len(_st):] == _st
                }
        # ジャグラー末尾台番セット（Step2のジャグラー優秀台から除外・Step3には影響しない）
        jug_sue_bans: set[int] = set()
        for _jt in jug_suebangai_tails:
            _jts = _jt.strip()
            if _jts == "ゾロ目":
                jug_sue_bans |= {
                    int(b) for b in df["台番"]
                    if (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                }
            elif _jts.isdigit() and len(_jts) in (1, 2):
                jug_sue_bans |= {
                    int(b) for b in df["台番"]
                    if str(int(b))[-len(_jts):] == _jts
                }
        # バラエティ台番をsonota除外セットに追加（run_step3_otherの_ex_bansに含める）
        suebangai_bans |= variety_bans

        log("① 全台系PNG ＋ 全台プラス機種別JPG")
        f1, zen_dai_list = run_step1_main(df, diff_raw, output_dir, stem, cfg, log, article_mode=article_mode)

        log("② ジャグラーシリーズ優秀台")
        _jug_series = cfg["juggler_series"]
        _zen_dai_jug = {item["name"] for item in zen_dai_list if item["name"] in _jug_series}
        f2, ov_df, ov_diff, jug_hr, jug_excellent, jug_pool_df = run_step2_juggler(df, diff_raw, output_dir, cfg, narabi_bans, log, recommended_machines, suebangai_bans | jug_sue_bans, zen_dai_juggler_machines=_zen_dai_jug, article_mode=article_mode, sonota_exclude=sonota_exclude)

        log("③ その他の優秀台ピックアップ")
        f3, oth_hr, sonota_excellent = run_step3_other(df, diff_raw, output_dir, cfg, narabi_bans, ov_df, ov_diff, log, recommended_machines, suebangai_bans, article_mode=article_mode, sonota_exclude=sonota_exclude)
        _ex_seen: set[tuple] = set()
        excellent_list = []
        for _ex_item in jug_excellent + sonota_excellent:
            _ex_key = (_ex_item["name"], _ex_item["ban"])
            if _ex_key not in _ex_seen:
                _ex_seen.add(_ex_key)
                excellent_list.append(_ex_item)

        # 並び(nami)データ: narabi_ranges が指定されている場合に差枚を集計
        nami_list: list[dict] = []
        if narabi_ranges:
            for ban_list in narabi_ranges:
                bans = set(ban_list)
                sub = df[df["台番"].isin(bans)]
                if sub.empty:
                    continue
                dr_sub = diff_raw_original.loc[sub.index]
                machines = list(dict.fromkeys(str(m) for m in sub["機種名"].tolist()))
                if len(machines) == 1:
                    machine_label = machines[0]
                elif len(machines) == 2:
                    machine_label = f"{machines[0]}+{machines[1]}"
                else:
                    machine_label = f"{machines[0]}～{machines[-1]}"
                title = f"{machine_label}({len(sub)}台並び)"
                ban_list_s = sorted(ban_list)
                _is_consec = len(ban_list_s) >= 2 and (ban_list_s[-1] - ban_list_s[0] + 1 == len(ban_list_s))
                if _is_consec:
                    ban_range = f"{ban_list_s[0]}-{ban_list_s[-1]}"
                elif len(ban_list_s) == 1:
                    ban_range = str(ban_list_s[0])
                else:
                    ban_range = "+".join(str(b) for b in ban_list_s)
                nami_list.append({
                    "title":    title,
                    "count":    len(sub),
                    "avg_diff": int(round(dr_sub.mean())),
                    "machine":  machine_label,
                    "ban_range": ban_range,
                    "bans":     ban_list_s,
                })

        # 日付をファイル名から取得
        date_obj = None
        parts = stem.split("_")
        if parts and len(parts[0]) == 8 and parts[0].isdigit():
            d = parts[0]
            try:
                import datetime as _dt
                date_obj = _dt.date(int(d[:4]), int(d[4:6]), int(d[6:8]))
            except ValueError:
                pass

        return {
            "ok":             True,
            "files":          f1 + f2 + f3,
            "error":          None,
            "zen_dai_list":   zen_dai_list,
            "high_ratio_list": jug_hr + oth_hr,
            "nami_list":      nami_list,
            "excellent_list":        excellent_list,
            "sonota_excellent_list": sonota_excellent,
            "jug_excellent_list":    jug_excellent,
            "jug_pool_df":           jug_pool_df,
            "jug_overflow_df":       ov_df,
            "date":                  date_obj,
            "df":             df,
            "diff_raw":       diff_raw,
        }
    except Exception:
        return {"ok": False, "files": [], "error": traceback.format_exc()}


def parse_ranges(text: str) -> list[list[int]]:
    """台番範囲パーサー。

    【テーブル形式】「台番 TAB 機種名 数値…」形式の行が2行以上ある場合:
      空行をグループ区切りとして台番（第1列）だけを抽出する。
      例: Excelからコピペしたデータをそのまま貼り付け可能。

    【範囲形式】上記以外:
      区切り: カンマ(,/、)・スペース(半角/全角)・改行
      '409-413' : 連番展開（- / ~ / ～）
      '508+424' : スポット指定（非連続並び）
      '316'     : 単番
      非数値テキスト（機種名など）は読み飛ばす
    """
    # テーブル形式判定: 「数字 + 空白 + 非数字」で始まる行が2行以上
    _table_lines = [l for l in text.splitlines() if re.match(r"^\d+[\t ]+\D", l.strip())]
    if len(_table_lines) >= 2:
        # 台番を出現順に収集（重複除外）。テーブル形式でない行は範囲形式として後処理。
        bans_ordered: list[int] = []
        seen_bans: set[int] = set()
        non_table_parts: list[str] = []
        for line in text.splitlines():
            m = re.match(r"^(\d+)[\t ]+\D", line.strip())
            if m:
                ban = int(m.group(1))
                if ban not in seen_bans:
                    bans_ordered.append(ban)
                    seen_bans.add(ban)
            else:
                stripped = line.strip()
                if stripped:
                    non_table_parts.append(stripped)
        # 連番グループ化: 台番[i+1] = 台番[i]+1 なら同グループ、2台以上のみnarabi
        result: list[list[int]] = []
        if bans_ordered:
            grp = [bans_ordered[0]]
            for ban in bans_ordered[1:]:
                if ban == grp[-1] + 1:
                    grp.append(ban)
                else:
                    if len(grp) >= 2:
                        result.append(grp)
                    grp = [ban]
            if len(grp) >= 2:
                result.append(grp)
        # テーブル行でない行（"440-441+511" など）を範囲形式でパース
        for part in re.split(r"[,、\s　]+", " ".join(non_table_parts)):
            part = part.strip().lstrip("・")
            part = re.sub(r"[番号台]+$", "", part)  # 末尾の「番」「号」「台」を除去
            part = re.sub(r"[（(][^）)]*[）)]", "", part).strip()  # 末尾の（機種名）等を除去
            # 先頭の■機種名：等を除去（残余が3桁以上の数字で始まる場合のみ適用）
            _s = re.sub(r"^[^\d]+", "", part)
            if _s and re.match(r"\d{3}", _s):
                part = _s
            if not part:
                continue
            m2 = re.match(r"(\d+)\s*[-~～]\s*(\d+)$", part)
            if m2:
                s2, e2 = int(m2.group(1)), int(m2.group(2))
                if s2 <= e2:
                    result.append(list(range(s2, e2 + 1)))
                continue
            plus_parts = re.split(r"\+", part)
            if len(plus_parts) >= 2:
                bans = []
                for p in plus_parts:
                    p = p.strip()
                    m3 = re.match(r"(\d+)\s*[-~～]\s*(\d+)$", p)
                    if m3:
                        bans.extend(range(int(m3.group(1)), int(m3.group(2)) + 1))
                    elif p.isdigit():
                        bans.append(int(p))
                if bans:
                    result.append(bans)
                continue
            if part.isdigit():
                result.append([int(part)])
        return result

    # 範囲形式
    result = []
    for part in re.split(r"[,、\s　]+", text):
        part = part.strip().lstrip("・")
        part = re.sub(r"[番号台]+$", "", part)  # 末尾の「番」「号」「台」を除去
        part = re.sub(r"[（(][^）)]*[）)]", "", part).strip()  # 末尾の（機種名）等を除去
        # 先頭の■機種名：等を除去（残余が3桁以上の数字で始まる場合のみ適用）
        _s = re.sub(r"^[^\d]+", "", part)
        if _s and re.match(r"\d{3}", _s):
            part = _s
        if not part:
            continue
        m = re.match(r"(\d+)\s*[-~～]\s*(\d+)$", part)
        if m:
            s, e = int(m.group(1)), int(m.group(2))
            if s <= e:
                result.append(list(range(s, e + 1)))
            continue
        plus_parts = re.split(r"\+", part)
        if len(plus_parts) >= 2:
            bans = []
            for p in plus_parts:
                p = p.strip()
                m2 = re.match(r"(\d+)\s*[-~～]\s*(\d+)$", p)
                if m2:
                    bans.extend(range(int(m2.group(1)), int(m2.group(2)) + 1))
                elif p.isdigit():
                    bans.append(int(p))
            if bans:
                result.append(bans)
            continue
        if part.isdigit():
            result.append([int(part)])
    return result


def ranges_to_bans(ranges: list[list[int]]) -> set[int]:
    """台番リストのリストを台番セットに展開する"""
    bans: set[int] = set()
    for ban_list in ranges:
        bans.update(ban_list)
    return bans


def _patch_and_run_narabi(
    script_path: str, input_path: str, split_dir: str, ranges: list
) -> tuple[bool, str, str]:
    """並びスクリプト専用: INPUT/SPLIT_DIR/RANGES を書き換えて実行する"""
    with open(script_path, encoding="utf-8") as f:
        code = f.read()

    for var, val in [("INPUT", input_path), ("SPLIT_DIR", split_dir)]:
        code = re.sub(
            rf'^{re.escape(var)}\s*=\s*r?"[^"]*"',
            lambda m, _var=var, _val=val: f'{_var} = r"{_val}"',
            code, flags=re.MULTILINE,
        )

    # 機種名変換.xlsxのパスをBASE_DIRベースに書き換え（Cloud対応）
    _name_map_path = os.path.join(BASE_DIR, "機種名変換.xlsx").replace("\\", "/")
    code = re.sub(
        r'r?"[^"]*機種名変換\.xlsx"',
        f'r"{_name_map_path}"',
        code,
    )

    ranges_str = repr(ranges)
    m_ranges = re.search(r'^RANGES\s*=\s*', code, re.MULTILINE)
    if m_ranges:
        bracket_pos = code.index('[', m_ranges.end())
        depth, end_pos = 0, bracket_pos
        for i, c in enumerate(code[bracket_pos:]):
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    end_pos = bracket_pos + i + 1
                    break
        code = code[:m_ranges.start()] + f'RANGES = {ranges_str}\n' + code[end_pos:].lstrip('\n')

    # Cloud環境ではPlaywright不可のためimgkit(wkhtmltopdf)で代替
    if _IS_CLOUD:
        cloud_patch = (
            "try:\n"
            "    import imgkit as _imgkit, dataframe_image as _dfi_m\n"
            "    def _imgkit_export(obj, filename, *, fontsize=14, **kw):\n"
            "        _css = f'<style>body{{font-size:{fontsize}pt;font-family:sans-serif;}} td,th{{white-space:nowrap;}}</style>'\n"
            "        _imgkit.from_string(_css + obj.to_html(), filename,\n"
            "                            options={'quiet':'','zoom':'1.5','width':'1600'})\n"
            "    _dfi_m.export = _imgkit_export\n"
            "except Exception:\n"
            "    pass\n"
        )
        code = cloud_patch + code

    with tempfile.NamedTemporaryFile(
        suffix=".py", mode="w", encoding="utf-8", delete=False, dir=BASE_DIR
    ) as f:
        f.write(code)
        tmp_path = f.name

    try:
        result = subprocess.run(
            [sys.executable, tmp_path],
            capture_output=True, text=True, encoding="utf-8", errors="replace", cwd=BASE_DIR,
        )
        return result.returncode == 0, result.stdout, result.stderr
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# =============================================================================
# ■ 店舗別設定の永続化（オススメ機種ピックアップ用）
# =============================================================================

def load_store_settings(store: str) -> dict:
    """店舗固有設定を store_settings/<store>.json から読み込む。"""
    path = os.path.join(STORE_SETTINGS_DIR, f"{store}.json")
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_store_settings(store: str, data: dict) -> None:
    """店舗固有設定を store_settings/<store>.json に保存する。"""
    os.makedirs(STORE_SETTINGS_DIR, exist_ok=True)
    path = os.path.join(STORE_SETTINGS_DIR, f"{store}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _save_rec_titles(store: str) -> None:
    """タイトル・抽出条件変更時に即JSON保存。他の設定（機種名など）は上書きしない。"""
    prev = load_store_settings(store)
    for n in range(1, 7):
        tk = f"rec_title_{n}_{store}"
        fk = f"rec_f_{n}_{store}"
        if tk in st.session_state:
            prev[f"recommended_title_{n}"]  = st.session_state[tk]
        if fk in st.session_state:
            prev[f"recommended_filter_{n}"] = st.session_state[fk]
    save_store_settings(store, prev)


def _save_rec_enabled(store: str) -> None:
    """チェックボックス変更時に即JSON保存。"""
    prev = load_store_settings(store)
    prev["rec_enabled"] = bool(st.session_state.get(f"rec_enabled_{store}", False))
    save_store_settings(store, prev)


# ── 自動処理ページ入力値の永続化 ──────────────────────────────────────────────
_AUTO_INPUTS_JSON = os.path.join(BASE_DIR, "auto_page_inputs.json")
_AUTO_PERSISTENT_JSON = os.path.join(BASE_DIR, "auto_page_persistent_inputs.json")
_ARTICLE_INPUTS_JSON = os.path.join(BASE_DIR, "article_page_inputs.json")


def _persistent_keys(store: str) -> set[str]:
    """Excel切り替えをまたいで保持するキー（機種名・台番範囲など）。"""
    return {f"kojin_y_{i}_{store}" for i in range(8)} | {f"variety_range_{store}"}


def _load_persistent_json() -> dict:
    if os.path.exists(_AUTO_PERSISTENT_JSON):
        try:
            with open(_AUTO_PERSISTENT_JSON, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_persistent_inputs(store: str) -> None:
    """永続キーを別ファイルに保存（Excelファイル名に依存しない）。
    ウィジェット未描画（session_stateにない）のキーは既存値を維持する。"""
    pk   = _persistent_keys(store)
    data = _load_persistent_json()
    merged = dict(data.get(store, {}))  # 既存値をベースに
    for k in pk:
        if k in st.session_state and st.session_state[k]:
            merged[k] = st.session_state[k]  # 描画済み・非空 → 上書き
        # 未描画 or 空 → 既存値をそのまま維持（消さない）
    if merged:
        data[store] = merged
        try:
            with open(_AUTO_PERSISTENT_JSON, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass


_KOJIN_PICK_COUNT = 3

def _collect_kojin_pick(store: str, prefix: str = "") -> list[tuple[str, set[int]]]:
    """個別機種の優秀台ピックアップの (タイトル, 台番set) を返す（台番非空のみ）。"""
    out: list[tuple[str, set[int]]] = []
    for i in range(_KOJIN_PICK_COUNT):
        title = str(st.session_state.get(f"{prefix}kojin_pick_title_{i}_{store}", "")).strip()
        bans_text = str(st.session_state.get(f"{prefix}kojin_pick_bans_{i}_{store}", "")).strip()
        if not bans_text:
            continue
        bans = set(expand_machine_numbers(bans_text))
        if not bans:
            continue
        out.append((title or "個別機種の優秀台ピックアップ", bans))
    return out

def _kojin_pick_machines_from_df(picks: list[tuple[str, set[int]]], df) -> set[str]:
    """ピックアップ台番から df で機種を逆引きした集合。"""
    all_bans: set[int] = set()
    for _t, b in picks:
        all_bans |= b
    if not all_bans or df is None or df.empty:
        return set()
    _rows = df[df["台番"].apply(lambda b: int(b) in all_bans)]
    return {str(m).strip() for m in _rows["機種名"] if str(m).strip()}

def _kojin_pick_suppressed_machines(uploaded, store: str, prefix: str = "") -> set[str]:
    """アップロード Excel を読み、ピックアップ台番の機種名 set を返す（パイプライン前の抑制用）。"""
    picks = _collect_kojin_pick(store, prefix)
    if not picks:
        return set()
    try:
        _raw = _read_uploaded_df(uploaded)
        _df0, _ = normalize_df(_raw)
        _df0 = apply_name_conversion(_df0)
    except Exception:
        return set()
    finally:
        try:
            uploaded.seek(0)
        except Exception:
            pass
    return _kojin_pick_machines_from_df(picks, _df0)


def _auto_input_keys(store: str) -> list[str]:
    keys = ["kojin_enabled", "narabi_enabled", "narabi_ranges_input",
            "suebangai_enabled",
            "suebangai_tail_input_1", "suebangai_tail_input_2", "suebangai_tail_input_3",
            "suebangai_mode"]
    keys += ["jug_sue_enabled",
             "jug_sue_tail_input_1", "jug_sue_tail_input_2", "jug_sue_tail_input_3",
             "jug_sue_mode"]
    for i in range(12):
        keys += [f"kojin_z_{i}_{store}"]
    for i in range(21):
        keys += [f"kojin_y_{i}_{store}"]
    keys += [
        f"kojin_narabi_range_{store}", f"kojin_narabi_title_{store}",
        f"kojin_narabi2_range_{store}", f"kojin_narabi2_title_{store}",
        f"sonota_extra_title_{store}", f"sonota_extra_text_{store}",
        "variety_enabled", f"variety_range_{store}", "variety_mode",
    ]
    for i in range(_KOJIN_PICK_COUNT):
        keys += [f"kojin_pick_title_{i}_{store}", f"kojin_pick_bans_{i}_{store}"]
    return keys


def _load_auto_inputs_json() -> dict:
    if os.path.exists(_AUTO_INPUTS_JSON):
        try:
            with open(_AUTO_INPUTS_JSON, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_auto_inputs(store: str) -> None:
    """自動処理ページの入力値を Excelファイル名をキーにして JSON 保存。"""
    excel_name = st.session_state.get("auto_current_excel")
    if not excel_name:
        return
    data = _load_auto_inputs_json()
    data[excel_name] = {k: st.session_state[k] for k in _auto_input_keys(store) if k in st.session_state}
    try:
        with open(_AUTO_INPUTS_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    # 永続キー（機種名・台番範囲）を別ファイルにも保存
    _save_persistent_inputs(store)


def _restore_auto_inputs(excel_name: str, store: str) -> None:
    """JSON から入力値を復元して session_state に反映する。
    保存データにないキーはデフォルト値にリセット（ただし永続キーは別ファイルの値を使う）。
    永続キーは保存値が空の場合も永続値を優先する。"""
    saved      = _load_auto_inputs_json().get(excel_name, {})
    persistent = _load_persistent_json().get(store, {})
    pk         = _persistent_keys(store)
    for k in _auto_input_keys(store):
        if k not in saved:
            if k in pk and k in persistent:
                st.session_state[k] = persistent[k]  # 永続値を優先
            else:
                st.session_state[k] = False if k.endswith("_enabled") else ""
    for k, v in saved.items():
        # 永続キーは保存値が空でも永続値を維持（空で上書きしない）
        if k in pk and not v and k in persistent:
            st.session_state[k] = persistent[k]
        else:
            st.session_state[k] = v
    # 秋葉原は個別画像を毎回デフォルトONにする
    if store == "秋葉原":
        st.session_state["kojin_enabled"] = True


# ── 記事用ページ入力値の永続化 ────────────────────────────────────────────────

def _article_input_keys(store: str) -> list[str]:
    keys = ["art_kojin_enabled", "art_narabi_enabled", "art_narabi_ranges_input",
            "art_suebangai_enabled", "art_suebangai_tail_input"]
    for i in range(6):
        keys += [f"art_kojin_z_{i}_{store}", f"art_kojin_y_{i}_{store}"]
    keys += [
        f"art_kojin_narabi_range_{store}", f"art_kojin_narabi_title_{store}",
        f"art_kojin_narabi2_range_{store}", f"art_kojin_narabi2_title_{store}",
    ]
    for i in range(_KOJIN_PICK_COUNT):
        keys += [f"art_kojin_pick_title_{i}_{store}", f"art_kojin_pick_bans_{i}_{store}"]
    return keys


def _load_article_inputs_json() -> dict:
    if os.path.exists(_ARTICLE_INPUTS_JSON):
        try:
            with open(_ARTICLE_INPUTS_JSON, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_article_inputs(store: str) -> None:
    excel_name = st.session_state.get("art_current_excel")
    if not excel_name:
        return
    data = _load_article_inputs_json()
    data[excel_name] = {k: st.session_state[k] for k in _article_input_keys(store) if k in st.session_state}
    try:
        with open(_ARTICLE_INPUTS_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _restore_article_inputs(excel_name: str, store: str) -> None:
    saved = _load_article_inputs_json().get(excel_name, {})
    for k in _article_input_keys(store):
        if k not in saved:
            st.session_state[k] = False if k.endswith("_enabled") else ""
    for k, v in saved.items():
        st.session_state[k] = v


def _init_recommended_settings(store: str) -> None:
    """オススメ機種設定を session_state に初期化する（JSON → デフォルト値の順）。"""
    saved = load_store_settings(store)
    m1 = saved.get("recommended_machines_1", [""] * 5)
    m2 = saved.get("recommended_machines_2", [""] * 5)
    m3 = saved.get("recommended_machines_3", [""] * 5)
    m4 = saved.get("recommended_machines_4", [""] * 5)
    m5 = saved.get("recommended_machines_5", [""] * 5)
    m6 = saved.get("recommended_machines_6", [""] * 5)
    defaults: dict = {
        f"rec_enabled_{store}":  False,
        f"rec_title_1_{store}": saved.get("recommended_title_1", "月間オススメ機種"),
        f"rec_title_2_{store}": saved.get("recommended_title_2", "週間オススメ機種"),
        f"rec_title_3_{store}": saved.get("recommended_title_3", ""),
        f"rec_title_4_{store}": saved.get("recommended_title_4", ""),
        f"rec_title_5_{store}": saved.get("recommended_title_5", ""),
        f"rec_title_6_{store}": saved.get("recommended_title_6", ""),
        f"rec_f_1_{store}": saved.get("recommended_filter_1", "+1,000枚以上"),
        f"rec_f_2_{store}": saved.get("recommended_filter_2", "プラス台"),
        f"rec_f_3_{store}": saved.get("recommended_filter_3", "プラス台"),
        f"rec_f_4_{store}": saved.get("recommended_filter_4", "プラス台"),
        f"rec_f_5_{store}": saved.get("recommended_filter_5", "プラス台"),
        f"rec_f_6_{store}": saved.get("recommended_filter_6", "プラス台"),
        f"result_extra_note_{store}": "",
    }
    for i in range(9):
        defaults[f"rec_m1_{i}_{store}"] = m1[i] if i < len(m1) else ""
        defaults[f"rec_m2_{i}_{store}"] = m2[i] if i < len(m2) else ""
        defaults[f"rec_m3_{i}_{store}"] = m3[i] if i < len(m3) else ""
        defaults[f"rec_m4_{i}_{store}"] = m4[i] if i < len(m4) else ""
        defaults[f"rec_m5_{i}_{store}"] = m5[i] if i < len(m5) else ""
        defaults[f"rec_m6_{i}_{store}"] = m6[i] if i < len(m6) else ""
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def load_machine_candidates() -> list[str]:
    """変換マスタの短縮機種名一覧を返す（重複なし・ソート済み）。
    load_name_map() がキャッシュ済みなので2回目以降は高速。"""
    name_map, _ = load_name_map()
    return sorted(set(name_map.values()))


def suggest_machine_candidates(
    query: str, candidates: list[str], max_results: int = 15
) -> list[str]:
    """クエリに対して前方一致優先・部分一致の順で候補リストを返す。
    大文字小文字・全角半角・スペースの差は _normalize_key で吸収する。"""
    if not query:
        return []
    nq = _normalize_key(query).lower()
    front: list[str] = []
    partial: list[str] = []
    for c in candidates:
        nc = _normalize_key(c).lower()
        if nc.startswith(nq):
            front.append(c)
        elif nq in nc:
            partial.append(c)
    return (front + partial)[:max_results]


def render_machine_autocomplete_input(
    label: str, key: str, candidates: list[str], default: str = "",
    on_change=None, on_change_args: tuple = (),
) -> None:
    """テキスト入力 + 候補ボタンによるオートコンプリートUI。
    ・1文字以上入力 かつ 完全一致でない場合のみ候補ボタンを表示する。
    ・on_click コールバックで session_state[key] を更新する
      （widget 描画後の直接代入は Streamlit が禁止するため）。"""
    if key not in st.session_state:
        st.session_state[key] = default
    _kw: dict = {"on_change": on_change}
    if on_change_args:
        _kw["args"] = on_change_args
    st.text_input(label, value=st.session_state[key], key=key, placeholder="機種名を入力", **_kw)
    query: str = st.session_state.get(key, "")

    # 未入力 or すでに候補と完全一致 → 候補を非表示
    if len(query) < 1 or query in candidates:
        return

    matches = suggest_machine_candidates(query, candidates)
    if not matches:
        st.caption("　候補なし")
        return

    st.caption(f"　↓ {len(matches)}件の候補（クリックで入力）")
    n_cols = min(len(matches), 3)
    cols = st.columns(n_cols)
    for i, cand in enumerate(matches):
        with cols[i % n_cols]:
            def _cand_click(c=cand, k=key, cb=on_change, ca=on_change_args):
                st.session_state[k] = c
                if cb is not None:
                    cb(*ca)
            st.button(
                cand,
                key=f"{key}_sug_{i}",
                use_container_width=True,
                on_click=_cand_click,
            )


def filter_recommended_machines(
    machines: list[str],
    df: pd.DataFrame,
    zen_dai_names: set[str],
    high_ratio_names: set[str],
) -> tuple[list[str], list[str]]:
    """オススメ機種リストを絞り込む。
    - 空白・重複を除去
    - Excel に存在しない機種をスキップ
    - 全台系・高配分判定済み機種を除外
    戻り値: (掲載対象機種リスト, 除外ログリスト)"""
    all_in_excel = set(df["機種名"].astype(str).unique())
    valid: list[str] = []
    logs: list[str] = []
    seen: set[str] = set()
    for m in machines:
        m = m.strip()
        if not m or m in seen:
            continue
        seen.add(m)
        if m in zen_dai_names:
            logs.append(f"「{m}」は全台系に該当したためオススメ機種ピックアップから除外しました")
        elif m in high_ratio_names:
            logs.append(f"「{m}」は高配分機種に該当したため除外しました")
        elif m not in all_in_excel:
            logs.append(f"「{m}」はExcelに存在しないためスキップしました")
        else:
            valid.append(m)
    return valid, logs


def generate_recommended_block_image(
    title: str,
    machines: list[str],
    df: pd.DataFrame,
    diff_raw: pd.Series,
    narabi_bans: set[int] = set(),
    min_diff: int = 1,
    juggler_cfg: "dict | None" = None,
) -> "Image.Image | None":
    """オススメ機種ブロックの優秀台統合画像を生成する。
    min_diff 以上の台を抽出。narabi_bans に含まれる台番は除外する。
    juggler_cfg を渡すとジャグラー機種に専用抽出条件を適用する。"""
    dfs = []
    _jug_series  = juggler_cfg.get("series", set())  if juggler_cfg else set()
    _jug_map     = juggler_cfg.get("jobs_map", {})   if juggler_cfg else {}
    _jug_g_min   = juggler_cfg.get("g_min", 2000)    if juggler_cfg else 2000
    _jug_d_bonus = juggler_cfg.get("diff_bonus", 1000) if juggler_cfg else 1000
    for machine in machines:
        grp = df[df["機種名"] == machine].copy()
        if narabi_bans:
            grp = grp[~grp["台番"].isin(narabi_bans)]
        if grp.empty:
            continue
        dr_m = diff_raw.loc[grp.index]
        if juggler_cfg and machine in _jug_series:
            # ジャグラー専用条件: G数>=2000 かつ (確率<=閾値 AND diff>=0) OR diff>=1000
            g_mask = grp["ゲーム数_rounded"] >= _jug_g_min
            grp   = grp[g_mask].copy()
            dr_m  = dr_m[g_mask]
            if grp.empty:
                continue
            prob_thr = _jug_map.get(machine)
            if prob_thr is not None and "合算確率_num" in grp.columns:
                mask = ((grp["合算確率_num"] <= prob_thr) & (dr_m >= 0)) | (dr_m >= _jug_d_bonus)
            else:
                mask = dr_m >= 0
            good = grp[mask].copy().reset_index(drop=True)
        else:
            good = grp[dr_m >= min_diff].copy().reset_index(drop=True)
        if good.empty:
            continue
        dfs.append(good)
    if not dfs:
        return None
    combined = pd.concat(dfs, ignore_index=True)
    # 機種ごとの最小台番でグループ順を決め、グループ内は台番昇順
    combined["_grp_order"] = combined.groupby("機種名")["台番"].transform("min")
    combined = (combined
                .sort_values(["_grp_order", "台番"])
                .drop(columns=["_grp_order"])
                .reset_index(drop=True))
    return _build_machine_img(combined, title, None)


def _kojin_yushu_filter(km: str, grp_all: pd.DataFrame, dr_all: pd.Series, cfg: dict,
                        force_1k: bool = False) -> pd.DataFrame:
    """個別優秀台のマスク: Step3高配分と同じ条件でフィルター（index は元のまま）。
    force_1k=True のとき差枚+1,000枚以上のみに絞る（秋葉原スランプ付き専用）。"""
    if force_1k:
        return grp_all[dr_all.values >= 1000].copy()
    prob_jobs_map      = cfg["prob_jobs_map"]
    rb_thresh_machines = cfg["rb_threshold_machines"]
    rb_min             = cfg["rb_min"]
    has_g = "ゲーム数_rounded" in grp_all.columns
    juggler_series   = cfg.get("juggler_series", set())
    juggler_jobs_map = {m: (p, d) for m, p, d in cfg.get("juggler_jobs", [])}
    juggler_g_min    = cfg.get("juggler_g_min", 2000)
    if km in juggler_series and km in juggler_jobs_map:
        prob_thr, diff_bon = juggler_jobs_map[km]
        has_prob = "合算確率_num" in grp_all.columns
        if has_g and has_prob:
            mask = (grp_all["ゲーム数_rounded"] >= juggler_g_min) & (
                ((grp_all["合算確率_num"] <= prob_thr) & (dr_all >= 0)) | (dr_all >= diff_bon)
            )
        elif has_g:
            mask = (grp_all["ゲーム数_rounded"] >= juggler_g_min) & (dr_all >= diff_bon)
        else:
            mask = dr_all >= diff_bon
    elif km in prob_jobs_map:
        prob_thr, diff_bon = prob_jobs_map[km]
        has_prob = "合算確率_num" in grp_all.columns
        if has_g and has_prob:
            mask = (grp_all["ゲーム数_rounded"] >= 2000) & (
                ((grp_all["合算確率_num"] <= prob_thr) & (dr_all >= 0)) | (dr_all >= diff_bon)
            )
        elif has_g:
            mask = (grp_all["ゲーム数_rounded"] >= 2000) & (dr_all >= diff_bon)
        else:
            mask = dr_all >= diff_bon
    elif km in rb_thresh_machines:
        _rb_col = next((c for c in ["RB", "REG"] if c in grp_all.columns), None)
        if _rb_col and has_g:
            mask = (dr_all >= 1000) | (
                (grp_all["ゲーム数_rounded"] >= 2000) & (grp_all[_rb_col] >= rb_min) & (dr_all >= 0)
            )
        else:
            mask = dr_all >= 1000
    else:
        if has_g:
            _cnt_1k = int((dr_all >= 1000).sum())
            if _cnt_1k >= 2:
                mask = (dr_all >= 1000) | ((dr_all >= 0) & (grp_all["ゲーム数_rounded"] >= 2000))
            else:
                mask = (dr_all >= 1000) | ((dr_all > 0) & (grp_all["ゲーム数_rounded"] >= 5000))
        else:
            mask = dr_all >= 1000
    return grp_all[mask.values].copy()


def _build_pision_like_html(title: str, summary: dict, rows: list) -> str:
    """pision記事風のサマリーボックス＋機種別テーブルのHTMLを返す。
    summary: {total_diff, avg_diff, avg_games, plus, total}
    rows: [(機種名, 台数, 勝台数, 総差枚, 平均差枚, 平均G数), ...]（バラエティ行は末尾）"""
    def sdiff(v: int) -> str:
        if v > 0:
            return f"+{v:,}"
        if v < 0:
            return f"{v:,}"
        return "±0"

    sum_rows = (
        f'<tr><th>総差枚</th><td>{sdiff(summary["total_diff"])}</td></tr>'
        f'<tr><th>平均差枚</th><td>{sdiff(summary["avg_diff"])}</td></tr>'
        f'<tr><th>平均G数</th><td>{summary["avg_games"]:,}</td></tr>'
        f'<tr><th>勝率</th><td>{summary["plus"]}/{summary["total"]}</td></tr>'
    )
    body = []
    for name, n, w, td, ad, ag in rows:
        is_var = (name == "バラエティ")
        wr = f"{w}/{n} ({round(w / n * 100)}%)" if n else "－"
        body.append(
            f'<tr class="{"variety" if is_var else ""}">'
            f'<td class="name">{name}</td>'
            f'<td class="num">{sdiff(ad)}</td>'
            f'<td class="num">{sdiff(td)}</td>'
            f'<td class="num">{ag:,}</td>'
            f'<td class="wr">{wr}</td>'
            f'</tr>'
        )
    return f'''
<style>
.pis-box {{ font-family:"Meiryo","Yu Gothic",sans-serif; color:#2b3a42; }}
.pis-title {{ font-size:20px; font-weight:600; margin:2px 0 10px; }}
.pis-sum {{ border-collapse:collapse; margin-bottom:14px; }}
.pis-sum th {{ background:#f4f6f7; border:1px solid #dde3e6; padding:6px 14px;
              text-align:left; font-weight:600; width:90px; white-space:nowrap; }}
.pis-sum td {{ border:1px solid #dde3e6; padding:6px 18px; min-width:120px;
              font-variant-numeric:tabular-nums; }}
.pis-sec {{ font-size:16px; font-weight:700; margin:6px 0 8px; }}
.pis-wrap {{ max-height:560px; overflow:auto; border:1px solid #e0e0e0; border-radius:4px; }}
.pis-tbl {{ border-collapse:collapse; width:100%; font-size:13px; }}
.pis-tbl th {{ background:#eceff1; color:#455a64; padding:8px 10px;
              border-bottom:2px solid #cfd8dc; position:sticky; top:0; z-index:1; white-space:nowrap; }}
.pis-tbl th.name {{ text-align:left; }}
.pis-tbl td {{ padding:7px 10px; border-bottom:1px solid #eceff1; }}
.pis-tbl td.name {{ text-align:left; color:#1565c0; }}
.pis-tbl td.num {{ text-align:right; color:#263238; font-variant-numeric:tabular-nums; white-space:nowrap; }}
.pis-tbl td.wr {{ text-align:center; color:#263238; white-space:nowrap; }}
.pis-tbl tbody tr:hover td {{ background:#f5f9ff; }}
.pis-tbl tbody tr.variety td.name {{ color:#5f6368; font-weight:600; }}
</style>
<div class="pis-box">
<div class="pis-title">{title}</div>
<table class="pis-sum"><tbody>{sum_rows}</tbody></table>
<div class="pis-sec">機種別データ</div>
<div class="pis-wrap"><table class="pis-tbl">
<thead><tr><th class="name">機種</th><th>平均差枚</th><th>総差枚</th><th>平均G数</th><th>勝率</th></tr></thead>
<tbody>{''.join(body)}</tbody>
</table></div>
</div>
'''


def _build_pision_detail_html(name: str, df: pd.DataFrame) -> str:
    """機種クリック時の台別詳細（pision準拠：台番/差枚/BB/RB/合算/ART/G数＋平均行）。
    df: その機種の台別行（列: 台番/差枚/BB/RB/AT/ゲーム数 のうち存在するもの）"""
    def dcell(v: int) -> str:
        cls = "pos" if v > 0 else ("neg" if v < 0 else "z")
        s = f"+{v:,}" if v > 0 else (f"{v:,}" if v < 0 else "±0")
        return f'<td class="dn {cls}">{s}</td>'

    def prob(cnt: int, g: int) -> str:
        return f"{cnt} (1/{round(g / cnt)})" if cnt > 0 else f"{cnt}"

    has_bb = "BB" in df.columns
    has_rb = "RB" in df.columns
    has_at = "AT" in df.columns
    has_g  = "ゲーム数" in df.columns
    body = []
    for _, r in df.sort_values("台番").iterrows():
        dai = int(r["台番"]); dv = int(r["差枚"])
        bb = int(r["BB"]) if has_bb else 0
        rb = int(r["RB"]) if has_rb else 0
        at = int(r["AT"]) if has_at else 0
        g  = int(r["ゲーム数"]) if has_g else 0
        tot = bb + rb
        gou = f"1/{round(g / tot)}" if tot > 0 and g else "─"
        body.append(
            f'<tr><td class="c">{dai}</td>{dcell(dv)}'
            f'<td class="c">{prob(bb, g)}</td><td class="c">{prob(rb, g)}</td>'
            f'<td class="c">{gou}</td><td class="c">{at}</td>'
            f'<td class="r">{g:,}</td></tr>'
        )
    n = len(df)
    avg_row = ""
    if n:
        tg = int(df["ゲーム数"].sum()) if has_g else 0
        tbb = int(df["BB"].sum()) if has_bb else 0
        trb = int(df["RB"].sum()) if has_rb else 0
        avg_d = int(round(df["差枚"].mean()))
        avg_at = int(round(df["AT"].mean())) if has_at else 0
        avg_g = int(round(df["ゲーム数"].mean())) if has_g else 0
        tot = tbb + trb
        avg_gou = f"1/{round(tg / tot)}" if tot > 0 and tg else "─"
        avg_row = (
            f'<tr class="avg"><td class="c">平均</td>{dcell(avg_d)}'
            f'<td class="c"></td><td class="c"></td>'
            f'<td class="c">{avg_gou}</td><td class="c">{avg_at}</td>'
            f'<td class="r">{avg_g:,}</td></tr>'
        )
    return f'''
<style>
.pisd {{ font-family:"Meiryo","Yu Gothic",sans-serif; margin:4px 0 10px; }}
.pisd-title {{ font-size:15px; font-weight:700; color:#2b3a42; margin:2px 0 6px; }}
.pisd-tbl {{ border-collapse:collapse; width:100%; font-size:13px; }}
.pisd-tbl th {{ background:#eceff1; color:#455a64; padding:7px 10px;
               border-bottom:2px solid #cfd8dc; white-space:nowrap; }}
.pisd-tbl td {{ padding:6px 10px; border-bottom:1px solid #eceff1; }}
.pisd-tbl td.c {{ text-align:center; color:#263238; white-space:nowrap; }}
.pisd-tbl td.r {{ text-align:right; color:#263238; font-variant-numeric:tabular-nums; white-space:nowrap; }}
.pisd-tbl td.dn {{ text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }}
.pisd-tbl td.pos {{ color:#2f9e44; font-weight:700; }}
.pisd-tbl td.neg {{ color:#e03131; }}
.pisd-tbl td.z {{ color:#868e96; }}
.pisd-tbl tr.avg td {{ border-top:2px solid #cfd8dc; background:#f7fafb; font-weight:600; }}
</style>
<div class="pisd">
<div class="pisd-title">{name}（{n}台）</div>
<table class="pisd-tbl">
<thead><tr><th>台番</th><th>差枚</th><th>BB</th><th>RB</th><th>合算</th><th>ART</th><th>G数</th></tr></thead>
<tbody>{''.join(body)}{avg_row}</tbody>
</table>
</div>
'''


def _build_pision_interactive_html(title: str, summary: "dict | None", rows: list,
                                    units_df, single_names=None) -> str:
    """機種名クリックで台別詳細を右パネルに表示する自己完結型HTMLを返す。
    units_df: 台番/機種名/差枚/BB/RB/AT/ゲーム数 のうち存在する列のDataFrame。
    single_names: バラエティ集約対象の機種名set（Noneなら追加しない）。
    summary=None のときサマリーボックスを非表示にする。"""
    import json as _json

    def sdiff(v: int) -> str:
        if v > 0: return f"+{v:,}"
        if v < 0: return f"{v:,}"
        return "±0"

    sum_rows = "" if not summary else (
        f'<tr><th>総差枚</th><td>{sdiff(summary["total_diff"])}</td></tr>'
        f'<tr><th>平均差枚</th><td>{sdiff(summary["avg_diff"])}</td></tr>'
        f'<tr><th>平均G数</th><td>{summary["avg_games"]:,}</td></tr>'
        f'<tr><th>勝率</th><td>{summary["plus"]}/{summary["total"]}</td></tr>'
    )

    body = []
    for name, n, w, td, ad, ag in rows:
        is_var = (name == "バラエティ")
        wr = f"{w}/{n} ({round(w / n * 100)}%)" if n else "－"
        body.append(
            f'<tr class="{"variety " if is_var else ""}mac-row" data-name="{name}">'
            f'<td class="name">{name}</td>'
            f'<td class="num">{sdiff(ad)}</td>'
            f'<td class="num">{sdiff(td)}</td>'
            f'<td class="num">{ag:,}</td>'
            f'<td class="wr">{wr}</td>'
            f'</tr>'
        )

    detail_map: dict = {}
    if units_df is not None and len(units_df) > 0:
        for name, *_ in rows:
            if name == "バラエティ":
                if single_names:
                    _df = units_df[units_df["機種名"].isin(single_names)]
                    detail_map[name] = _build_pision_detail_html("バラエティ（1台機種すべて）", _df)
            else:
                _df = units_df[units_df["機種名"] == name]
                if not _df.empty:
                    detail_map[name] = _build_pision_detail_html(name, _df)

    detail_json = _json.dumps(detail_map, ensure_ascii=False)
    hint_init = '<div class="pis-sec">&nbsp;</div><p class="hint-txt">← 機種名をクリックすると台別詳細が表示されます</p>'

    return f'''<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{box-sizing:border-box;}}
body{{margin:0;padding:0;font-family:"Meiryo","Yu Gothic",sans-serif;color:#2b3a42;}}
.layout{{display:flex;gap:20px;align-items:flex-start;}}
.col-l{{flex:5;min-width:0;}}
.col-r{{flex:7;min-width:0;overflow-y:auto;}}
.pis-title{{font-size:20px;font-weight:600;margin:2px 0 10px;}}
.pis-sum{{border-collapse:collapse;margin-bottom:14px;}}
.pis-sum th{{background:#f4f6f7;border:1px solid #dde3e6;padding:6px 14px;
             text-align:left;font-weight:600;width:90px;white-space:nowrap;}}
.pis-sum td{{border:1px solid #dde3e6;padding:6px 18px;min-width:120px;
             font-variant-numeric:tabular-nums;}}
.pis-sec{{font-size:16px;font-weight:700;margin:6px 0 8px;}}
.hint{{font-size:11px;font-weight:400;color:#888;}}
.hint-txt{{color:#888;font-size:13px;margin-top:50px;}}
.pis-wrap{{max-height:520px;overflow:auto;border:1px solid #e0e0e0;border-radius:4px;}}
.pis-tbl{{border-collapse:collapse;width:100%;font-size:13px;}}
.pis-tbl th{{background:#eceff1;color:#455a64;padding:8px 10px;
             border-bottom:2px solid #cfd8dc;position:sticky;top:0;z-index:1;white-space:nowrap;}}
.pis-tbl th.name{{text-align:left;}}
.pis-tbl td{{padding:7px 10px;border-bottom:1px solid #eceff1;}}
.pis-tbl td.name{{text-align:left;color:#1565c0;cursor:pointer;text-decoration:underline dotted;}}
.pis-tbl td.num{{text-align:right;color:#263238;font-variant-numeric:tabular-nums;white-space:nowrap;}}
.pis-tbl td.wr{{text-align:center;color:#263238;white-space:nowrap;}}
.pis-tbl tbody tr.mac-row:hover td{{background:#f5f9ff;}}
.pis-tbl tbody tr.mac-row:hover td.name{{background:#e3f2fd;}}
.pis-tbl tbody tr.mac-row.active td{{background:#e3f2fd;}}
.pis-tbl tbody tr.variety td.name{{color:#5f6368;font-weight:600;}}
.pisd{{margin:4px 0 10px;}}
.pisd-title{{font-size:15px;font-weight:700;color:#2b3a42;margin:2px 0 6px;}}
.pisd-tbl{{border-collapse:collapse;width:100%;font-size:13px;}}
.pisd-tbl th{{background:#eceff1;color:#455a64;padding:7px 10px;
              border-bottom:2px solid #cfd8dc;white-space:nowrap;}}
.pisd-tbl td{{padding:6px 10px;border-bottom:1px solid #eceff1;}}
.pisd-tbl td.c{{text-align:center;color:#263238;white-space:nowrap;}}
.pisd-tbl td.r{{text-align:right;color:#263238;font-variant-numeric:tabular-nums;white-space:nowrap;}}
.pisd-tbl td.dn{{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap;}}
.pisd-tbl td.pos{{color:#2f9e44;font-weight:700;}}
.pisd-tbl td.neg{{color:#e03131;}}
.pisd-tbl td.z{{color:#868e96;}}
.pisd-tbl tr.avg td{{border-top:2px solid #cfd8dc;background:#f7fafb;font-weight:600;}}
</style>
</head><body>
<div class="pis-title">{title}</div>
<table class="pis-sum"><tbody>{sum_rows}</tbody></table>
<div class="layout">
  <div class="col-l">
    <div class="pis-sec">機種別データ <span class="hint">（機種名をクリックで台別詳細）</span></div>
    <div class="pis-wrap"><table class="pis-tbl">
    <thead><tr><th class="name">機種</th><th>平均差枚</th><th>総差枚</th><th>平均G数</th><th>勝率</th></tr></thead>
    <tbody>{"".join(body)}</tbody>
    </table></div>
  </div>
  <div class="col-r" id="detail-panel">{hint_init}</div>
</div>
<script>
var details={detail_json};
var activeRow=null;
var hintHtml={_json.dumps(hint_init, ensure_ascii=False)};
document.querySelectorAll('.mac-row').forEach(function(row){{
  row.addEventListener('click',function(){{
    var name=this.getAttribute('data-name');
    var panel=document.getElementById('detail-panel');
    if(activeRow===this){{
      this.classList.remove('active');
      activeRow=null;
      panel.innerHTML=hintHtml;
      return;
    }}
    if(activeRow)activeRow.classList.remove('active');
    this.classList.add('active');
    activeRow=this;
    panel.innerHTML=details[name]||'<p style="color:#888">台別データなし</p>';
  }});
}});
</script>
</body></html>'''


def render_pision_data_view(vt_df: pd.DataFrame, title: str, sel_key: str) -> None:
    """正規化済みDataFrame（列: 台番/機種名/差枚/ゲーム数/BB/RB/AT のうち存在するもの）から
    pision記事風の機種別サマリー＋台別詳細＋台別データexpanderを描画する。
    結果ポスト用ページとスランプグラフ生成ページで共用。"""
    if vt_df is None or vt_df.empty:
        return
    # 台別（詳細照合用）
    _disp = pd.DataFrame()
    if "台番" in vt_df.columns:    _disp["台番"]    = vt_df["台番"]
    if "機種名" in vt_df.columns:  _disp["機種名"]  = vt_df["機種名"]
    if "ゲーム数" in vt_df.columns: _disp["G数"]     = vt_df["ゲーム数"]
    if "BB" in vt_df.columns:      _disp["BB"]      = vt_df["BB"]
    if "RB" in vt_df.columns:      _disp["RB"]      = vt_df["RB"]
    if "AT" in vt_df.columns:      _disp["ART"]     = vt_df["AT"]
    if {"BB", "RB", "ゲーム数"} <= set(vt_df.columns):
        _tot = (vt_df["BB"] + vt_df["RB"]).replace(0, pd.NA)
        _disp["合算確率"] = (vt_df["ゲーム数"] / _tot).map(
            lambda v: f"1/{v:.1f}" if pd.notna(v) else "─")
    if "差枚" in vt_df.columns:    _disp["差枚"]    = vt_df["差枚"]
    if "台番" in _disp.columns:
        _disp = _disp.sort_values("台番").reset_index(drop=True)

    # 機種別集計（全機種・素データ）
    _agg = None
    _meta = None
    if {"機種名", "差枚"} <= set(vt_df.columns):
        _g = vt_df.groupby("機種名", sort=False)
        _agg = pd.DataFrame({
            "機種名":   list(_g.groups.keys()),
            "台数":     _g["差枚"].size().values,
            "勝台数":   _g["差枚"].apply(lambda s: int((s > 0).sum())).values,
            "総差枚":   _g["差枚"].sum().astype(int).values,
            "平均差枚": _g["差枚"].mean().round().astype(int).values,
        })
        if "ゲーム数" in vt_df.columns:
            _agg["平均G数"] = _g["ゲーム数"].mean().round().astype(int).values
        else:
            _agg["平均G数"] = 0
        _tot_all = len(vt_df)
        _td_all  = int(vt_df["差枚"].sum())
        _meta = {
            "total":      _tot_all,
            "plus":       int((vt_df["差枚"] > 0).sum()),
            "total_diff": _td_all,
            "avg_diff":   int(round(_td_all / _tot_all)) if _tot_all else 0,
            "avg_games":  (int(round(vt_df["ゲーム数"].mean()))
                           if "ゲーム数" in vt_df.columns and _tot_all else 0),
        }

    # 機種クリック詳細用の台別素データ
    _ucols = [c for c in ["台番", "機種名", "差枚", "BB", "RB", "AT", "ゲーム数"]
              if c in vt_df.columns]
    _units_df = vt_df[_ucols].copy() if _ucols else None

    if _agg is not None and not _agg.empty and _meta is not None:
        _multi  = _agg[_agg["台数"] >= 2].sort_values("平均差枚", ascending=False)
        _single = _agg[_agg["台数"] == 1]
        _rows = [
            (r["機種名"], int(r["台数"]), int(r["勝台数"]),
             int(r["総差枚"]), int(r["平均差枚"]), int(r["平均G数"]))
            for _, r in _multi.iterrows()
        ]
        if not _single.empty:
            _vn  = int(_single["台数"].sum())
            _vw  = int(_single["勝台数"].sum())
            _vtd = int(_single["総差枚"].sum())
            _vad = int(round(_vtd / _vn)) if _vn else 0
            _vg  = int(round((_single["平均G数"] * _single["台数"]).sum() / _vn)) if _vn else 0
            _rows.append(("バラエティ", _vn, _vw, _vtd, _vad, _vg))
        st.caption("📋 pisionの代わりに照合用（2台以上を平均差枚順・1台機種はバラエティに集約／数値はpisionの生データと一致）")
        _snames = set(_single["機種名"].tolist()) if not _single.empty else None
        _comp_h = max(480, min(820, len(_rows) * 42 + 350))
        components.html(
            _build_pision_interactive_html(title, _meta, _rows, _units_df, _snames),
            height=_comp_h, scrolling=True,
        )
    if not _disp.empty:
        with st.expander(f"📋 台別データ（全{len(_disp)}台）", expanded=False):
            st.dataframe(_disp, use_container_width=True, hide_index=True, height=520)


def show_auto_page(with_slump: bool = False) -> None:
    """自動処理ページ: PIL パイプラインで全画像を生成する"""
    store = st.session_state.selected_store
    # その他の優秀台ピックアップ分割（秋葉原=①②③・上野新館/上野本館=①②）
    _sonota_split = with_slump and store in ("秋葉原", "上野新館", "上野本館")
    _sonota_extra_thrs = [(2000, "その他の優秀台+2,000枚以上.jpg")]
    if store == "秋葉原":
        _sonota_extra_thrs.append((3000, "その他の優秀台+3,000枚以上.jpg"))
    if with_slump:
        st.markdown(f"## 【{store}】スランプ付き結果ポスト用")
    else:
        st.markdown(f"## 【{store}】結果ポスト用")
    st.caption("全台系・ジャグラー優秀台・その他の優秀台を一括生成します。")
    st.markdown("---")

    # ── ⓪ pision.io から日付データを自動取得（全店舗・その店舗のデータ）──────
    _tb_uploaded = None
    if store in STORES:
        st.markdown(f"### 📈 日付からデータを自動取得（{store}）")
        api_key = _get_pision_api_key()
        if not api_key:
            st.caption("⚠️ PISION_API_KEY が未設定のため利用できません。①から手動でアップロードしてください。")
        else:
            _tb_mode = st.radio(
                "データ種別",
                ["確定データ", "速報データ（当日・営業中）"],
                horizontal=True,
                key=f"auto_tb_mode_{store}",
                help="確定データ＝前日まで（X-Api-Key）。速報データ＝当日の営業中データ（realtimeログインが必要）。",
            )
            _tb_is_rt = _tb_mode.startswith("速報")
            _tb_rt_ok = True
            if _tb_is_rt:
                _rt_user, _rt_pass = _get_pision_rt_credentials()
                if not _rt_user or not _rt_pass:
                    st.error("❌ 速報データには realtime のログイン情報が必要です。"
                             ".env に PISION_RT_USER / PISION_RT_PASS を設定してください。")
                    _tb_rt_ok = False

            if _tb_is_rt:
                import datetime as _dt
                _jst = _dt.timezone(_dt.timedelta(hours=9))
                _now = _dt.datetime.now(_jst)
                _rt_default = _now.date() if _now.hour >= 7 else _now.date() - _dt.timedelta(days=1)
                # 確定→速報切替時にカウンタを増やして新キーを生成し当日にリセット
                # （ReactのWidget内部キャッシュを確実に破棄するためキー自体を変える）
                _tb_prev_mode_key = f"_auto_tb_prev_mode_{store}"
                _tb_cnt_key = f"_auto_tb_cnt_{store}"
                if st.session_state.get(_tb_prev_mode_key) != "rt":
                    st.session_state[_tb_cnt_key] = st.session_state.get(_tb_cnt_key, 0) + 1
                st.session_state[_tb_prev_mode_key] = "rt"
                _tb_cnt = st.session_state.get(_tb_cnt_key, 0)
                _tb_date = st.date_input(
                    "日付を選択（速報は営業日に合わせて選択・日付変更では自動取得しません）",
                    value=_rt_default,
                    key=f"auto_tb_date_rt_{store}_{_tb_cnt}",
                )
            else:
                st.session_state[f"_auto_tb_prev_mode_{store}"] = "fix"
                _tb_date = st.date_input(
                    "日付を選択",
                    value=datetime.date.today() - datetime.timedelta(days=1),
                    key=f"auto_tb_date_{store}",
                )
            _tb_date_str = _tb_date.strftime("%Y-%m-%d")

            _tb_mode_tag       = "rt" if _tb_is_rt else "fix"
            _tb_seen_key       = f"_auto_tb_seen_{_tb_mode_tag}_{store}"
            _tb_bytes_key      = f"_auto_tb_file_bytes_{_tb_mode_tag}_{store}"
            _tb_name_key       = f"_auto_tb_file_name_{_tb_mode_tag}_{store}"
            _tb_count_key      = f"_auto_tb_count_{_tb_mode_tag}_{store}"
            _tb_fetched_key    = f"_auto_tb_fetched_{_tb_mode_tag}_{store}"
            _tb_collecting_key    = f"_auto_tb_collecting_rt_{store}"
            _tb_rt_items_key      = f"_auto_tb_rt_items_{store}"
            _tb_rt_items_date_key = f"_auto_tb_rt_items_date_{store}"
            _tb_baseline_artid_key = f"_auto_tb_baseline_artid_{store}"  # 収集開始時点の article_id

            # 収集中かどうか（速報で、今表示中の日付に対して収集を開始済み）
            _is_collecting = _tb_is_rt and st.session_state.get(_tb_collecting_key) == _tb_date_str

            # ── ボタン描画 ───────────────────────────────────────────────
            _do_rt_check    = False
            _do_rt_existing = False
            if _is_collecting:
                _btn_c1, _btn_c2 = st.columns(2)
                with _btn_c1:
                    _tb_refetch = st.button("⏳ 収集中...", key=f"auto_tb_refetch_{store}",
                                            disabled=True, use_container_width=True)
                with _btn_c2:
                    _do_rt_check = st.button("🔍 今すぐ確認", key=f"auto_tb_rt_check_{store}",
                                             use_container_width=True,
                                             help="30秒待たずに今すぐ収集完了を確認します。")
            elif _tb_is_rt and _tb_rt_ok:
                _btn_c1, _btn_c2 = st.columns(2)
                with _btn_c1:
                    _tb_refetch = st.button("⚡ 速報を取得", key=f"auto_tb_refetch_{store}",
                                            use_container_width=True, type="primary")
                with _btn_c2:
                    _do_rt_existing = st.button("📂 既存のデータを取得", key=f"auto_tb_rt_existing_{store}",
                                                use_container_width=True,
                                                help="新しい収集を開始せず、過去に取得済みの直近データを読み込みます。")
            else:
                _tb_refetch = st.button("🔄 取得", key=f"auto_tb_refetch_{store}")

            # 確定データのみ日付変更で自動取得。速報はボタン押下のみトリガー。
            _tb_seen = st.session_state.get(_tb_seen_key)
            _tb_date_changed = (not _tb_is_rt) and _tb_seen is not None and _tb_seen != _tb_date_str
            st.session_state[_tb_seen_key] = _tb_date_str

            def _save_rt_items_to_session(items: list) -> None:
                """速報 items を変換・セッション保存する共通処理。"""
                _slump_apply_names(items)
                _rows = [
                    {
                        "台番":     it.get("unitId"),
                        "機種名":   it.get("_machineName") or it.get("_convertedName") or it.get("displayName") or "",
                        "差枚":     it.get("diff", 0),
                        "ゲーム数": it.get("games", 0),
                        "BB":       it.get("bb", 0),
                        "RB":       it.get("rb", 0),
                        "AT":       it.get("art", 0),
                    }
                    for it in items
                ]
                _df  = pd.DataFrame(_rows)
                _buf = io.BytesIO()
                _df.to_excel(_buf, index=False)
                st.session_state[_tb_bytes_key]         = _buf.getvalue()
                st.session_state[_tb_name_key]          = f"{_tb_date.strftime('%Y%m%d')}_{store}_20S.xlsx"
                st.session_state[_tb_count_key]         = len(_df)
                st.session_state[_tb_fetched_key]       = _tb_date_str
                st.session_state[_tb_rt_items_key]      = items
                st.session_state[_tb_rt_items_date_key] = _tb_date_str
                st.session_state.pop(_tb_collecting_key, None)

            def _is_new_artid(poll_result: dict) -> bool:
                """ベースラインより新しい article_id かどうか判定する。"""
                _baseline = st.session_state.get(_tb_baseline_artid_key)
                _new_id   = poll_result.get("article_id")
                if _baseline is None:
                    return True  # ベースランなし → 最初に取れたものが新データ
                try:
                    return int(_new_id) > int(_baseline)
                except (TypeError, ValueError):
                    return False

            # ── 手動確認（「今すぐ確認」ボタン）────────────────────────────
            if _do_rt_check:
                with st.spinner("収集状況を確認中..."):
                    _chk = fetch_pision_realtime(store, _tb_date_str, trigger=False)
                if _chk["ok"] and _is_new_artid(_chk):
                    _save_rt_items_to_session(_chk["items"])
                    st.rerun()
                elif _chk["ok"]:
                    st.info("⏳ 収集はまだ完了していません（前回と同じスナップショット）。自動確認に戻ります。")
                    st.rerun()
                else:
                    st.rerun()  # 完了していない → 自動ポーリングに戻す

            # ── 自動ポーリング（収集中かつ手動確認なし）─────────────────
            if _is_collecting and not _do_rt_check:
                import time as _time
                _ap_ph = st.empty()
                _ap_ph.info("⏳ 速報データを収集中... 30秒後に自動で確認します。")
                _time.sleep(30)
                _ap_ph.empty()
                with st.spinner("収集状況を自動確認中..."):
                    _auto_poll = fetch_pision_realtime(store, _tb_date_str, trigger=False)
                if _auto_poll["ok"] and _is_new_artid(_auto_poll):
                    _save_rt_items_to_session(_auto_poll["items"])
                    st.rerun()
                elif _auto_poll.get("running") or (_auto_poll["ok"] and not _is_new_artid(_auto_poll)):
                    st.rerun()  # まだ収集中 or 古いデータ → 30秒ループを継続
                else:
                    st.warning("⚠️ 収集が完了しましたがデータが取得できませんでした。「⚡ 速報を取得」をもう一度押してください。")
                    st.session_state.pop(_tb_collecting_key, None)
                    st.rerun()

            # ── 既存データ取得（新規収集なし）────────────────────────────
            if _do_rt_existing:
                with st.spinner("既存の速報データを確認中（新しい収集は開始しません）..."):
                    _exist = fetch_pision_realtime(store, _tb_date_str, trigger=False)
                if _exist["ok"]:
                    _save_rt_items_to_session(_exist["items"])
                    st.rerun()
                else:
                    st.warning("⚠️ 既存の速報データが見つかりませんでした。「⚡ 速報を取得」で新しい収集を開始してください。")

            # ── 新規取得（速報を取得ボタン or 確定取得ボタン or 日付変更）──
            if (_tb_date_changed or _tb_refetch) and _tb_rt_ok:
                with st.spinner(f"{_tb_date_str} のデータを取得中..."):
                    _tb_fetched = None
                    if _tb_is_rt:
                        _rt = fetch_pision_realtime(store, _tb_date_str)
                        if not _rt["ok"]:
                            st.error(f"❌ {_rt['error']}")
                            if _rt.get("collect_started"):
                                # 既存スナップショットなし → ベースライン=None で収集待ち
                                st.session_state[_tb_collecting_key]     = _tb_date_str
                                st.session_state[_tb_baseline_artid_key] = None
                        else:
                            # 既存スナップショットあり → ベースラインを保存して新収集を待つ
                            st.session_state[_tb_collecting_key]     = _tb_date_str
                            st.session_state[_tb_baseline_artid_key] = _rt.get("article_id")
                    else:
                        st.session_state.pop(_tb_collecting_key, None)
                        st.session_state.pop(_tb_rt_items_key, None)
                        st.session_state.pop(_tb_rt_items_date_key, None)
                        try:
                            _tb_halls = fetch_pision_halls(api_key)
                        except Exception as e:
                            st.error(f"❌ ホール一覧取得失敗: {e}")
                            _tb_halls = []
                        _tb_hall_id = None
                        for h in _tb_halls:
                            _hn = h.get("name") or h.get("displayName") or ""
                            if store in _hn and "エスパス" in _hn:
                                _tb_hall_id = str(h.get("id") or h.get("hallId") or "")
                                break
                        if _tb_hall_id is not None:
                            try:
                                _tb_fetched = fetch_pision_results(api_key, _tb_hall_id, _tb_date_str)
                            except Exception as e:
                                st.error(f"❌ データ取得失敗: {e}")
                if not _tb_fetched:
                    st.session_state[_tb_bytes_key] = None
                else:
                    _tb_rows = [
                        {
                            "台番":     item.get("unitId"),
                            "機種名":   item.get("_machineName") or item.get("_convertedName") or item.get("displayName") or "",
                            "差枚":     item.get("diff", 0),
                            "ゲーム数": item.get("games", 0),
                            "BB":       item.get("bb", 0),
                            "RB":       item.get("rb", 0),
                            "AT":       item.get("art", 0),
                        }
                        for item in _tb_fetched
                    ]
                    _tb_df    = pd.DataFrame(_tb_rows)
                    _tb_fname = f"{_tb_date.strftime('%Y%m%d')}_{store}_20S.xlsx"
                    _tb_buf   = io.BytesIO()
                    _tb_df.to_excel(_tb_buf, index=False)
                    st.session_state[_tb_bytes_key] = _tb_buf.getvalue()
                    st.session_state[_tb_name_key]  = _tb_fname
                    st.session_state[_tb_count_key] = len(_tb_df)
                st.session_state[_tb_fetched_key] = _tb_date_str
                st.rerun()

            if st.session_state.get(_tb_fetched_key) == _tb_date_str:
                _tb_data = st.session_state.get(_tb_bytes_key)
                if _tb_data:
                    _tb_uploaded = io.BytesIO(_tb_data)
                    _tb_uploaded.name = st.session_state.get(_tb_name_key, f"{_tb_date.strftime('%Y%m%d')}_{store}_20S.xlsx")
                    _tb_label = "速報" if _tb_is_rt else "確定"
                    st.success(f"✅ {_tb_date_str} の{_tb_label}データ（{st.session_state.get(_tb_count_key, '?')}台）を取得し、①にセットしました。")
                elif not _is_collecting:
                    st.info(f"📭 {_tb_date_str} のデータを取得できませんでした（404 / 未公開 / 店休日の可能性があります）。①から手動でアップロードしてください。")

    st.markdown("---")

    # ── ① Excel アップロード（常に描画）─────────────────────────────
    st.markdown("### ① Excelファイルをアップロード")
    st.caption("ファイル名は `YYYYMMDD_店舗名_20S.xlsx` の形式を想定しています。")
    uploaded = st.file_uploader("xlsx / csv を選択", type=["xlsx", "xls", "csv"], key="auto_upload")
    if uploaded is None and _tb_uploaded is not None:
        uploaded = _tb_uploaded
        st.caption(f"📈 自動取得データを使用中: `{uploaded.name}`（手動でアップロードすると優先されます）")

    # 同じExcelが再アップロードされたら前回の入力値を復元（rerunなし・同レンダリング内で反映）
    if uploaded is not None:
        st.session_state["auto_current_excel"] = uploaded.name
        if st.session_state.get("_auto_prev_excel") != uploaded.name:
            # 切り替え前の入力値を旧ファイル名で先に保存（再フェッチ時の値消え防止）
            _prev_excel = st.session_state.get("_auto_prev_excel")
            if _prev_excel:
                _cur_save = _load_auto_inputs_json()
                _cur_save[_prev_excel] = {k: st.session_state[k] for k in _auto_input_keys(store) if k in st.session_state}
                try:
                    with open(_AUTO_INPUTS_JSON, "w", encoding="utf-8") as _sf:
                        json.dump(_cur_save, _sf, ensure_ascii=False, indent=2)
                except Exception:
                    pass
            _restore_auto_inputs(uploaded.name, store)
            st.session_state["_auto_prev_excel"] = uploaded.name
        # 毎レンダーで自動保存（ボタン未押下でブラウザを閉じても値が残るようにする）
        _save_auto_inputs(store)

    # ── Excel由来の機種名候補リストを抽出・キャッシュ ─────────────────
    _em_ss_key = f"_auto_excel_machines_{store}"
    _em_fn_key = f"_auto_excel_machines_fn_{store}"
    if uploaded is not None and st.session_state.get(_em_fn_key) != uploaded.name:
        try:
            uploaded.seek(0)
            _raw_em = _read_uploaded_df(uploaded)
            uploaded.seek(0)
            _df_em, _ = normalize_df(_raw_em)
            _df_em = apply_name_conversion(_df_em)
            st.session_state[_em_ss_key] = sorted(
                set(_df_em["機種名"].dropna().astype(str).str.strip().tolist())
            )
            st.session_state[_em_fn_key] = uploaded.name
        except Exception:
            pass
    _excel_candidates: list[str] = st.session_state.get(_em_ss_key) or load_machine_candidates()

    # ── 📋 取得データを作業画面内に表示（pisionを開かず照合するため）──────
    if uploaded is not None:
        _vt_key       = f"_auto_view_df_{store}"
        _vt_sum_key   = f"_auto_view_summary_{store}"
        _vt_meta_key  = f"_auto_view_meta_{store}"
        _vt_units_key = f"_auto_view_units_{store}"
        _vt_fn_key    = f"_auto_view_df_fn_{store}"
        if st.session_state.get(_vt_fn_key) != uploaded.name:
            try:
                uploaded.seek(0)
                _vt_raw = _read_uploaded_df(uploaded)
                uploaded.seek(0)
                _vt_df, _ = normalize_df(_vt_raw)
                _vt_df = apply_name_conversion(_vt_df)
                # 台別（詳細照合用）
                _disp = pd.DataFrame()
                if "台番" in _vt_df.columns:    _disp["台番"]    = _vt_df["台番"]
                if "機種名" in _vt_df.columns:  _disp["機種名"]  = _vt_df["機種名"]
                if "ゲーム数" in _vt_df.columns: _disp["G数"]     = _vt_df["ゲーム数"]
                if "BB" in _vt_df.columns:      _disp["BB"]      = _vt_df["BB"]
                if "RB" in _vt_df.columns:      _disp["RB"]      = _vt_df["RB"]
                if "AT" in _vt_df.columns:      _disp["ART"]     = _vt_df["AT"]
                if {"BB", "RB", "ゲーム数"} <= set(_vt_df.columns):
                    _tot = (_vt_df["BB"] + _vt_df["RB"]).replace(0, pd.NA)
                    _disp["合算確率"] = (_vt_df["ゲーム数"] / _tot).map(
                        lambda v: f"1/{v:.1f}" if pd.notna(v) else "─")
                if "差枚" in _vt_df.columns:    _disp["差枚"]    = _vt_df["差枚"]
                if "台番" in _disp.columns:
                    _disp = _disp.sort_values("台番").reset_index(drop=True)
                st.session_state[_vt_key] = _disp
                # 機種別集計（全機種・素データ。表示時に2台以上/バラエティへ振り分け）
                _agg = None
                _meta = None
                if {"機種名", "差枚"} <= set(_vt_df.columns):
                    _g = _vt_df.groupby("機種名", sort=False)
                    _agg = pd.DataFrame({
                        "機種名":   list(_g.groups.keys()),
                        "台数":     _g["差枚"].size().values,
                        "勝台数":   _g["差枚"].apply(lambda s: int((s > 0).sum())).values,
                        "総差枚":   _g["差枚"].sum().astype(int).values,
                        "平均差枚": _g["差枚"].mean().round().astype(int).values,
                    })
                    if "ゲーム数" in _vt_df.columns:
                        _agg["平均G数"] = _g["ゲーム数"].mean().round().astype(int).values
                    else:
                        _agg["平均G数"] = 0
                    _tot_all = len(_vt_df)
                    _td_all  = int(_vt_df["差枚"].sum())
                    _meta = {
                        "total":      _tot_all,
                        "plus":       int((_vt_df["差枚"] > 0).sum()),
                        "total_diff": _td_all,
                        "avg_diff":   int(round(_td_all / _tot_all)) if _tot_all else 0,
                        "avg_games":  (int(round(_vt_df["ゲーム数"].mean()))
                                       if "ゲーム数" in _vt_df.columns and _tot_all else 0),
                    }
                # 機種クリック詳細用の台別素データ（必要列のみ）
                _ucols = [c for c in ["台番", "機種名", "差枚", "BB", "RB", "AT", "ゲーム数"]
                          if c in _vt_df.columns]
                st.session_state[_vt_units_key] = _vt_df[_ucols].copy() if _ucols else None
                st.session_state[_vt_sum_key]   = _agg
                st.session_state[_vt_meta_key]  = _meta
                st.session_state[_vt_fn_key]    = uploaded.name
            except Exception:
                st.session_state[_vt_key]       = None
                st.session_state[_vt_sum_key]   = None
                st.session_state[_vt_meta_key]  = None
                st.session_state[_vt_units_key] = None
        _view_df = st.session_state.get(_vt_key)
        _agg_df  = st.session_state.get(_vt_sum_key)
        _meta    = st.session_state.get(_vt_meta_key)
        if _agg_df is not None and not _agg_df.empty and _meta is not None:
            # 2台以上＝平均差枚の降順ランキング、1台機種＝バラエティに集約（pision準拠）
            _multi  = _agg_df[_agg_df["台数"] >= 2].sort_values("平均差枚", ascending=False)
            _single = _agg_df[_agg_df["台数"] == 1]
            _rows = [
                (r["機種名"], int(r["台数"]), int(r["勝台数"]),
                 int(r["総差枚"]), int(r["平均差枚"]), int(r["平均G数"]))
                for _, r in _multi.iterrows()
            ]
            if not _single.empty:
                _vn  = int(_single["台数"].sum())
                _vw  = int(_single["勝台数"].sum())
                _vtd = int(_single["総差枚"].sum())
                _vad = int(round(_vtd / _vn)) if _vn else 0
                _vg  = int(round((_single["平均G数"] * _single["台数"]).sum() / _vn)) if _vn else 0
                _rows.append(("バラエティ", _vn, _vw, _vtd, _vad, _vg))
            _m = re.match(r"(\d{4})(\d{2})(\d{2})", os.path.basename(uploaded.name))
            _title = (f"{int(_m.group(1))}/{int(_m.group(2))}/{int(_m.group(3))} エスパス{store}"
                      if _m else f"エスパス{store}")
            st.caption("📋 pisionの代わりに照合用（2台以上を平均差枚順・1台機種はバラエティに集約／数値はpisionの生データと一致）")
            _units_df = st.session_state.get(_vt_units_key)
            _snames = set(_single["機種名"].tolist()) if not _single.empty else None
            _comp_h = max(480, min(820, len(_rows) * 42 + 350))
            components.html(
                _build_pision_interactive_html(_title, _meta, _rows, _units_df, _snames),
                height=_comp_h, scrolling=True,
            )
        if _view_df is not None and not _view_df.empty:
            with st.expander(f"📋 台別データ（全{len(_view_df)}台）", expanded=False):
                st.dataframe(_view_df, use_container_width=True, hide_index=True, height=520)

    # ── ② 処理内容（常に描画）────────────────────────────────────────
    st.caption("処理内容：① 全台系PNG ＋ 全台プラス機種別JPG　② ジャグラーシリーズ優秀台JPG　③ その他の優秀台ピックアップJPG")

    # ── ② 個別画像（常に描画）────────────────────────────────────────
    kojin_zentai_machines: list[str] = []
    kojin_yushu_machines:  list[str] = []
    kojin_narabi_ranges_text: str = ""
    kojin_narabi_title: str = ""
    kojin_narabi2_ranges_text: str = ""
    kojin_narabi2_title: str = ""
    sonota_extra_title: str = ""
    sonota_extra_text: str = ""
    st.markdown("### ② 個別画像")
    variety_enabled: bool = False
    variety_ranges_text: str = ""
    variety_mode: str = "全台"
    kojin_enabled = st.checkbox("個別画像も生成する", key="kojin_enabled",
                                on_change=_save_auto_inputs, args=(store,))
    if kojin_enabled:
        _kojin_candidates = _excel_candidates
        st.caption("指定した機種の個別画像を生成します。ここに入力した機種はその他の優秀台ピックアップから除外されます。")
        col_kz, col_ky = st.columns(2, gap="large")
        with col_kz:
            st.markdown("**全台**")
            _kz_rows = [st.columns(3) for _ in range(4)]
            for _i, _col in enumerate([c for row in _kz_rows for c in row]):
                with _col:
                    render_machine_autocomplete_input(str(_i + 1), f"kojin_z_{_i}_{store}", _kojin_candidates,
                                                      on_change=_save_auto_inputs, on_change_args=(store,))
            kojin_zentai_machines = [st.session_state.get(f"kojin_z_{_i}_{store}", "") for _i in range(12)]
        with col_ky:
            st.markdown("**優秀台**")
            _akihab_slump = with_slump and store == "秋葉原"
            _ky_count = 21 if _akihab_slump else 12
            _ky_rows = [st.columns(3) for _ in range(_ky_count // 3)]
            for _i, _col in enumerate([c for row in _ky_rows for c in row]):
                with _col:
                    render_machine_autocomplete_input(str(_i + 1), f"kojin_y_{_i}_{store}", _kojin_candidates,
                                                      on_change=_save_auto_inputs, on_change_args=(store,))
            kojin_yushu_machines = [st.session_state.get(f"kojin_y_{_i}_{store}", "") for _i in range(_ky_count)]
        if not (with_slump and store == "秋葉原") and store != "溝の口新館":
            st.markdown("**並び台番範囲 優秀台**")
            _col_nr, _col_nt = st.columns([2, 3])
            with _col_nr:
                st.text_input(
                    "台番範囲（例: 409-413）　ピンクバーあり",
                    key=f"kojin_narabi_range_{store}",
                    placeholder="例: 409-413",
                    on_change=_save_auto_inputs, args=(store,),
                )
            with _col_nt:
                st.text_input(
                    "タイトル（省略時は台番範囲をそのまま使用）",
                    key=f"kojin_narabi_title_{store}",
                    placeholder="例: 4・5列目の優秀台",
                    on_change=_save_auto_inputs, args=(store,),
                )
            _col_nr2, _col_nt2 = st.columns([2, 3])
            with _col_nr2:
                st.text_input(
                    "台番範囲（例: 409-413）　ピンクバーなし",
                    key=f"kojin_narabi2_range_{store}",
                    placeholder="例: 409-413",
                    on_change=_save_auto_inputs, args=(store,),
                )
            with _col_nt2:
                st.text_input(
                    "タイトル（省略時は台番範囲をそのまま使用）",
                    key=f"kojin_narabi2_title_{store}",
                    placeholder="例: 4・5列目の優秀台",
                    on_change=_save_auto_inputs, args=(store,),
                )
        kojin_narabi_ranges_text  = st.session_state.get(f"kojin_narabi_range_{store}", "")
        kojin_narabi_title        = st.session_state.get(f"kojin_narabi_title_{store}", "")
        kojin_narabi2_ranges_text = st.session_state.get(f"kojin_narabi2_range_{store}", "")
        kojin_narabi2_title       = st.session_state.get(f"kojin_narabi2_title_{store}", "")
        if True:  # その他の優秀台ピックアップ（全店舗）
            st.markdown("**その他の優秀台ピックアップ**")
            _col_set, _col_seb = st.columns([2, 3])
            with _col_set:
                st.text_input(
                    "タイトル",
                    value=st.session_state.get(f"sonota_extra_title_{store}", "") or "その他の優秀台ピックアップ",
                    key=f"sonota_extra_title_{store}",
                    placeholder="例: その他の優秀台ピックアップ",
                    on_change=_save_auto_inputs, args=(store,),
                )
            with _col_seb:
                st.text_area(
                    "台番テキスト（台番を含むテキストをそのまま貼り付け）",
                    value=st.session_state.get(f"sonota_extra_text_{store}", ""),
                    key=f"sonota_extra_text_{store}",
                    height=80,
                    on_change=_save_auto_inputs, args=(store,),
                )
        sonota_extra_title = st.session_state.get(f"sonota_extra_title_{store}", "")
        sonota_extra_text  = st.session_state.get(f"sonota_extra_text_{store}", "")

    # ── ③ 並び画像オプション（常に描画）──────────────────────────────
    narabi_ok     = False
    narabi_ranges: list[list[int]] = []
    if store in STORE_NARABI_SCRIPT:
        st.markdown("### ③ 並び画像")
        narabi_enabled = st.checkbox("並び画像も生成する", key="narabi_enabled",
                                     on_change=_save_auto_inputs, args=(store,))
        if narabi_enabled:
            ranges_text = st.text_area(
                "台番範囲　連番: '409-413'、スポット: '508+424'、複数: カンマ/スペース/改行区切り　Excelからのコピペ（台番・機種名・数値の表）もそのまま貼り付け可",
                value="",
                key="narabi_ranges_input",
                height=120,
                on_change=_save_auto_inputs, args=(store,),
            )
            if ranges_text.strip():
                try:
                    _parsed_ranges = parse_ranges(ranges_text.strip())
                    if _parsed_ranges:
                        _prev_key    = f"narabi_previews_{store}"
                        _prev_rt_key = f"narabi_prev_rt_{store}"
                        # 範囲テキストが変わったらプレビューをクリア
                        if st.session_state.get(_prev_rt_key, "") != ranges_text.strip():
                            st.session_state.pop(_prev_key, None)
                            for _ci in range(30):
                                st.session_state.pop(f"narabi_ck_{store}_{_ci}", None)
                        _previews = st.session_state.get(_prev_key)

                        if _previews is None:
                            # プレビュー未生成
                            st.caption(f"並び指定: {_parsed_ranges}")
                            if uploaded is not None:
                                if st.button("🔍 プレビュー生成", key="narabi_preview_btn"):
                                    with st.spinner("プレビュー生成中..."):
                                        _raw_p = _read_uploaded_df(uploaded)
                                        _df_p, _ = normalize_df(_raw_p)
                                        _df_p = apply_name_conversion(_df_p)
                                        _ban_map_p = {int(row["台番"]): i for i, row in _df_p.iterrows()}
                                        _prev_list = []
                                        for _bans in _parsed_ranges:
                                            _idxs = [_ban_map_p[b] for b in _bans if b in _ban_map_p]
                                            if not _idxs:
                                                _prev_list.append(None)
                                                continue
                                            _grp = _df_p.loc[_idxs].copy().reset_index(drop=True)
                                            _ds  = _grp["差枚"]
                                            _ms  = list(dict.fromkeys(str(m) for m in _grp["機種名"]))
                                            _n   = len(_grp)
                                            if len(_ms) == 1:
                                                _tit = f"{_ms[0]}({_n}台並び)"
                                            elif len(_ms) == 2:
                                                _tit = f"{_ms[0]}+{_ms[1]}({_n}台並び)"
                                            else:
                                                _tit = f"{_ms[0]}～{_ms[-1]}({_n}台並び)"
                                            _stat_p = {
                                                "total_diff":  int(_ds.sum()),
                                                "avg_diff":    int(round(_ds.mean())),
                                                "win_count":   int((_ds > 0).sum()),
                                                "total_count": _n,
                                            }
                                            _prev_list.append((_tit, _build_machine_img(_grp, _tit, _stat_p)))
                                        st.session_state[_prev_key]    = _prev_list
                                        st.session_state[_prev_rt_key] = ranges_text.strip()
                                        for _ci in range(len(_prev_list)):
                                            st.session_state[f"narabi_ck_{store}_{_ci}"] = True
                                    st.rerun()
                            narabi_ranges = _parsed_ranges
                            narabi_ok = uploaded is not None
                        else:
                            # プレビュー表示 + チェックボックス
                            st.caption(f"📋 {len(_previews)}件のプレビュー　チェックした並びのみ生成されます")
                            for _ci, _item in enumerate(_previews):
                                _ck_key = f"narabi_ck_{store}_{_ci}"
                                if _item is None:
                                    st.warning(f"⚠️ 範囲 {_parsed_ranges[_ci]} がExcelに見つかりませんでした")
                                    continue
                                _tit, _img = _item
                                _col_ck, _col_img = st.columns([1, 12])
                                with _col_ck:
                                    if _ck_key not in st.session_state:
                                        st.session_state[_ck_key] = True
                                    st.checkbox("", key=_ck_key, label_visibility="collapsed")
                                with _col_img:
                                    st.image(_img, caption=_tit, use_container_width=True)
                            # チェックされた範囲だけ処理対象に
                            narabi_ranges = [
                                _parsed_ranges[_ci]
                                for _ci, _item in enumerate(_previews)
                                if _item is not None and st.session_state.get(f"narabi_ck_{store}_{_ci}", True)
                            ]
                            if narabi_ranges:
                                st.caption(f"✅ {len(narabi_ranges)}件を処理対象に設定")
                                narabi_ok = True
                            else:
                                st.warning("⚠️ 1件以上チェックしてください。")
                    else:
                        st.warning("範囲を正しく認識できませんでした。例: 409-413, 315-317, 508+424")
                except Exception:
                    st.warning("台番範囲の形式が正しくありません。例: 409-413, 315-317")
            else:
                st.info("台番範囲を入力してください。")

    # ── ④ 末尾画像オプション（末尾画像を持つ店舗）──────────────────────
    if "末尾画像" in STORES.get(store, []):
        st.markdown("### ④ 末尾画像")
        suebangai_enabled = st.checkbox("末尾画像も生成する", key="suebangai_enabled",
                                        on_change=_save_auto_inputs, args=(store,))
        if suebangai_enabled:
            _stc1, _stc2, _stc3 = st.columns(3)
            with _stc1:
                _sue_ti1 = st.text_input("末尾①", value="", key="suebangai_tail_input_1",
                                          placeholder="例: 5",
                                          on_change=_save_auto_inputs, args=(store,))
            with _stc2:
                _sue_ti2 = st.text_input("末尾②", value="", key="suebangai_tail_input_2",
                                          placeholder="例: 7",
                                          on_change=_save_auto_inputs, args=(store,))
            with _stc3:
                _sue_ti3 = st.text_input("末尾③", value="", key="suebangai_tail_input_3",
                                          placeholder="ゾロ目",
                                          on_change=_save_auto_inputs, args=(store,))
            _sue_tails_ui = [t.strip() for t in [_sue_ti1, _sue_ti2, _sue_ti3] if t.strip()]
            if _sue_tails_ui:
                _sue_mode_opts = (["全台", "+1,000枚以上の優秀台", "プラス台"]
                                  if with_slump and store == "秋葉原"
                                  else ["全台", "プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）", "プラス台（ピンクバーなし）", "優秀台（ピンクバーなし）"])
                _sue_mode = st.radio("モード", _sue_mode_opts, key="suebangai_mode",
                                     horizontal=True, on_change=_save_auto_inputs, args=(store,))
                _sue_prev_key    = f"sue_preview_{store}"
                _sue_prev_rt_key = f"sue_prev_tails_{store}"
                _sue_cur_rt      = ",".join(_sue_tails_ui) + "|" + _sue_mode
                if st.session_state.get(_sue_prev_rt_key, "") != _sue_cur_rt:
                    st.session_state.pop(_sue_prev_key, None)
                    for _ci in range(20):
                        st.session_state.pop(f"sue_ck_{store}_{_ci}", None)
                _sue_previews = st.session_state.get(_sue_prev_key)
                if _sue_previews is None:
                    if uploaded is not None:
                        if st.button("🔍 プレビュー生成", key="sue_preview_btn", use_container_width=True):
                            with st.spinner("末尾画像のプレビュー生成中..."):
                                try:
                                    _raw = _read_uploaded_df(uploaded)
                                    _df_s, _ = normalize_df(_raw)
                                    nm_s, nm_norm_s = load_name_map()
                                    if nm_s:
                                        _df_s, _ = _apply_map(_df_s, nm_s, nm_norm_s)
                                    _sue_circle_map = {"0":"⓪","1":"①","2":"②","3":"③","4":"④",
                                                       "5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
                                    _sprev_list = []
                                    for _tail in _sue_tails_ui:
                                        if _tail == "ゾロ目":
                                            _filtered = _df_s[_df_s["台番"].apply(
                                                lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                                            )].copy()
                                            _base_label = "末尾ゾロ目の台"
                                        elif _tail.isdigit() and len(_tail) in (1, 2):
                                            _filtered = _df_s[_df_s["台番"].astype(str).str[-len(_tail):] == _tail].copy()
                                            _circle = _sue_circle_map.get(_tail, _tail)
                                            _base_label = f"末尾{_circle}番台"
                                        else:
                                            st.error(f"❌ 「{_tail}」は不正な値です（例: 5、ゾロ目）。")
                                            continue
                                        if _filtered.empty:
                                            st.error(f"❌ {_base_label} の台が見つかりません。")
                                            continue
                                        _is_plus_s   = _sue_mode in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）", "プラス台")
                                        _is_yushu_s  = _sue_mode in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                                        _is_1k_s     = _sue_mode == "+1,000枚以上の優秀台"
                                        if _is_plus_s or _is_yushu_s or _is_1k_s:
                                            _sue_total      = len(_filtered)
                                            _sue_total_diff = int(_filtered["差枚"].sum())
                                            _sue_avg_diff   = int(round(_filtered["差枚"].mean()))
                                            _sue_win_count  = int((_filtered["差枚"] > 0).sum())
                                            if _is_plus_s:
                                                _filtered = _filtered[_filtered["差枚"] > 0].copy()
                                                _img_title = f"{_base_label}のプラス台"
                                            elif _is_1k_s:
                                                _filtered = _filtered[_filtered["差枚"] >= 1000].copy()
                                                _img_title = f"{_base_label}の優秀台"
                                            else:
                                                _sg_col = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _filtered.columns), None)
                                                if _sg_col:
                                                    _smask = (_filtered["差枚"] >= 1000) | ((_filtered[_sg_col] >= 1800) & (_filtered["差枚"] > 0))
                                                else:
                                                    _smask = _filtered["差枚"] >= 1000
                                                _filtered = _filtered[_smask].copy()
                                                _img_title = f"{_base_label}の優秀台"
                                            if _filtered.empty:
                                                st.error(f"❌ {_base_label}で条件を満たす台がありません。")
                                                continue
                                            _has_bar_s = _sue_mode in ("プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）")
                                            if _has_bar_s:
                                                _stat = {
                                                    "total_diff":  _sue_total_diff,
                                                    "avg_diff":    _sue_avg_diff,
                                                    "win_count":   _sue_win_count,
                                                    "total_count": _sue_total,
                                                }
                                            else:
                                                _stat = None
                                        else:
                                            _img_title = _base_label
                                            _stat = {
                                                "total_diff":  int(_filtered["差枚"].sum()),
                                                "avg_diff":    int(round(_filtered["差枚"].mean())),
                                                "win_count":   int((_filtered["差枚"] > 0).sum()),
                                                "total_count": len(_filtered),
                                            }
                                        _img_s = _build_machine_img(_filtered, _img_title, _stat)
                                        _sprev_list.append((f"{_make_safe_fn(_img_title)}.jpg", _img_s))
                                    if _sprev_list:
                                        st.session_state[_sue_prev_key]    = _sprev_list
                                        st.session_state[_sue_prev_rt_key] = _sue_cur_rt
                                        for _ci in range(len(_sprev_list)):
                                            st.session_state[f"sue_ck_{store}_{_ci}"] = True
                                        st.rerun()
                                except Exception as _e:
                                    st.error(f"❌ エラー: {_e}")
                                    with st.expander("詳細"):
                                        st.code(traceback.format_exc())
                else:
                    st.caption(f"📋 {len(_sue_previews)}件のプレビュー　チェックした末尾のみ生成されます")
                    for _ci, (_sfn, _simg) in enumerate(_sue_previews):
                        _sck_key = f"sue_ck_{store}_{_ci}"
                        _scol_ck, _scol_img = st.columns([1, 12])
                        with _scol_ck:
                            if _sck_key not in st.session_state:
                                st.session_state[_sck_key] = True
                            st.checkbox("", key=_sck_key, label_visibility="collapsed")
                        with _scol_img:
                            _stit = os.path.splitext(_sfn)[0]
                            st.image(_simg, caption=_stit, use_container_width=True)
            else:
                st.info("末尾を入力してください。")

        # ジャグラー専用末尾画像
        st.markdown("##### ジャグラー末尾画像")
        jug_sue_enabled = st.checkbox("ジャグラー機種の末尾画像も生成する", key="jug_sue_enabled",
                                       on_change=_save_auto_inputs, args=(store,))
        if jug_sue_enabled:
            _jtc1, _jtc2, _jtc3 = st.columns(3)
            with _jtc1:
                _jt1 = st.text_input("末尾①（ジャグラー）", value="", key="jug_sue_tail_input_1",
                                      placeholder="例: 7",
                                      on_change=_save_auto_inputs, args=(store,))
            with _jtc2:
                _jt2 = st.text_input("末尾②（ジャグラー）", value="", key="jug_sue_tail_input_2",
                                      placeholder="例: 3",
                                      on_change=_save_auto_inputs, args=(store,))
            with _jtc3:
                _jt3 = st.text_input("末尾③（ジャグラー）", value="", key="jug_sue_tail_input_3",
                                      placeholder="ゾロ目",
                                      on_change=_save_auto_inputs, args=(store,))
            _jug_tails_ui = [t.strip() for t in [_jt1, _jt2, _jt3] if t.strip()]
            if _jug_tails_ui:
                _jug_sue_mode = st.radio("モード（ジャグラー）", ["全台", "プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）", "プラス台（ピンクバーなし）", "優秀台（ピンクバーなし）"], key="jug_sue_mode",
                                         horizontal=True, on_change=_save_auto_inputs, args=(store,))
                _jsue_prev_key    = f"jug_sue_preview_{store}"
                _jsue_prev_rt_key = f"jug_sue_prev_tails_{store}"
                _jsue_cur_rt      = ",".join(_jug_tails_ui) + "|" + _jug_sue_mode
                if st.session_state.get(_jsue_prev_rt_key, "") != _jsue_cur_rt:
                    st.session_state.pop(_jsue_prev_key, None)
                    for _ci in range(20):
                        st.session_state.pop(f"jug_sue_ck_{store}_{_ci}", None)
                _jsue_previews = st.session_state.get(_jsue_prev_key)
                if _jsue_previews is None:
                    if uploaded is not None:
                        if st.button("🔍 プレビュー生成（ジャグラー末尾）", key="jug_sue_preview_btn", use_container_width=True):
                            with st.spinner("ジャグラー末尾画像のプレビュー生成中..."):
                                try:
                                    _raw_j = _read_uploaded_df(uploaded)
                                    _df_j, _ = normalize_df(_raw_j)
                                    _nm_j, _nm_norm_j = load_name_map()
                                    if _nm_j:
                                        _df_j, _ = _apply_map(_df_j, _nm_j, _nm_norm_j)
                                    _df_j["ゲーム数_rounded"] = _df_j["ゲーム数"].apply(round_games)
                                    _df_j["合算確率_num"] = _df_j.apply(
                                        lambda r: (r["ゲーム数_rounded"] / (r["BB"] + r["RB"])
                                                   if (r["BB"] + r["RB"]) > 0 else float("inf")),
                                        axis=1,
                                    )
                                    _jcfg      = get_store_config(store)
                                    _jug_ser   = set(_jcfg["juggler_series"])
                                    _jug_g_min = _jcfg["juggler_g_min"]
                                    _jug_thresh = {m: (p, d) for m, p, d in _jcfg["juggler_jobs"]}
                                    _circle_map = {"0":"⓪","1":"①","2":"②","3":"③","4":"④",
                                                   "5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
                                    _jprev_list = []
                                    for _jtail in _jug_tails_ui:
                                        if _jtail == "ゾロ目":
                                            _jfilt = _df_j[_df_j["台番"].apply(
                                                lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                                            )].copy()
                                            _jbase_label = "末尾ゾロ目"
                                            _jcircle = "ゾロ目"
                                        elif _jtail.isdigit() and len(_jtail) in (1, 2):
                                            _jfilt = _df_j[_df_j["台番"].astype(str).str[-len(_jtail):] == _jtail].copy()
                                            _jbase_label = f"末尾{_jtail}"
                                            _jcircle = _circle_map.get(_jtail, _jtail)
                                        else:
                                            st.error(f"❌ 「{_jtail}」は不正な値です（例: 7、ゾロ目）。")
                                            continue
                                        _jfilt = _jfilt[_jfilt["機種名"].isin(_jug_ser)].copy()
                                        _jfilt = _jfilt[_jfilt["ゲーム数_rounded"] >= _jug_g_min].copy()
                                        if _jfilt.empty:
                                            st.error(f"❌ {_jbase_label}番台でジャグラー機種が見つかりません。")
                                            continue
                                        _is_jplus  = _jug_sue_mode in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）")
                                        _is_jyushu = _jug_sue_mode in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                                        if _is_jplus or _is_jyushu:
                                            _jsue_total      = len(_jfilt)
                                            _jsue_total_diff = int(_jfilt["差枚"].sum())
                                            _jsue_avg_diff   = int(round(_jfilt["差枚"].mean()))
                                            _jsue_win_count  = int((_jfilt["差枚"] > 0).sum())
                                            if _is_jplus:
                                                _jfilt = _jfilt[_jfilt["差枚"] > 0].copy()
                                                _jimg_title = f"ジャグラーの末尾{_jcircle}番台のプラス台"
                                            else:
                                                # 優秀台: 確率フィルター + 差枚 > 0
                                                if not _jfilt.empty:
                                                    _p_ser = _jfilt["機種名"].map(
                                                        lambda m: _jug_thresh.get(m, (float("inf"), float("inf")))[0])
                                                    _d_ser = _jfilt["機種名"].map(
                                                        lambda m: _jug_thresh.get(m, (float("inf"), float("inf")))[1])
                                                    _jmask = ((_jfilt["合算確率_num"] <= _p_ser) & (_jfilt["差枚"] >= 0)) | (_jfilt["差枚"] >= _d_ser)
                                                    _jfilt = _jfilt[_jmask].copy().reset_index(drop=True)
                                                _jfilt = _jfilt[_jfilt["差枚"] > 0].copy()
                                                _jimg_title = f"ジャグラーの末尾{_jcircle}番台の優秀台"
                                            if _jfilt.empty:
                                                st.error(f"❌ {_jbase_label}番台のジャグラー条件を満たす台がありません。")
                                                continue
                                            _has_jbar = _jug_sue_mode in ("プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）")
                                            if _has_jbar:
                                                _jstat = {
                                                    "total_diff":  _jsue_total_diff,
                                                    "avg_diff":    _jsue_avg_diff,
                                                    "win_count":   _jsue_win_count,
                                                    "total_count": _jsue_total,
                                                }
                                            else:
                                                _jstat = None
                                        else:
                                            # 全台: G数フィルターのみ（確率・差枚条件なし）
                                            _jimg_title = f"ジャグラーの末尾{_jcircle}番台"
                                            _jstat = {
                                                "total_diff":  int(_jfilt["差枚"].sum()),
                                                "avg_diff":    int(round(_jfilt["差枚"].mean())),
                                                "win_count":   int((_jfilt["差枚"] > 0).sum()),
                                                "total_count": len(_jfilt),
                                            }
                                        _jimg_s = _build_machine_img(_jfilt, _jimg_title, _jstat)
                                        _jprev_list.append((f"{_make_safe_fn(_jimg_title)}.jpg", _jimg_s))
                                    if _jprev_list:
                                        st.session_state[_jsue_prev_key]    = _jprev_list
                                        st.session_state[_jsue_prev_rt_key] = _jsue_cur_rt
                                        for _ci in range(len(_jprev_list)):
                                            st.session_state[f"jug_sue_ck_{store}_{_ci}"] = True
                                        st.rerun()
                                except Exception as _e:
                                    st.error(f"❌ エラー: {_e}")
                                    with st.expander("詳細"):
                                        st.code(traceback.format_exc())
                else:
                    st.caption(f"📋 {len(_jsue_previews)}件のプレビュー　チェックしたジャグラー末尾のみ生成されます")
                    for _ci, (_jfn, _jimg) in enumerate(_jsue_previews):
                        _jck_key = f"jug_sue_ck_{store}_{_ci}"
                        _jcol_ck, _jcol_img = st.columns([1, 12])
                        with _jcol_ck:
                            if _jck_key not in st.session_state:
                                st.session_state[_jck_key] = True
                            st.checkbox("", key=_jck_key, label_visibility="collapsed")
                        with _jcol_img:
                            _jtit = os.path.splitext(_jfn)[0]
                            st.image(_jimg, caption=_jtit, use_container_width=True)
            else:
                st.info("末尾を入力してください。")

    # ── ⑤ バラエティ画像（秋葉原スランプ付きのみ）──────────────────────
    if with_slump and store == "秋葉原":
        st.markdown("### ⑤ バラエティ画像")
        variety_enabled = st.checkbox("個別画像も生成する", key="variety_enabled",
                                      on_change=_save_auto_inputs, args=(store,))
        if variety_enabled:
            st.markdown("**バラエティの台番範囲**")
            _vr_key = f"variety_range_{store}"
            if _vr_key not in st.session_state:
                st.session_state[_vr_key] = ""
            variety_ranges_text = st.text_input(
                "台番範囲（例: 1-50）",
                value=st.session_state[_vr_key],
                key=_vr_key,
                placeholder="例: 1-50",
                on_change=_save_auto_inputs, args=(store,),
            )
            variety_mode = st.radio(
                "モード", ["全台", "+1,000枚以上の優秀台", "プラス台"],
                key="variety_mode",
                horizontal=True,
                on_change=_save_auto_inputs, args=(store,),
            )

    # ── ⑤ オススメ機種ピックアップ（拡張機能店舗）──────────────────────

    recommended_blocks: list[dict] = []
    if store in EXTENDED_FEATURE_STORES:
        _init_recommended_settings(store)
        st.markdown("### ⑤ オススメ機種ピックアップ")
        rec_enabled = st.checkbox("オススメ機種ピックアップを使用する", key=f"rec_enabled_{store}",
                                   on_change=_save_rec_enabled, args=(store,))
        if rec_enabled:
            st.caption(
                "入力した機種のうち、全台系・高配分に該当しない機種の優秀台（プラス台）を画像化します。"
                "　タイトルと機種名は次回起動後も保持されます。"
            )

            _machine_candidates = _excel_candidates
            col_b1, col_b2 = st.columns(2, gap="large")

            with col_b1:
                st.markdown("**ブロック1**")
                title_1 = st.text_input(
                    "タイトル（青バー）",
                    value=st.session_state.get(f"rec_title_1_{store}", ""),
                    key=f"rec_title_1_{store}",
                    placeholder="例: 月間オススメ機種",
                    on_change=_save_rec_titles,
                    args=(store,),
                )
                _b1_top = st.columns(3)
                _b1_mid = st.columns(3)
                _b1_bot = st.columns(3)
                for _i, _col in enumerate(list(_b1_top) + list(_b1_mid) + list(_b1_bot)):
                    with _col:
                        render_machine_autocomplete_input(
                            str(_i + 1), f"rec_m1_{_i}_{store}", _machine_candidates
                        )
                machines_1 = [st.session_state.get(f"rec_m1_{_i}_{store}", "") for _i in range(9)]
                _opts_f1 = ["プラス台", "+1,000枚以上", "+2,000枚以上"]
                _f1_val = st.session_state.get(f"rec_f_1_{store}", "+1,000枚以上")
                _f1_idx = _opts_f1.index(_f1_val) if _f1_val in _opts_f1 else 1
                _sel1 = st.radio("抽出条件", _opts_f1, index=_f1_idx, key=f"rec_f_1_{store}", horizontal=True, on_change=_save_rec_titles, args=(store,))
                thresholds_1 = [{"プラス台": 1, "+1,000枚以上": 1000, "+2,000枚以上": 2000}.get(_sel1, 1)]

            with col_b2:
                st.markdown("**ブロック2**")
                title_2 = st.text_input(
                    "タイトル（青バー）",
                    value=st.session_state.get(f"rec_title_2_{store}", ""),
                    key=f"rec_title_2_{store}",
                    placeholder="例: 週間オススメ機種",
                    on_change=_save_rec_titles,
                    args=(store,),
                )
                _b2_top = st.columns(3)
                _b2_mid = st.columns(3)
                _b2_bot = st.columns(3)
                for _i, _col in enumerate(list(_b2_top) + list(_b2_mid) + list(_b2_bot)):
                    with _col:
                        render_machine_autocomplete_input(
                            str(_i + 1), f"rec_m2_{_i}_{store}", _machine_candidates
                        )
                machines_2 = [st.session_state.get(f"rec_m2_{_i}_{store}", "") for _i in range(9)]
                _sel2 = st.radio("抽出条件", ["プラス台", "+1,000枚以上", "+2,000枚以上"], key=f"rec_f_2_{store}", horizontal=True, on_change=_save_rec_titles, args=(store,))
                thresholds_2 = [{"プラス台": 1, "+1,000枚以上": 1000, "+2,000枚以上": 2000}.get(_sel2, 1)]

            st.markdown("")

            col_b3, col_b4 = st.columns(2, gap="large")

            with col_b3:
                st.markdown("**ブロック3**")
                title_3 = st.text_input(
                    "タイトル（青バー）",
                    value=st.session_state.get(f"rec_title_3_{store}", ""),
                    key=f"rec_title_3_{store}",
                    placeholder="例: 注目機種",
                    on_change=_save_rec_titles,
                    args=(store,),
                )
                _b3_top = st.columns(3)
                _b3_mid = st.columns(3)
                _b3_bot = st.columns(3)
                for _i, _col in enumerate(list(_b3_top) + list(_b3_mid) + list(_b3_bot)):
                    with _col:
                        render_machine_autocomplete_input(
                            str(_i + 1), f"rec_m3_{_i}_{store}", _machine_candidates
                        )
                machines_3 = [st.session_state.get(f"rec_m3_{_i}_{store}", "") for _i in range(9)]
                _sel3 = st.radio("抽出条件", ["プラス台", "+1,000枚以上", "+2,000枚以上"], key=f"rec_f_3_{store}", horizontal=True, on_change=_save_rec_titles, args=(store,))
                thresholds_3 = [{"プラス台": 1, "+1,000枚以上": 1000, "+2,000枚以上": 2000}.get(_sel3, 1)]

            with col_b4:
                st.markdown("**ブロック4**")
                title_4 = st.text_input(
                    "タイトル（青バー）",
                    value=st.session_state.get(f"rec_title_4_{store}", ""),
                    key=f"rec_title_4_{store}",
                    placeholder="例: 特選機種",
                    on_change=_save_rec_titles,
                    args=(store,),
                )
                _b4_top = st.columns(3)
                _b4_mid = st.columns(3)
                _b4_bot = st.columns(3)
                for _i, _col in enumerate(list(_b4_top) + list(_b4_mid) + list(_b4_bot)):
                    with _col:
                        render_machine_autocomplete_input(
                            str(_i + 1), f"rec_m4_{_i}_{store}", _machine_candidates
                        )
                machines_4 = [st.session_state.get(f"rec_m4_{_i}_{store}", "") for _i in range(9)]
                _sel4 = st.radio("抽出条件", ["プラス台", "+1,000枚以上", "+2,000枚以上"], key=f"rec_f_4_{store}", horizontal=True, on_change=_save_rec_titles, args=(store,))
                thresholds_4 = [{"プラス台": 1, "+1,000枚以上": 1000, "+2,000枚以上": 2000}.get(_sel4, 1)]

            st.markdown("")

            col_b5, col_b6 = st.columns(2, gap="large")

            with col_b5:
                st.markdown("**ブロック5**")
                title_5 = st.text_input(
                    "タイトル（青バー）",
                    value=st.session_state.get(f"rec_title_5_{store}", ""),
                    key=f"rec_title_5_{store}",
                    placeholder="例: 注目機種",
                    on_change=_save_rec_titles,
                    args=(store,),
                )
                _b5_top = st.columns(3)
                _b5_mid = st.columns(3)
                _b5_bot = st.columns(3)
                for _i, _col in enumerate(list(_b5_top) + list(_b5_mid) + list(_b5_bot)):
                    with _col:
                        render_machine_autocomplete_input(
                            str(_i + 1), f"rec_m5_{_i}_{store}", _machine_candidates
                        )
                machines_5 = [st.session_state.get(f"rec_m5_{_i}_{store}", "") for _i in range(9)]
                _sel5 = st.radio("抽出条件", ["プラス台", "+1,000枚以上", "+2,000枚以上"], key=f"rec_f_5_{store}", horizontal=True, on_change=_save_rec_titles, args=(store,))
                thresholds_5 = [{"プラス台": 1, "+1,000枚以上": 1000, "+2,000枚以上": 2000}.get(_sel5, 1)]

            with col_b6:
                st.markdown("**ブロック6**")
                title_6 = st.text_input(
                    "タイトル（青バー）",
                    value=st.session_state.get(f"rec_title_6_{store}", ""),
                    key=f"rec_title_6_{store}",
                    placeholder="例: 特選機種",
                    on_change=_save_rec_titles,
                    args=(store,),
                )
                _b6_top = st.columns(3)
                _b6_mid = st.columns(3)
                _b6_bot = st.columns(3)
                for _i, _col in enumerate(list(_b6_top) + list(_b6_mid) + list(_b6_bot)):
                    with _col:
                        render_machine_autocomplete_input(
                            str(_i + 1), f"rec_m6_{_i}_{store}", _machine_candidates
                        )
                machines_6 = [st.session_state.get(f"rec_m6_{_i}_{store}", "") for _i in range(9)]
                _sel6 = st.radio("抽出条件", ["プラス台", "+1,000枚以上", "+2,000枚以上"], key=f"rec_f_6_{store}", horizontal=True, on_change=_save_rec_titles, args=(store,))
                thresholds_6 = [{"プラス台": 1, "+1,000枚以上": 1000, "+2,000枚以上": 2000}.get(_sel6, 1)]

            recommended_blocks = [
                {"title": title_1, "machines": machines_1, "thresholds": thresholds_1},
                {"title": title_2, "machines": machines_2, "thresholds": thresholds_2},
                {"title": title_3, "machines": machines_3, "thresholds": thresholds_3},
                {"title": title_4, "machines": machines_4, "thresholds": thresholds_4},
                {"title": title_5, "machines": machines_5, "thresholds": thresholds_5},
                {"title": title_6, "machines": machines_6, "thresholds": thresholds_6},
            ]

    # ── ⑥ 結果テキスト素材メモ（拡張機能店舗）─────────────────────────
    if store in EXTENDED_FEATURE_STORES:
        st.markdown("### ⑥ 結果テキスト素材メモ")
        memo_enabled = st.checkbox("結果テキスト素材メモを使用する", key=f"memo_enabled_{store}")
        if memo_enabled:
            st.caption(
                "機種名（・で始まる行）＋ ①②… セクション ＋ 台番リストを入力すると、"
                "差枚データを付与した文章を「🎁その他の優秀台」の直前に自動挿入します。"
                "　空欄なら何も挿入しません。"
            )
            st.text_area(
                "素材メモ（機種名・台番・コメントを入力）",
                key=f"result_extra_note_{store}",
                height=200,
                placeholder=(
                    "①塊（ブドウ図柄沢山）\n"
                    "2027-2029　2045-2048\n"
                    "\n"
                    "②上げ（ツノッチが階段を上がる）\n"
                    "2024.2029.2037.2039\n"
                    "\n"
                    "③カド2（つの2本）\n"
                    "2024.2027.2047"
                ),
            )

    # ── ⑦ プレビュー ────────────────────────────────────────────────
    st.markdown("### ⑦ プレビュー")
    if uploaded is not None:
        _aprev_key    = f"auto_preview_imgs_{store}"
        _aprev_fname  = f"auto_preview_fname_{store}"
        _aprev_df_key = f"auto_preview_df_{store}"
        _aprev_di_key = f"auto_preview_diff_{store}"
        _aprev_ex_key = f"auto_preview_ex_{store}"
        _aprev_hr_key       = f"auto_preview_hr_{store}"
        _aprev_zen_key      = f"auto_preview_zen_{store}"
        _aprev_jug_ex_key   = f"auto_preview_jug_ex_{store}"
        _aprev_jug_pool_key = f"auto_preview_jug_pool_{store}"
        _aprev_jug_ov_key   = f"auto_preview_jug_ov_{store}"
        _aprev_narabi_key   = f"auto_preview_narabi_{store}"
        _aprev_hr_img_key   = f"auto_preview_hr_img_{store}"
        # Excel が変わったらプレビューをクリア
        if st.session_state.get(_aprev_fname) != uploaded.name:
            for _k in (_aprev_key, _aprev_df_key, _aprev_di_key, _aprev_ex_key, _aprev_hr_key, _aprev_zen_key, _aprev_jug_ex_key, _aprev_jug_pool_key, _aprev_jug_ov_key, _aprev_narabi_key):
                st.session_state.pop(_k, None)
            st.session_state.pop(f"sue_preview_{store}", None)
            st.session_state.pop(f"sue_prev_tails_{store}", None)
            st.session_state.pop(f"jug_sue_preview_{store}", None)
            st.session_state.pop(f"jug_sue_prev_tails_{store}", None)
            for _ci in range(20):
                st.session_state.pop(f"sue_ck_{store}_{_ci}", None)
                st.session_state.pop(f"jug_sue_ck_{store}_{_ci}", None)
            st.session_state[_aprev_fname] = uploaded.name

        _auto_previews = st.session_state.get(_aprev_key)
        if _auto_previews is None:
            if not with_slump:
                _mc1, _mc2 = st.columns(2)
                with _mc1:
                    _full_prev_btn = st.button("🔍 プレビュー生成", key="auto_preview_btn", use_container_width=True)
                with _mc2:
                    _manual_prev_btn = st.button("📝 記入部分のみプレビュー作成", key="manual_only_preview_btn", use_container_width=True)
            else:
                _full_prev_btn = st.button("🔍 プレビュー生成", key="auto_preview_btn")
                _manual_prev_btn = False
            if _full_prev_btn:
                st.session_state.pop(f"_manual_preview_mode_{store}", None)
                _save_auto_inputs(store)
                with st.spinner("画像を生成中（しばらくお待ちください）…"):
                    import tempfile as _tempfile
                    _excel_bytes = uploaded.getvalue()
                    uploaded.seek(0)
                    _prev_narabi_bans: set[int] = ranges_to_bans(narabi_ranges) if narabi_ok else set()
                    if kojin_enabled and kojin_narabi_ranges_text.strip():
                        try:
                            _prev_narabi_bans |= ranges_to_bans(parse_ranges(kojin_narabi_ranges_text.strip()))
                        except Exception:
                            pass
                    if kojin_enabled and kojin_narabi2_ranges_text.strip():
                        try:
                            _prev_narabi_bans |= ranges_to_bans(parse_ranges(kojin_narabi2_ranges_text.strip()))
                        except Exception:
                            pass
                    with _tempfile.TemporaryDirectory() as _tmpdir:
                        _tmp_excel = os.path.join(_tmpdir, uploaded.name)
                        with open(_tmp_excel, "wb") as _tf:
                            _tf.write(_excel_bytes)
                        _prev_rec_names: set[str] = {
                            m.strip()
                            for block in recommended_blocks
                            for m in block["machines"]
                            if m.strip()
                        }
                        if kojin_enabled:
                            _prev_rec_names |= {m.strip() for m in kojin_zentai_machines if m.strip()}
                            _prev_rec_names |= {m.strip() for m in kojin_yushu_machines if m.strip()}
                        _prev_sue_tails: list[str] = []
                        if st.session_state.get("suebangai_enabled", False):
                            _prev_sue_tails += [t for i in range(1, 4) if (t := st.session_state.get(f"suebangai_tail_input_{i}", "").strip())]
                        _prev_jug_sue_tails: list[str] = []
                        if st.session_state.get("jug_sue_enabled", False):
                            _prev_jug_sue_tails += [t for i in range(1, 4) if (t := st.session_state.get(f"jug_sue_tail_input_{i}", "").strip())]
                        _prev_result = run_auto_pipeline(
                            _tmp_excel, _tmpdir, store, _prev_narabi_bans,
                            lambda _m: None,
                            narabi_ranges=narabi_ranges if narabi_ok else None,
                            recommended_machines=_prev_rec_names,
                            suebangai_tails=_prev_sue_tails,
                            sonota_exclude={m.strip() for block in recommended_blocks for m in block["machines"] if m.strip()},
                            jug_suebangai_tails=_prev_jug_sue_tails,
                            variety_bans=(ranges_to_bans(parse_ranges(variety_ranges_text.strip())) if (with_slump and store == "秋葉原" and variety_enabled and variety_ranges_text.strip()) else set()),
                        )
                        # スランプ付き: その他の優秀台ピックアップ①②(③)生成（プレビュー用・秋葉原/上野新館）
                        if _sonota_split and _prev_result.get("ok"):
                            _s3_old_pv = os.path.join(_tmpdir, "その他の優秀台ピックアップ.jpg")
                            _s3_1_pv   = os.path.join(_tmpdir, "その他の優秀台+1,000枚以上.jpg")
                            if os.path.exists(_s3_old_pv):
                                os.replace(_s3_old_pv, _s3_1_pv)
                                _rfl_pv = _prev_result["files"]
                                for _ri_pv in range(len(_rfl_pv)):
                                    if os.path.basename(_rfl_pv[_ri_pv]) == "その他の優秀台ピックアップ.jpg":
                                        _rfl_pv[_ri_pv] = _s3_1_pv
                                        break
                            _pv_s3_df_g = _prev_result.get("df")
                            _pv_s3_dr_g = _prev_result.get("diff_raw")
                            _pv_1_bans  = sorted({int(_e["ban"]) for _e in _prev_result.get("sonota_excellent_list", []) if "ban" in _e})
                            if _pv_s3_df_g is not None and _pv_s3_dr_g is not None and _pv_1_bans:
                                for _thr_pv, _fn_pv in _sonota_extra_thrs:
                                    _out_pv = os.path.join(_tmpdir, _fn_pv)
                                    _s3_k_pv = [
                                        _b for _b in _pv_1_bans
                                        if not (_pv_s3_df_g[_pv_s3_df_g["台番"] == _b]).empty
                                        and int(_pv_s3_dr_g.loc[_pv_s3_df_g[_pv_s3_df_g["台番"] == _b].index[0]]) >= _thr_pv
                                    ]
                                    if _s3_k_pv:
                                        _s3_k_df_pv  = _pv_s3_df_g[_pv_s3_df_g["台番"].isin(_s3_k_pv)].copy().reset_index(drop=True)
                                        _s3_k_img_pv = _build_machine_img(_s3_k_df_pv, "その他の優秀台ピックアップ", None)
                                        _save_jpeg(_s3_k_img_pv, _out_pv, target_kb=800)
                                        _prev_result["files"].append(_out_pv)
                        _prev_img_list: list[tuple[str, "Image.Image"]] = []
                        if _prev_result["ok"]:
                            # パイプライン出力JPGをbasename→(name, Image)辞書に読み込む
                            _fp_map: dict[str, tuple[str, "Image.Image"]] = {}
                            for _fp in _prev_result["files"]:
                                if os.path.exists(_fp) and _fp.lower().endswith((".jpg", ".jpeg")):
                                    _bn = os.path.basename(_fp)
                                    _fp_map[_bn] = (_bn, Image.open(_fp).copy())

                            _pv_df   = _prev_result.get("df")
                            _pv_diff = _prev_result.get("diff_raw")

                            # ─ ① 全台系（平均差枚 大→小・個別全台系を含む）─
                            _zen_pv_items: list[tuple[int, str, object]] = []
                            for _item in _prev_result.get("zen_dai_list", []):
                                _zen_pv_items.append((_item.get("all_avg_diff", 0), "pipeline", _item))
                            if kojin_enabled and _pv_df is not None and _pv_diff is not None:
                                for _km in kojin_zentai_machines:
                                    _km = _km.strip()
                                    if not _km:
                                        continue
                                    _kgrp = _pv_df[_pv_df["機種名"] == _km].copy().reset_index(drop=True)
                                    if _kgrp.empty:
                                        continue
                                    _kdr = _pv_diff.loc[_pv_df[_pv_df["機種名"] == _km].index].reset_index(drop=True)
                                    _kavg = int(round(_kdr.mean()))
                                    _zen_pv_items.append((_kavg, "kojin", (_km, _kgrp, _kdr)))
                            for _avg, _typ, _data in sorted(_zen_pv_items, key=lambda x: x[0], reverse=True):
                                if _typ == "pipeline":
                                    _fn = f"{_make_safe_fn(_data['name'])}.jpg"
                                    if _fn in _fp_map:
                                        _prev_img_list.append(_fp_map[_fn])
                                else:
                                    _km, _kgrp, _kdr = _data
                                    _prev_img_list.append((f"{_km}.jpg", _build_machine_img(_kgrp, _km, _stat_from_diff(_kdr))))

                            # ─ ② 高配分（平均差枚 大→小・個別優秀台を含む）─
                            def _hr_sort_key(x):
                                if "all_avg_diff" in x:
                                    return x["all_avg_diff"]
                                return int(round(sum(x["diffs"]) / len(x["diffs"]))) if x.get("diffs") else 0
                            _hr_pv_items: list[tuple[int, str, object]] = []
                            for _item in _prev_result.get("high_ratio_list", []):
                                _hr_pv_items.append((_hr_sort_key(_item), "pipeline", _item))
                            if kojin_enabled and _pv_df is not None and _pv_diff is not None:
                                for _km in kojin_yushu_machines:
                                    _km = _km.strip()
                                    if not _km:
                                        continue
                                    _kgrp_all = _pv_df[_pv_df["機種名"] == _km]
                                    if _kgrp_all.empty:
                                        continue
                                    _kdr_all = _pv_diff.loc[_kgrp_all.index]
                                    _kgrp_p = _kojin_yushu_filter(_km, _kgrp_all, _kdr_all, get_store_config(store), force_1k=(with_slump and store == "秋葉原")).reset_index(drop=True)
                                    if _kgrp_p.empty:
                                        continue
                                    _kavg = int(round(_kdr_all.mean()))
                                    _hr_pv_items.append((_kavg, "kojin", (_km, _kgrp_p)))
                            for _avg, _typ, _data in sorted(_hr_pv_items, key=lambda x: x[0], reverse=True):
                                if _typ == "pipeline":
                                    _fn = f"{_make_safe_fn(_data['name'])}_高配分.jpg"
                                    if _fn in _fp_map:
                                        _prev_img_list.append(_fp_map[_fn])
                                else:
                                    _km, _kgrp_p = _data
                                    _ktitle = f"{_km}（優秀台）"
                                    _prev_img_list.append((f"{_ktitle}.jpg", _build_machine_img(_kgrp_p, _ktitle, None)))

                            # ─ ③ 並び画像（narabi_ranges）+ 個別並び ─
                            if narabi_ok and narabi_ranges and _pv_df is not None:
                                _ban_map_pv = {int(row["台番"]): i for i, row in _pv_df.iterrows()}
                                # narabi スクリプトと同じ重複タイトル検出
                                def _make_ntit(nms, nn):
                                    if len(nms) == 1:
                                        return f"{nms[0]}({nn}台並び)"
                                    elif len(nms) == 2:
                                        return f"{nms[0]}+{nms[1]}({nn}台並び)"
                                    return f"{nms[0]}～{nms[-1]}({nn}台並び)"
                                _nb_infos = []
                                for _bans in narabi_ranges:
                                    _idxs = [_ban_map_pv[b] for b in _bans if b in _ban_map_pv]
                                    if not _idxs:
                                        continue
                                    _ngrp = _pv_df.loc[_idxs].copy().reset_index(drop=True)
                                    _nms  = list(dict.fromkeys(str(m) for m in _ngrp["機種名"]))
                                    _nb_infos.append((_ngrp, _nms, _make_ntit(_nms, len(_ngrp))))
                                from collections import Counter as _Ctr
                                _dup_tits = {t for t, c in _Ctr(i[2] for i in _nb_infos).items() if c > 1}
                                _narabi_ban_map: dict[str, list[int]] = {}
                                for _ngrp, _nms, _ntit in _nb_infos:
                                    _nds = _ngrp["差枚"]
                                    _nn  = len(_ngrp)
                                    if _ntit in _dup_tits:
                                        _bs = int(_ngrp.iloc[0]["台番"])
                                        _be = int(_ngrp.iloc[-1]["台番"])
                                        _file_tit = f"{_ntit}（{_bs}～{_be}）"
                                    else:
                                        _file_tit = _ntit
                                    _narabi_ban_map[f"{_file_tit}.jpg"] = [int(b) for b in _ngrp["台番"].tolist()]
                                    _nstat = {
                                        "total_diff":  int(_nds.sum()),
                                        "avg_diff":    int(round(_nds.mean())),
                                        "win_count":   int((_nds > 0).sum()),
                                        "total_count": _nn,
                                    }
                                    _prev_img_list.append((f"{_file_tit}.jpg", _build_machine_img(_ngrp, _ntit, _nstat)))
                            if kojin_enabled and _pv_df is not None and _pv_diff is not None:
                                if kojin_narabi_ranges_text.strip():
                                    try:
                                        _rng_bans = ranges_to_bans(parse_ranges(kojin_narabi_ranges_text.strip()))
                                        _rng_df   = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _rng_bans)].copy()
                                        _rng_diff = _pv_diff.loc[_rng_df.index]
                                        _rng_p    = _rng_df.copy().reset_index(drop=True)
                                        if not _rng_p.empty:
                                            _base = kojin_narabi_title.strip() or f"{kojin_narabi_ranges_text.strip()}の優秀台"
                                            _prev_img_list.append((f"{_base}.jpg", _build_machine_img(_rng_p, _base, _stat_from_diff(_rng_diff))))
                                    except Exception:
                                        pass
                                if kojin_narabi2_ranges_text.strip():
                                    try:
                                        _rng2_bans = ranges_to_bans(parse_ranges(kojin_narabi2_ranges_text.strip()))
                                        _rng2_df   = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _rng2_bans)].copy()
                                        _rng2_diff = _pv_diff.loc[_rng2_df.index]
                                        _rng2_p    = _rng2_df.copy().reset_index(drop=True)
                                        if not _rng2_p.empty:
                                            _base2 = kojin_narabi2_title.strip() or f"{kojin_narabi2_ranges_text.strip()}の優秀台"
                                            _prev_img_list.append((f"{_base2}.jpg", _build_machine_img(_rng2_p, _base2, None)))
                                    except Exception:
                                        pass

                            # ─ ⑤ バラエティ画像（秋葉原スランプ付きのみ）─
                            _variety_ban_map: dict[str, list[int]] = {}
                            if with_slump and store == "秋葉原" and variety_enabled and variety_ranges_text.strip() and _pv_df is not None and _pv_diff is not None:
                                try:
                                    _var_bans = ranges_to_bans(parse_ranges(variety_ranges_text.strip()))
                                    _var_df = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _var_bans)].copy()
                                    if not _var_df.empty:
                                        _var_dr = _pv_diff.loc[_var_df.index]
                                        if variety_mode == "プラス台":
                                            _vmask = _var_dr > 0
                                            _var_df = _var_df[_vmask.values].copy()
                                            _var_dr = _var_dr[_vmask]
                                            _var_title = "バラエティのプラス台"
                                        elif variety_mode == "+1,000枚以上の優秀台":
                                            _vmask = _var_dr >= 1000
                                            _var_df = _var_df[_vmask.values].copy()
                                            _var_dr = _var_dr[_vmask]
                                            _var_title = "バラエティの優秀台"
                                        else:
                                            _var_title = "バラエティ"
                                        if not _var_df.empty:
                                            _var_sort = _var_df["台番"].argsort().values
                                            _var_df = _var_df.iloc[_var_sort].reset_index(drop=True)
                                            _var_dr = _var_dr.iloc[_var_sort].reset_index(drop=True)
                                            _var_stat = {"total_diff": int(_var_dr.sum()), "avg_diff": int(round(_var_dr.mean())), "win_count": int((_var_dr > 0).sum()), "total_count": len(_var_df)} if variety_mode == "全台" else None
                                            _var_fn = f"{_make_safe_fn(_var_title)}.jpg"
                                            _prev_img_list.append((_var_fn, _build_machine_img(_var_df, _var_title, _var_stat)))
                                            _variety_ban_map[_var_fn] = [int(b) for b in _var_df["台番"].dropna()]
                                except Exception:
                                    pass

                            # ─ ④ 末尾・ジャグラー末尾画像（プレビュー済みはチェック済みのみ、未生成は直接生成）─
                            def _gen_sue_imgs_on_fly(tails, mode, is_juggler=False):
                                _imgs = []
                                try:
                                    _raw_of = _read_uploaded_df(uploaded)
                                    _df_of, _ = normalize_df(_raw_of)
                                    _nm_of, _nm_norm_of = load_name_map()
                                    if _nm_of:
                                        _df_of, _ = _apply_map(_df_of, _nm_of, _nm_norm_of)
                                    _circle_of = {"0":"⓪","1":"①","2":"②","3":"③","4":"④",
                                                  "5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
                                    if is_juggler:
                                        _df_of["ゲーム数_rounded"] = _df_of["ゲーム数"].apply(round_games)
                                        _df_of["合算確率_num"] = _df_of.apply(
                                            lambda r: r["ゲーム数_rounded"] / (r["BB"] + r["RB"])
                                                      if (r["BB"] + r["RB"]) > 0 else float("inf"), axis=1)
                                        _jcfg_of = get_store_config(store)
                                        _jug_ser_of = set(_jcfg_of["juggler_series"])
                                        _jug_g_min_of = _jcfg_of["juggler_g_min"]
                                        _jug_thresh_of = {m: (p, d) for m, p, d in _jcfg_of["juggler_jobs"]}
                                    for _t in tails:
                                        if _t == "ゾロ目":
                                            _filt = _df_of[_df_of["台番"].apply(
                                                lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                                            )].copy()
                                            _circ = "ゾロ目"
                                            _lbl_base = "末尾ゾロ目" if is_juggler else "末尾ゾロ目の台"
                                        elif _t.isdigit() and len(_t) in (1, 2):
                                            _filt = _df_of[_df_of["台番"].astype(str).str[-len(_t):] == _t].copy()
                                            _circ = _circle_of.get(_t, _t)
                                            _lbl_base = f"末尾{_circ}" if is_juggler else f"末尾{_circ}番台"
                                        else:
                                            continue
                                        if _filt.empty:
                                            continue
                                        if is_juggler:
                                            _filt = _filt[_filt["機種名"].isin(_jug_ser_of)].copy()
                                            _filt = _filt[_filt["ゲーム数_rounded"] >= _jug_g_min_of].copy()
                                            if _filt.empty:
                                                continue
                                            _jp_of  = mode in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）")
                                            _jy_of  = mode in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                                            _jb_of  = mode in ("プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）")
                                            if _jp_of or _jy_of:
                                                _jof_total = len(_filt); _jof_td = int(_filt["差枚"].sum()); _jof_ad = int(round(_filt["差枚"].mean())); _jof_wc = int((_filt["差枚"] > 0).sum())
                                                if _jp_of:
                                                    _filt = _filt[_filt["差枚"] > 0].copy()
                                                    _title = f"ジャグラーの{_lbl_base}番台のプラス台"
                                                else:
                                                    # 優秀台: 確率フィルター + 差枚 > 0
                                                    if not _filt.empty:
                                                        _ps = _filt["機種名"].map(
                                                            lambda m: _jug_thresh_of.get(m, (float("inf"), float("inf")))[0])
                                                        _ds = _filt["機種名"].map(
                                                            lambda m: _jug_thresh_of.get(m, (float("inf"), float("inf")))[1])
                                                        _filt = _filt[((_filt["合算確率_num"] <= _ps) & (_filt["差枚"] >= 0)) |
                                                                      (_filt["差枚"] >= _ds)].copy().reset_index(drop=True)
                                                    _filt = _filt[_filt["差枚"] > 0].copy()
                                                    _title = f"ジャグラーの{_lbl_base}番台の優秀台"
                                                if _filt.empty:
                                                    continue
                                                _stat_of = {"total_diff": _jof_td, "avg_diff": _jof_ad, "win_count": _jof_wc, "total_count": _jof_total} if _jb_of else None
                                            else:
                                                # 全台: G数フィルターのみ（確率・差枚条件なし）
                                                _title = f"ジャグラーの{_lbl_base}番台"
                                                _stat_of = {"total_diff": int(_filt["差枚"].sum()),
                                                            "avg_diff": int(round(_filt["差枚"].mean())),
                                                            "win_count": int((_filt["差枚"] > 0).sum()),
                                                            "total_count": len(_filt)}
                                        else:
                                            _p_of  = mode in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）", "プラス台")
                                            _y_of  = mode in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                                            _1k_of = mode == "+1,000枚以上の優秀台"
                                            _b_of  = mode in ("プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）")
                                            if _p_of or _y_of or _1k_of:
                                                _of_total = len(_filt); _of_td = int(_filt["差枚"].sum()); _of_ad = int(round(_filt["差枚"].mean())); _of_wc = int((_filt["差枚"] > 0).sum())
                                                if _p_of:
                                                    _filt = _filt[_filt["差枚"] > 0].copy()
                                                    _title = f"{_lbl_base}のプラス台"
                                                elif _1k_of:
                                                    _filt = _filt[_filt["差枚"] >= 1000].copy()
                                                    _title = f"{_lbl_base}の優秀台"
                                                else:
                                                    _og_col = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _filt.columns), None)
                                                    _ofm = (_filt["差枚"] >= 1000) | ((_filt[_og_col] >= 1800) & (_filt["差枚"] > 0)) if _og_col else (_filt["差枚"] >= 1000)
                                                    _filt = _filt[_ofm].copy()
                                                    _title = f"{_lbl_base}の優秀台"
                                                if _filt.empty:
                                                    continue
                                                _stat_of = {"total_diff": _of_td, "avg_diff": _of_ad, "win_count": _of_wc, "total_count": _of_total} if _b_of else None
                                            else:
                                                _title = _lbl_base
                                                _stat_of = {"total_diff": int(_filt["差枚"].sum()),
                                                            "avg_diff": int(round(_filt["差枚"].mean())),
                                                            "win_count": int((_filt["差枚"] > 0).sum()),
                                                            "total_count": len(_filt)}
                                        _imgs.append((f"{_make_safe_fn(_title)}.jpg",
                                                      _build_machine_img(_filt, _title, _stat_of)))
                                except Exception:
                                    pass
                                return _imgs

                            _sue_prevs = st.session_state.get(f"sue_preview_{store}", [])
                            if _sue_prevs:
                                for _ci, (_sm_fn, _sm_img) in enumerate(_sue_prevs):
                                    if st.session_state.get(f"sue_ck_{store}_{_ci}", True):
                                        _prev_img_list.append((_sm_fn, _sm_img))
                            elif st.session_state.get("suebangai_enabled", False):
                                _sue_tails_of = [t for i in range(1, 4)
                                                 if (t := st.session_state.get(f"suebangai_tail_input_{i}", "").strip())]
                                _sue_mode_of = st.session_state.get("suebangai_mode", "全台")
                                for _item in _gen_sue_imgs_on_fly(_sue_tails_of, _sue_mode_of, is_juggler=False):
                                    _prev_img_list.append(_item)

                            _jsue_prevs = st.session_state.get(f"jug_sue_preview_{store}", [])
                            if _jsue_prevs:
                                for _ci, (_jm_fn, _jm_img) in enumerate(_jsue_prevs):
                                    if st.session_state.get(f"jug_sue_ck_{store}_{_ci}", True):
                                        _prev_img_list.append((_jm_fn, _jm_img))
                            elif st.session_state.get("jug_sue_enabled", False):
                                _jug_tails_of = [t for i in range(1, 4)
                                                 if (t := st.session_state.get(f"jug_sue_tail_input_{i}", "").strip())]
                                _jug_sue_mode_of = st.session_state.get("jug_sue_mode", "全台")
                                for _item in _gen_sue_imgs_on_fly(_jug_tails_of, _jug_sue_mode_of, is_juggler=True):
                                    _prev_img_list.append(_item)

                            # ─ ⑤ ジャグラーシリーズ優秀台（秋葉原スランプ付きは除外）─
                            if "ジャグラーシリーズ優秀台.jpg" in _fp_map and not (with_slump and store == "秋葉原"):
                                _prev_img_list.append(_fp_map["ジャグラーシリーズ優秀台.jpg"])

                            # ─ ⑥ その他の優秀台ピックアップ + オススメ ─
                            if _sonota_split:
                                for _s3fn in ["その他の優秀台+1,000枚以上.jpg"] + [_f for _, _f in _sonota_extra_thrs]:
                                    if _s3fn in _fp_map:
                                        _prev_img_list.append(_fp_map[_s3fn])
                            else:
                                if "その他の優秀台ピックアップ.jpg" in _fp_map:
                                    _prev_img_list.append(_fp_map["その他の優秀台ピックアップ.jpg"])
                            # 手動入力の「その他の優秀台ピックアップ」（全店舗・⑧実行と同ロジック）
                            if sonota_extra_text.strip() and _pv_df is not None:
                                _se_bans_pv = set(expand_machine_numbers(sonota_extra_text))
                                if _se_bans_pv:
                                    _se_df_pv = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _se_bans_pv)].copy().reset_index(drop=True)
                                    if not _se_df_pv.empty:
                                        _se_tit_pv = sonota_extra_title.strip() or "その他の優秀台ピックアップ"
                                        _se_fn_pv  = f"{_make_safe_fn(_se_tit_pv)}.jpg"
                                        # 同名のパイプライン版が既にあれば置き換える
                                        _prev_img_list = [(_n, _im) for (_n, _im) in _prev_img_list if _n != _se_fn_pv]
                                        _prev_img_list.append((_se_fn_pv, _build_machine_img(_se_df_pv, _se_tit_pv, None)))
                            _rec_ban_map: dict[str, list[int]] = {}
                            if recommended_blocks and _pv_df is not None and _pv_diff is not None:
                                _pv_scfg = get_store_config(store)
                                _pv_jug_cfg = {
                                    "series":     _pv_scfg["juggler_series"],
                                    "jobs_map":   {j[0]: j[1] for j in _pv_scfg["juggler_jobs"]},
                                    "g_min":      _pv_scfg["juggler_g_min"],
                                    "diff_bonus": _pv_scfg["diff_bonus"],
                                }
                                _pv_zen   = {item["name"] for item in _prev_result.get("zen_dai_list", [])}
                                _pv_high  = {item["name"] for item in _prev_result.get("high_ratio_list", []) if item.get("has_image", True)}
                                if kojin_enabled:
                                    _pv_zen  |= {m.strip() for m in kojin_zentai_machines if m.strip()}
                                    _pv_high |= {m.strip() for m in kojin_yushu_machines if m.strip()}
                                _sfx_map  = {1: "プラス台", 1000: "1000枚以上", 2000: "2000枚以上"}
                                for _block in recommended_blocks:
                                    _bt = _block["title"].strip()
                                    _bm = _block["machines"]
                                    if not _bt and not any(m.strip() for m in _bm):
                                        continue
                                    if not _bt:
                                        _bt = "オススメ機種"
                                    _valid, _ = filter_recommended_machines(_bm, _pv_df, _pv_zen, _pv_high)
                                    if not _valid:
                                        continue
                                    for _thr in _block.get("thresholds", [1]):
                                        _rec_img = generate_recommended_block_image(
                                            _bt, _valid, _pv_df, _pv_diff, _prev_narabi_bans,
                                            min_diff=_thr, juggler_cfg=_pv_jug_cfg
                                        )
                                        if _rec_img is None:
                                            continue
                                        _sfx = _sfx_map.get(_thr, str(_thr))
                                        _rec_fn = f"オススメ_{_make_safe_fn(_bt)}_{_sfx}.jpg"
                                        _prev_img_list.append((_rec_fn, _rec_img))
                                        # スランプグラフ合成用：画像に含まれる台番を収集
                                        _rec_bans: list[int] = []
                                        for _rvm in _valid:
                                            _rgrp = _pv_df[_pv_df["機種名"] == _rvm].copy()
                                            if _prev_narabi_bans:
                                                _rgrp = _rgrp[~_rgrp["台番"].isin(_prev_narabi_bans)]
                                            if _rgrp.empty:
                                                continue
                                            _rdr = _pv_diff.loc[_rgrp.index]
                                            if _pv_jug_cfg and _rvm in _pv_jug_cfg.get("series", set()):
                                                _jg_col = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _rgrp.columns), None)
                                                if _jg_col:
                                                    _jgmask = _rgrp[_jg_col] >= _pv_jug_cfg.get("g_min", 2000)
                                                    _rgrp = _rgrp[_jgmask]
                                                    _rdr  = _rdr[_jgmask]
                                                _jpthr = _pv_jug_cfg["jobs_map"].get(_rvm)
                                                if _jpthr is not None and "合算確率_num" in _rgrp.columns:
                                                    _rmask = ((_rgrp["合算確率_num"] <= _jpthr) & (_rdr >= 0)) | (_rdr >= _pv_jug_cfg.get("diff_bonus", 1000))
                                                else:
                                                    _rmask = _rdr >= 0
                                            else:
                                                _rmask = _rdr >= _thr
                                            _rec_bans.extend([int(b) for b in _rgrp[_rmask]["台番"].dropna()])
                                        if _rec_bans:
                                            _rec_ban_map[_rec_fn] = sorted(_rec_bans)

                    # ── with_slump=True の場合: pisionデータ取得 → スランプグラフ合成 ──
                    if with_slump and _prev_result.get("ok"):
                        _pv_ban_map: dict[str, list[int]] = {}
                        for _zd2 in _prev_result.get("zen_dai_list", []):
                            _pv_ban_map[f"{_make_safe_fn(_zd2['name'])}.jpg"] = _zd2.get("bans", [])
                        for _hr2 in _prev_result.get("high_ratio_list", []):
                            if _hr2.get("has_image", False):
                                _pv_ban_map[f"{_make_safe_fn(_hr2['name'])}_高配分.jpg"] = _hr2.get("bans", [])
                        _jug_pool_pv = _prev_result.get("jug_pool_df")
                        if _jug_pool_pv is not None and not _jug_pool_pv.empty:
                            if with_slump and store == "秋葉原":
                                pass  # ジャグラー統合画像なし → sonota側に追加
                            else:
                                _pv_ban_map["ジャグラーシリーズ優秀台.jpg"] = [
                                    int(str(_b2).split(".")[0]) for _b2 in _jug_pool_pv["台番"].dropna()
                                    if str(_b2).split(".")[0].lstrip("-").isdigit()
                                ]
                        _son_bans_pv = sorted({int(_e2["ban"]) for _e2 in _prev_result.get("sonota_excellent_list", []) if "ban" in _e2})
                        # 秋葉原スランプ付き: jug_pool +1000枚台をsonota bansに追加
                        if with_slump and store == "秋葉原" and _jug_pool_pv is not None and not _jug_pool_pv.empty:
                            _pv_diff_jp = _prev_result.get("diff_raw")
                            _pv_df_jp   = _prev_result.get("df")
                            if _pv_df_jp is not None and _pv_diff_jp is not None:
                                _jp_bns_pv = {int(str(b).split(".")[0]) for b in _jug_pool_pv["台番"].dropna()
                                              if str(b).split(".")[0].lstrip("-").isdigit()}
                                _jp_rows_pv = _pv_df_jp[_pv_df_jp["台番"].apply(lambda b: int(b) in _jp_bns_pv)]
                                if not _jp_rows_pv.empty:
                                    _jp_dr_pv = _pv_diff_jp.loc[_jp_rows_pv.index]
                                    _jp_1k_pv = [int(b) for b in _jp_rows_pv[_jp_dr_pv.values >= 1000]["台番"].dropna()]
                                    if _jp_1k_pv:
                                        _son_bans_pv = sorted(set(_son_bans_pv) | set(_jp_1k_pv))
                        if _son_bans_pv:
                            if _sonota_split:
                                _pv_ban_map["その他の優秀台+1,000枚以上.jpg"] = _son_bans_pv
                                # ②(③): ①の台番を diff_raw で 2000+ (/3000+) に絞る
                                _pv_s3_df2  = _prev_result.get("df")
                                _pv_s3_dr2  = _prev_result.get("diff_raw")
                                if _pv_s3_df2 is not None and _pv_s3_dr2 is not None:
                                    for _thr_bm, _fn_bm in _sonota_extra_thrs:
                                        _s3_k_pv = [
                                            _b for _b in _son_bans_pv
                                            if not (_pv_s3_df2[_pv_s3_df2["台番"] == _b]).empty
                                            and int(_pv_s3_dr2.loc[_pv_s3_df2[_pv_s3_df2["台番"] == _b].index[0]]) >= _thr_bm
                                        ]
                                        if _s3_k_pv:
                                            _pv_ban_map[_fn_bm] = _s3_k_pv
                            else:
                                _pv_ban_map["その他の優秀台ピックアップ.jpg"] = _son_bans_pv
                        for _nami_pv in _prev_result.get("nami_list", []):
                            _nt_pv = _nami_pv.get("title", "")
                            if _nt_pv and _nami_pv.get("bans"):
                                _pv_ban_map[f"{_make_safe_fn(_nt_pv)}.jpg"] = [int(_b3) for _b3 in _nami_pv["bans"]]
                        # ファイル名 → 表示タイトル（秋葉原専用・_prev_resultから構築）
                        _pv_title_map: dict[str, str] = {}
                        for _zd_t in _prev_result.get("zen_dai_list", []):
                            _pv_title_map[f"{_make_safe_fn(_zd_t['name'])}.jpg"] = _zd_t['name']
                        for _hr_t in _prev_result.get("high_ratio_list", []):
                            if _hr_t.get("has_image", False):
                                _pv_title_map[f"{_make_safe_fn(_hr_t['name'])}_高配分.jpg"] = _hr_t['name'] + "（優秀台）"
                        _pv_title_map["ジャグラーシリーズ優秀台.jpg"] = "ジャグラーシリーズ優秀台"
                        if _sonota_split:
                            _pv_title_map["その他の優秀台+1,000枚以上.jpg"] = "その他の優秀台ピックアップ"
                            for _, _fn_tt in _sonota_extra_thrs:
                                _pv_title_map[_fn_tt] = "その他の優秀台ピックアップ"
                        else:
                            _pv_title_map["その他の優秀台ピックアップ.jpg"] = "その他の優秀台ピックアップ"
                        for _nami_t in _prev_result.get("nami_list", []):
                            _nt_t = _nami_t.get("title", "")
                            if _nt_t:
                                _pv_title_map[f"{_make_safe_fn(_nt_t)}.jpg"] = _nt_t
                        # 並び画像（UI直接入力から生成）の台番・タイトルを追加
                        # _narabi_ban_map は ③並び画像セクションで構築。NameError は narabi_ok=False 時
                        try:
                            _pv_ban_map.update(_narabi_ban_map)
                            for _fn_nb in _narabi_ban_map:
                                _pv_title_map.setdefault(_fn_nb, os.path.splitext(_fn_nb)[0])
                        except NameError:
                            pass
                        # バラエティ画像の台番・タイトルを追加（_variety_ban_map はプレビュー生成時に構築）
                        try:
                            _pv_ban_map.update(_variety_ban_map)
                            for _fn_var in _variety_ban_map:
                                _pv_title_map.setdefault(_fn_var, os.path.splitext(_fn_var)[0])
                        except NameError:
                            pass
                        if kojin_enabled and _pv_df is not None:
                            for _km_pv in kojin_zentai_machines:
                                _km_pv = _km_pv.strip()
                                if not _km_pv:
                                    continue
                                _kgr_pv = _pv_df[_pv_df["機種名"] == _km_pv]
                                if not _kgr_pv.empty:
                                    _pv_ban_map[f"{_make_safe_fn(_km_pv)}.jpg"] = [int(b) for b in _kgr_pv["台番"].tolist()]
                        if kojin_enabled and _pv_df is not None and _pv_diff is not None:
                            for _km_pv in kojin_yushu_machines:
                                _km_pv = _km_pv.strip()
                                if not _km_pv:
                                    continue
                                _kgr_all_pv = _pv_df[_pv_df["機種名"] == _km_pv]
                                if _kgr_all_pv.empty:
                                    continue
                                _kdr_all_pv = _pv_diff.loc[_kgr_all_pv.index]
                                _kgrp_p_pv = _kojin_yushu_filter(_km_pv, _kgr_all_pv, _kdr_all_pv, get_store_config(store), force_1k=(with_slump and store == "秋葉原")).reset_index(drop=True)
                                if not _kgrp_p_pv.empty:
                                    _pv_ban_map[f"{_make_safe_fn(_km_pv)}（優秀台）.jpg"] = [int(b) for b in _kgrp_p_pv["台番"].tolist()]
                        # 末尾画像の台番を _pv_ban_map に追加（モードと末尾入力から算出）
                        if st.session_state.get("suebangai_enabled", False) and _pv_df is not None:
                            _sue_tails_bm = [t for i in range(1, 4)
                                             if (t := st.session_state.get(f"suebangai_tail_input_{i}", "").strip())]
                            _sue_mode_bm  = st.session_state.get("suebangai_mode", "全台")
                            _sue_circle_bm = {"0":"⓪","1":"①","2":"②","3":"③","4":"④",
                                              "5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
                            for _tail_bm in _sue_tails_bm:
                                try:
                                    if _tail_bm == "ゾロ目":
                                        _filt_bm = _pv_df[_pv_df["台番"].apply(
                                            lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                                        )]
                                        _base_bm = "末尾ゾロ目の台"
                                    elif _tail_bm.isdigit() and len(_tail_bm) in (1, 2):
                                        _filt_bm = _pv_df[_pv_df["台番"].astype(str).str[-len(_tail_bm):] == _tail_bm]
                                        _base_bm = f"末尾{_sue_circle_bm.get(_tail_bm, _tail_bm)}番台"
                                    else:
                                        continue
                                    if _filt_bm.empty:
                                        continue
                                    _is_plus_bm  = _sue_mode_bm in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）", "プラス台")
                                    _is_yushu_bm = _sue_mode_bm in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                                    _is_1k_bm    = _sue_mode_bm == "+1,000枚以上の優秀台"
                                    if _is_plus_bm:
                                        _filt_bm = _filt_bm[_filt_bm["差枚"] > 0]
                                        _title_bm = f"{_base_bm}のプラス台"
                                    elif _is_1k_bm:
                                        _filt_bm = _filt_bm[_filt_bm["差枚"] >= 1000]
                                        _title_bm = f"{_base_bm}の優秀台"
                                    elif _is_yushu_bm:
                                        _sg_bm = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _filt_bm.columns), None)
                                        _smask_bm = (_filt_bm["差枚"] >= 1000) | ((_filt_bm[_sg_bm] >= 1800) & (_filt_bm["差枚"] > 0)) if _sg_bm else (_filt_bm["差枚"] >= 1000)
                                        _filt_bm = _filt_bm[_smask_bm]
                                        _title_bm = f"{_base_bm}の優秀台"
                                    else:
                                        _title_bm = _base_bm
                                    if not _filt_bm.empty:
                                        _fn_bm = f"{_make_safe_fn(_title_bm)}.jpg"
                                        _pv_ban_map[_fn_bm] = [int(b) for b in _filt_bm["台番"].dropna()]
                                        _pv_title_map[_fn_bm] = _title_bm
                                except Exception:
                                    pass
                        # オススメ機種の台番を ban_map に追加
                        _pv_ban_map.update(_rec_ban_map)
                        _ig_api_key_pv = _get_pision_api_key()
                        if _ig_api_key_pv and _pv_ban_map:
                            _ig_date_pv = st.session_state.get(f"_inagawa_date_{store}", "")
                            if not _ig_date_pv:
                                _ig_rd_pv = _prev_result.get("date")
                                _ig_dt_key_pv = st.session_state.get(f"auto_tb_date_{store}")
                                _ig_date_pv = (
                                    _ig_rd_pv.strftime("%Y-%m-%d") if hasattr(_ig_rd_pv, "strftime") else str(_ig_rd_pv)
                                ) if _ig_rd_pv is not None else (
                                    _ig_dt_key_pv.strftime("%Y-%m-%d") if hasattr(_ig_dt_key_pv, "strftime") else str(_ig_dt_key_pv or "")
                                )
                            try:
                                # 速報モードで取得済みの items（points 込み）を優先使用
                                _rt_cached_pv = st.session_state.get(f"_auto_tb_rt_items_{store}")
                                _rt_cached_date_pv = st.session_state.get(f"_auto_tb_rt_items_date_{store}", "")
                                if _rt_cached_pv and _rt_cached_date_pv == _ig_date_pv:
                                    _ig_pision_items_pv = _rt_cached_pv  # _slump_apply_names は取得時に適用済み
                                else:
                                    _ig_halls_pv = fetch_pision_halls(_ig_api_key_pv)
                                    _ig_hall_id_pv = None
                                    for _igh_pv in _ig_halls_pv:
                                        _ighn_pv = _igh_pv.get("name") or _igh_pv.get("displayName") or ""
                                        if store in _ighn_pv and "エスパス" in _ighn_pv:
                                            _ig_hall_id_pv = str(_igh_pv.get("id") or _igh_pv.get("hallId") or "")
                                            break
                                    _ig_pision_items_pv = fetch_pision_results(_ig_api_key_pv, _ig_hall_id_pv, _ig_date_pv) if _ig_hall_id_pv else None
                                    if _ig_pision_items_pv:
                                        _slump_apply_names(_ig_pision_items_pv)
                                if _ig_pision_items_pv:
                                    _ig_by_uid_pv = {str(_it.get("unitId", "")): _it for _it in _ig_pision_items_pv}
                                    st.session_state[f"_slump_by_uid_{store}"] = _ig_by_uid_pv
                                    _ig_tmpl_pv = find_slump_template()
                                    _ig_bbb_pv  = _find_slump_bg()
                                    _ig_ban2mac_pv: dict[str, str] = {}
                                    if _pv_df is not None:
                                        for _, _igr_pv in _pv_df.iterrows():
                                            _bs0_pv = str(_igr_pv.get("台番", "")).split(".")[0]
                                            if _bs0_pv.lstrip("-").isdigit():
                                                _ig_ban2mac_pv[_bs0_pv] = str(_igr_pv.get("機種名", ""))
                                    _merged_pv: list[tuple[str, "Image.Image"]] = []
                                    for (_fn_pv, _img_pv) in _prev_img_list:
                                        _bare_pv = re.sub(r"^\d{2}_", "", _fn_pv)
                                        _bans_pv = _pv_ban_map.get(_bare_pv, [])
                                        if not _bans_pv or _ig_tmpl_pv is None:
                                            if store != "秋葉原":
                                                _merged_pv.append((_fn_pv, _img_pv))
                                            continue
                                        _g_imgs_pv: list["Image.Image"] = []
                                        _show_mn_pv = (_bare_pv in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg",
                                                                       "その他の優秀台+1,000枚以上.jpg", "その他の優秀台+2,000枚以上.jpg", "その他の優秀台+3,000枚以上.jpg")
                                                       or _bare_pv.startswith("末尾") or _bare_pv.startswith("バラエティ"))
                                        _is_zentai_pv = (not _bare_pv.endswith("_高配分.jpg") and
                                                         _bare_pv not in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg",
                                                                          "その他の優秀台+1,000枚以上.jpg", "その他の優秀台+2,000枚以上.jpg", "その他の優秀台+3,000枚以上.jpg"))
                                        _ban2diff_pv: dict[str, int] = {}
                                        if _pv_df is not None and _pv_diff is not None:
                                            for _idx_p, _row_p in _pv_df.iterrows():
                                                _bp = str(_row_p.get("台番", "")).split(".")[0]
                                                if _bp.lstrip("-").isdigit():
                                                    try:
                                                        _ban2diff_pv[_bp] = int(_pv_diff.loc[_idx_p])
                                                    except Exception:
                                                        pass
                                        for _b_pv in _bans_pv:
                                            _it_pv = _ig_by_uid_pv.get(str(_b_pv))
                                            if _it_pv is None or not _it_pv.get("points"):
                                                continue
                                            _dn_pv = (_it_pv.get("_convertedName")
                                                      or _it_pv.get("displayName")
                                                      or _ig_ban2mac_pv.get(str(_b_pv), str(_b_pv)))
                                            _sd_pv = not (_is_zentai_pv and _ban2diff_pv.get(str(_b_pv), 0) < 0)
                                            try:
                                                _g_imgs_pv.append(draw_slump_graph(
                                                    _ig_tmpl_pv, str(_b_pv), _dn_pv,
                                                    _it_pv["points"], diff=_it_pv.get("diff"),
                                                    machine_name=_dn_pv if _show_mn_pv else None,
                                                    show_diff=_sd_pv,
                                                ))
                                            except Exception:
                                                pass
                                        if store == "秋葉原":
                                            _pv_title = _pv_title_map.get(_bare_pv, os.path.splitext(_bare_pv)[0])
                                            _pv_slump = _build_slump_title_img(_pv_title, _g_imgs_pv, _ig_bbb_pv)
                                            if _pv_slump is not None:
                                                _merged_pv.append((_fn_pv, _pv_slump))
                                        else:
                                            _merged_pv.append((_fn_pv, _attach_slump_to_table(_img_pv, _g_imgs_pv, _ig_bbb_pv)))
                                        if len(_g_imgs_pv) >= 16 and store != "秋葉原":
                                            try:
                                                _side_fn_pv = os.path.splitext(_fn_pv)[0] + "_side.jpg"
                                                _merged_pv.append((_side_fn_pv, _attach_slump_to_table_side(_img_pv, _g_imgs_pv, _ig_bbb_pv)))
                                            except Exception:
                                                pass
                                    _prev_img_list = _merged_pv
                            except Exception:
                                pass  # pision取得失敗時は表のみ画像のままプレビュー表示

                    st.session_state[_aprev_key]    = _prev_img_list
                    st.session_state[_aprev_df_key] = _prev_result.get("df")
                    st.session_state[_aprev_di_key] = _prev_result.get("diff_raw")
                    st.session_state[_aprev_ex_key] = _prev_result.get("sonota_excellent_list", [])
                    st.session_state[_aprev_hr_key] = {
                        f"{_make_safe_fn(item['name'])}_高配分.jpg": item["name"]
                        for item in _prev_result.get("high_ratio_list", [])
                    }
                    st.session_state[_aprev_zen_key] = {
                        f"{_make_safe_fn(item['name'])}.jpg": item["name"]
                        for item in _prev_result.get("zen_dai_list", [])
                    }
                    st.session_state[_aprev_jug_ex_key]   = _prev_result.get("jug_excellent_list", [])
                    st.session_state[_aprev_jug_pool_key] = _prev_result.get("jug_pool_df")
                    st.session_state[_aprev_jug_ov_key]   = _prev_result.get("jug_overflow_df")
                    st.session_state[_aprev_hr_img_key]   = {
                        item["name"]
                        for item in _prev_result.get("high_ratio_list", [])
                        if item.get("has_image", True)
                    }
                    st.session_state[_aprev_narabi_key]   = _narabi_ban_map if narabi_ok and narabi_ranges else {}
                st.rerun()
            if _manual_prev_btn:
                _save_auto_inputs(store)
                with st.spinner("記入部分のみプレビュー生成中…"):
                    try:
                        uploaded.seek(0)
                        _raw_m = _read_uploaded_df(uploaded)
                        _df_m, _ = normalize_df(_raw_m)
                        _df_m = apply_name_conversion(_df_m)
                        if "差枚" in _df_m.columns:
                            _df_m["差枚"] = _df_m["差枚"].apply(_pipeline_calc_d)
                        _diff_m = _df_m["差枚"].copy()
                        _manual_imgs: list[tuple[str, "Image.Image"]] = []

                        # ② 個別画像 - 全台
                        if kojin_enabled:
                            for _km in kojin_zentai_machines:
                                _km = _km.strip()
                                if not _km:
                                    continue
                                _mg = _df_m[_df_m["機種名"] == _km].copy().reset_index(drop=True)
                                if _mg.empty:
                                    continue
                                _md = _diff_m.loc[_df_m[_df_m["機種名"] == _km].index].reset_index(drop=True)
                                _manual_imgs.append((f"{_make_safe_fn(_km)}.jpg", _build_machine_img(_mg, _km, _stat_from_diff(_md))))

                            # ② 個別画像 - 優秀台
                            _m_cfg = get_store_config(store)
                            for _km in kojin_yushu_machines:
                                _km = _km.strip()
                                if not _km:
                                    continue
                                _mga = _df_m[_df_m["機種名"] == _km]
                                if _mga.empty:
                                    continue
                                _mda = _diff_m.loc[_mga.index]
                                _mgp = _kojin_yushu_filter(_km, _mga, _mda, _m_cfg).reset_index(drop=True)
                                if _mgp.empty:
                                    continue
                                _mtit = f"{_km}（優秀台）"
                                _manual_imgs.append((f"{_make_safe_fn(_mtit)}.jpg", _build_machine_img(_mgp, _mtit, None)))

                            # ② その他の優秀台ピックアップ（溝の口新館専用）
                            if sonota_extra_text.strip():
                                _se_bans_m = set(expand_machine_numbers(sonota_extra_text))
                                if _se_bans_m:
                                    _se_df_m = _df_m[_df_m["台番"].apply(lambda b: int(b) in _se_bans_m)].copy().reset_index(drop=True)
                                    if not _se_df_m.empty:
                                        _se_tit_m = sonota_extra_title.strip() or "その他の優秀台ピックアップ"
                                        _manual_imgs.append((f"{_make_safe_fn(_se_tit_m)}.jpg", _build_machine_img(_se_df_m, _se_tit_m, None)))

                        # ③ 並び画像
                        if narabi_ok and narabi_ranges:
                            _ban_map_m = {int(row["台番"]): i for i, row in _df_m.iterrows()}
                            # 重複タイトルを事前検出（フルプレビューと同じ方式）
                            _nb_infos_m = []
                            for _n_bans in narabi_ranges:
                                _n_idxs = [_ban_map_m[b] for b in _n_bans if b in _ban_map_m]
                                if not _n_idxs:
                                    continue
                                _ngrp = _df_m.loc[_n_idxs].copy().reset_index(drop=True)
                                _nms  = list(dict.fromkeys(str(m) for m in _ngrp["機種名"]))
                                _nn   = len(_ngrp)
                                if len(_nms) == 1: _ntit = f"{_nms[0]}({_nn}台並び)"
                                elif len(_nms) == 2: _ntit = f"{_nms[0]}+{_nms[1]}({_nn}台並び)"
                                else: _ntit = f"{_nms[0]}～{_nms[-1]}({_nn}台並び)"
                                _nb_infos_m.append((_ngrp, _ntit, list(_n_bans)))
                            from collections import Counter as _CtrM
                            _dup_tits_m = {t for t, c in _CtrM(i[1] for i in _nb_infos_m).items() if c > 1}
                            for _ngrp, _ntit, _n_blist_m in _nb_infos_m:
                                _nds  = _ngrp["差枚"]
                                _nstat = {"total_diff": int(_nds.sum()), "avg_diff": int(round(_nds.mean())), "win_count": int((_nds > 0).sum()), "total_count": len(_ngrp)}
                                if _ntit in _dup_tits_m:
                                    _nb_s = int(_ngrp.iloc[0]["台番"]); _nb_e = int(_ngrp.iloc[-1]["台番"])
                                    _file_tit_m = f"{_ntit}（{_nb_s}～{_nb_e}）"
                                else:
                                    _file_tit_m = _ntit
                                _manual_imgs.append((f"{_make_safe_fn(_file_tit_m)}.jpg", _build_machine_img(_ngrp, _ntit, _nstat)))

                        # ④ 末尾画像
                        if st.session_state.get("suebangai_enabled", False):
                            _m_sue_tails = [t for _i in range(1, 4) if (t := st.session_state.get(f"suebangai_tail_input_{_i}", "").strip())]
                            _m_sue_mode = st.session_state.get("suebangai_mode", "全台")
                            for _item in _gen_sue_imgs_on_fly(_m_sue_tails, _m_sue_mode, is_juggler=False):
                                _manual_imgs.append(_item)
                        if st.session_state.get("jug_sue_enabled", False):
                            _m_jug_tails = [t for _i in range(1, 4) if (t := st.session_state.get(f"jug_sue_tail_input_{_i}", "").strip())]
                            _m_jug_mode = st.session_state.get("jug_sue_mode", "全台")
                            for _item in _gen_sue_imgs_on_fly(_m_jug_tails, _m_jug_mode, is_juggler=True):
                                _manual_imgs.append(_item)

                        # ⑤ オススメ機種ピックアップ
                        if recommended_blocks:
                            _ms_cfg = get_store_config(store)
                            _mj_cfg = {
                                "series":     _ms_cfg["juggler_series"],
                                "jobs_map":   {j[0]: j[1] for j in _ms_cfg["juggler_jobs"]},
                                "g_min":      _ms_cfg["juggler_g_min"],
                                "diff_bonus": _ms_cfg["diff_bonus"],
                            }
                            _ms_sfx = {1: "プラス台", 1000: "1000枚以上", 2000: "2000枚以上"}
                            for _blk in recommended_blocks:
                                _mbt = _blk["title"].strip() or "オススメ機種"
                                _mbm = [m.strip() for m in _blk["machines"] if m.strip()]
                                if not _mbm:
                                    continue
                                for _mthr in _blk.get("thresholds", [1]):
                                    _mrimg = generate_recommended_block_image(
                                        _mbt, _mbm, _df_m, _diff_m, set(),
                                        min_diff=_mthr, juggler_cfg=_mj_cfg,
                                    )
                                    if _mrimg is None:
                                        continue
                                    _ms_sfxv = _ms_sfx.get(_mthr, str(_mthr))
                                    _manual_imgs.append((f"オススメ_{_make_safe_fn(_mbt)}_{_ms_sfxv}.jpg", _mrimg))

                        if _manual_imgs:
                            st.session_state[_aprev_key] = _manual_imgs
                            st.session_state[f"_manual_preview_mode_{store}"] = True
                        else:
                            st.warning("⚠️ 記入された項目がないか、該当台が見つかりませんでした。")
                    except Exception as _me:
                        st.error(f"❌ エラー: {_me}")
                        with st.expander("詳細"):
                            st.code(traceback.format_exc())
                st.rerun()
        else:
            st.caption(f"📋 {len(_auto_previews)}枚の画像プレビュー　チェックした画像のみ生成されます")
            for _row_start in range(0, len(_auto_previews), 3):
                _grid_cols = st.columns(3)
                for _col_idx, _ci in enumerate(range(_row_start, min(_row_start + 3, len(_auto_previews)))):
                    _ptitle, _pimg = _auto_previews[_ci]
                    _ck_key = f"auto_prev_ck_{store}_{_ci}"
                    if _ck_key not in st.session_state:
                        st.session_state[_ck_key] = True
                    with _grid_cols[_col_idx]:
                        _sub_ck, _sub_img = st.columns([1, 10])
                        with _sub_ck:
                            st.checkbox("", key=_ck_key, label_visibility="collapsed")
                        with _sub_img:
                            st.image(_pimg, caption=_ptitle, use_container_width=True)
            _btn_upd, _btn_clr = st.columns(2)
            with _btn_upd:
                if st.button("🔄 その他を更新", key="auto_preview_update_btn", use_container_width=True):
                    _pv_df     = st.session_state.get(_aprev_df_key)
                    _pv_diff   = st.session_state.get(_aprev_di_key)
                    _pv_ex     = st.session_state.get(_aprev_ex_key, [])
                    _pv_jug_ex   = st.session_state.get(_aprev_jug_ex_key, [])
                    _pv_jug_pool = st.session_state.get(_aprev_jug_pool_key)
                    _pv_jug_ov   = st.session_state.get(_aprev_jug_ov_key)
                    _pv_hr       = st.session_state.get(_aprev_hr_key, {})
                    _pv_zen    = st.session_state.get(_aprev_zen_key, {})
                    _pv_narabi = st.session_state.get(_aprev_narabi_key, {})
                    if _pv_df is not None and _pv_diff is not None:
                        _jug_series_set = set(get_store_config(store)["juggler_series"])
                        # kojin個別画像がある機種はチェック外し時に統合画像へ追加しない
                        _kojin_machines_set = {m.strip() for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip()}
                        # オススメ機種セットとチェック外し追跡
                        _rec_machines_set: set[str] = {m.strip() for block in recommended_blocks for m in block["machines"] if m.strip()}
                        _rec_unchecked_machines: set[str] = set()
                        # 末尾台番セット（その他の優秀台への重複追加を防止）
                        _sue_bans_upd: set[int] = set()
                        _stails_upd = [t for i in range(1, 4)
                                       if (t := st.session_state.get(f"suebangai_tail_input_{i}", "").strip())]
                        for _st_upd in _stails_upd:
                            if _st_upd == "ゾロ目":
                                _sue_bans_upd |= {
                                    int(b) for b in _pv_df["台番"]
                                    if (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                                }
                            elif _st_upd.isdigit() and len(_st_upd) in (1, 2):
                                _sue_bans_upd |= {
                                    int(b) for b in _pv_df["台番"]
                                    if str(int(b))[-len(_st_upd):] == _st_upd
                                }
                        _upd_extra_dfs:   list[pd.DataFrame] = []
                        _upd_extra_diffs: list[pd.Series]    = []
                        _jug_extra_dfs:   list[pd.DataFrame] = []
                        # チェック済みの高配分・全台系画像がある機種セット（並び外しでその他に重複追加しないため）
                        _checked_img_machines: set[str] = set()
                        for _si, (_spn, _) in enumerate(_auto_previews):
                            if st.session_state.get(f"auto_prev_ck_{store}_{_si}", True):
                                _hm = _pv_hr.get(_spn)
                                if _hm:
                                    _checked_img_machines.add(_hm)
                                _zm = _pv_zen.get(_spn)
                                if _zm:
                                    _checked_img_machines.add(_zm)
                        for _ci, (_pname, _) in enumerate(_auto_previews):
                            if not st.session_state.get(f"auto_prev_ck_{store}_{_ci}", True):
                                # 高配分：kojin済みはスキップ・ジャグラーならジャグラー優秀台へ、それ以外はその他へ
                                _m = _pv_hr.get(_pname)
                                if _m and _m not in _kojin_machines_set:
                                    _mrows = _pv_df[_pv_df["機種名"] == _m]
                                    if not _mrows.empty:
                                        _mdiff = _pv_diff.loc[_mrows.index]
                                        if _m in _jug_series_set:
                                            # ジャグラー：juggler_jobsの合算確率条件で判定
                                            _jcfg_chk = get_store_config(store)
                                            _jg_min_chk = _jcfg_chk["juggler_g_min"]
                                            _jthresh_chk = {_jm: (_jp, _jd) for _jm, _jp, _jd in _jcfg_chk["juggler_jobs"]}
                                            _mrows_j = _mrows.copy().reset_index(drop=True)
                                            _mdiff_j = _mdiff.reset_index(drop=True)
                                            if "ゲーム数_rounded" not in _mrows_j.columns:
                                                _mrows_j["ゲーム数_rounded"] = _mrows_j["ゲーム数"].apply(round_games)
                                            if "合算確率_num" not in _mrows_j.columns:
                                                _mrows_j["合算確率_num"] = _mrows_j.apply(
                                                    lambda r: r["ゲーム数_rounded"] / (r["BB"] + r["RB"])
                                                              if (r["BB"] + r["RB"]) > 0 else float("inf"), axis=1)
                                            _g_ok_chk = _mrows_j["ゲーム数_rounded"] >= _jg_min_chk
                                            _mrows_j = _mrows_j[_g_ok_chk.values].reset_index(drop=True)
                                            _mdiff_j = _mdiff_j[_g_ok_chk.values].reset_index(drop=True)
                                            if not _mrows_j.empty:
                                                _ps_chk = _mrows_j["機種名"].apply(lambda _mm: _jthresh_chk.get(_mm, (float("inf"), float("inf")))[0])
                                                _ds_chk = _mrows_j["機種名"].apply(lambda _mm: _jthresh_chk.get(_mm, (float("inf"), float("inf")))[1])
                                                _jcond_chk = ((_mrows_j["合算確率_num"] <= _ps_chk.values) & (_mdiff_j.values >= 0)) | (_mdiff_j.values >= _ds_chk.values)
                                                _mgood = _mrows_j[_jcond_chk].copy().reset_index(drop=True)
                                                _mgood_diff = _mdiff_j[_jcond_chk].reset_index(drop=True)
                                            else:
                                                _mgood = pd.DataFrame()
                                                _mgood_diff = pd.Series(dtype=float)
                                        else:
                                            _mmask = _mdiff >= 1000
                                            _mgood = _mrows[_mmask.values].copy().reset_index(drop=True)
                                            _mgood_diff = _mdiff[_mmask].reset_index(drop=True)
                                        # チェック済み並び画像に含まれる台番は除外（並びと優秀台の重複防止）
                                        _narabi_excl_m: set[int] = set()
                                        for _nci2, (_npname2, _) in enumerate(_auto_previews):
                                            if st.session_state.get(f"auto_prev_ck_{store}_{_nci2}", True):
                                                _nb2 = _pv_narabi.get(_npname2)
                                                if _nb2:
                                                    _nb2_set = set(_nb2)
                                                    _nb2_m = _pv_df[
                                                        (_pv_df["台番"].apply(lambda b: int(b) in _nb2_set)) &
                                                        (_pv_df["機種名"] == _m)
                                                    ]
                                                    _narabi_excl_m.update(_nb2_m["台番"].apply(int).tolist())
                                        if _narabi_excl_m:
                                            _keep_m = ~_mgood["台番"].apply(int).isin(_narabi_excl_m)
                                            _mgood = _mgood[_keep_m.values].reset_index(drop=True)
                                            _mgood_diff = _mgood_diff[_keep_m.values].reset_index(drop=True)
                                        if not _mgood.empty:
                                            if _sue_bans_upd:
                                                _mg_keep = ~_mgood["台番"].apply(int).isin(_sue_bans_upd)
                                                _mgood      = _mgood[_mg_keep.values].copy().reset_index(drop=True)
                                                _mgood_diff = _mgood_diff[_mg_keep.values].reset_index(drop=True)
                                        if not _mgood.empty:
                                            if _m in _jug_series_set and not (with_slump and store == "秋葉原"):
                                                _jug_extra_dfs.append(_mgood)
                                            elif _m in _jug_series_set:
                                                # 秋葉原: diff>=1000でその他へ
                                                _jm_1k = _mgood_diff >= 1000
                                                if _jm_1k.any():
                                                    _upd_extra_dfs.append(_mgood[_jm_1k.values].copy().reset_index(drop=True))
                                                    _upd_extra_diffs.append(_mgood_diff[_jm_1k].reset_index(drop=True))
                                            elif _m in _rec_machines_set:
                                                _rec_unchecked_machines.add(_m)
                                            else:
                                                _upd_extra_dfs.append(_mgood)
                                                _upd_extra_diffs.append(_mgood_diff)
                                # 全台系画像：kojin済みはスキップ・同様にジャグラー判定
                                _mz = _pv_zen.get(_pname)
                                if _mz and not _m and _mz not in _kojin_machines_set:
                                    _mzrows = _pv_df[_pv_df["機種名"] == _mz]
                                    if not _mzrows.empty:
                                        _mzdiff = _pv_diff.loc[_mzrows.index]
                                        _mzmask = _mzdiff >= 1000
                                        _mzgood = _mzrows[_mzmask.values].copy().reset_index(drop=True)
                                        _mzgood_diff = _mzdiff[_mzmask].reset_index(drop=True)
                                        # チェック済み並び画像に含まれる台番は除外（並びと優秀台の重複防止）
                                        _narabi_excl_z: set[int] = set()
                                        for _nci3, (_npname3, _) in enumerate(_auto_previews):
                                            if st.session_state.get(f"auto_prev_ck_{store}_{_nci3}", True):
                                                _nb3 = _pv_narabi.get(_npname3)
                                                if _nb3:
                                                    _nb3_set = set(_nb3)
                                                    _nb3_m = _pv_df[
                                                        (_pv_df["台番"].apply(lambda b: int(b) in _nb3_set)) &
                                                        (_pv_df["機種名"] == _mz)
                                                    ]
                                                    _narabi_excl_z.update(_nb3_m["台番"].apply(int).tolist())
                                        if _narabi_excl_z:
                                            _keep_mz = ~_mzgood["台番"].apply(int).isin(_narabi_excl_z)
                                            _mzgood = _mzgood[_keep_mz.values].reset_index(drop=True)
                                            _mzgood_diff = _mzgood_diff[_keep_mz.values].reset_index(drop=True)
                                        if not _mzgood.empty:
                                            if _sue_bans_upd:
                                                _mzg_keep = ~_mzgood["台番"].apply(int).isin(_sue_bans_upd)
                                                _mzgood      = _mzgood[_mzg_keep.values].copy().reset_index(drop=True)
                                                _mzgood_diff = _mzgood_diff[_mzg_keep.values].reset_index(drop=True)
                                        if not _mzgood.empty:
                                            if _mz in _jug_series_set and not (with_slump and store == "秋葉原"):
                                                _jug_extra_dfs.append(_mzgood)
                                            elif _mz in _jug_series_set:
                                                _mzm_1k = _mzgood_diff >= 1000
                                                if _mzm_1k.any():
                                                    _upd_extra_dfs.append(_mzgood[_mzm_1k.values].copy().reset_index(drop=True))
                                                    _upd_extra_diffs.append(_mzgood_diff[_mzm_1k].reset_index(drop=True))
                                            else:
                                                _upd_extra_dfs.append(_mzgood)
                                                _upd_extra_diffs.append(_mzgood_diff)
                                # 並び画像：機種ごとにジャグラー/非ジャグラーへ振り分け
                                _nb_bans = _pv_narabi.get(_pname)
                                if _nb_bans:
                                    _nb_ban_set = set(_nb_bans)
                                    _nbrows = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _nb_ban_set)].copy()
                                    if not _nbrows.empty:
                                        _nbdiff = _pv_diff.loc[_nbrows.index]
                                        # 非ジャグラー: diff >= 1000 のみその他へ
                                        _nb_non_jug_m = ~_nbrows["機種名"].isin(_jug_series_set)
                                        _nb_oth_part = _nbrows[(_nb_non_jug_m) & (_nbdiff >= 1000).values].copy().reset_index(drop=True)
                                        _nb_oth_diff = _nbdiff[(_nb_non_jug_m) & (_nbdiff >= 1000).values].reset_index(drop=True)
                                        if not _nb_oth_part.empty:
                                            if _sue_bans_upd:
                                                _nb_oth_keep = ~_nb_oth_part["台番"].apply(int).isin(_sue_bans_upd)
                                                _nb_oth_part = _nb_oth_part[_nb_oth_keep.values].copy().reset_index(drop=True)
                                                _nb_oth_diff = _nb_oth_diff[_nb_oth_keep.values].reset_index(drop=True)
                                        # 高配分・全台系画像がチェック済みの機種はその他へ追加しない
                                        if not _nb_oth_part.empty and _checked_img_machines:
                                            _nb_no_img = ~_nb_oth_part["機種名"].isin(_checked_img_machines)
                                            _nb_oth_part = _nb_oth_part[_nb_no_img.values].copy().reset_index(drop=True)
                                            _nb_oth_diff = _nb_oth_diff[_nb_no_img.values].reset_index(drop=True)
                                        if not _nb_oth_part.empty:
                                            _upd_extra_dfs.append(_nb_oth_part)
                                            _upd_extra_diffs.append(_nb_oth_diff)
                                        # ジャグラー: 合算確率条件（G数>=g_min AND (確率<=閾値かつdiff>=0) OR diff>=_ds）
                                        _nb_jug_base = _nbrows[_nbrows["機種名"].isin(_jug_series_set)].copy()
                                        if not _nb_jug_base.empty:
                                            _nb_jug_diff_base = _nbdiff.loc[_nb_jug_base.index].reset_index(drop=True)
                                            _nb_jug_base = _nb_jug_base.reset_index(drop=True)
                                            _jcfg_nb = get_store_config(store)
                                            _jg_min_nb = _jcfg_nb["juggler_g_min"]
                                            _jthresh_nb = {m: (p, d) for m, p, d in _jcfg_nb["juggler_jobs"]}
                                            if "ゲーム数_rounded" not in _nb_jug_base.columns:
                                                _nb_jug_base["ゲーム数_rounded"] = _nb_jug_base["ゲーム数"].apply(round_games)
                                            if "合算確率_num" not in _nb_jug_base.columns:
                                                _nb_jug_base["合算確率_num"] = _nb_jug_base.apply(
                                                    lambda r: r["ゲーム数_rounded"] / (r["BB"] + r["RB"])
                                                              if (r["BB"] + r["RB"]) > 0 else float("inf"), axis=1)
                                            _g_ok = _nb_jug_base["ゲーム数_rounded"] >= _jg_min_nb
                                            _nb_jug_base = _nb_jug_base[_g_ok.values].reset_index(drop=True)
                                            _nb_jug_diff_base = _nb_jug_diff_base[_g_ok.values].reset_index(drop=True)
                                            if not _nb_jug_base.empty:
                                                _ps_nb = _nb_jug_base["機種名"].apply(
                                                    lambda m: _jthresh_nb.get(m, (float("inf"), float("inf")))[0])
                                                _ds_nb = _nb_jug_base["機種名"].apply(
                                                    lambda m: _jthresh_nb.get(m, (float("inf"), float("inf")))[1])
                                                _jug_cond_nb = ((_nb_jug_base["合算確率_num"] <= _ps_nb.values) &
                                                                (_nb_jug_diff_base.values >= 0)) | \
                                                               (_nb_jug_diff_base.values >= _ds_nb.values)
                                                _nb_jug_part = _nb_jug_base[_jug_cond_nb].copy().reset_index(drop=True)
                                                # 高配分・全台系画像がチェック済みの機種はジャグラー優秀台へ追加しない
                                                if not _nb_jug_part.empty and _checked_img_machines:
                                                    _nb_jug_no_img = ~_nb_jug_part["機種名"].isin(_checked_img_machines)
                                                    _nb_jug_part = _nb_jug_part[_nb_jug_no_img.values].copy().reset_index(drop=True)
                                                if not _nb_jug_part.empty:
                                                    if with_slump and store == "秋葉原":
                                                        _nb_jp_d = _nb_jug_diff_base[_jug_cond_nb].reset_index(drop=True)
                                                        _nb_1k = _nb_jp_d >= 1000
                                                        if _nb_1k.any():
                                                            _upd_extra_dfs.append(_nb_jug_part[_nb_1k.values].copy().reset_index(drop=True))
                                                            _upd_extra_diffs.append(_nb_jp_d[_nb_1k].reset_index(drop=True))
                                                    else:
                                                        _jug_extra_dfs.append(_nb_jug_part)
                                # kojin_yushu 画像（{機種名}（優秀台）.jpg）がチェック外された場合
                                _YUSHU_SFX = "（優秀台）.jpg"
                                if _pname.endswith(_YUSHU_SFX) and _m is None and not _nb_bans:
                                    _km_y = _pname[:-len(_YUSHU_SFX)]
                                    _kojin_y_set = {m.strip() for m in kojin_yushu_machines if m.strip()}
                                    if _km_y in _kojin_y_set:
                                        _myrows = _pv_df[_pv_df["機種名"] == _km_y]
                                        if not _myrows.empty:
                                            _mydiff = _pv_diff.loc[_myrows.index]
                                            _mymask = _mydiff >= 1000
                                            _mygood = _myrows[_mymask.values].copy().reset_index(drop=True)
                                            _mygood_diff = _mydiff[_mymask].reset_index(drop=True)
                                            if not _mygood.empty and _sue_bans_upd:
                                                _my_keep = ~_mygood["台番"].apply(int).isin(_sue_bans_upd)
                                                _mygood      = _mygood[_my_keep.values].copy().reset_index(drop=True)
                                                _mygood_diff = _mygood_diff[_my_keep.values].reset_index(drop=True)
                                            if not _mygood.empty:
                                                _has_jug_ex_img = any(_pn == "ジャグラーシリーズ優秀台.jpg" for _pn, _ in _auto_previews)
                                                if _km_y in _jug_series_set and _has_jug_ex_img and not (with_slump and store == "秋葉原"):
                                                    _jug_extra_dfs.append(_mygood)
                                                else:
                                                    _upd_extra_dfs.append(_mygood)
                                                    _upd_extra_diffs.append(_mygood_diff)
                        _new_prev = list(_auto_previews)
                        _updated  = False
                        # overflowデータ（ジャグラー）がジャグラーシリーズ優秀台に移る台番を事前計算
                        _jug_ov_used_bans: set[int] = set()
                        if _jug_extra_dfs:
                            _pool_empty = _pv_jug_pool is None or (_pv_jug_pool is not None and _pv_jug_pool.empty)
                            if _pool_empty and _pv_jug_ov is not None and not _pv_jug_ov.empty:
                                _jug_hr_pre = set(_pv_hr.values())
                                _ov_pre = _pv_jug_ov[~_pv_jug_ov["機種名"].isin(_jug_hr_pre)]
                                _jug_ov_used_bans = {int(b) for b in _ov_pre["台番"].dropna()}
                        # その他の優秀台ピックアップ更新（非ジャグラーチェック外し or overflowジャグラー台を除去する場合）
                        _ex_bans_for_son = {item["ban"] for item in _pv_ex} - _jug_ov_used_bans
                        # 秋葉原スランプ付き: jug_pool +1000枚台を基礎bansに追加
                        # （excellent_listはG数条件あり台のみのため、G数未達でもdiff>=1000の台が脱落しないよう補完）
                        if with_slump and store == "秋葉原" and _pv_df is not None:
                            _pv_jp_upd = st.session_state.get(_aprev_jug_pool_key)
                            if _pv_jp_upd is not None and not _pv_jp_upd.empty:
                                _jp_bns_upd = {int(str(b).split(".")[0]) for b in _pv_jp_upd["台番"].dropna()
                                               if str(b).split(".")[0].lstrip("-").isdigit()}
                                _jp_rows_upd = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _jp_bns_upd)]
                                if not _jp_rows_upd.empty:
                                    _jp_dr_upd = _pv_diff.loc[_jp_rows_upd.index]
                                    _jp_1k_upd = {int(b) for b in _jp_rows_upd[_jp_dr_upd.values >= 1000]["台番"].dropna()}
                                    _ex_bans_for_son |= _jp_1k_upd
                        # 秋葉原スランプ付き: ジャグラー機種は diff >= 1000 のみ含める
                        # _pv_ex の item["diff"] を直接参照（_pv_diff.loc 経由だとインデックスずれで誤判定するため）
                        if with_slump and store == "秋葉原":
                            _jug_low_bans = {item["ban"] for item in _pv_ex
                                             if item.get("diff", 0) < 1000 and item.get("name", "") in _jug_series_set}
                            if _jug_low_bans:
                                _ex_bans_for_son -= _jug_low_bans
                        if _upd_extra_dfs or _jug_ov_used_bans:
                            if _ex_bans_for_son:
                                _ex_rows = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _ex_bans_for_son)].copy().reset_index(drop=True)
                                _ex_diff = _pv_diff.loc[_pv_df[_pv_df["台番"].apply(lambda b: int(b) in _ex_bans_for_son)].index].reset_index(drop=True)
                                _all_dfs   = [_ex_rows] + _upd_extra_dfs
                                _all_diffs = [_ex_diff] + _upd_extra_diffs
                            else:
                                _all_dfs   = _upd_extra_dfs
                                _all_diffs = _upd_extra_diffs
                        else:
                            _all_dfs = []
                        if _all_dfs:
                            _son_comb  = pd.concat(_all_dfs,   ignore_index=True)
                            _son_comb  = _son_comb.drop_duplicates(subset=["台番"])
                            _son_order = _son_comb["台番"].argsort()
                            _son_comb  = _son_comb.iloc[_son_order].reset_index(drop=True)
                            _son_img   = _build_machine_img(_son_comb, "その他の優秀台ピックアップ", None)
                            # 秋葉原スランプ付きは①.jpgキーで管理
                            _son_pv_key = "その他の優秀台+1,000枚以上.jpg" if _sonota_split else "その他の優秀台ピックアップ.jpg"
                            for _ci, (_pname, _) in enumerate(_new_prev):
                                if _pname == _son_pv_key:
                                    _new_prev[_ci] = (_pname, _son_img)
                                    break
                            else:
                                _new_prev.append((_son_pv_key, _son_img))
                                st.session_state[f"auto_prev_ck_{store}_{len(_new_prev)-1}"] = True
                            # ②(③)も素の表を再生成して差し替える（後段のスランプ合成が
                            # 合成済み画像に二重合成して崩れるのを防ぐ）
                            if _sonota_split and _pv_diff is not None and _pv_df is not None:
                                _son_bans_bt = [int(str(b).split(".")[0]) for b in _son_comb["台番"].dropna()
                                                if str(b).split(".")[0].lstrip("-").isdigit()]
                                for _thr_bt, _fn_bt in _sonota_extra_thrs:
                                    _sk_bt = {b for b in _son_bans_bt
                                              if not _pv_df[_pv_df["台番"] == b].empty
                                              and int(_pv_diff.loc[_pv_df[_pv_df["台番"] == b].index[0]]) >= _thr_bt}
                                    if not _sk_bt:
                                        continue
                                    _sk_df_bt = _son_comb[_son_comb["台番"].apply(
                                        lambda b: int(str(b).split(".")[0]) in _sk_bt)].copy().reset_index(drop=True)
                                    _sk_img_bt = _build_machine_img(_sk_df_bt, "その他の優秀台ピックアップ", None)
                                    for _ci_bt, (_pn_bt, _) in enumerate(_new_prev):
                                        if _pn_bt == _fn_bt:
                                            _new_prev[_ci_bt] = (_pn_bt, _sk_img_bt)
                                            break
                            _updated = True
                        # ジャグラーシリーズ優秀台更新（秋葉原スランプ付きは生成しない）
                        if _jug_extra_dfs and not (with_slump and store == "秋葉原"):
                            # 優先順: jug_pool_df（通常生成時） → jug_overflow_df（overflow時） → jug_excellent_list(+1000台)フォールバック
                            if _pv_jug_pool is not None and not _pv_jug_pool.empty:
                                _jug_base = [_pv_jug_pool.copy()]
                            elif _pv_jug_ov is not None and not _pv_jug_ov.empty:
                                _jug_hr_names_pv = set(_pv_hr.values())
                                _jug_ov_filt = _pv_jug_ov[~_pv_jug_ov["機種名"].isin(_jug_hr_names_pv)].copy().reset_index(drop=True)
                                _jug_base = [_jug_ov_filt] if not _jug_ov_filt.empty else []
                            else:
                                _jug_hr_names_pv = set(_pv_hr.values())
                                _jug_ex_bans = {item["ban"] for item in _pv_jug_ex if item["name"] not in _jug_hr_names_pv}
                                _jug_base = [_pv_df[_pv_df["台番"].apply(lambda b: int(b) in _jug_ex_bans)].copy().reset_index(drop=True)] if _jug_ex_bans else []
                            _jug_all_dfs = _jug_base + _jug_extra_dfs
                            _jug_comb = pd.concat(_jug_all_dfs, ignore_index=True)
                            _jug_ord  = _jug_comb["台番"].argsort()
                            _jug_comb = _jug_comb.iloc[_jug_ord].reset_index(drop=True)
                            if len(_jug_comb) <= 5:
                                # 5台以下 → overflowと同じ扱い: その他の優秀台ピックアップへ
                                _ov_ex_bans = {item["ban"] for item in _pv_ex}
                                _ov_rows = _pv_df[_pv_df["台番"].apply(lambda b: int(b) in _ov_ex_bans)].copy().reset_index(drop=True) if _pv_df is not None else pd.DataFrame()
                                _ov_dfs = ([_ov_rows] if not _ov_rows.empty else []) + _jug_extra_dfs
                                _ov_son = pd.concat(_ov_dfs, ignore_index=True)
                                _ov_son = _ov_son.drop_duplicates(subset=["台番"])
                                _ov_son = _ov_son.iloc[_ov_son["台番"].argsort()].reset_index(drop=True)
                                _ov_img = _build_machine_img(_ov_son, "その他の優秀台ピックアップ", None)
                                for _ci2, (_pn2, _) in enumerate(_new_prev):
                                    if _pn2 == "その他の優秀台ピックアップ.jpg":
                                        _new_prev[_ci2] = (_pn2, _ov_img)
                                        break
                                else:
                                    _new_prev.append(("その他の優秀台ピックアップ.jpg", _ov_img))
                                    st.session_state[f"auto_prev_ck_{store}_{len(_new_prev)-1}"] = True
                                _updated = True
                            else:
                                _has_kojin_jug = any(m.strip() in _jug_series_set for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip())
                                _still_jug_other = _has_kojin_jug or any(
                                    st.session_state.get(f"auto_prev_ck_{store}_{_si}", True) and
                                    (_pv_hr.get(_sp) in _jug_series_set or _pv_zen.get(_sp) in _jug_series_set)
                                    for _si, (_sp, _) in enumerate(_auto_previews)
                                )
                                _jug_title = "その他のジャグラーシリーズの優秀台" if _still_jug_other else "ジャグラーシリーズの優秀台"
                                _jug_img = _build_machine_img(_jug_comb, _jug_title, None)
                                for _jpi, (_jpn, _) in enumerate(_new_prev):
                                    if _jpn == "ジャグラーシリーズ優秀台.jpg":
                                        _new_prev[_jpi] = (_jpn, _jug_img)
                                        break
                                else:
                                    _new_prev.append(("ジャグラーシリーズ優秀台.jpg", _jug_img))
                                    st.session_state[f"auto_prev_ck_{store}_{len(_new_prev)-1}"] = True
                                _updated = True
                        # オススメ機種ピックアップ再生成（チェック外し機種がオススメに含まれる場合）
                        if _rec_unchecked_machines and _pv_df is not None and _pv_diff is not None:
                            _pv_scfg_r = get_store_config(store)
                            _pv_jcfg_r = {
                                "series":     _pv_scfg_r["juggler_series"],
                                "jobs_map":   {j[0]: j[1] for j in _pv_scfg_r["juggler_jobs"]},
                                "g_min":      _pv_scfg_r["juggler_g_min"],
                                "diff_bonus": _pv_scfg_r["diff_bonus"],
                            }
                            _pv_zen_r  = set(st.session_state.get(_aprev_zen_key, {}).values())
                            _pv_hr_img = st.session_state.get(_aprev_hr_img_key, set())
                            _pv_high_r = (_pv_hr_img - _rec_unchecked_machines)
                            if kojin_enabled:
                                _pv_zen_r  |= {m.strip() for m in kojin_zentai_machines if m.strip()}
                                _pv_high_r |= {m.strip() for m in kojin_yushu_machines if m.strip()}
                            _upd_nb: set[int] = ranges_to_bans(narabi_ranges) if narabi_ok else set()
                            if kojin_enabled and kojin_narabi_ranges_text.strip():
                                try: _upd_nb |= ranges_to_bans(parse_ranges(kojin_narabi_ranges_text.strip()))
                                except Exception: pass
                            if kojin_enabled and kojin_narabi2_ranges_text.strip():
                                try: _upd_nb |= ranges_to_bans(parse_ranges(kojin_narabi2_ranges_text.strip()))
                                except Exception: pass
                            _sfx_map_r = {1: "プラス台", 1000: "1000枚以上", 2000: "2000枚以上"}
                            for _block_r in recommended_blocks:
                                _bt_r = _block_r["title"].strip() or "オススメ機種"
                                _bm_r = _block_r["machines"]
                                if not any(m.strip() in _rec_unchecked_machines for m in _bm_r if m.strip()):
                                    continue
                                _valid_r, _ = filter_recommended_machines(_bm_r, _pv_df, _pv_zen_r, _pv_high_r)
                                if not _valid_r:
                                    continue
                                for _thr_r in _block_r.get("thresholds", [1]):
                                    _rec_img_r = generate_recommended_block_image(
                                        _bt_r, _valid_r, _pv_df, _pv_diff, _upd_nb,
                                        min_diff=_thr_r, juggler_cfg=_pv_jcfg_r
                                    )
                                    if _rec_img_r is None:
                                        continue
                                    _sfx_r = _sfx_map_r.get(_thr_r, str(_thr_r))
                                    _tgt_r = f"オススメ_{_make_safe_fn(_bt_r)}_{_sfx_r}.jpg"
                                    for _ri, (_rpn, _) in enumerate(_new_prev):
                                        if _rpn == _tgt_r:
                                            _new_prev[_ri] = (_rpn, _rec_img_r)
                                            break
                                    else:
                                        _new_prev.append((_tgt_r, _rec_img_r))
                                        st.session_state[f"auto_prev_ck_{store}_{len(_new_prev)-1}"] = True
                            _updated = True
                        # with_slump=True の場合、更新された画像にスランプグラフを合成
                        if _updated and with_slump:
                            _s_uid_upd = st.session_state.get(f"_slump_by_uid_{store}")
                            if _s_uid_upd:
                                _upd_tmpl = find_slump_template()
                                _upd_bbb  = _find_slump_bg()
                                _upd_ban2mac: dict[str, str] = {}
                                if _pv_df is not None:
                                    for _, _r_upd in _pv_df.iterrows():
                                        _bs_upd = str(_r_upd.get("台番", "")).split(".")[0]
                                        if _bs_upd.lstrip("-").isdigit():
                                            _upd_ban2mac[_bs_upd] = str(_r_upd.get("機種名", ""))
                                # 更新対象画像のban_mapを「その他を更新」時のDataFrameから動的に構築
                                _upd_dyn_ban_map: dict[str, list[int]] = {}
                                if _jug_extra_dfs and not (with_slump and store == "秋葉原"):
                                    _jug_bans_upd = []
                                    for _jdf in ([_pv_jug_pool.copy()] if _pv_jug_pool is not None and not _pv_jug_pool.empty else []) + _jug_extra_dfs:
                                        _jug_bans_upd += [int(str(b).split(".")[0]) for b in _jdf["台番"].dropna() if str(b).split(".")[0].lstrip("-").isdigit()]
                                    _upd_dyn_ban_map["ジャグラーシリーズ優秀台.jpg"] = sorted(dict.fromkeys(_jug_bans_upd))
                                if _all_dfs:
                                    _son_bans_upd = [int(str(b).split(".")[0]) for b in _son_comb["台番"].dropna() if str(b).split(".")[0].lstrip("-").isdigit()]
                                    if _sonota_split:
                                        _upd_dyn_ban_map["その他の優秀台+1,000枚以上.jpg"] = _son_bans_upd
                                        # ②(③)も更新（+2000枚以上 (/+3000枚以上)）
                                        if _pv_diff is not None and _pv_df is not None:
                                            for _thr_upd, _fn_upd in _sonota_extra_thrs:
                                                _sk_upd = [b for b in _son_bans_upd
                                                           if not _pv_df[_pv_df["台番"] == b].empty
                                                           and int(_pv_diff.loc[_pv_df[_pv_df["台番"] == b].index[0]]) >= _thr_upd]
                                                if _sk_upd:
                                                    _upd_dyn_ban_map[_fn_upd] = _sk_upd
                                    else:
                                        _upd_dyn_ban_map["その他の優秀台ピックアップ.jpg"] = _son_bans_upd
                                for _ui, (_ufn, _uimg) in enumerate(_new_prev):
                                    if _ufn not in _upd_dyn_ban_map:
                                        continue
                                    _bans_u2 = _upd_dyn_ban_map[_ufn]
                                    if not _bans_u2 or _upd_tmpl is None:
                                        continue
                                    _g_imgs_u2: list["Image.Image"] = []
                                    for _b_u2 in _bans_u2:
                                        _it_u2 = _s_uid_upd.get(str(_b_u2))
                                        if _it_u2 is None or not _it_u2.get("points"):
                                            continue
                                        _dn_u2 = (_it_u2.get("_convertedName") or _it_u2.get("displayName") or _upd_ban2mac.get(str(_b_u2), str(_b_u2)))
                                        try:
                                            _g_imgs_u2.append(draw_slump_graph(_upd_tmpl, str(_b_u2), _dn_u2, _it_u2["points"], diff=_it_u2.get("diff"),
                                                                               machine_name=_dn_u2))
                                        except Exception:
                                            pass
                                    if store == "秋葉原":
                                        _u2_bare  = re.sub(r"^\d{2}_", "", _ufn)
                                        if _u2_bare.startswith("その他の優秀台"):
                                            # ①②③分割ファイル名(その他の優秀台+N,000枚以上)は青バーを「その他の優秀台ピックアップ」に統一
                                            _u2_title = "その他の優秀台ピックアップ"
                                        else:
                                            _u2_title = (st.session_state.get(f"_inagawa_title_map_{store}", {}).get(_u2_bare)
                                                         or re.sub(r"[①②③④⑤⑥⑦⑧⑨⑩]", "", os.path.splitext(_u2_bare)[0]))
                                        _u2_slump = _build_slump_title_img(_u2_title, _g_imgs_u2, _upd_bbb)
                                        if _u2_slump is not None:
                                            _new_prev[_ui] = (_ufn, _u2_slump)
                                    else:
                                        _new_prev[_ui] = (_ufn, _attach_slump_to_table(_uimg, _g_imgs_u2, _upd_bbb))
                                    if len(_g_imgs_u2) >= 16 and store != "秋葉原":
                                        try:
                                            _side_ufn = os.path.splitext(_ufn)[0] + "_side.jpg"
                                            _side_u2  = _attach_slump_to_table_side(_uimg, _g_imgs_u2, _upd_bbb)
                                            for _si2, (_spn2, _) in enumerate(_new_prev):
                                                if _spn2 == _side_ufn:
                                                    _new_prev[_si2] = (_spn2, _side_u2)
                                                    break
                                            else:
                                                _new_prev.append((_side_ufn, _side_u2))
                                                st.session_state[f"auto_prev_ck_{store}_{len(_new_prev)-1}"] = True
                                        except Exception:
                                            pass
                        if _updated:
                            st.session_state[_aprev_key] = _new_prev
                            st.rerun()
            with _btn_clr:
                if st.button("🔄 プレビューをクリア", key="auto_preview_clear_btn", use_container_width=True):
                    for _k in (_aprev_key, _aprev_df_key, _aprev_di_key, _aprev_ex_key, _aprev_hr_key, _aprev_zen_key, _aprev_jug_ex_key, _aprev_jug_pool_key, _aprev_narabi_key):
                        st.session_state.pop(_k, None)
                    st.session_state.pop(f"sue_preview_{store}", None)
                    st.session_state.pop(f"sue_prev_tails_{store}", None)
                    st.session_state.pop(f"jug_sue_preview_{store}", None)
                    st.session_state.pop(f"jug_sue_prev_tails_{store}", None)
                    for _ci in range(20):
                        st.session_state.pop(f"sue_ck_{store}_{_ci}", None)
                        st.session_state.pop(f"jug_sue_ck_{store}_{_ci}", None)
                    for _ci in range(len(_auto_previews)):
                        st.session_state.pop(f"auto_prev_ck_{store}_{_ci}", None)
                    st.rerun()

    # ── ⑧ 実行ボタン（常に描画・ファイル未選択時は disabled）─────────
    st.markdown("### ⑧ 実行")
    run_clicked = st.button(
        "▶▶ 自動処理を開始",
        type="primary",
        use_container_width=True,
        disabled=(uploaded is None),
        key="auto_run",
    )
    # Cloud のみ：ボタン直下にZIPダウンロード用スロットを確保
    _auto_zip_slot = st.empty() if _IS_CLOUD else None

    if uploaded is None:
        st.info("⬆️ まずExcelをアップロードしてください。")
    elif run_clicked:
        # 実行時に必ず保存（on_changeが発火しなかった場合のフォールバック）
        _save_auto_inputs(store)
        # オススメ機種設定をJSON保存（次回起動時に復元する）
        if store in EXTENDED_FEATURE_STORES:
            _s = load_store_settings(store)
            _s["rec_enabled"] = bool(st.session_state.get(f"rec_enabled_{store}", False))
            if recommended_blocks:
                for _bi, _bk in enumerate(["1","2","3","4","5","6"]):
                    _s[f"recommended_title_{_bk}"]    = recommended_blocks[_bi]["title"]
                    _s[f"recommended_machines_{_bk}"] = recommended_blocks[_bi]["machines"]
                    _fk = f"rec_f_{_bk}_{store}"
                    if _fk in st.session_state:
                        _s[f"recommended_filter_{_bk}"] = st.session_state[_fk]
            save_store_settings(store, _s)

        stem     = os.path.splitext(uploaded.name)[0]
        dir_stem = stem.replace("_20S", "")
        if _IS_CLOUD:
            _run_tmpdir = tempfile.mkdtemp()
            excel_path  = os.path.join(_run_tmpdir, uploaded.name)
            output_dir  = os.path.join(_run_tmpdir, dir_stem)
        else:
            _run_tmpdir = tempfile.mkdtemp()
            excel_path  = os.path.join(_run_tmpdir, uploaded.name)
            output_dir  = os.path.join(_DESKTOP, dir_stem)
        narabi_dir = os.path.join(output_dir, "並び画像")
        narabi_bans = ranges_to_bans(narabi_ranges) if narabi_ok else set()
        if kojin_enabled and kojin_narabi_ranges_text.strip():
            try:
                narabi_bans |= ranges_to_bans(parse_ranges(kojin_narabi_ranges_text.strip()))
            except Exception:
                pass
        if kojin_enabled and kojin_narabi2_ranges_text.strip():
            try:
                narabi_bans |= ranges_to_bans(parse_ranges(kojin_narabi2_ranges_text.strip()))
            except Exception:
                pass

        with open(excel_path, "wb") as f:
            f.write(uploaded.getvalue())
        os.makedirs(output_dir, exist_ok=True)

        _is_manual_mode = (not with_slump) and st.session_state.get(f"_manual_preview_mode_{store}", False)
        if _is_manual_mode:
            with st.status("記入部分のみ処理を実行中…", expanded=True) as _m_status:
                def _m_log(msg: str) -> None:
                    st.write(msg)
                try:
                    uploaded.seek(0)
                    _raw_exec_m = _read_uploaded_df(uploaded)
                    _df_exec_m, _ = normalize_df(_raw_exec_m)
                    _df_exec_m = apply_name_conversion(_df_exec_m)
                    if "差枚" in _df_exec_m.columns:
                        _df_exec_m["差枚"] = _df_exec_m["差枚"].apply(_pipeline_calc_d)
                    _diff_exec_m = _df_exec_m["差枚"].copy()
                    _exec_order: list[str] = []
                    _used_fns_e: set[str] = set()
                    _m_zen:  list[dict] = []
                    _m_high: list[dict] = []
                    _m_nami: list[dict] = []

                    def _unique_fn_e(base: str) -> str:
                        stem, ext = os.path.splitext(base)
                        candidate = base
                        ctr = 2
                        while candidate in _used_fns_e:
                            candidate = f"{stem}_{ctr}{ext}"
                            ctr += 1
                        _used_fns_e.add(candidate)
                        return candidate

                    # ② 個別画像 - 全台
                    if kojin_enabled:
                        for _km_e in kojin_zentai_machines:
                            _km_e = _km_e.strip()
                            if not _km_e:
                                continue
                            _mg_e = _df_exec_m[_df_exec_m["機種名"] == _km_e].copy().reset_index(drop=True)
                            if _mg_e.empty:
                                continue
                            _md_e = _diff_exec_m.loc[_df_exec_m[_df_exec_m["機種名"] == _km_e].index].reset_index(drop=True)
                            _mfn_e = _unique_fn_e(f"{_make_safe_fn(_km_e)}.jpg")
                            _mout = os.path.join(output_dir, _mfn_e)
                            _save_jpeg(_build_machine_img(_mg_e, _km_e, _stat_from_diff(_md_e)), _mout)
                            _exec_order.append(_mfn_e)
                            _m_log(f"  ✅ 全台「{_km_e}」({len(_mg_e)}台)")
                            _m_zen.append({"name": _km_e, "count": int((_md_e > 0).sum()), "total": len(_mg_e), "diffs": sorted([int(d) for d in _md_e.tolist() if int(d) >= 1000], reverse=True), "all_avg_diff": int(round(_md_e.mean()))})

                        # ② 個別画像 - 優秀台
                        _me_cfg = get_store_config(store)
                        for _km_e in kojin_yushu_machines:
                            _km_e = _km_e.strip()
                            if not _km_e:
                                continue
                            _mga_e = _df_exec_m[_df_exec_m["機種名"] == _km_e]
                            if _mga_e.empty:
                                continue
                            _mda_e = _diff_exec_m.loc[_mga_e.index]
                            _mgp_e = _kojin_yushu_filter(_km_e, _mga_e, _mda_e, _me_cfg).reset_index(drop=True)
                            if _mgp_e.empty:
                                continue
                            _metit = f"{_km_e}（優秀台）"
                            _mefn_e = _unique_fn_e(f"{_make_safe_fn(_metit)}.jpg")
                            _meout = os.path.join(output_dir, _mefn_e)
                            _save_jpeg(_build_machine_img(_mgp_e, _metit, None), _meout)
                            _exec_order.append(_mefn_e)
                            _m_log(f"  ✅ 優秀台「{_metit}」({len(_mgp_e)}台)")
                            _mga_all_e = _df_exec_m[_df_exec_m["機種名"] == _km_e]
                            _mda_all_e = _diff_exec_m.loc[_mga_all_e.index]
                            _m_high.append({"name": _km_e, "count": int((_mda_all_e > 0).sum()), "total": len(_mga_all_e), "diffs": sorted([int(d) for d in _mda_e.tolist() if int(d) >= 1000], reverse=True), "all_avg_diff": int(round(_mda_all_e.mean())), "has_image": True})

                        # ② その他の優秀台ピックアップ
                        if sonota_extra_text.strip():
                            _se_bans_e = set(expand_machine_numbers(sonota_extra_text))
                            if _se_bans_e:
                                _se_df_e = _df_exec_m[_df_exec_m["台番"].apply(lambda b: int(b) in _se_bans_e)].copy().reset_index(drop=True)
                                if not _se_df_e.empty:
                                    _se_tit_e = sonota_extra_title.strip() or "その他の優秀台ピックアップ"
                                    _sefn_e = _unique_fn_e(f"{_make_safe_fn(_se_tit_e)}.jpg")
                                    _se_out_e = os.path.join(output_dir, _sefn_e)
                                    _save_jpeg(_build_machine_img(_se_df_e, _se_tit_e, None), _se_out_e)
                                    _exec_order.append(_sefn_e)
                                    _m_log(f"  ✅ その他の優秀台ピックアップ「{_se_tit_e}」({len(_se_df_e)}台)")

                    # ③ 並び画像（重複タイトルは台番範囲サフィックスで区別）
                    if narabi_ok and narabi_ranges:
                        _ban_map_e = {int(row["台番"]): i for i, row in _df_exec_m.iterrows()}
                        _nb_infos_e = []
                        for _n_bans_e in narabi_ranges:
                            _n_idxs_e = [_ban_map_e[b] for b in _n_bans_e if b in _ban_map_e]
                            if not _n_idxs_e:
                                continue
                            _ngrp_e = _df_exec_m.loc[_n_idxs_e].copy().reset_index(drop=True)
                            _nms_e  = list(dict.fromkeys(str(m) for m in _ngrp_e["機種名"]))
                            _nn_e   = len(_ngrp_e)
                            if len(_nms_e) == 1: _ntit_e = f"{_nms_e[0]}({_nn_e}台並び)"
                            elif len(_nms_e) == 2: _ntit_e = f"{_nms_e[0]}+{_nms_e[1]}({_nn_e}台並び)"
                            else: _ntit_e = f"{_nms_e[0]}～{_nms_e[-1]}({_nn_e}台並び)"
                            _nb_infos_e.append((_ngrp_e, _ntit_e, list(_n_bans_e)))
                        from collections import Counter as _CtrE
                        _dup_tits_e = {t for t, c in _CtrE(i[1] for i in _nb_infos_e).items() if c > 1}
                        for _ngrp_e, _ntit_e, _n_blist_e in _nb_infos_e:
                            _nds_e = _ngrp_e["差枚"]
                            _nstat_e = {"total_diff": int(_nds_e.sum()), "avg_diff": int(round(_nds_e.mean())), "win_count": int((_nds_e > 0).sum()), "total_count": len(_ngrp_e)}
                            if _ntit_e in _dup_tits_e:
                                _nb_s_e = int(_ngrp_e.iloc[0]["台番"]); _nb_e_e = int(_ngrp_e.iloc[-1]["台番"])
                                _file_tit_e = f"{_ntit_e}（{_nb_s_e}～{_nb_e_e}）"
                            else:
                                _file_tit_e = _ntit_e
                            _nfn_e = _unique_fn_e(f"{_make_safe_fn(_file_tit_e)}.jpg")
                            _nout_e = os.path.join(output_dir, _nfn_e)
                            _save_jpeg(_build_machine_img(_ngrp_e, _ntit_e, _nstat_e), _nout_e)
                            _exec_order.append(_nfn_e)
                            _m_log(f"  ✅ 並び「{_file_tit_e}」")
                            _nms_e2 = list(dict.fromkeys(str(m) for m in _ngrp_e["機種名"]))
                            if len(_nms_e2) == 1: _mach_e = _nms_e2[0]
                            elif len(_nms_e2) == 2: _mach_e = f"{_nms_e2[0]}+{_nms_e2[1]}"
                            else: _mach_e = f"{_nms_e2[0]}～{_nms_e2[-1]}"
                            _bls_e = sorted(_n_blist_e)
                            _is_consec_e = len(_bls_e) >= 2 and (_bls_e[-1] - _bls_e[0] + 1 == len(_bls_e))
                            _br_e = (f"{_bls_e[0]}-{_bls_e[-1]}" if _is_consec_e else (str(_bls_e[0]) if len(_bls_e)==1 else "+".join(str(b) for b in _bls_e)))
                            _m_nami.append({"title": _ntit_e, "count": _nstat_e["total_count"], "avg_diff": _nstat_e["avg_diff"], "machine": _mach_e, "ban_range": _br_e, "bans": _bls_e})

                    # ④ 末尾画像
                    if st.session_state.get("suebangai_enabled", False):
                        _m_sue_tails_e = [t for _i in range(1, 4) if (t := st.session_state.get(f"suebangai_tail_input_{_i}", "").strip())]
                        _m_sue_mode_e = st.session_state.get("suebangai_mode", "全台")
                        for _fn_e, _img_e in _gen_sue_imgs_on_fly(_m_sue_tails_e, _m_sue_mode_e, is_juggler=False):
                            _fn_e = _unique_fn_e(_fn_e)
                            _sout_e = os.path.join(output_dir, _fn_e)
                            _save_jpeg(_img_e, _sout_e)
                            _exec_order.append(_fn_e)
                            _m_log(f"  ✅ 末尾画像「{_fn_e}」")
                    if st.session_state.get("jug_sue_enabled", False):
                        _m_jt_e = [t for _i in range(1, 4) if (t := st.session_state.get(f"jug_sue_tail_input_{_i}", "").strip())]
                        _m_jm_e = st.session_state.get("jug_sue_mode", "全台")
                        for _fn_e, _img_e in _gen_sue_imgs_on_fly(_m_jt_e, _m_jm_e, is_juggler=True):
                            _fn_e = _unique_fn_e(_fn_e)
                            _sout_e = os.path.join(output_dir, _fn_e)
                            _save_jpeg(_img_e, _sout_e)
                            _exec_order.append(_fn_e)
                            _m_log(f"  ✅ ジャグラー末尾画像「{_fn_e}」")

                    # ⑤ オススメ機種ピックアップ
                    if recommended_blocks:
                        _ms_cfg_e = get_store_config(store)
                        _mj_cfg_e = {"series": _ms_cfg_e["juggler_series"], "jobs_map": {j[0]: j[1] for j in _ms_cfg_e["juggler_jobs"]}, "g_min": _ms_cfg_e["juggler_g_min"], "diff_bonus": _ms_cfg_e["diff_bonus"]}
                        _ms_sfx_e = {1: "プラス台", 1000: "1000枚以上", 2000: "2000枚以上"}
                        for _blk_e in recommended_blocks:
                            _mbt_e = _blk_e["title"].strip() or "オススメ機種"
                            _mbm_e = [m.strip() for m in _blk_e["machines"] if m.strip()]
                            if not _mbm_e:
                                continue
                            for _mthr_e in _blk_e.get("thresholds", [1]):
                                _mrimg_e = generate_recommended_block_image(_mbt_e, _mbm_e, _df_exec_m, _diff_exec_m, set(), min_diff=_mthr_e, juggler_cfg=_mj_cfg_e)
                                if _mrimg_e is None:
                                    continue
                                _ms_sfxv_e = _ms_sfx_e.get(_mthr_e, str(_mthr_e))
                                _mrfn_e = f"オススメ_{_make_safe_fn(_mbt_e)}_{_ms_sfxv_e}.jpg"
                                _mrfn_e = _unique_fn_e(_mrfn_e)
                                _mrout_e = os.path.join(output_dir, _mrfn_e)
                                _save_jpeg(_mrimg_e, _mrout_e)
                                _exec_order.append(_mrfn_e)
                                _m_log(f"  ✅ オススメ「{_mbt_e}」({_ms_sfxv_e})")

                    # 連番プレフィックス付与
                    _seq_e = 1
                    for _bfn_e in _exec_order:
                        _src_e = os.path.join(output_dir, _bfn_e)
                        if os.path.exists(_src_e):
                            os.replace(_src_e, os.path.join(output_dir, f"{_seq_e:02d}_{_bfn_e}"))
                            _seq_e += 1

                    # 結果テキスト生成
                    _m_report_text = ""
                    try:
                        import datetime as _dt_rt
                        _m_date = None
                        _m_stem_rt = os.path.splitext(uploaded.name)[0]
                        _m_dp = _m_stem_rt.split("_")[0]
                        if len(_m_dp) == 8 and _m_dp.isdigit():
                            _m_date = _dt_rt.date(int(_m_dp[:4]), int(_m_dp[4:6]), int(_m_dp[6:8]))
                        _m_excel: list[dict] = []
                        if sonota_extra_text.strip():
                            _se_bns_rt = set(expand_machine_numbers(sonota_extra_text))
                            if _se_bns_rt:
                                for _idx_rt, _row_rt in _df_exec_m.iterrows():
                                    if int(_row_rt["台番"]) in _se_bns_rt:
                                        _m_excel.append({"name": str(_row_rt["機種名"]), "diff": int(_diff_exec_m.loc[_idx_rt]), "ban": int(_row_rt["台番"])})
                        _m_sue_data: list[dict] = []
                        if st.session_state.get("suebangai_enabled", False):
                            for _t_rt in [t for _i in range(1, 4) if (t := st.session_state.get(f"suebangai_tail_input_{_i}", "").strip())]:
                                _sf_rt = _df_exec_m[_df_exec_m["台番"].apply(lambda b: (str(int(b))[-len(_t_rt):] == _t_rt) if _t_rt.isdigit() and len(_t_rt) in (1,2) else ((s:=str(int(b))) and len(s)>=2 and s[-2]==s[-1]))]
                                if _sf_rt.empty: continue
                                _sfd_rt = _diff_exec_m.loc[_sf_rt.index]
                                _m_sue_data.append({"tail": _t_rt, "total": len(_sf_rt), "win_count": int((_sfd_rt > 0).sum()), "avg_diff": int(round(_sfd_rt.mean()))})
                        _m_jug_data: list[dict] = []
                        if st.session_state.get("jug_sue_enabled", False):
                            _jug_ser_rt = set(get_store_config(store)["juggler_series"])
                            for _t_rt in [t for _i in range(1, 4) if (t := st.session_state.get(f"jug_sue_tail_input_{_i}", "").strip())]:
                                _sf_rt = _df_exec_m[_df_exec_m["台番"].apply(lambda b: (str(int(b))[-len(_t_rt):] == _t_rt) if _t_rt.isdigit() and len(_t_rt) in (1,2) else ((s:=str(int(b))) and len(s)>=2 and s[-2]==s[-1]))]
                                _sf_rt = _sf_rt[_sf_rt["機種名"].isin(_jug_ser_rt)]
                                if _sf_rt.empty: continue
                                _sfd_rt = _diff_exec_m.loc[_sf_rt.index]
                                _m_jug_data.append({"tail": _t_rt, "total": len(_sf_rt), "win_count": int((_sfd_rt > 0).sum()), "avg_diff": int(round(_sfd_rt.mean()))})
                        _m_report_text = generate_report_text(
                            store_name=store, date=_m_date,
                            zen_dai_list=_m_zen, high_ratio_list=_m_high,
                            nami_list=_m_nami, excellent_list=_m_excel,
                            diff_raw=_diff_exec_m, df=_df_exec_m,
                            suebangai_data=_m_sue_data or None,
                            jug_sue_data=_m_jug_data or None,
                        )
                        for _old_rt, _new_rt in STORE_RESULT_TRANSFORMS.get(store, []):
                            _m_report_text = _m_report_text.replace(_old_rt, _new_rt)
                        _txt_name_m = (f"{_m_date.month:02d}{_m_date.day:02d}_結果.txt" if _m_date else "結果.txt")
                        with open(os.path.join(output_dir, _txt_name_m), "w", encoding="utf-8") as _f_rt:
                            _f_rt.write(_m_report_text)
                    except Exception as _rte:
                        _m_log(f"  ⚠️ 結果テキスト生成エラー: {_rte}")
                        import traceback as _tbe
                        _m_log(_tbe.format_exc())
                    _m_status.update(label="✅ 全処理完了！", state="complete", expanded=False)

                    pass  # ZIP は status 外で生成
                except Exception as _me:
                    _m_status.update(label="⚠️ エラーあり", state="error", expanded=True)
                    st.error(f"❌ エラー: {_me}")
                    with st.expander("詳細"):
                        st.code(traceback.format_exc())
            st.markdown("### 生成されたファイル")
            if not _IS_CLOUD:
                st.info(f"📁 `{output_dir}`")
            if os.path.isdir(output_dir):
                _m_imgs = sorted(
                    f for f in os.listdir(output_dir)
                    if f.lower().endswith((".png", ".jpg", ".jpeg"))
                )
                if _m_imgs:
                    for _mf in _m_imgs:
                        with st.expander(_mf, expanded=False):
                            st.image(os.path.join(output_dir, _mf), use_container_width=True)
                else:
                    st.warning("画像ファイルが見つかりませんでした。")
            if _m_report_text:
                st.markdown("---")
                if _IS_CLOUD:
                    st.caption(f"📄 {_txt_name_m} をZIPに含めます")
                else:
                    st.caption(f"📄 {_txt_name_m} を保存しました")
                import html as _html_m
                _safe_m = _html_m.escape(_m_report_text)
                _lines_m = _m_report_text.count("\n") + 1
                _h_m = min(600, max(200, _lines_m * 20 + 110))
                st.iframe(f"""
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:8px;">
              <span style="font-size:1.4rem;font-weight:700;">結果報告</span>
              <button id="cb_m" onclick="
                var t=document.getElementById('rt_m');
                t.select();t.setSelectionRange(0,99999);
                document.execCommand('copy');
                this.textContent='✅ コピー済み';this.style.background='#4CAF50';
                var b=this;setTimeout(function(){{b.textContent='📋 コピー';b.style.background='#2F559E';}},2000);
              " style="padding:5px 14px;background:#2F559E;color:#fff;
                       border:none;border-radius:4px;cursor:pointer;font-size:14px;">
                📋 コピー
              </button>
            </div>
            <textarea id="rt_m" readonly
              style="width:100%;height:{_h_m - 70}px;font-family:monospace;font-size:13px;
                     border:1px solid #ccc;padding:8px;box-sizing:border-box;resize:vertical;"
            >{_safe_m}</textarea>
            """, height=_h_m)
            if os.path.isdir(output_dir):
                try:
                    _m_zip_data = _make_zip_bytes(output_dir)
                    if _IS_CLOUD:
                        if _auto_zip_slot is not None:
                            with _auto_zip_slot:
                                st.download_button(
                                    "⬇️ 生成画像をダウンロード (ZIP)",
                                    data=_m_zip_data,
                                    file_name=f"{dir_stem}.zip",
                                    mime="application/zip",
                                    key="manual_zip_dl",
                                )
                    else:
                        st.download_button(
                            label="📥 画像・テキストをZIPでダウンロード",
                            data=_m_zip_data,
                            file_name=f"{dir_stem}.zip",
                            mime="application/zip",
                            key="manual_zip_dl",
                            type="secondary",
                        )
                except Exception as _mze:
                    st.warning(f"ZIP生成に失敗: {_mze}")
            st.stop()

        log_lines: list[str] = []
        recommended_exclusion_logs: list[str] = []

        with st.status("自動処理を実行中…", expanded=True) as status_widget:

            def _log(msg: str) -> None:
                log_lines.append(msg)
                st.write(msg)

            _rec_names: set[str] = {
                m.strip()
                for block in recommended_blocks
                for m in block["machines"]
                if m.strip()
            } | {
                m.strip()
                for m in (kojin_zentai_machines + kojin_yushu_machines)
                if m.strip()
            }
            _sue_tails_run: list[str] = []
            if st.session_state.get("suebangai_enabled", False):
                _sue_tails_run += [t for i in range(1, 4) if (t := st.session_state.get(f"suebangai_tail_input_{i}", "").strip())]
            _jug_sue_tails_run: list[str] = []
            if st.session_state.get("jug_sue_enabled", False):
                _jug_sue_tails_run += [t for i in range(1, 4) if (t := st.session_state.get(f"jug_sue_tail_input_{i}", "").strip())]
            result = run_auto_pipeline(
                excel_path, output_dir, store, narabi_bans, _log,
                narabi_ranges=narabi_ranges if narabi_ok else None,
                recommended_machines=_rec_names,
                suebangai_tails=_sue_tails_run,
                sonota_exclude={m.strip() for block in recommended_blocks for m in block["machines"] if m.strip()},
                jug_suebangai_tails=_jug_sue_tails_run,
                variety_bans=(ranges_to_bans(parse_ranges(variety_ranges_text.strip())) if (with_slump and store == "秋葉原" and variety_enabled and variety_ranges_text.strip()) else set()),
            )

            # スランプ付き: その他の優秀台ピックアップ①②(③)生成（秋葉原/上野新館）
            if _sonota_split and result.get("ok"):
                _s3_old = os.path.join(output_dir, "その他の優秀台ピックアップ.jpg")
                _s3_1   = os.path.join(output_dir, "その他の優秀台+1,000枚以上.jpg")
                if os.path.exists(_s3_old):
                    os.replace(_s3_old, _s3_1)
                    _rfl = result["files"]
                    for _ri in range(len(_rfl)):
                        if os.path.basename(_rfl[_ri]) == "その他の優秀台ピックアップ.jpg":
                            _rfl[_ri] = _s3_1
                            break
                # ②(③): ①の台番を diff_raw で +2000 (/+3000) に絞る
                _s3_df_r   = result.get("df")
                _s3_diff_r = result.get("diff_raw")
                _s3_1_bans = sorted({int(_e["ban"]) for _e in result.get("sonota_excellent_list", []) if "ban" in _e})
                if _s3_df_r is not None and _s3_diff_r is not None and _s3_1_bans:
                    for _thr, _fn_ex in _sonota_extra_thrs:
                        _out = os.path.join(output_dir, _fn_ex)
                        _s3_k = [
                            _b for _b in _s3_1_bans
                            if not (_s3_df_r[_s3_df_r["台番"] == _b]).empty
                            and int(_s3_diff_r.loc[_s3_df_r[_s3_df_r["台番"] == _b].index[0]]) >= _thr
                        ]
                        if _s3_k:
                            _s3_k_df = _s3_df_r[_s3_df_r["台番"].isin(_s3_k)].copy().reset_index(drop=True)
                            _s3_k_img = _build_machine_img(_s3_k_df, "その他の優秀台ピックアップ", None)
                            _save_jpeg(_s3_k_img, _out, target_kb=800)
                            result["files"].append(_out)

            # スランプ付き結果ポスト用：合成用データをsession_stateに保存
            if with_slump and result.get("ok"):
                _df_ig = result.get("df")
                # プレビューに出た台番セットを収集（各リストの "bans" フィールドを使用）
                _ig_preview_bans: set[int] = set()
                # Step1 全台プラス機種別 JPG の台番
                for _zd in result.get("zen_dai_list", []):
                    for _b in _zd.get("bans", []):
                        _ig_preview_bans.add(_b)
                # Step2/3 高配分個別 JPG の台番（has_image=True のもの）
                for _hr in result.get("high_ratio_list", []):
                    if _hr.get("has_image", False):
                        for _b in _hr.get("bans", []):
                            _ig_preview_bans.add(_b)
                # 統合画像（ジャグラーシリーズ優秀台）: jug_pool_df の全台番
                # ※ jug_excellent_list は diff>=1000 のみで、確率条件通過の diff<1000 台が漏れるため pool_df を使う
                _jug_pool = result.get("jug_pool_df")
                if _jug_pool is not None and not _jug_pool.empty:
                    for _b in _jug_pool["台番"].dropna():
                        try:
                            _ig_preview_bans.add(int(str(_b).split(".")[0]))
                        except (ValueError, TypeError):
                            pass
                # 統合画像（その他の優秀台ピックアップ）: excellent_list の台番
                for _eld in result.get("excellent_list", []):
                    try:
                        _ig_preview_bans.add(int(_eld["ban"]))
                    except (KeyError, ValueError, TypeError):
                        pass
                # ファイル名 → 台番 マッピング（表＋スランプグラフ合成用）
                _ig_ban_map: dict[str, list[int]] = {}
                _ig_title_map: dict[str, str] = {}  # ファイル名 → 表示タイトル
                for _zd2 in result.get("zen_dai_list", []):
                    _fn2 = f"{_make_safe_fn(_zd2['name'])}.jpg"
                    _ig_ban_map[_fn2]   = _zd2.get("bans", [])
                    _ig_title_map[_fn2] = _zd2['name']
                for _hr2 in result.get("high_ratio_list", []):
                    if _hr2.get("has_image", False):
                        _fn2 = f"{_make_safe_fn(_hr2['name'])}_高配分.jpg"
                        _ig_ban_map[_fn2]   = _hr2.get("bans", [])
                        _ig_title_map[_fn2] = _hr2['name'] + "（優秀台）"
                if _jug_pool is not None and not _jug_pool.empty:
                    _ig_ban_map["ジャグラーシリーズ優秀台.jpg"] = [
                        int(str(_b2).split(".")[0]) for _b2 in _jug_pool["台番"].dropna()
                        if str(_b2).split(".")[0].lstrip("-").isdigit()
                    ]
                    _ig_title_map["ジャグラーシリーズ優秀台.jpg"] = "ジャグラーシリーズ優秀台"
                _sonota_bans_ig = sorted({int(_e2["ban"]) for _e2 in result.get("sonota_excellent_list", []) if "ban" in _e2})
                if _sonota_bans_ig:
                    if _sonota_split:
                        _ig_ban_map["その他の優秀台+1,000枚以上.jpg"]   = _sonota_bans_ig
                        _ig_title_map["その他の優秀台+1,000枚以上.jpg"] = "その他の優秀台ピックアップ"
                        _ig_s3_df2  = result.get("df")
                        _ig_s3_dr2  = result.get("diff_raw")
                        if _ig_s3_df2 is not None and _ig_s3_dr2 is not None:
                            for _thr_ig, _fn_ig in _sonota_extra_thrs:
                                _ig_k = [
                                    _b for _b in _sonota_bans_ig
                                    if not (_ig_s3_df2[_ig_s3_df2["台番"] == _b]).empty
                                    and int(_ig_s3_dr2.loc[_ig_s3_df2[_ig_s3_df2["台番"] == _b].index[0]]) >= _thr_ig
                                ]
                                if _ig_k:
                                    _ig_ban_map[_fn_ig]   = _ig_k
                                    _ig_title_map[_fn_ig] = "その他の優秀台ピックアップ"
                    else:
                        _ig_ban_map["その他の優秀台ピックアップ.jpg"]   = _sonota_bans_ig
                        _ig_title_map["その他の優秀台ピックアップ.jpg"] = "その他の優秀台ピックアップ"
                for _nami2 in result.get("nami_list", []):
                    _nt2 = _nami2.get("title", "")
                    if _nt2 and _nami2.get("bans"):
                        _fn2 = f"{_make_safe_fn(_nt2)}.jpg"
                        _ig_ban_map[_fn2]   = [int(_b3) for _b3 in _nami2["bans"]]
                        _ig_title_map[_fn2] = _nt2
                st.session_state[f"_inagawa_ban_map_{store}"]   = _ig_ban_map
                st.session_state[f"_inagawa_title_map_{store}"] = _ig_title_map
                # JPGファイルをバイトとして保存（後でnarabi含む全ファイルを再収集）
                _ig_jpgs_save = []
                for _ig_fp in result.get("files", []):
                    if os.path.exists(_ig_fp) and _ig_fp.lower().endswith((".jpg", ".jpeg")):
                        with open(_ig_fp, "rb") as _ig_fh:
                            _ig_jpgs_save.append((os.path.basename(_ig_fp), _ig_fh.read()))
                st.session_state[f"_inagawa_jpgs_{store}"]  = _ig_jpgs_save
                st.session_state[f"_inagawa_df_{store}"]    = _df_ig
                st.session_state[f"_inagawa_bans_{store}"]  = _ig_preview_bans
                _ig_rd = result.get("date")
                _ig_dt_key = st.session_state.get(f"auto_tb_date_{store}")
                st.session_state[f"_inagawa_date_{store}"] = (
                    _ig_rd.strftime("%Y-%m-%d") if hasattr(_ig_rd, "strftime") else str(_ig_rd)
                ) if _ig_rd is not None else (
                    _ig_dt_key.strftime("%Y-%m-%d") if hasattr(_ig_dt_key, "strftime") else str(_ig_dt_key or "")
                )

            # ── ④ 末尾・ジャグラー末尾の保存（プレビュー済みはチェック済みのみ、未生成はオンザフライ）──
            def _compute_sue_stats(tails, mode, is_juggler, df_run) -> list[dict]:
                """末尾ごとの統計（total/win_count/avg_diff）を返す。
                is_juggler=False: 末尾一致する全台（機種不問）
                is_juggler=True:  末尾一致するジャグラー機種のみ（G数・確率フィルターなし）"""
                _stats = []
                _jug_ser2: set[str] = set()
                if is_juggler:
                    _jug_ser2 = set(get_store_config(store)["juggler_series"])
                for _t in tails:
                    if _t == "ゾロ目":
                        _f2 = df_run[df_run["台番"].apply(
                            lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                        )].copy()
                    elif _t.isdigit() and len(_t) in (1, 2):
                        _f2 = df_run[df_run["台番"].astype(str).str[-len(_t):] == _t].copy()
                    else:
                        continue
                    if is_juggler:
                        _f2 = _f2[_f2["機種名"].isin(_jug_ser2)].copy()
                    if _f2.empty:
                        continue
                    _stats.append({
                        "tail": _t,
                        "total": len(_f2),
                        "win_count": int((_f2["差枚"] > 0).sum()),
                        "avg_diff": int(round(_f2["差枚"].mean())) if len(_f2) > 0 else 0,
                    })
                return _stats

            def _save_sue_imgs_run(tails, mode, is_juggler, df_run) -> list[str]:
                _saved = []
                circle_map = {"0":"⓪","1":"①","2":"②","3":"③","4":"④",
                              "5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
                if is_juggler:
                    df_run = df_run.copy()
                    df_run["ゲーム数_rounded"] = df_run["ゲーム数"].apply(round_games)
                    df_run["合算確率_num"] = df_run.apply(
                        lambda r: r["ゲーム数_rounded"] / (r["BB"] + r["RB"])
                                  if (r["BB"] + r["RB"]) > 0 else float("inf"), axis=1)
                    _jcfg = get_store_config(store)
                    _jug_ser = set(_jcfg["juggler_series"])
                    _jug_g_min = _jcfg["juggler_g_min"]
                    _jug_thresh = {m: (p, d) for m, p, d in _jcfg["juggler_jobs"]}
                for _t in tails:
                    if _t == "ゾロ目":
                        _filt = df_run[df_run["台番"].apply(
                            lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                        )].copy()
                        _circ = "ゾロ目"
                        _lbl = "末尾ゾロ目" if is_juggler else "末尾ゾロ目の台"
                    elif _t.isdigit() and len(_t) in (1, 2):
                        _filt = df_run[df_run["台番"].astype(str).str[-len(_t):] == _t].copy()
                        _circ = circle_map.get(_t, _t)
                        _lbl = f"末尾{_circ}" if is_juggler else f"末尾{_circ}番台"
                    else:
                        continue
                    if _filt.empty:
                        continue
                    if is_juggler:
                        _filt = _filt[_filt["機種名"].isin(_jug_ser)].copy()
                        _filt = _filt[_filt["ゲーム数_rounded"] >= _jug_g_min].copy()
                        if _filt.empty:
                            continue
                        _jp_r = mode in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）")
                        _jy_r = mode in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                        _jb_r = mode in ("プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）")
                        if _jp_r or _jy_r:
                            _jrun_total = len(_filt); _jrun_td = int(_filt["差枚"].sum()); _jrun_ad = int(round(_filt["差枚"].mean())); _jrun_wc = int((_filt["差枚"] > 0).sum())
                            if _jp_r:
                                _filt = _filt[_filt["差枚"] > 0].copy()
                                _title = f"ジャグラーの{_lbl}番台のプラス台"
                            else:
                                # 優秀台: 確率フィルター + 差枚 > 0
                                if not _filt.empty:
                                    _ps = _filt["機種名"].map(lambda m: _jug_thresh.get(m, (float("inf"), float("inf")))[0])
                                    _ds = _filt["機種名"].map(lambda m: _jug_thresh.get(m, (float("inf"), float("inf")))[1])
                                    _filt = _filt[((_filt["合算確率_num"] <= _ps) & (_filt["差枚"] >= 0)) |
                                                  (_filt["差枚"] >= _ds)].copy().reset_index(drop=True)
                                _filt = _filt[_filt["差枚"] > 0].copy()
                                _title = f"ジャグラーの{_lbl}番台の優秀台"
                            if _filt.empty:
                                continue
                            _stat = {"total_diff": _jrun_td, "avg_diff": _jrun_ad, "win_count": _jrun_wc, "total_count": _jrun_total} if _jb_r else None
                        else:
                            # 全台: G数フィルターのみ（確率・差枚条件なし）
                            _title = f"ジャグラーの{_lbl}番台"
                            _stat = {"total_diff": int(_filt["差枚"].sum()),
                                     "avg_diff": int(round(_filt["差枚"].mean())),
                                     "win_count": int((_filt["差枚"] > 0).sum()),
                                     "total_count": len(_filt)}
                    else:
                        _p_r = mode in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）")
                        _y_r = mode in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                        _b_r = mode in ("プラス台（ピンクバー付き）", "優秀台（ピンクバー付き）")
                        if _p_r or _y_r:
                            _run_total = len(_filt); _run_td = int(_filt["差枚"].sum()); _run_ad = int(round(_filt["差枚"].mean())); _run_wc = int((_filt["差枚"] > 0).sum())
                            if _p_r:
                                _filt = _filt[_filt["差枚"] > 0].copy()
                                _title = f"{_lbl}のプラス台"
                            else:
                                _rg_col = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _filt.columns), None)
                                _rm = (_filt["差枚"] >= 1000) | ((_filt[_rg_col] >= 1800) & (_filt["差枚"] > 0)) if _rg_col else (_filt["差枚"] >= 1000)
                                _filt = _filt[_rm].copy()
                                _title = f"{_lbl}の優秀台"
                            if _filt.empty:
                                continue
                            _stat = {"total_diff": _run_td, "avg_diff": _run_ad, "win_count": _run_wc, "total_count": _run_total} if _b_r else None
                        else:
                            _title = _lbl
                            _stat = {"total_diff": int(_filt["差枚"].sum()),
                                     "avg_diff": int(round(_filt["差枚"].mean())),
                                     "win_count": int((_filt["差枚"] > 0).sum()),
                                     "total_count": len(_filt)}
                    _img = _build_machine_img(_filt, _title, _stat)
                    _fn = f"{_make_safe_fn(_title)}.jpg"
                    _save_jpeg(_img, os.path.join(output_dir, _fn))
                    _log(f"  ✅ {'ジャグラー末尾' if is_juggler else '末尾'}画像保存: {_fn}")
                    _saved.append(_fn)
                return _saved

            _sue_saved_fns: list[str] = []  # リネーム順序に追加するため収集
            _sue_stats_data: list[dict] = []      # 結果テキスト用・通常末尾
            _jug_sue_stats_data: list[dict] = []  # 結果テキスト用・ジャグラー末尾
            _sue_tails_r: list[str] = []   # banmap追加で参照するため先に初期化
            _sue_mode_r: str = "全台"
            _run_df_for_sue = result.get("df")
            if _run_df_for_sue is not None:
                _sue_tails_r = [t for i in range(1, 4) if (t := st.session_state.get(f"suebangai_tail_input_{i}", "").strip())]
                _sue_mode_r = st.session_state.get("suebangai_mode", "全台")
                _sue_prevs_run = st.session_state.get(f"sue_preview_{store}", [])
                if _sue_prevs_run:
                    # プレビュー生成時のモードを sue_prev_tails_{store} から復元
                    _cached_sue_rt = st.session_state.get(f"sue_prev_tails_{store}", "")
                    if "|" in _cached_sue_rt:
                        _sue_mode_r = _cached_sue_rt.split("|", 1)[1]
                    for _ci, (_sm_fn, _sm_img) in enumerate(_sue_prevs_run):
                        if st.session_state.get(f"sue_ck_{store}_{_ci}", True):
                            _save_jpeg(_sm_img, os.path.join(output_dir, _sm_fn))
                            _log(f"  ✅ 末尾画像保存: {_sm_fn}")
                            _sue_saved_fns.append(_sm_fn)
                    if _sue_tails_r:
                        _sue_stats_data = _compute_sue_stats(_sue_tails_r, _sue_mode_r, False, _run_df_for_sue)
                elif st.session_state.get("suebangai_enabled", False):
                    if _sue_tails_r:
                        _sue_saved_fns += _save_sue_imgs_run(_sue_tails_r, _sue_mode_r, False, _run_df_for_sue)
                        _sue_stats_data = _compute_sue_stats(_sue_tails_r, _sue_mode_r, False, _run_df_for_sue)

                _jug_tails_r = [t for i in range(1, 4) if (t := st.session_state.get(f"jug_sue_tail_input_{i}", "").strip())]
                _jug_sue_mode_r = st.session_state.get("jug_sue_mode", "全台")
                _jsue_prevs_run = st.session_state.get(f"jug_sue_preview_{store}", [])
                if _jsue_prevs_run:
                    for _ci, (_jm_fn, _jm_img) in enumerate(_jsue_prevs_run):
                        if st.session_state.get(f"jug_sue_ck_{store}_{_ci}", True):
                            _save_jpeg(_jm_img, os.path.join(output_dir, _jm_fn))
                            _log(f"  ✅ ジャグラー末尾画像保存: {_jm_fn}")
                            _sue_saved_fns.append(_jm_fn)
                    if _jug_tails_r:
                        _jug_sue_stats_data = _compute_sue_stats(_jug_tails_r, _jug_sue_mode_r, True, _run_df_for_sue)
                elif st.session_state.get("jug_sue_enabled", False):
                    if _jug_tails_r:
                        _sue_saved_fns += _save_sue_imgs_run(_jug_tails_r, _jug_sue_mode_r, True, _run_df_for_sue)
                        _jug_sue_stats_data = _compute_sue_stats(_jug_tails_r, _jug_sue_mode_r, True, _run_df_for_sue)

            # ── プレビューでチェックを外した画像を削除 / 高配分・並び外しはその他/ジャグラーに追加 ──
            _sonota_extra_bans: list[int] = []  # チェック外し再生成後の全台番（スランプ合成に渡す）
            _unchecked_kojin_y: set[str] = set()  # プレビューでチェック外しされたkojin優秀台機種（全店舗共通）
            _aprev_imgs = st.session_state.get(f"auto_preview_imgs_{store}")
            if _aprev_imgs and result["ok"]:
                _df_res   = result.get("df")
                _diff_res = result.get("diff_raw")
                _ex_jug_series_set = set(get_store_config(store)["juggler_series"])
                # kojin個別画像がある機種はチェック外し時に統合画像へ追加しない
                _kojin_ex_set = {m.strip() for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip()}
                # 高配分ファイル名 → 機種名マップ
                _hr_map = {
                    f"{_make_safe_fn(item['name'])}_高配分.jpg": item["name"]
                    for item in result.get("high_ratio_list", [])
                }
                # 全台系ファイル名 → 機種名マップ
                _zen_map = {
                    f"{_make_safe_fn(item['name'])}.jpg": item["name"]
                    for item in result.get("zen_dai_list", [])
                }
                # 並び画像ファイル名 → 台番リストマップ（プレビュー生成時に保存済み）
                _narabi_ban_map_ex = st.session_state.get(_aprev_narabi_key, {})
                _extra_dfs:   list[pd.DataFrame] = []
                _extra_diffs: list[pd.Series]    = []
                _jug_ex_dfs:  list[pd.DataFrame] = []

                # チェック済みの高配分・全台系画像がある機種セット（並び外しで重複追加しないため）
                _checked_img_machines_ex: set[str] = set()
                for _si2, (_spn2, _) in enumerate(_aprev_imgs):
                    if st.session_state.get(f"auto_prev_ck_{store}_{_si2}", True):
                        _hm2 = _hr_map.get(_spn2)
                        if _hm2:
                            _checked_img_machines_ex.add(_hm2)
                        _zm2 = _zen_map.get(_spn2)
                        if _zm2:
                            _checked_img_machines_ex.add(_zm2)

                for _ci, (_pname, _) in enumerate(_aprev_imgs):
                    if not st.session_state.get(f"auto_prev_ck_{store}_{_ci}", True):
                        _del_path = os.path.join(output_dir, _pname)
                        if os.path.exists(_del_path):
                            os.remove(_del_path)
                            _log(f"  🗑️ スキップ: {_pname}")
                        if _df_res is None or _diff_res is None:
                            continue
                        # 高配分をチェック外した → kojin済みはスキップ・ジャグラーならジャグラー優秀台へ、他はその他へ
                        _m = _hr_map.get(_pname)
                        if _m and _m not in _kojin_ex_set:
                            _m_rows = _df_res[_df_res["機種名"] == _m]
                            if not _m_rows.empty:
                                _m_diff = _diff_res.loc[_m_rows.index]
                                _m_mask = _m_diff >= 1000
                                _m_good = _m_rows[_m_mask.values].copy().reset_index(drop=True)
                                _m_good_diff = _m_diff[_m_mask].reset_index(drop=True)
                                if not _m_good.empty:
                                    if _m in _ex_jug_series_set and not (with_slump and store == "秋葉原"):
                                        _jug_ex_dfs.append(_m_good)
                                    else:
                                        _extra_dfs.append(_m_good)
                                        _extra_diffs.append(_m_good_diff)
                        # 全台系をチェック外した → kojin済みはスキップ・同様に振り分け
                        _mz = _zen_map.get(_pname)
                        if _mz and not _m and _mz not in _kojin_ex_set:
                            _mz_rows = _df_res[_df_res["機種名"] == _mz]
                            if not _mz_rows.empty:
                                _mz_diff = _diff_res.loc[_mz_rows.index]
                                _mz_mask = _mz_diff >= 1000
                                _mz_good = _mz_rows[_mz_mask.values].copy().reset_index(drop=True)
                                _mz_good_diff = _mz_diff[_mz_mask].reset_index(drop=True)
                                if not _mz_good.empty:
                                    if _mz in _ex_jug_series_set and not (with_slump and store == "秋葉原"):
                                        _jug_ex_dfs.append(_mz_good)
                                    else:
                                        _extra_dfs.append(_mz_good)
                                        _extra_diffs.append(_mz_good_diff)
                        # 並び画像をチェック外した → 機種ごとにジャグラー/非ジャグラーへ振り分け
                        _nb_bans = _narabi_ban_map_ex.get(_pname)
                        if _nb_bans:
                            _nb_set = set(_nb_bans)
                            _nb_rows = _df_res[_df_res["台番"].apply(lambda b: int(b) in _nb_set)].copy()
                            if not _nb_rows.empty:
                                _nb_diff = _diff_res.loc[_nb_rows.index]
                                # 非ジャグラー: diff >= 1000 のみその他へ
                                _nb_non_jug_m = ~_nb_rows["機種名"].isin(_ex_jug_series_set)
                                _nb_oth_part = _nb_rows[(_nb_non_jug_m) & (_nb_diff >= 1000).values].copy().reset_index(drop=True)
                                _nb_oth_diff = _nb_diff[(_nb_non_jug_m) & (_nb_diff >= 1000).values].reset_index(drop=True)
                                # 高配分・全台系画像がチェック済みの機種はその他へ追加しない
                                if not _nb_oth_part.empty and _checked_img_machines_ex:
                                    _nb_oth_no_img = ~_nb_oth_part["機種名"].isin(_checked_img_machines_ex)
                                    _nb_oth_part = _nb_oth_part[_nb_oth_no_img.values].copy().reset_index(drop=True)
                                    _nb_oth_diff = _nb_oth_diff[_nb_oth_no_img.values].reset_index(drop=True)
                                if not _nb_oth_part.empty:
                                    _extra_dfs.append(_nb_oth_part)
                                    _extra_diffs.append(_nb_oth_diff)
                                # ジャグラー: 合算確率条件（G数>=g_min AND (確率<=閾値かつdiff>=0) OR diff>=diff_bonus）
                                _nb_jug_base = _nb_rows[_nb_rows["機種名"].isin(_ex_jug_series_set)].copy()
                                if not _nb_jug_base.empty:
                                    _nb_jug_diff2 = _nb_diff.loc[_nb_jug_base.index].reset_index(drop=True)
                                    _nb_jug_base = _nb_jug_base.reset_index(drop=True)
                                    _jcfg_ex = get_store_config(store)
                                    _jg_min_ex = _jcfg_ex["juggler_g_min"]
                                    _jthresh_ex = {m: (p, d) for m, p, d in _jcfg_ex["juggler_jobs"]}
                                    if "ゲーム数_rounded" not in _nb_jug_base.columns:
                                        _nb_jug_base["ゲーム数_rounded"] = _nb_jug_base["ゲーム数"].apply(round_games)
                                    if "合算確率_num" not in _nb_jug_base.columns:
                                        _nb_jug_base["合算確率_num"] = _nb_jug_base.apply(
                                            lambda r: r["ゲーム数_rounded"] / (r["BB"] + r["RB"])
                                                      if (r["BB"] + r["RB"]) > 0 else float("inf"), axis=1)
                                    _g_ok_ex = _nb_jug_base["ゲーム数_rounded"] >= _jg_min_ex
                                    _nb_jug_base = _nb_jug_base[_g_ok_ex.values].reset_index(drop=True)
                                    _nb_jug_diff2 = _nb_jug_diff2[_g_ok_ex.values].reset_index(drop=True)
                                    if not _nb_jug_base.empty:
                                        _ps_ex = _nb_jug_base["機種名"].apply(
                                            lambda m: _jthresh_ex.get(m, (float("inf"), float("inf")))[0])
                                        _ds_ex = _nb_jug_base["機種名"].apply(
                                            lambda m: _jthresh_ex.get(m, (float("inf"), float("inf")))[1])
                                        _jug_cond_ex = ((_nb_jug_base["合算確率_num"] <= _ps_ex.values) &
                                                        (_nb_jug_diff2.values >= 0)) | \
                                                       (_nb_jug_diff2.values >= _ds_ex.values)
                                        _nb_jug_part = _nb_jug_base[_jug_cond_ex].copy().reset_index(drop=True)
                                        # 高配分・全台系画像がチェック済みの機種はジャグラー優秀台へ追加しない
                                        if not _nb_jug_part.empty and _checked_img_machines_ex:
                                            _nb_jug_no_img_ex = ~_nb_jug_part["機種名"].isin(_checked_img_machines_ex)
                                            _nb_jug_part = _nb_jug_part[_nb_jug_no_img_ex.values].copy().reset_index(drop=True)
                                        if not _nb_jug_part.empty:
                                            if with_slump and store == "秋葉原":
                                                # 秋葉原: diff>=1000でその他へ
                                                _nb_jug_part_d = _nb_jug_diff2[_jug_cond_ex].reset_index(drop=True)
                                                _nb_jug_1k = _nb_jug_part[_nb_jug_part_d.values >= 1000].copy().reset_index(drop=True)
                                                _nb_jug_1k_d = _nb_jug_part_d[_nb_jug_part_d.values >= 1000].reset_index(drop=True)
                                                if not _nb_jug_1k.empty:
                                                    _extra_dfs.append(_nb_jug_1k)
                                                    _extra_diffs.append(_nb_jug_1k_d)
                                            else:
                                                _jug_ex_dfs.append(_nb_jug_part)
                        # （優秀台）.jpgをチェック外した → kojin_yushu機種をジャグラー/その他へ
                        _YUSHU_SFX_EX = "（優秀台）.jpg"
                        if _pname.endswith(_YUSHU_SFX_EX) and not _m and not _mz and not _nb_bans:
                            _km_y_ex = _pname[:-len(_YUSHU_SFX_EX)]
                            _kojin_y_ex_set = {m.strip() for m in kojin_yushu_machines if m.strip()}
                            if _km_y_ex in _kojin_y_ex_set:
                                _unchecked_kojin_y.add(_km_y_ex)  # 再生成スキップ用に記録
                                _myrows_ex = _df_res[_df_res["機種名"] == _km_y_ex]
                                if not _myrows_ex.empty:
                                    _mydiff_ex = _diff_res.loc[_myrows_ex.index]
                                    _mymask_ex = _mydiff_ex >= 1000
                                    _mygood_ex = _myrows_ex[_mymask_ex.values].copy().reset_index(drop=True)
                                    _mygood_ex_diff = _mydiff_ex[_mymask_ex].reset_index(drop=True)
                                    if not _mygood_ex.empty:
                                        _has_jug_img_ex = any(_pn == "ジャグラーシリーズ優秀台.jpg" for _pn, _ in _aprev_imgs)
                                        if _km_y_ex in _ex_jug_series_set and _has_jug_img_ex and not (with_slump and store == "秋葉原"):
                                            _jug_ex_dfs.append(_mygood_ex)
                                        else:
                                            _extra_dfs.append(_mygood_ex)
                                            _extra_diffs.append(_mygood_ex_diff)

                # 秋葉原スランプ付き: ジャグラー統合画像は不要 → jug_pool +1000枚をその他へ
                if with_slump and store == "秋葉原" and result["ok"] and _df_res is not None and _diff_res is not None:
                    _jug_pool_ex = result.get("jug_pool_df")
                    if _jug_pool_ex is not None and not _jug_pool_ex.empty:
                        _jp_bans_ex = {int(str(b).split(".")[0]) for b in _jug_pool_ex["台番"].dropna()
                                       if str(b).split(".")[0].lstrip("-").isdigit()}
                        _jp_rows_ex = _df_res[_df_res["台番"].apply(lambda b: int(b) in _jp_bans_ex)].copy()
                        if not _jp_rows_ex.empty:
                            _jp_dr_ex = _diff_res.loc[_jp_rows_ex.index]
                            _jp_mask  = _jp_dr_ex >= 1000
                            _jp_good  = _jp_rows_ex[_jp_mask.values].copy().reset_index(drop=True)
                            _jp_good_d = _jp_dr_ex[_jp_mask].reset_index(drop=True)
                            if not _jp_good.empty:
                                _extra_dfs.append(_jp_good)
                                _extra_diffs.append(_jp_good_d)
                    # ジャグラーシリーズ優秀台.jpg をファイルから削除
                    _jug_del = os.path.join(output_dir, "ジャグラーシリーズ優秀台.jpg")
                    if os.path.exists(_jug_del):
                        os.remove(_jug_del)
                        _log("  🗑️ 秋葉原: ジャグラーシリーズ優秀台.jpg を除外")

                # その他の優秀台ピックアップを再生成
                if _extra_dfs and _df_res is not None and _diff_res is not None:
                    # スランプ付き分割店舗は①.jpgに保存（その他の優秀台ピックアップ.jpgを作らない）
                    if _sonota_split:
                        _sonota_path = os.path.join(output_dir, "その他の優秀台+1,000枚以上.jpg")
                    else:
                        _sonota_path = os.path.join(output_dir, "その他の優秀台ピックアップ.jpg")
                    _ex_bans = {item["ban"] for item in result.get("sonota_excellent_list", [])}
                    if _ex_bans:
                        _ex_rows = _df_res[_df_res["台番"].apply(lambda b: int(b) in _ex_bans)].copy().reset_index(drop=True)
                        _ex_diff = _diff_res.loc[_df_res[_df_res["台番"].apply(lambda b: int(b) in _ex_bans)].index].reset_index(drop=True)
                        all_dfs   = [_ex_rows]   + _extra_dfs
                        all_diffs = [_ex_diff]   + _extra_diffs
                    else:
                        all_dfs   = _extra_dfs
                        all_diffs = _extra_diffs
                    _son_combined = pd.concat(all_dfs,   ignore_index=True)
                    _son_order    = _son_combined["台番"].argsort()
                    _son_combined = _son_combined.iloc[_son_order].reset_index(drop=True)
                    _son_img = _build_machine_img(_son_combined, "その他の優秀台ピックアップ", None)
                    _save_jpeg(_son_img, _sonota_path, target_kb=800)
                    _log(f"  ✅ その他の優秀台ピックアップ再生成: {len(_son_combined)}台")
                    _sonota_extra_bans = [int(str(b).split(".")[0]) for b in _son_combined["台番"].dropna()
                                          if str(b).split(".")[0].lstrip("-").isdigit()]
                    # スランプ付き分割店舗: ②.jpg（+2000枚以上）(③.jpg（+3000枚以上）)も更新
                    if _sonota_split:
                        _son_bans_set2 = set(_son_combined["台番"].dropna().astype(int))
                        _s2_mask = _df_res["台番"].apply(lambda b: int(b) in _son_bans_set2)
                        _s2_diff_vals = _diff_res.loc[_df_res[_s2_mask].index]
                        for _thr_rg, _fn_rg in _sonota_extra_thrs:
                            _sk_path_regen = os.path.join(output_dir, _fn_rg)
                            _sk_bans = set(_df_res[_s2_mask][(_s2_diff_vals.values >= _thr_rg)]["台番"].astype(int))
                            if _sk_bans:
                                _sk_df_regen = _son_combined[_son_combined["台番"].apply(lambda b: int(b) in _sk_bans)].copy().reset_index(drop=True)
                                _save_jpeg(_build_machine_img(_sk_df_regen, "その他の優秀台ピックアップ", None), _sk_path_regen, target_kb=800)
                            elif os.path.exists(_sk_path_regen):
                                os.remove(_sk_path_regen)
                # ジャグラーシリーズ優秀台を再生成（秋葉原スランプ付きは生成しない）
                if _jug_ex_dfs and _df_res is not None and not (with_slump and store == "秋葉原"):
                    _jug_path = os.path.join(output_dir, "ジャグラーシリーズ優秀台.jpg")
                    _jug_pool_res = result.get("jug_pool_df")
                    _jug_ov_res   = result.get("jug_overflow_df")
                    if _jug_pool_res is not None and not _jug_pool_res.empty:
                        _jug_base = [_jug_pool_res.copy()]
                    elif _jug_ov_res is not None and not _jug_ov_res.empty:
                        _jug_hr_names_r = {item["name"] for item in result.get("high_ratio_list", []) if item.get("has_image", False)}
                        _jug_ov_filt = _jug_ov_res[~_jug_ov_res["機種名"].isin(_jug_hr_names_r)].copy().reset_index(drop=True)
                        _jug_base = [_jug_ov_filt] if not _jug_ov_filt.empty else []
                    else:
                        _jug_hr_names_r = {item["name"] for item in result.get("high_ratio_list", []) if item.get("has_image", False)}
                        _jug_ex_bans = {item["ban"] for item in result.get("jug_excellent_list", []) if item["name"] not in _jug_hr_names_r}
                        _jug_base = [_df_res[_df_res["台番"].apply(lambda b: int(b) in _jug_ex_bans)].copy().reset_index(drop=True)] if _jug_ex_bans else []
                    _jug_all = _jug_base + _jug_ex_dfs
                    _jug_comb = pd.concat(_jug_all, ignore_index=True)
                    _jug_ord  = _jug_comb["台番"].argsort()
                    _jug_comb = _jug_comb.iloc[_jug_ord].reset_index(drop=True)
                    if len(_jug_comb) <= 5:
                        # 5台以下 → overflowと同じ扱い: その他の優秀台ピックアップへ
                        _sonota_path_ov = os.path.join(output_dir, "その他の優秀台ピックアップ.jpg")
                        _ex_bans_ov = {item["ban"] for item in result.get("sonota_excellent_list", [])}
                        _ex_rows_ov = _df_res[_df_res["台番"].apply(lambda b: int(b) in _ex_bans_ov)].copy().reset_index(drop=True) if _ex_bans_ov else pd.DataFrame()
                        _ov_dfs = ([_ex_rows_ov] if not _ex_rows_ov.empty else []) + _jug_ex_dfs
                        _ov_son = pd.concat(_ov_dfs, ignore_index=True)
                        _ov_son = _ov_son.drop_duplicates(subset=["台番"])
                        _ov_son = _ov_son.iloc[_ov_son["台番"].argsort()].reset_index(drop=True)
                        _save_jpeg(_build_machine_img(_ov_son, "その他の優秀台ピックアップ", None), _sonota_path_ov, target_kb=800)
                        _log(f"  ✅ ジャグラー{len(_jug_comb)}台→overflow: その他の優秀台ピックアップに追加({len(_ov_son)}台)")
                    else:
                        _has_kojin_jug_ex = any(m.strip() in _ex_jug_series_set for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip())
                        _jug_has_other = _has_kojin_jug_ex or bool(result.get("high_ratio_list")) or bool(result.get("zen_dai_list") and any(
                            item["name"] in _ex_jug_series_set for item in result.get("zen_dai_list", [])))
                        _jug_t = "その他のジャグラーシリーズの優秀台" if _jug_has_other else "ジャグラーシリーズの優秀台"
                        _jug_img_r = _build_machine_img(_jug_comb, _jug_t, None)
                        _save_jpeg(_jug_img_r, _jug_path, target_kb=800)
                        _log(f"  ✅ ジャグラーシリーズ優秀台再生成: {len(_jug_comb)}台")

            # ── 並び画像（subprocess）────────────────────────────────
            narabi_result: dict | None = None
            _moved_narabi: list[str] = []
            if narabi_ok:
                st.write(f"⏳ 並び画像スクリプトを実行中…")
                os.makedirs(narabi_dir, exist_ok=True)
                ok_n, out_n, err_n = _patch_and_run_narabi(
                    STORE_NARABI_SCRIPT[store], excel_path, narabi_dir, narabi_ranges
                )
                narabi_result = {"ok": ok_n, "stdout": out_n, "stderr": err_n}
                st.write(f"{'✅' if ok_n else '❌'} 並び画像{'完了' if ok_n else 'エラー'}")
                # ⑦プレビューでチェックを外した並び画像を narabi_dir から削除（完全一致）
                # narabiスクリプトはmake_safeでASCIIコロン→全角コロンに変換するため両方試みる
                if ok_n and _aprev_imgs and os.path.isdir(narabi_dir):
                    for _ci, (_pname, _) in enumerate(_aprev_imgs):
                        if not st.session_state.get(f"auto_prev_ck_{store}_{_ci}", True):
                            _del_n = os.path.join(narabi_dir, _pname)
                            if not os.path.exists(_del_n):
                                _del_n = os.path.join(narabi_dir, _pname.replace(":", "："))
                            if os.path.exists(_del_n):
                                os.remove(_del_n)
                                _log(f"  🗑️ スキップ(並び): {_pname}")
                # 残りの並び画像を output_dir に移動してサブフォルダを削除
                if ok_n and os.path.isdir(narabi_dir):
                    for _nf in sorted(os.listdir(narabi_dir)):
                        if _nf.lower().endswith((".jpg", ".jpeg")):
                            os.rename(os.path.join(narabi_dir, _nf),
                                      os.path.join(output_dir, _nf))
                            _moved_narabi.append(_nf)
                    try:
                        os.rmdir(narabi_dir)
                    except OSError:
                        pass

            # ── オススメ機種ピックアップ画像（拡張機能店舗）─────────────
            _exec_rec_ban_map: dict[str, list[int]] = {}
            if store in EXTENDED_FEATURE_STORES and result["ok"] and recommended_blocks:
                df_pipe   = result.get("df")
                diff_pipe = result.get("diff_raw")
                if df_pipe is not None and diff_pipe is not None:
                    zen_dai_names    = {item["name"] for item in result.get("zen_dai_list", [])}
                    high_ratio_names = {item["name"] for item in result.get("high_ratio_list", []) if item.get("has_image", True)}
                    if kojin_enabled:
                        zen_dai_names    |= {m.strip() for m in kojin_zentai_machines if m.strip()}
                        high_ratio_names |= {m.strip() for m in kojin_yushu_machines if m.strip()}
                    _scfg = get_store_config(store)
                    _juggler_cfg = {
                        "series":     _scfg["juggler_series"],
                        "jobs_map":   {j[0]: j[1] for j in _scfg["juggler_jobs"]},
                        "g_min":      _scfg["juggler_g_min"],
                        "diff_bonus": _scfg["diff_bonus"],
                    }
                    for block in recommended_blocks:
                        b_title    = block["title"].strip()
                        b_machines = block["machines"]
                        # タイトルも機種名もすべて空なら丸ごとスキップ
                        if not b_title and not any(m.strip() for m in b_machines):
                            continue
                        if not b_title:
                            b_title = "オススメ機種"
                        valid, exc_logs = filter_recommended_machines(
                            b_machines, df_pipe, zen_dai_names, high_ratio_names
                        )
                        recommended_exclusion_logs.extend(exc_logs)
                        if not valid:
                            _log(f"  オススメ「{b_title}」: 除外後に掲載対象機種なし")
                            continue
                        _threshold_suffix = {1: "プラス台", 1000: "1000枚以上", 2000: "2000枚以上"}
                        for _thr in block.get("thresholds", [1]):
                            img = generate_recommended_block_image(
                                b_title, valid, df_pipe, diff_pipe, narabi_bans,
                                min_diff=_thr, juggler_cfg=_juggler_cfg
                            )
                            if img is None:
                                _log(f"  オススメ「{b_title}」({_threshold_suffix.get(_thr, str(_thr))}): 該当台なし")
                                continue
                            _sfx = _threshold_suffix.get(_thr, str(_thr))
                            _rec_fn_exec = f"オススメ_{_make_safe_fn(b_title)}_{_sfx}.jpg"
                            out_rec = os.path.join(output_dir, _rec_fn_exec)
                            _save_jpeg(img, out_rec)
                            result["files"].append(out_rec)
                            _log(f"  ✅ オススメ「{b_title}」({_sfx}/{len(valid)}機種)")
                            # スランプグラフ合成用：画像に含まれる台番を収集
                            _exec_bans: list[int] = []
                            for _rvm in valid:
                                _rgrp = df_pipe[df_pipe["機種名"] == _rvm].copy()
                                if narabi_bans:
                                    _rgrp = _rgrp[~_rgrp["台番"].isin(narabi_bans)]
                                if _rgrp.empty:
                                    continue
                                _rdr = diff_pipe.loc[_rgrp.index]
                                if _juggler_cfg and _rvm in _juggler_cfg.get("series", set()):
                                    _jg_col = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _rgrp.columns), None)
                                    if _jg_col:
                                        _jgmask = _rgrp[_jg_col] >= _juggler_cfg.get("g_min", 2000)
                                        _rgrp = _rgrp[_jgmask]
                                        _rdr  = _rdr[_jgmask]
                                    _jpthr = _juggler_cfg["jobs_map"].get(_rvm)
                                    if _jpthr is not None and "合算確率_num" in _rgrp.columns:
                                        _rmask = ((_rgrp["合算確率_num"] <= _jpthr) & (_rdr >= 0)) | (_rdr >= _juggler_cfg.get("diff_bonus", 1000))
                                    else:
                                        _rmask = _rdr >= 0
                                else:
                                    _rmask = _rdr >= _thr
                                _exec_bans.extend([int(b) for b in _rgrp[_rmask]["台番"].dropna()])
                            if _exec_bans:
                                _exec_rec_ban_map[_rec_fn_exec] = sorted(_exec_bans)

            # ── 個別画像生成 ─────────────────────────────────────────
            if kojin_enabled and result["ok"]:
                df_k   = result.get("df")
                diff_k = result.get("diff_raw")
                if df_k is not None and diff_k is not None:
                    for _km in kojin_zentai_machines:
                        _km = _km.strip()
                        if not _km:
                            continue
                        _kgrp = df_k[df_k["機種名"] == _km].copy().reset_index(drop=True)
                        if _kgrp.empty:
                            _log(f"  個別(全台)「{_km}」: 該当台なし")
                            continue
                        _kdr = diff_k.loc[df_k[df_k["機種名"] == _km].index].reset_index(drop=True)
                        _kimg = _build_machine_img(_kgrp, _km, _stat_from_diff(_kdr))
                        _kout = os.path.join(output_dir, f"{_make_safe_fn(_km)}.jpg")
                        _save_jpeg(_kimg, _kout)
                        result["files"].append(_kout)
                        result["zen_dai_list"].append({
                            "name":         _km,
                            "count":        int((_kdr > 0).sum()),
                            "total":        len(_kgrp),
                            "diffs":        sorted([int(d) for d in _kdr.tolist() if int(d) >= 1000], reverse=True),
                            "all_avg_diff": int(round(_kdr.mean())),
                            "bans":         [int(b) for b in _kgrp["台番"].tolist()],
                        })
                        _log(f"  ✅ 個別(全台)「{_km}」({len(_kgrp)}台)")
                    for _km in kojin_yushu_machines:
                        _km = _km.strip()
                        if not _km:
                            continue
                        if _km in _unchecked_kojin_y:
                            _log(f"  ⏭️ 個別(優秀台)「{_km}」: プレビューでチェック外し済みのためスキップ")
                            continue
                        _kgrp_all = df_k[df_k["機種名"] == _km]
                        if _kgrp_all.empty:
                            _log(f"  個別(優秀台)「{_km}」: 該当台なし")
                            continue
                        _kdr_all  = diff_k.loc[_kgrp_all.index]
                        _kgrp_p   = _kojin_yushu_filter(_km, _kgrp_all, _kdr_all, get_store_config(store), force_1k=(with_slump and store == "秋葉原"))
                        if _kgrp_p.empty:
                            _log(f"  個別(優秀台)「{_km}」: 条件を満たす台なし")
                            continue
                        _kdr_p    = _kdr_all.loc[_kgrp_p.index]
                        _kgrp_p   = _kgrp_p.reset_index(drop=True)
                        _ktitle = f"{_km}（優秀台）"
                        _kimg   = _build_machine_img(_kgrp_p, _ktitle, None)
                        _kout   = os.path.join(output_dir, f"{_make_safe_fn(_km)}（優秀台）.jpg")
                        _save_jpeg(_kimg, _kout)
                        result["files"].append(_kout)
                        result["high_ratio_list"].append({
                            "name":         _km,
                            "count":        len(_kgrp_p),
                            "total":        len(_kgrp_all),
                            "diffs":        sorted([int(d) for d in _kdr_p.tolist() if int(d) >= 1000], reverse=True),
                            "all_avg_diff": int(round(_kdr_all.mean())),
                            "has_image":    True,
                            "is_yushu":     True,
                            "bans":         [int(b) for b in _kgrp_p["台番"].tolist()],
                        })
                        _log(f"  ✅ 個別(優秀台)「{_km}」({len(_kgrp_p)}台)")

                    # 台番範囲 優秀台（ピンクバーあり）
                    if kojin_narabi_ranges_text.strip():
                        try:
                            _rng_list = parse_ranges(kojin_narabi_ranges_text.strip())
                            _rng_bans = ranges_to_bans(_rng_list)
                            if _rng_bans:
                                _rng_df   = df_k[df_k["台番"].apply(lambda b: int(b) in _rng_bans)].copy()
                                _rng_diff = diff_k.loc[_rng_df.index]
                                _rng_p    = _rng_df.copy().reset_index(drop=True)
                                if not _rng_p.empty:
                                    _base = kojin_narabi_title.strip() if kojin_narabi_title.strip() else f"{kojin_narabi_ranges_text.strip()}の優秀台"
                                    _rng_stat  = _stat_from_diff(_rng_diff)
                                    _rng_img   = _build_machine_img(_rng_p, _base, _rng_stat)
                                    _rng_out   = os.path.join(output_dir, f"{_make_safe_fn(_base)}.jpg")
                                    _save_jpeg(_rng_img, _rng_out)
                                    result["files"].append(_rng_out)
                                    _log(f"  ✅ 台番範囲(優秀台・ピンクバーあり)「{_base}」({len(_rng_p)}台)")
                                else:
                                    _log(f"  台番範囲(優秀台): 台番 {sorted(_rng_bans)} に台なし")
                        except Exception:
                            _log(f"  ❌ 台番範囲優秀台エラー: {traceback.format_exc()}")

                    # 台番範囲 優秀台（ピンクバーなし）
                    if kojin_narabi2_ranges_text.strip():
                        try:
                            _rng2_list = parse_ranges(kojin_narabi2_ranges_text.strip())
                            _rng2_bans = ranges_to_bans(_rng2_list)
                            if _rng2_bans:
                                _rng2_df   = df_k[df_k["台番"].apply(lambda b: int(b) in _rng2_bans)].copy()
                                _rng2_diff = diff_k.loc[_rng2_df.index]
                                _rng2_p    = _rng2_df.copy().reset_index(drop=True)
                                if not _rng2_p.empty:
                                    _base2 = kojin_narabi2_title.strip() if kojin_narabi2_title.strip() else f"{kojin_narabi2_ranges_text.strip()}の優秀台"
                                    _rng2_img = _build_machine_img(_rng2_p, _base2, None)
                                    _rng2_out = os.path.join(output_dir, f"{_make_safe_fn(_base2)}.jpg")
                                    _save_jpeg(_rng2_img, _rng2_out)
                                    result["files"].append(_rng2_out)
                                    _log(f"  ✅ 台番範囲(優秀台・ピンクバーなし)「{_base2}」({len(_rng2_p)}台)")
                                else:
                                    _log(f"  台番範囲(優秀台・ピンクバーなし): 台番 {sorted(_rng2_bans)} に台なし")
                        except Exception:
                            _log(f"  ❌ 台番範囲優秀台(ピンクバーなし)エラー: {traceback.format_exc()}")

                    # その他の優秀台ピックアップ（全店舗）
                    if sonota_extra_text.strip():
                        try:
                            _se_bans = set(expand_machine_numbers(sonota_extra_text))
                            if _se_bans:
                                _se_df = df_k[df_k["台番"].apply(lambda b: int(b) in _se_bans)].copy().reset_index(drop=True)
                                if not _se_df.empty:
                                    _se_title = sonota_extra_title.strip() or "その他の優秀台ピックアップ"
                                    _se_img = _build_machine_img(_se_df, _se_title, None)
                                    _se_out = os.path.join(output_dir, f"{_make_safe_fn(_se_title)}.jpg")
                                    _save_jpeg(_se_img, _se_out)
                                    result["files"].append(_se_out)
                                    _log(f"  ✅ その他の優秀台ピックアップ「{_se_title}」({len(_se_df)}台)")
                                else:
                                    _log(f"  その他の優秀台ピックアップ: 台番 {sorted(_se_bans)} に台なし")
                        except Exception:
                            _log(f"  ❌ その他の優秀台ピックアップエラー: {traceback.format_exc()}")

            # ── バラエティ画像生成（秋葉原スランプ付きのみ）──────────────────────
            if with_slump and store == "秋葉原" and variety_enabled and variety_ranges_text.strip() and result["ok"]:
                df_v = result.get("df"); diff_v = result.get("diff_raw")
                if df_v is not None and diff_v is not None:
                    try:
                        _var_bans_ex = ranges_to_bans(parse_ranges(variety_ranges_text.strip()))
                        _var_df_ex = df_v[df_v["台番"].apply(lambda b: int(b) in _var_bans_ex)].copy()
                        if not _var_df_ex.empty:
                            _var_dr_ex = diff_v.loc[_var_df_ex.index]
                            if variety_mode == "プラス台":
                                _vm_ex = _var_dr_ex > 0
                                _var_df_ex = _var_df_ex[_vm_ex.values].copy(); _var_dr_ex = _var_dr_ex[_vm_ex]
                                _var_title_ex = "バラエティのプラス台"
                            elif variety_mode == "+1,000枚以上の優秀台":
                                _vm_ex = _var_dr_ex >= 1000
                                _var_df_ex = _var_df_ex[_vm_ex.values].copy(); _var_dr_ex = _var_dr_ex[_vm_ex]
                                _var_title_ex = "バラエティの優秀台"
                            else:
                                _var_title_ex = "バラエティ"
                            if not _var_df_ex.empty:
                                _vs_ex = _var_df_ex["台番"].argsort().values
                                _var_df_ex = _var_df_ex.iloc[_vs_ex].reset_index(drop=True)
                                _var_dr_ex = _var_dr_ex.iloc[_vs_ex].reset_index(drop=True)
                                _var_stat_ex = {"total_diff": int(_var_dr_ex.sum()), "avg_diff": int(round(_var_dr_ex.mean())), "win_count": int((_var_dr_ex > 0).sum()), "total_count": len(_var_df_ex)} if variety_mode == "全台" else None
                                _var_img_ex = _build_machine_img(_var_df_ex, _var_title_ex, _var_stat_ex)
                                _var_out_ex = os.path.join(output_dir, f"{_make_safe_fn(_var_title_ex)}.jpg")
                                _save_jpeg(_var_img_ex, _var_out_ex)
                                result["files"].append(_var_out_ex)
                                _log(f"  ✅ バラエティ「{_var_title_ex}」({len(_var_df_ex)}台)")
                    except Exception:
                        _log(f"  ❌ バラエティ画像エラー: {traceback.format_exc()}")

            # ── 全台系→高配分→並び→ジャグラー優秀台→その他 の順にリネーム ──
            if result["ok"]:
                def _hr_avg_k(x):
                    if "all_avg_diff" in x:
                        return x["all_avg_diff"]
                    return int(round(sum(x["diffs"]) / len(x["diffs"]))) if x.get("diffs") else 0
                _order: list[str] = []
                _df_res = result.get("df")
                _dr_res = result.get("diff_raw")
                # ① 全台系（avg_diff 降順・個別全台系を含む）
                _zen_ord: list[tuple[int, str]] = []
                for _oi in result.get("zen_dai_list", []):
                    _zen_ord.append((_oi.get("all_avg_diff", 0), f"{_make_safe_fn(_oi['name'])}.jpg"))
                if kojin_enabled and _df_res is not None and _dr_res is not None:
                    for _km in kojin_zentai_machines:
                        _km = _km.strip()
                        if not _km:
                            continue
                        _kgi = _df_res[_df_res["機種名"] == _km].index
                        if len(_kgi) == 0:
                            continue
                        _kavg = int(round(_dr_res.loc[_kgi].mean()))
                        _zen_ord.append((_kavg, f"{_make_safe_fn(_km)}.jpg"))
                for _, _fn in sorted(_zen_ord, key=lambda x: x[0], reverse=True):
                    if _fn not in _order:
                        _order.append(_fn)
                # ② 高配分（avg_diff 降順・個別優秀台を含む）
                _kyushu_set_ord = {m.strip() for m in kojin_yushu_machines if m.strip()} if kojin_enabled else set()
                for _oi in sorted(result.get("high_ratio_list", []), key=_hr_avg_k, reverse=True):
                    _nm = _oi["name"]
                    _fn = f"{_make_safe_fn(_nm)}（優秀台）.jpg" if _nm in _kyushu_set_ord else f"{_make_safe_fn(_nm)}_高配分.jpg"
                    if _fn not in _order:
                        _order.append(_fn)
                # ③ 並び画像（プレビュー順優先 → プレビューにないものはmoved_narabiで補完）
                # narabiスクリプトのmake_safeはASCIIコロン→全角に変換するため、_pnameも同様に変換して実ファイル名に揃える
                _nb_map_for_order = st.session_state.get(_aprev_narabi_key, {})
                if _aprev_imgs and _nb_map_for_order:
                    for _ci, (_pname, _) in enumerate(_aprev_imgs):
                        if _pname in _nb_map_for_order and st.session_state.get(f"auto_prev_ck_{store}_{_ci}", True):
                            _pname_safe = _pname.replace(":", "：")
                            if _pname_safe not in _order:
                                _order.append(_pname_safe)
                # プレビューにないファイル（プレビュー後追加・ファイル名不一致）もmoved_narabiで補完
                for _nbn in _moved_narabi:
                    if _nbn not in _order:
                        _order.append(_nbn)
                # ④ 個別並び（output_dir 直下のみ）
                if kojin_enabled:
                    for _nt, _nr in [(kojin_narabi_title, kojin_narabi_ranges_text),
                                     (kojin_narabi2_title, kojin_narabi2_ranges_text)]:
                        if _nr.strip():
                            _b = _nt.strip() or f"{_nr.strip()}の優秀台"
                            _fn = f"{_make_safe_fn(_b)}.jpg"
                            if _fn not in _order:
                                _order.append(_fn)
                # ④ 末尾・ジャグラー末尾画像
                for _sfn in _sue_saved_fns:
                    if _sfn not in _order:
                        _order.append(_sfn)
                # ⑤ バラエティ画像
                if with_slump and store == "秋葉原" and variety_enabled and variety_ranges_text.strip():
                    _var_title_ord = "バラエティのプラス台" if variety_mode == "プラス台" else ("バラエティの優秀台" if variety_mode == "+1,000枚以上の優秀台" else "バラエティ")
                    _vfn_ord = f"{_make_safe_fn(_var_title_ord)}.jpg"
                    if _vfn_ord not in _order:
                        _order.append(_vfn_ord)
                # ⑤ ジャグラーシリーズ優秀台
                _order.append("ジャグラーシリーズ優秀台.jpg")
                # ⑤ その他の優秀台ピックアップ
                if _sonota_split:
                    _order.append("その他の優秀台+1,000枚以上.jpg")
                    for _, _fn_ord in _sonota_extra_thrs:
                        _order.append(_fn_ord)
                else:
                    _order.append("その他の優秀台ピックアップ.jpg")
                # オススメ
                _sfx_r = {1: "プラス台", 1000: "1000枚以上", 2000: "2000枚以上"}
                for _blk in recommended_blocks:
                    _bt2 = _blk["title"].strip() or "オススメ機種"
                    for _thr in _blk.get("thresholds", [1]):
                        _order.append(f"オススメ_{_make_safe_fn(_bt2)}_{_sfx_r.get(_thr, str(_thr))}.jpg")
                # 再実行時に前回の連番プレフィックスが残っていると衝突するため、先に剥がしておく
                for _ef in os.listdir(output_dir):
                    _pm = re.match(r"^\d{2}_(.+)$", _ef)
                    if _pm and os.path.isfile(os.path.join(output_dir, _ef)):
                        os.replace(os.path.join(output_dir, _ef),
                                   os.path.join(output_dir, _pm.group(1)))
                # 実在するファイルに 01_ 02_ … プレフィックスを付与
                _seq = 1
                for _bn in _order:
                    _src = os.path.join(output_dir, _bn)
                    if os.path.exists(_src):
                        os.replace(_src, os.path.join(output_dir, f"{_seq:02d}_{_bn}"))
                        _seq += 1
                # スランプ付き：narabi含む全JPGを最終収集してセッション更新
                if with_slump and os.path.isdir(output_dir):
                    _ig_late = []
                    for _lf in sorted(os.listdir(output_dir)):
                        _lfp = os.path.join(output_dir, _lf)
                        if os.path.isfile(_lfp) and _lf.lower().endswith((".jpg", ".jpeg")):
                            with open(_lfp, "rb") as _lfh:
                                _ig_late.append((_lf, _lfh.read()))
                    if _ig_late:
                        st.session_state[f"_inagawa_jpgs_{store}"] = _ig_late
                    # ban_map を final result（kojin追加後）ベースで完全再構築
                    _ig_bm_u: dict[str, list[int]] = {}
                    for _zd2u in result.get("zen_dai_list", []):
                        _ig_bm_u[f"{_make_safe_fn(_zd2u['name'])}.jpg"] = _zd2u.get("bans", [])
                    for _hr2u in result.get("high_ratio_list", []):
                        if not _hr2u.get("has_image", False):
                            continue
                        if _hr2u.get("is_yushu", False):
                            _fn2u = f"{_make_safe_fn(_hr2u['name'])}（優秀台）.jpg"
                        else:
                            _fn2u = f"{_make_safe_fn(_hr2u['name'])}_高配分.jpg"
                        _ig_bm_u[_fn2u] = _hr2u.get("bans", [])
                    if _jug_pool is not None and not _jug_pool.empty:
                        if with_slump and store == "秋葉原":
                            pass  # ジャグラー統合画像なし → その他bansへは _sonota_bans_ig2 側で追加
                        else:
                            _ig_bm_u["ジャグラーシリーズ優秀台.jpg"] = [
                                int(str(_b2).split(".")[0]) for _b2 in _jug_pool["台番"].dropna()
                                if str(_b2).split(".")[0].lstrip("-").isdigit()
                            ]
                    _sonota_bans_ig2 = sorted({int(_e2["ban"]) for _e2 in result.get("sonota_excellent_list", []) if "ban" in _e2})
                    if _sonota_extra_bans:
                        _sonota_bans_ig2 = sorted(set(_sonota_bans_ig2) | set(_sonota_extra_bans))
                    # 秋葉原スランプ付き: jug_pool +1000枚台をsonota bansに追加
                    if with_slump and store == "秋葉原" and _jug_pool is not None and not _jug_pool.empty:
                        _jp_dr_bm = result.get("diff_raw")
                        _jp_df_bm = result.get("df")
                        if _jp_dr_bm is not None and _jp_df_bm is not None:
                            _jp_bans_bm = {int(str(b).split(".")[0]) for b in _jug_pool["台番"].dropna()
                                           if str(b).split(".")[0].lstrip("-").isdigit()}
                            _jp_rows_bm = _jp_df_bm[_jp_df_bm["台番"].apply(lambda b: int(b) in _jp_bans_bm)]
                            if not _jp_rows_bm.empty:
                                _jp_dr2 = _jp_dr_bm.loc[_jp_rows_bm.index]
                                _jp_1k_bm = [int(b) for b in _jp_rows_bm[_jp_dr2.values >= 1000]["台番"].dropna()]
                                if _jp_1k_bm:
                                    _sonota_bans_ig2 = sorted(set(_sonota_bans_ig2) | set(_jp_1k_bm))
                    if _sonota_bans_ig2:
                        if _sonota_split:
                            _ig_bm_u["その他の優秀台+1,000枚以上.jpg"] = _sonota_bans_ig2
                            _ig_bm_u2_df = result.get("df")
                            _ig_bm_u2_dr = result.get("diff_raw")
                            if _ig_bm_u2_df is not None and _ig_bm_u2_dr is not None:
                                for _thr_bmu, _fn_bmu in _sonota_extra_thrs:
                                    _sk_bm = [
                                        _b for _b in _sonota_bans_ig2
                                        if not (_ig_bm_u2_df[_ig_bm_u2_df["台番"] == _b]).empty
                                        and int(_ig_bm_u2_dr.loc[_ig_bm_u2_df[_ig_bm_u2_df["台番"] == _b].index[0]]) >= _thr_bmu
                                    ]
                                    if _sk_bm:
                                        _ig_bm_u[_fn_bmu] = _sk_bm
                        else:
                            _ig_bm_u["その他の優秀台ピックアップ.jpg"] = _sonota_bans_ig2
                    for _nami2u in result.get("nami_list", []):
                        _nt2u = _nami2u.get("title", "")
                        if _nt2u and _nami2u.get("bans"):
                            _ig_bm_u[f"{_make_safe_fn(_nt2u)}.jpg"] = [int(_b3) for _b3 in _nami2u["bans"]]
                    for _nfn_u, _nbns_u in st.session_state.get(_aprev_narabi_key, {}).items():
                        _bare_u = re.sub(r"^\d{2}_", "", _nfn_u)
                        if _bare_u not in _ig_bm_u:
                            _ig_bm_u[_bare_u] = [int(_b4) for _b4 in _nbns_u]
                    # 末尾画像のbansを追加
                    # 末尾桁+現在モードで直接フィルターし、保存済みファイル名・モード由来ファイル名の両方にセット
                    if st.session_state.get("suebangai_enabled", False) and _sue_tails_r and result.get("df") is not None:
                        _run_df_ig = result.get("df")
                        _run_dr_ig = result.get("diff_raw")
                        _sue_cir_ig = {"0":"⓪","1":"①","2":"②","3":"③","4":"④","5":"⑤","6":"⑥","7":"⑦","8":"⑧","9":"⑨"}
                        _is_plus_ig  = _sue_mode_r in ("プラス台（ピンクバー付き）", "プラス台（ピンクバーなし）", "プラス台")
                        _is_1k_ig    = _sue_mode_r == "+1,000枚以上の優秀台"
                        _is_yushu_ig = _sue_mode_r in ("優秀台（ピンクバー付き）", "優秀台（ピンクバーなし）")
                        for _tail_ig in _sue_tails_r:
                            try:
                                if _tail_ig == "ゾロ目":
                                    _filt_ig = _run_df_ig[_run_df_ig["台番"].apply(
                                        lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1])].copy()
                                    _base_ig = "末尾ゾロ目の台"
                                elif _tail_ig.isdigit() and len(_tail_ig) in (1, 2):
                                    _filt_ig = _run_df_ig[_run_df_ig["台番"].astype(str).str[-len(_tail_ig):] == _tail_ig].copy()
                                    _base_ig = f"末尾{_sue_cir_ig.get(_tail_ig, _tail_ig)}番台"
                                else:
                                    continue
                                if _filt_ig.empty:
                                    continue
                                if _is_plus_ig and _run_dr_ig is not None:
                                    _filt_ig = _filt_ig[_run_dr_ig.loc[_filt_ig.index] > 0]
                                    _title_ig = f"{_base_ig}のプラス台"
                                elif _is_1k_ig and _run_dr_ig is not None:
                                    _filt_ig = _filt_ig[_run_dr_ig.loc[_filt_ig.index] >= 1000]
                                    _title_ig = f"{_base_ig}の優秀台"
                                elif _is_yushu_ig and _run_dr_ig is not None:
                                    _sg_ig = next((c for c in ["ゲーム数_rounded", "ゲーム数"] if c in _filt_ig.columns), None)
                                    if _sg_ig:
                                        _sm_ig = (_run_dr_ig.loc[_filt_ig.index] >= 1000) | ((_filt_ig[_sg_ig] >= 1800) & (_run_dr_ig.loc[_filt_ig.index] > 0))
                                        _filt_ig = _filt_ig[_sm_ig.values]
                                    else:
                                        _filt_ig = _filt_ig[_run_dr_ig.loc[_filt_ig.index] >= 1000]
                                    _title_ig = f"{_base_ig}の優秀台"
                                else:
                                    _title_ig = _base_ig
                                if not _filt_ig.empty:
                                    _bans_ig = [int(b) for b in _filt_ig["台番"].dropna()]
                                    # モードから導出されるファイル名にセット
                                    _fn_from_mode = f"{_make_safe_fn(_title_ig)}.jpg"
                                    if _fn_from_mode not in _ig_bm_u:
                                        _ig_bm_u[_fn_from_mode] = _bans_ig
                                        _log(f"  📌 末尾banmap追加: {_fn_from_mode} ({len(_filt_ig)}台)")
                                    # 実際に保存されたファイル名にも同じbansをセット（モード不一致でも正しく合成）
                                    for _sfn_saved in _sue_saved_fns:
                                        _m_saved = re.match(r"末尾(\d+)番台", _sfn_saved)
                                        if _m_saved and _m_saved.group(1) == _tail_ig and _sfn_saved not in _ig_bm_u:
                                            _ig_bm_u[_sfn_saved] = _bans_ig
                                            _log(f"  📌 末尾banmap追加(saved): {_sfn_saved} → bans={len(_bans_ig)}台")
                            except Exception as _eig:
                                _log(f"  ⚠️ 末尾banmap追加エラー: {_eig}")
                    # バラエティ画像のbansを追加
                    if with_slump and store == "秋葉原" and variety_enabled and variety_ranges_text.strip():
                        _var_title_bm = "バラエティのプラス台" if variety_mode == "プラス台" else ("バラエティの優秀台" if variety_mode == "+1,000枚以上の優秀台" else "バラエティ")
                        _vfn_bm = f"{_make_safe_fn(_var_title_bm)}.jpg"
                        if _vfn_bm not in _ig_bm_u:
                            try:
                                _vbns_bm = ranges_to_bans(parse_ranges(variety_ranges_text.strip()))
                                _vdf_bm = result.get("df")
                                if _vdf_bm is not None:
                                    _vdr_bm = result.get("diff_raw")
                                    _vfilt = _vdf_bm[_vdf_bm["台番"].apply(lambda b: int(b) in _vbns_bm)]
                                    if not _vfilt.empty and _vdr_bm is not None:
                                        if variety_mode == "プラス台":
                                            _vfilt = _vfilt[_vdr_bm.loc[_vfilt.index] > 0]
                                        elif variety_mode == "+1,000枚以上の優秀台":
                                            _vfilt = _vfilt[_vdr_bm.loc[_vfilt.index] >= 1000]
                                        _ig_bm_u[_vfn_bm] = [int(b) for b in _vfilt["台番"].dropna()]
                            except Exception:
                                pass
                    # オススメ機種の台番を ban_map に追加
                    _ig_bm_u.update(_exec_rec_ban_map)
                    st.session_state[f"_inagawa_ban_map_{store}"] = _ig_bm_u

                    # ── pisionデータ取得・スランプグラフ合成・output_dir上書き ──
                    _ig_api_key_exec = _get_pision_api_key()
                    if not _ig_api_key_exec:
                        _log("⚠️ スランプ: PISION_API_KEY が未設定のためスランプグラフをスキップ")
                    else:
                        _ig_date_exec = st.session_state.get(f"_inagawa_date_{store}", "")
                        _log(f"📡 スランプ: pisionデータ取得中（日付={_ig_date_exec}）")
                        try:
                            _ig_halls_exec = fetch_pision_halls(_ig_api_key_exec)
                            _ig_hall_id_exec = None
                            _ig_all_hall_names: list[str] = []
                            for _igh_exec in _ig_halls_exec:
                                _ighn_exec = _igh_exec.get("name") or _igh_exec.get("displayName") or ""
                                _ig_all_hall_names.append(_ighn_exec)
                                if store in _ighn_exec and "エスパス" in _ighn_exec:
                                    _ig_hall_id_exec = str(_igh_exec.get("id") or _igh_exec.get("hallId") or "")
                                    break
                            if not _ig_hall_id_exec:
                                _log(f"⚠️ スランプ: '{store}' に対応するホールが見つかりません。pisionホール一覧: {_ig_all_hall_names}")
                            else:
                                _log(f"✅ スランプ: ホール発見（hall_id={_ig_hall_id_exec}）")
                                # 速報キャッシュが同日付で存在する場合は優先使用（確定APIより新鮮なデータ）
                                _rt_cached = st.session_state.get(f"_auto_tb_rt_items_{store}")
                                _rt_cached_date = st.session_state.get(f"_auto_tb_rt_items_date_{store}", "")
                                if _rt_cached and _rt_cached_date == _ig_date_exec:
                                    _ig_pision_exec = _rt_cached
                                    _log(f"✅ スランプ: 速報キャッシュ使用（{len(_ig_pision_exec)}台）")
                                else:
                                    _ig_pision_exec = fetch_pision_results(_ig_api_key_exec, _ig_hall_id_exec, _ig_date_exec)
                                    if not _ig_pision_exec:
                                        _log(f"⚠️ スランプ: {_ig_date_exec} の確定データなし（404/未公開）")
                                    else:
                                        _log(f"✅ スランプ: 確定データ使用")
                                if _ig_pision_exec:
                                    _log(f"✅ スランプ: {len(_ig_pision_exec)}台分のデータ取得")
                                    _slump_apply_names(_ig_pision_exec)
                                    _ig_by_uid_exec = {str(_it.get("unitId", "")): _it for _it in _ig_pision_exec}
                                    _ig_tmpl_exec = find_slump_template()
                                    _ig_bbb_exec  = _find_slump_bg()
                                    if _ig_tmpl_exec is None:
                                        _log("⚠️ スランプ: テンプレート画像(base_3000_bk.png)が見つかりません")
                                    _ig_ban2mac_exec: dict[str, str] = {}
                                    _ban2diff_exec:  dict[str, int]  = {}
                                    _df_exec = result.get("df")
                                    _dr_exec = result.get("diff_raw")
                                    if _df_exec is not None:
                                        for _idx_e, _igr_exec in _df_exec.iterrows():
                                            _bs0_exec = str(_igr_exec.get("台番", "")).split(".")[0]
                                            if _bs0_exec.lstrip("-").isdigit():
                                                _ig_ban2mac_exec[_bs0_exec] = str(_igr_exec.get("機種名", ""))
                                                if _dr_exec is not None:
                                                    try:
                                                        _ban2diff_exec[_bs0_exec] = int(_dr_exec.loc[_idx_e])
                                                    except Exception:
                                                        pass
                                    _ig_composite: list[tuple[str, bytes]] = []
                                    _ig_slump_cnt = 0
                                    for _lfn_exec, _lfb_exec in st.session_state.get(f"_inagawa_jpgs_{store}", []):
                                        _bare_exec = re.sub(r"^\d{2}_", "", _lfn_exec)
                                        _bans_exec = _ig_bm_u.get(_bare_exec, [])
                                        try:
                                            _t_img_exec = Image.open(io.BytesIO(_lfb_exec)).convert("RGB")
                                        except Exception:
                                            if store != "秋葉原":
                                                _ig_composite.append((_lfn_exec, _lfb_exec))
                                            continue
                                        if not _bans_exec or _ig_tmpl_exec is None:
                                            if store != "秋葉原":
                                                _ig_composite.append((_lfn_exec, _lfb_exec))
                                            continue
                                        _g_imgs_exec: list["Image.Image"] = []
                                        _show_mn_exec = (_bare_exec in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg",
                                                                        "その他の優秀台+1,000枚以上.jpg", "その他の優秀台+2,000枚以上.jpg", "その他の優秀台+3,000枚以上.jpg")
                                                        or _bare_exec.startswith("末尾") or _bare_exec.startswith("バラエティ"))
                                        _is_zentai_exec = (not _bare_exec.endswith("_高配分.jpg") and
                                                           _bare_exec not in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg",
                                                                              "その他の優秀台+1,000枚以上.jpg", "その他の優秀台+2,000枚以上.jpg", "その他の優秀台+3,000枚以上.jpg"))
                                        for _b_exec in _bans_exec:
                                            _it_exec = _ig_by_uid_exec.get(str(_b_exec))
                                            if _it_exec is None or not _it_exec.get("points"):
                                                continue
                                            _dn_exec = (_it_exec.get("_convertedName")
                                                        or _it_exec.get("displayName")
                                                        or _ig_ban2mac_exec.get(str(_b_exec), str(_b_exec)))
                                            _sd_exec = not (_is_zentai_exec and _ban2diff_exec.get(str(_b_exec), 0) < 0)
                                            try:
                                                _g_imgs_exec.append(draw_slump_graph(
                                                    _ig_tmpl_exec, str(_b_exec), _dn_exec,
                                                    _it_exec["points"], diff=_it_exec.get("diff"),
                                                    machine_name=_dn_exec if _show_mn_exec else None,
                                                    show_diff=_sd_exec,
                                                ))
                                            except Exception:
                                                pass
                                        if store == "秋葉原":
                                            if _g_imgs_exec:
                                                _ex_title = st.session_state.get(f"_inagawa_title_map_{store}", {}).get(_bare_exec, os.path.splitext(_bare_exec)[0])
                                                _ex_slump = _build_slump_title_img(_ex_title, _g_imgs_exec, _ig_bbb_exec)
                                                if _ex_slump is not None:
                                                    _ex_buf = io.BytesIO()
                                                    _ex_slump.save(_ex_buf, format="JPEG", quality=92)
                                                    _ig_composite.append((_lfn_exec, _ex_buf.getvalue()))
                                                    _ig_slump_cnt += 1
                                        else:
                                            _combined_exec = _attach_slump_to_table(_t_img_exec, _g_imgs_exec, _ig_bbb_exec)
                                            _cbuf_exec = io.BytesIO()
                                            _combined_exec.save(_cbuf_exec, format="JPEG", quality=92)
                                            _ig_composite.append((_lfn_exec, _cbuf_exec.getvalue()))
                                            if _g_imgs_exec:
                                                _ig_slump_cnt += 1
                                        # 横レイアウト（16台以上・秋葉原除く）
                                        if len(_g_imgs_exec) >= 16 and store != "秋葉原":
                                            try:
                                                _side_img_exec = _attach_slump_to_table_side(_t_img_exec, _g_imgs_exec, _ig_bbb_exec)
                                                _side_buf_exec = io.BytesIO()
                                                _side_img_exec.save(_side_buf_exec, format="JPEG", quality=92)
                                                _side_fn_exec = os.path.splitext(_lfn_exec)[0] + "_side.jpg"
                                                _ig_composite.append((_side_fn_exec, _side_buf_exec.getvalue()))
                                            except Exception:
                                                pass
                                    _log(f"✅ スランプ: {_ig_slump_cnt}枚にスランプグラフを合成")
                                    # output_dirに上書き保存 & session_stateを合成済み画像で更新
                                    for _cfn_exec, _cb_exec in _ig_composite:
                                        _cfp_exec = os.path.join(output_dir, _cfn_exec)
                                        if os.path.exists(_cfp_exec) or _cfn_exec.endswith("_side.jpg"):
                                            with open(_cfp_exec, "wb") as _cfh_exec:
                                                _cfh_exec.write(_cb_exec)
                                    st.session_state[f"_inagawa_jpgs_{store}"] = _ig_composite
                        except Exception as _ig_exc_exec:
                            _log(f"❌ スランプグラフ合成エラー: {_ig_exc_exec}")

            all_ok = result["ok"] and (narabi_result is None or narabi_result["ok"])
            if all_ok:
                status_widget.update(label="✅ 全処理完了！", state="complete", expanded=False)
            else:
                status_widget.update(label="⚠️ エラーあり", state="error", expanded=True)

        # ── エラー詳細 ──────────────────────────────────────────────
        if not result["ok"]:
            st.markdown("### エラー詳細")
            st.error(result["error"])
        if narabi_result and not narabi_result["ok"]:
            st.markdown("### 並び画像エラー")
            if narabi_result["stderr"]:
                st.error(narabi_result["stderr"])

        # ── オススメ機種除外ログ ─────────────────────────────────────
        if recommended_exclusion_logs:
            st.markdown("### オススメ機種 除外ログ")
            for _msg in recommended_exclusion_logs:
                st.info(f"ℹ️ {_msg}")

        # ── 生成ファイル一覧 ──────────────────────────────────────────
        st.markdown("### 生成されたファイル")
        if not _IS_CLOUD:
            st.info(f"📁 `{output_dir}`")
        if os.path.isdir(output_dir):
            imgs = sorted(
                f for f in os.listdir(output_dir)
                if f.lower().endswith((".png", ".jpg", ".jpeg"))
            )
            if imgs:
                for fname in imgs:
                    with st.expander(fname, expanded=False):
                        st.image(os.path.join(output_dir, fname), use_container_width=True)
            else:
                st.warning("画像ファイルが見つかりませんでした。")

        # ── 結果報告文章 ────────────────────────────────────────────
        if result["ok"]:
            st.markdown("---")
            nami_for_report = result.get("nami_list", [])
            # ⑦プレビューでチェックを外した並び画像に対応する nami_list エントリを除外
            if _aprev_imgs:
                _nb_map_r = st.session_state.get(_aprev_narabi_key, {})
                _unchecked_ban_sets: list[frozenset] = [
                    frozenset(st.session_state[_aprev_narabi_key][_pname])
                    for _ci, (_pname, _) in enumerate(_aprev_imgs)
                    if not st.session_state.get(f"auto_prev_ck_{store}_{_ci}", True)
                    and _pname in _nb_map_r
                ]
                if _unchecked_ban_sets:
                    nami_for_report = [
                        item for item in nami_for_report
                        if frozenset(item.get("bans", [])) not in _unchecked_ban_sets
                    ]
            _excellent_all = result.get("excellent_list", [])
            _memo_machines: set[str] = set()   # 素材メモの機種（台番ベースで絞る）
            _memo_bans:     set[int] = set()   # 素材メモに書かれた台番
            if store in EXTENDED_FEATURE_STORES:
                # ── 素材メモをパースして機種名・台番を収集 ──────────────
                _memo_note = st.session_state.get(f"result_extra_note_{store}", "").strip() \
                    if st.session_state.get(f"memo_enabled_{store}", False) else ""
                _memo_df_r = result.get("df")
                if _memo_note and _memo_df_r is not None:
                    _memo_nm, _ = load_name_map()
                    for _blk in parse_result_memo(_memo_note):
                        if _blk["machine"]:
                            _memo_machines.add(_memo_nm.get(_blk["machine"], _blk["machine"]))
                        else:
                            _first = [n for s in _blk["sections"] for n in s["numbers"]]
                            if _first:
                                _r = _memo_df_r[_memo_df_r["台番"].apply(
                                    lambda x: int(x) if pd.notna(x) else -1) == _first[0]]
                                if not _r.empty:
                                    _memo_machines.add(str(_r.iloc[0]["機種名"]))
                        for _s in _blk["sections"]:
                            _memo_bans.update(_s["numbers"])

                # ── ブロック機種（メモ以外）は全除外、メモ機種はメモ台番のみ除外 ──
                _block_machines = {
                    m.strip()
                    for block in recommended_blocks
                    for m in block.get("machines", [])
                    if m.strip()
                }
                _block_only = _block_machines - _memo_machines
                _excellent_all = [
                    x for x in _excellent_all
                    if x["name"] not in _block_only
                    and not (x["name"] in _memo_machines and x.get("ban") in _memo_bans)
                ]
            report_text = generate_report_text(
                store_name=store,
                date=result.get("date"),
                zen_dai_list=result.get("zen_dai_list", []),
                high_ratio_list=result.get("high_ratio_list", []),
                nami_list=nami_for_report,
                excellent_list=_excellent_all,
                diff_raw=result.get("diff_raw"),
                df=result.get("df"),
                suebangai_data=_sue_stats_data or None,
                jug_sue_data=_jug_sue_stats_data or None,
            )
            # オススメ機種ブロックの優秀台（+1000枚以上）を挿入（拡張機能店舗）
            if store in EXTENDED_FEATURE_STORES and recommended_blocks:
                _rec_df       = result.get("df")
                _rec_diff_raw = result.get("diff_raw")
                if _rec_df is not None and _rec_diff_raw is not None:
                    _rec_text = generate_recommended_result_text(
                        recommended_blocks, _rec_df, _rec_diff_raw,
                        exclude_machines=_memo_machines,
                        store_name=store,
                    )
                    if _rec_text:
                        report_text = insert_formatted_result_before_other_picks(report_text, _rec_text, store)

            # ④ 素材メモを差枚付きテキストに変換して「🎁その他の優秀台」直前に差し込む
            _extra_note = st.session_state.get(f"result_extra_note_{store}", "").strip() \
                if st.session_state.get(f"memo_enabled_{store}", False) else ""
            if _extra_note and store in EXTENDED_FEATURE_STORES:
                _memo_df       = result.get("df")
                _memo_diff_raw = result.get("diff_raw")
                if _memo_df is not None and _memo_diff_raw is not None:
                    _parsed    = parse_result_memo(_extra_note)
                    _formatted, _missing = format_result_memo_sections(_parsed, _memo_df, _memo_diff_raw)
                    if not _parsed:
                        st.warning("⚠️ 素材メモを解析できませんでした。機種名行は「・」で始めてください（例：・ネオアイムジャグラー）")
                    elif not _formatted:
                        st.warning("⚠️ 素材メモのフォーマット後テキストが空でした。")
                    if _missing:
                        st.warning(f"台番がExcelに見つかりません: {', '.join(str(n) for n in sorted(set(_missing)))}")
                    if _formatted:
                        report_text = insert_formatted_result_before_other_picks(report_text, _formatted, store)
                        st.caption(f"✅ 素材メモを挿入しました（{len(_parsed)}機種）")
                else:
                    st.warning("⚠️ df/diff_raw が result に含まれていません")
            # 結果.txt をフォルダに保存
            for _old, _new in STORE_RESULT_TRANSFORMS.get(store, []):
                report_text = report_text.replace(_old, _new)
            _date = result.get("date")
            if _date:
                _txt_name = f"{_date.month:02d}{_date.day:02d}_結果.txt"
            else:
                _txt_name = "結果.txt"
            _txt_path = os.path.join(output_dir, _txt_name)
            try:
                with open(_txt_path, "w", encoding="utf-8") as _f:
                    _f.write(report_text)
                if _IS_CLOUD:
                    st.caption(f"📄 {_txt_name} をZIPに含めます")
                else:
                    st.caption(f"📄 {_txt_name} を保存しました")
            except Exception as _e:
                st.warning(f"結果.txt の保存に失敗: {_e}")

            # ── ZIPデータをセッションに保存（ボタン直下スロットへ後で表示）──
            if os.path.isdir(output_dir):
                try:
                    _zip_data = _make_zip_bytes(output_dir)
                    if _IS_CLOUD:
                        st.session_state[f"_auto_zip_data_{store}"] = _zip_data
                        st.session_state[f"_auto_zip_stem_{store}"] = dir_stem
                    else:
                        st.download_button(
                            label="📥 画像・テキストをZIPでダウンロード",
                            data=_zip_data,
                            file_name=f"{dir_stem}.zip",
                            mime="application/zip",
                            key="auto_zip_dl",
                            type="secondary",
                        )
                except Exception as _ze:
                    st.warning(f"ZIP生成に失敗: {_ze}")

            import html as _html
            _safe = _html.escape(report_text)
            _lines = report_text.count("\n") + 1
            _h = min(600, max(200, _lines * 20 + 110))
            st.iframe(f"""
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:8px;">
              <span style="font-size:1.4rem;font-weight:700;">結果報告</span>
              <button id="cb" onclick="
                var t=document.getElementById('rt');
                t.select();t.setSelectionRange(0,99999);
                document.execCommand('copy');
                this.textContent='✅ コピー済み';this.style.background='#4CAF50';
                var b=this;setTimeout(function(){{b.textContent='📋 コピー';b.style.background='#2F559E';}},2000);
              " style="padding:5px 14px;background:#2F559E;color:#fff;
                       border:none;border-radius:4px;cursor:pointer;font-size:14px;">
                📋 コピー
              </button>
            </div>
            <textarea id="rt" readonly
              style="width:100%;height:{_h - 70}px;font-family:monospace;font-size:13px;
                     border:1px solid #ccc;padding:8px;box-sizing:border-box;resize:vertical;"
            >{_safe}</textarea>
            """, height=_h)

    # ── ボタン直下スロットにZIPダウンロードボタンを表示（Cloud のみ）──
    if _IS_CLOUD and _auto_zip_slot is not None and st.session_state.get(f"_auto_zip_data_{store}"):
        with _auto_zip_slot.container():
            st.success("✅ 処理が完了しました。ZIPをダウンロードしてください。")
            st.download_button(
                label="📥 画像・テキストをZIPでダウンロード",
                data=st.session_state[f"_auto_zip_data_{store}"],
                file_name=f"{st.session_state.get(f'_auto_zip_stem_{store}', 'output')}.zip",
                mime="application/zip",
                key="auto_zip_dl",
                type="primary",
            )

    # ── 戻るボタン（常に末尾に1つだけ描画）──────────────────────────
    st.markdown("---")
    if st.button("← 戻る", key="auto_back"):
        _navigate("image_type")


def show_auto_article_page() -> None:
    """記事用 自動処理ページ（高田馬場専用）: ①～④ + 実行ボタン + 結果"""
    store = st.session_state.selected_store
    st.markdown(f"## 【{store}】記事用")
    st.caption("全台系・ジャグラー優秀台・その他の優秀台を一括生成します（記事用）。")
    st.markdown("---")

    # ── ⓪ pision.io から日付データを自動取得（記事用）──────────────────
    _art_tb_uploaded = None
    if store in STORES:
        st.markdown(f"### 📈 日付からデータを自動取得（{store}）")
        api_key = _get_pision_api_key()
        if not api_key:
            st.caption("⚠️ PISION_API_KEY が未設定のため利用できません。①から手動でアップロードしてください。")
        else:
            _art_tb_mode = st.radio(
                "データ種別",
                ["確定データ", "速報データ（当日・営業中）"],
                horizontal=True,
                key=f"art_tb_mode_{store}",
                help="確定データ＝前日まで（X-Api-Key）。速報データ＝当日の営業中データ（realtimeログインが必要）。",
            )
            _art_tb_is_rt = _art_tb_mode.startswith("速報")
            _art_tb_rt_ok = True
            if _art_tb_is_rt:
                _rt_user, _rt_pass = _get_pision_rt_credentials()
                if not _rt_user or not _rt_pass:
                    st.error("❌ 速報データには realtime のログイン情報が必要です。"
                             ".env に PISION_RT_USER / PISION_RT_PASS を設定してください。")
                    _art_tb_rt_ok = False

            if _art_tb_is_rt:
                import datetime as _dt
                _jst = _dt.timezone(_dt.timedelta(hours=9))
                _now = _dt.datetime.now(_jst)
                _rt_default = _now.date() if _now.hour >= 7 else _now.date() - _dt.timedelta(days=1)
                _art_tb_prev_mode_key = f"_art_tb_prev_mode_{store}"
                _art_tb_cnt_key = f"_art_tb_cnt_{store}"
                if st.session_state.get(_art_tb_prev_mode_key) != "rt":
                    st.session_state[_art_tb_cnt_key] = st.session_state.get(_art_tb_cnt_key, 0) + 1
                st.session_state[_art_tb_prev_mode_key] = "rt"
                _art_tb_cnt = st.session_state.get(_art_tb_cnt_key, 0)
                _art_tb_date = st.date_input(
                    "日付を選択（速報は営業日に合わせて選択・日付変更では自動取得しません）",
                    value=_rt_default,
                    key=f"art_tb_date_rt_{store}_{_art_tb_cnt}",
                )
            else:
                st.session_state[f"_art_tb_prev_mode_{store}"] = "fix"
                _art_tb_date = st.date_input(
                    "日付を選択",
                    value=datetime.date.today() - datetime.timedelta(days=1),
                    key=f"art_tb_date_{store}",
                )
            _art_tb_date_str = _art_tb_date.strftime("%Y-%m-%d")

            _art_tb_mode_tag          = "rt" if _art_tb_is_rt else "fix"
            _art_tb_seen_key          = f"_art_tb_seen_{_art_tb_mode_tag}_{store}"
            _art_tb_bytes_key         = f"_art_tb_file_bytes_{_art_tb_mode_tag}_{store}"
            _art_tb_name_key          = f"_art_tb_file_name_{_art_tb_mode_tag}_{store}"
            _art_tb_count_key         = f"_art_tb_count_{_art_tb_mode_tag}_{store}"
            _art_tb_fetched_key       = f"_art_tb_fetched_{_art_tb_mode_tag}_{store}"
            _art_tb_collecting_key    = f"_art_tb_collecting_rt_{store}"
            _art_tb_rt_items_key      = f"_art_tb_rt_items_{store}"
            _art_tb_rt_items_date_key = f"_art_tb_rt_items_date_{store}"
            _art_tb_baseline_artid_key = f"_art_tb_baseline_artid_{store}"

            _art_is_collecting = _art_tb_is_rt and st.session_state.get(_art_tb_collecting_key) == _art_tb_date_str

            _art_do_rt_check    = False
            _art_do_rt_existing = False
            if _art_is_collecting:
                _art_btn_c1, _art_btn_c2 = st.columns(2)
                with _art_btn_c1:
                    _art_tb_refetch = st.button("⏳ 収集中...", key=f"art_tb_refetch_{store}",
                                                disabled=True, use_container_width=True)
                with _art_btn_c2:
                    _art_do_rt_check = st.button("🔍 今すぐ確認", key=f"art_tb_rt_check_{store}",
                                                 use_container_width=True,
                                                 help="30秒待たずに今すぐ収集完了を確認します。")
            elif _art_tb_is_rt and _art_tb_rt_ok:
                _art_btn_c1, _art_btn_c2 = st.columns(2)
                with _art_btn_c1:
                    _art_tb_refetch = st.button("⚡ 速報を取得", key=f"art_tb_refetch_{store}",
                                                use_container_width=True, type="primary")
                with _art_btn_c2:
                    _art_do_rt_existing = st.button("📂 既存のデータを取得", key=f"art_tb_rt_existing_{store}",
                                                    use_container_width=True,
                                                    help="新しい収集を開始せず、過去に取得済みの直近データを読み込みます。")
            else:
                _art_tb_refetch = st.button("🔄 取得", key=f"art_tb_refetch_{store}")

            _art_tb_seen = st.session_state.get(_art_tb_seen_key)
            _art_tb_date_changed = (not _art_tb_is_rt) and _art_tb_seen is not None and _art_tb_seen != _art_tb_date_str
            st.session_state[_art_tb_seen_key] = _art_tb_date_str

            def _art_save_rt_items_to_session(items: list) -> None:
                _slump_apply_names(items)
                _rows = [
                    {
                        "台番":     it.get("unitId"),
                        "機種名":   it.get("_machineName") or it.get("_convertedName") or it.get("displayName") or "",
                        "差枚":     it.get("diff", 0),
                        "ゲーム数": it.get("games", 0),
                        "BB":       it.get("bb", 0),
                        "RB":       it.get("rb", 0),
                        "AT":       it.get("art", 0),
                    }
                    for it in items
                ]
                _df  = pd.DataFrame(_rows)
                _buf = io.BytesIO()
                _df.to_excel(_buf, index=False)
                st.session_state[_art_tb_bytes_key]          = _buf.getvalue()
                st.session_state[_art_tb_name_key]           = f"{_art_tb_date.strftime('%Y%m%d')}_{store}_20S.xlsx"
                st.session_state[_art_tb_count_key]          = len(_df)
                st.session_state[_art_tb_fetched_key]        = _art_tb_date_str
                st.session_state[_art_tb_rt_items_key]       = items
                st.session_state[_art_tb_rt_items_date_key]  = _art_tb_date_str
                st.session_state.pop(_art_tb_collecting_key, None)

            def _art_is_new_artid(poll_result: dict) -> bool:
                _baseline = st.session_state.get(_art_tb_baseline_artid_key)
                _new_id   = poll_result.get("article_id")
                if _baseline is None:
                    return True
                try:
                    return int(_new_id) > int(_baseline)
                except (TypeError, ValueError):
                    return False

            if _art_do_rt_check:
                with st.spinner("収集状況を確認中..."):
                    _art_chk = fetch_pision_realtime(store, _art_tb_date_str, trigger=False)
                if _art_chk["ok"] and _art_is_new_artid(_art_chk):
                    _art_save_rt_items_to_session(_art_chk["items"])
                    st.rerun()
                elif _art_chk["ok"]:
                    st.info("⏳ 収集はまだ完了していません（前回と同じスナップショット）。自動確認に戻ります。")
                    st.rerun()
                else:
                    st.rerun()

            if _art_is_collecting and not _art_do_rt_check:
                import time as _time
                _art_ap_ph = st.empty()
                _art_ap_ph.info("⏳ 速報データを収集中... 30秒後に自動で確認します。")
                _time.sleep(30)
                _art_ap_ph.empty()
                with st.spinner("収集状況を自動確認中..."):
                    _art_auto_poll = fetch_pision_realtime(store, _art_tb_date_str, trigger=False)
                if _art_auto_poll["ok"] and _art_is_new_artid(_art_auto_poll):
                    _art_save_rt_items_to_session(_art_auto_poll["items"])
                    st.rerun()
                elif _art_auto_poll.get("running") or (_art_auto_poll["ok"] and not _art_is_new_artid(_art_auto_poll)):
                    st.rerun()
                else:
                    st.warning("⚠️ 収集が完了しましたがデータが取得できませんでした。「⚡ 速報を取得」をもう一度押してください。")
                    st.session_state.pop(_art_tb_collecting_key, None)
                    st.rerun()

            if _art_do_rt_existing:
                with st.spinner("既存の速報データを確認中（新しい収集は開始しません）..."):
                    _art_exist = fetch_pision_realtime(store, _art_tb_date_str, trigger=False)
                if _art_exist["ok"]:
                    _art_save_rt_items_to_session(_art_exist["items"])
                    st.rerun()
                else:
                    st.warning("⚠️ 既存の速報データが見つかりませんでした。「⚡ 速報を取得」で新しい収集を開始してください。")

            if (_art_tb_date_changed or _art_tb_refetch) and _art_tb_rt_ok:
                with st.spinner(f"{_art_tb_date_str} のデータを取得中..."):
                    _art_tb_fetched_data = None
                    if _art_tb_is_rt:
                        _art_rt = fetch_pision_realtime(store, _art_tb_date_str)
                        if not _art_rt["ok"]:
                            st.error(f"❌ {_art_rt['error']}")
                            if _art_rt.get("collect_started"):
                                st.session_state[_art_tb_collecting_key]     = _art_tb_date_str
                                st.session_state[_art_tb_baseline_artid_key] = None
                        else:
                            st.session_state[_art_tb_collecting_key]     = _art_tb_date_str
                            st.session_state[_art_tb_baseline_artid_key] = _art_rt.get("article_id")
                    else:
                        st.session_state.pop(_art_tb_collecting_key, None)
                        st.session_state.pop(_art_tb_rt_items_key, None)
                        st.session_state.pop(_art_tb_rt_items_date_key, None)
                        try:
                            _art_tb_halls = fetch_pision_halls(api_key)
                        except Exception as e:
                            st.error(f"❌ ホール一覧取得失敗: {e}")
                            _art_tb_halls = []
                        _art_tb_hall_id = None
                        for h in _art_tb_halls:
                            _hn = h.get("name") or h.get("displayName") or ""
                            if store in _hn and "エスパス" in _hn:
                                _art_tb_hall_id = str(h.get("id") or h.get("hallId") or "")
                                break
                        if _art_tb_hall_id is not None:
                            try:
                                _art_tb_fetched_data = fetch_pision_results(api_key, _art_tb_hall_id, _art_tb_date_str)
                            except Exception as e:
                                st.error(f"❌ データ取得失敗: {e}")
                if not _art_tb_fetched_data:
                    st.session_state[_art_tb_bytes_key] = None
                else:
                    _art_tb_rows = [
                        {
                            "台番":     item.get("unitId"),
                            "機種名":   item.get("_machineName") or item.get("_convertedName") or item.get("displayName") or "",
                            "差枚":     item.get("diff", 0),
                            "ゲーム数": item.get("games", 0),
                            "BB":       item.get("bb", 0),
                            "RB":       item.get("rb", 0),
                            "AT":       item.get("art", 0),
                        }
                        for item in _art_tb_fetched_data
                    ]
                    _art_tb_df    = pd.DataFrame(_art_tb_rows)
                    _art_tb_fname = f"{_art_tb_date.strftime('%Y%m%d')}_{store}_20S.xlsx"
                    _art_tb_buf   = io.BytesIO()
                    _art_tb_df.to_excel(_art_tb_buf, index=False)
                    st.session_state[_art_tb_bytes_key] = _art_tb_buf.getvalue()
                    st.session_state[_art_tb_name_key]  = _art_tb_fname
                    st.session_state[_art_tb_count_key] = len(_art_tb_df)
                st.session_state[_art_tb_fetched_key] = _art_tb_date_str
                st.rerun()

            if st.session_state.get(_art_tb_fetched_key) == _art_tb_date_str:
                _art_tb_data = st.session_state.get(_art_tb_bytes_key)
                if _art_tb_data:
                    _art_tb_uploaded = io.BytesIO(_art_tb_data)
                    _art_tb_uploaded.name = st.session_state.get(_art_tb_name_key, f"{_art_tb_date.strftime('%Y%m%d')}_{store}_20S.xlsx")
                    _art_tb_label = "速報" if _art_tb_is_rt else "確定"
                    st.success(f"✅ {_art_tb_date_str} の{_art_tb_label}データ（{st.session_state.get(_art_tb_count_key, '?')}台）を取得し、①にセットしました。")
                elif not _art_is_collecting:
                    st.info(f"📭 {_art_tb_date_str} のデータを取得できませんでした（404 / 未公開 / 店休日の可能性があります）。①から手動でアップロードしてください。")

    st.markdown("---")

    # ── ① Excel アップロード ─────────────────────────────────────────
    st.markdown("### ① Excelファイルをアップロード")
    st.caption("ファイル名は `YYYYMMDD_店舗名_20S.xlsx` の形式を想定しています。")
    uploaded = st.file_uploader("xlsx を選択", type=["xlsx", "xls"], key="art_upload")

    if uploaded is None and _art_tb_uploaded is not None:
        uploaded = _art_tb_uploaded
        st.caption(f"📈 自動取得データを使用中: `{uploaded.name}`（手動でアップロードすると優先されます）")

    if uploaded is not None:
        st.session_state["art_current_excel"] = uploaded.name
        if st.session_state.get("_art_prev_excel") != uploaded.name:
            _restore_article_inputs(uploaded.name, store)
            st.session_state["_art_prev_excel"] = uploaded.name

    # ── 📋 取得データを作業画面内に表示（pisionを開かず照合するため）──────
    if uploaded is not None:
        _art_vt_key       = f"_art_view_df_{store}"
        _art_vt_sum_key   = f"_art_view_summary_{store}"
        _art_vt_meta_key  = f"_art_view_meta_{store}"
        _art_vt_units_key = f"_art_view_units_{store}"
        _art_vt_fn_key    = f"_art_view_df_fn_{store}"
        if st.session_state.get(_art_vt_fn_key) != uploaded.name:
            try:
                uploaded.seek(0)
                _art_vt_raw = _read_uploaded_df(uploaded)
                uploaded.seek(0)
                _art_vt_df, _ = normalize_df(_art_vt_raw)
                _art_vt_df = apply_name_conversion(_art_vt_df)
                _art_disp = pd.DataFrame()
                if "台番" in _art_vt_df.columns:     _art_disp["台番"]    = _art_vt_df["台番"]
                if "機種名" in _art_vt_df.columns:   _art_disp["機種名"]  = _art_vt_df["機種名"]
                if "ゲーム数" in _art_vt_df.columns: _art_disp["G数"]     = _art_vt_df["ゲーム数"]
                if "BB" in _art_vt_df.columns:       _art_disp["BB"]      = _art_vt_df["BB"]
                if "RB" in _art_vt_df.columns:       _art_disp["RB"]      = _art_vt_df["RB"]
                if "AT" in _art_vt_df.columns:       _art_disp["ART"]     = _art_vt_df["AT"]
                if {"BB", "RB", "ゲーム数"} <= set(_art_vt_df.columns):
                    _art_tot = (_art_vt_df["BB"] + _art_vt_df["RB"]).replace(0, pd.NA)
                    _art_disp["合算確率"] = (_art_vt_df["ゲーム数"] / _art_tot).map(
                        lambda v: f"1/{v:.1f}" if pd.notna(v) else "─")
                if "差枚" in _art_vt_df.columns:     _art_disp["差枚"]    = _art_vt_df["差枚"]
                if "台番" in _art_disp.columns:
                    _art_disp = _art_disp.sort_values("台番").reset_index(drop=True)
                st.session_state[_art_vt_key] = _art_disp
                _art_agg = None
                _art_meta_v = None
                if {"機種名", "差枚"} <= set(_art_vt_df.columns):
                    _art_g = _art_vt_df.groupby("機種名", sort=False)
                    _art_agg = pd.DataFrame({
                        "機種名":   list(_art_g.groups.keys()),
                        "台数":     _art_g["差枚"].size().values,
                        "勝台数":   _art_g["差枚"].apply(lambda s: int((s > 0).sum())).values,
                        "総差枚":   _art_g["差枚"].sum().astype(int).values,
                        "平均差枚": _art_g["差枚"].mean().round().astype(int).values,
                    })
                    if "ゲーム数" in _art_vt_df.columns:
                        _art_agg["平均G数"] = _art_g["ゲーム数"].mean().round().astype(int).values
                    else:
                        _art_agg["平均G数"] = 0
                    _art_tot_all = len(_art_vt_df)
                    _art_td_all  = int(_art_vt_df["差枚"].sum())
                    _art_meta_v = {
                        "total":      _art_tot_all,
                        "plus":       int((_art_vt_df["差枚"] > 0).sum()),
                        "total_diff": _art_td_all,
                        "avg_diff":   int(round(_art_td_all / _art_tot_all)) if _art_tot_all else 0,
                        "avg_games":  (int(round(_art_vt_df["ゲーム数"].mean()))
                                       if "ゲーム数" in _art_vt_df.columns and _art_tot_all else 0),
                    }
                _art_ucols = [c for c in ["台番", "機種名", "差枚", "BB", "RB", "AT", "ゲーム数"]
                              if c in _art_vt_df.columns]
                st.session_state[_art_vt_units_key] = _art_vt_df[_art_ucols].copy() if _art_ucols else None
                st.session_state[_art_vt_sum_key]   = _art_agg
                st.session_state[_art_vt_meta_key]  = _art_meta_v
                st.session_state[_art_vt_fn_key]    = uploaded.name
            except Exception:
                st.session_state[_art_vt_key]       = None
                st.session_state[_art_vt_sum_key]   = None
                st.session_state[_art_vt_meta_key]  = None
                st.session_state[_art_vt_units_key] = None
        _art_view_df  = st.session_state.get(_art_vt_key)
        _art_agg_df   = st.session_state.get(_art_vt_sum_key)
        _art_meta_v   = st.session_state.get(_art_vt_meta_key)
        if _art_agg_df is not None and not _art_agg_df.empty and _art_meta_v is not None:
            _art_multi  = _art_agg_df[_art_agg_df["台数"] >= 2].sort_values("平均差枚", ascending=False)
            _art_single = _art_agg_df[_art_agg_df["台数"] == 1]
            _art_rows = [
                (r["機種名"], int(r["台数"]), int(r["勝台数"]),
                 int(r["総差枚"]), int(r["平均差枚"]), int(r["平均G数"]))
                for _, r in _art_multi.iterrows()
            ]
            if not _art_single.empty:
                _art_vn  = int(_art_single["台数"].sum())
                _art_vw  = int(_art_single["勝台数"].sum())
                _art_vtd = int(_art_single["総差枚"].sum())
                _art_vad = int(round(_art_vtd / _art_vn)) if _art_vn else 0
                _art_vg  = int(round((_art_single["平均G数"] * _art_single["台数"]).sum() / _art_vn)) if _art_vn else 0
                _art_rows.append(("バラエティ", _art_vn, _art_vw, _art_vtd, _art_vad, _art_vg))
            _art_fn_m = re.match(r"(\d{4})(\d{2})(\d{2})", os.path.basename(uploaded.name))
            _art_title = (f"{int(_art_fn_m.group(1))}/{int(_art_fn_m.group(2))}/{int(_art_fn_m.group(3))} エスパス{store}"
                          if _art_fn_m else f"エスパス{store}")
            st.caption("📋 pisionの代わりに照合用（2台以上を平均差枚順・1台機種はバラエティに集約／数値はpisionの生データと一致）")
            _art_units_df = st.session_state.get(_art_vt_units_key)
            _art_snames   = set(_art_single["機種名"].tolist()) if not _art_single.empty else None
            _art_comp_h   = max(480, min(820, len(_art_rows) * 42 + 350))
            components.html(
                _build_pision_interactive_html(_art_title, _art_meta_v, _art_rows, _art_units_df, _art_snames),
                height=_art_comp_h, scrolling=True,
            )
        if _art_view_df is not None and not _art_view_df.empty:
            with st.expander(f"📋 台別データ（全{len(_art_view_df)}台）", expanded=False):
                st.dataframe(_art_view_df, use_container_width=True, hide_index=True, height=520)

    st.caption("処理内容：① 全台系PNG ＋ 全台プラス機種別JPG　② ジャグラーシリーズ優秀台JPG　③ その他の優秀台ピックアップJPG")

    # ── ② 個別画像 ──────────────────────────────────────────────────
    kojin_zentai_machines: list[str] = []
    kojin_yushu_machines:  list[str] = []
    kojin_narabi_ranges_text: str = ""
    kojin_narabi_title: str = ""
    kojin_narabi2_ranges_text: str = ""
    kojin_narabi2_title: str = ""
    st.markdown("### ② 個別画像")
    kojin_enabled = st.checkbox("個別画像も生成する", key="art_kojin_enabled",
                                on_change=_save_article_inputs, args=(store,))
    if kojin_enabled:
        _kojin_candidates = load_machine_candidates()
        st.caption("指定した機種の個別画像を生成します。ここに入力した機種はその他の優秀台ピックアップから除外されます。")
        col_kz, col_ky = st.columns(2, gap="large")
        with col_kz:
            st.markdown("**全台**")
            _kz_top = st.columns(3)
            _kz_bot = st.columns(3)
            for _i, _col in enumerate(list(_kz_top) + list(_kz_bot)):
                with _col:
                    render_machine_autocomplete_input(str(_i + 1), f"art_kojin_z_{_i}_{store}", _kojin_candidates,
                                                      on_change=_save_article_inputs, on_change_args=(store,))
            kojin_zentai_machines = [st.session_state.get(f"art_kojin_z_{_i}_{store}", "") for _i in range(6)]
        with col_ky:
            st.markdown("**優秀台**")
            _ky_top = st.columns(3)
            _ky_bot = st.columns(3)
            for _i, _col in enumerate(list(_ky_top) + list(_ky_bot)):
                with _col:
                    render_machine_autocomplete_input(str(_i + 1), f"art_kojin_y_{_i}_{store}", _kojin_candidates,
                                                      on_change=_save_article_inputs, on_change_args=(store,))
            kojin_yushu_machines = [st.session_state.get(f"art_kojin_y_{_i}_{store}", "") for _i in range(6)]
        st.markdown("**並び台番範囲 優秀台**")
        _col_nr, _col_nt = st.columns([2, 3])
        with _col_nr:
            st.text_input(
                "台番範囲（例: 409-413）　ピンクバーあり",
                key=f"art_kojin_narabi_range_{store}",
                placeholder="例: 409-413",
                on_change=_save_article_inputs, args=(store,),
            )
        with _col_nt:
            st.text_input(
                "タイトル（省略時は台番範囲をそのまま使用）",
                key=f"art_kojin_narabi_title_{store}",
                placeholder="例: 4・5列目の優秀台",
                on_change=_save_article_inputs, args=(store,),
            )
        _col_nr2, _col_nt2 = st.columns([2, 3])
        with _col_nr2:
            st.text_input(
                "台番範囲（例: 409-413）　ピンクバーなし",
                key=f"art_kojin_narabi2_range_{store}",
                placeholder="例: 409-413",
                on_change=_save_article_inputs, args=(store,),
            )
        with _col_nt2:
            st.text_input(
                "タイトル（省略時は台番範囲をそのまま使用）",
                key=f"art_kojin_narabi2_title_{store}",
                placeholder="例: 4・5列目の優秀台",
                on_change=_save_article_inputs, args=(store,),
            )
        kojin_narabi_ranges_text  = st.session_state.get(f"art_kojin_narabi_range_{store}", "")
        kojin_narabi_title        = st.session_state.get(f"art_kojin_narabi_title_{store}", "")
        kojin_narabi2_ranges_text = st.session_state.get(f"art_kojin_narabi2_range_{store}", "")
        kojin_narabi2_title       = st.session_state.get(f"art_kojin_narabi2_title_{store}", "")

    # ── ③ 並び画像オプション ─────────────────────────────────────────
    narabi_ok     = False
    narabi_ranges: list[list[int]] = []
    if store in STORE_NARABI_SCRIPT:
        st.markdown("### ③ 並び画像")
        narabi_enabled = st.checkbox("並び画像も生成する", key="art_narabi_enabled",
                                     on_change=_save_article_inputs, args=(store,))
        if narabi_enabled:
            ranges_text = st.text_area(
                "台番範囲　連番: '409-413'、スポット: '508+424'、複数: カンマ/スペース/改行区切り　Excelからのコピペ（台番・機種名・数値の表）もそのまま貼り付け可",
                value="",
                key="art_narabi_ranges_input",
                height=120,
                on_change=_save_article_inputs, args=(store,),
            )
            if ranges_text.strip():
                try:
                    _parsed_ranges = parse_ranges(ranges_text.strip())
                    if _parsed_ranges:
                        _prev_key    = f"art_narabi_previews_{store}"
                        _prev_rt_key = f"art_narabi_prev_rt_{store}"
                        if st.session_state.get(_prev_rt_key, "") != ranges_text.strip():
                            st.session_state.pop(_prev_key, None)
                            for _ci in range(30):
                                st.session_state.pop(f"art_narabi_ck_{store}_{_ci}", None)
                        _previews = st.session_state.get(_prev_key)

                        if _previews is None:
                            st.caption(f"並び指定: {_parsed_ranges}")
                            if uploaded is not None:
                                if st.button("🔍 プレビュー生成", key="art_narabi_preview_btn"):
                                    with st.spinner("プレビュー生成中..."):
                                        _raw_p = pd.read_excel(uploaded)
                                        uploaded.seek(0)
                                        _df_p, _ = normalize_df(_raw_p)
                                        _df_p = apply_name_conversion(_df_p)
                                        _ban_map_p = {int(row["台番"]): i for i, row in _df_p.iterrows()}
                                        _prev_list = []
                                        for _bans in _parsed_ranges:
                                            _idxs = [_ban_map_p[b] for b in _bans if b in _ban_map_p]
                                            if not _idxs:
                                                _prev_list.append(None)
                                                continue
                                            _grp = _df_p.loc[_idxs].copy().reset_index(drop=True)
                                            _ds  = _grp["差枚"]
                                            _ms  = list(dict.fromkeys(str(m) for m in _grp["機種名"]))
                                            _n   = len(_grp)
                                            if len(_ms) == 1:
                                                _tit = f"{_ms[0]}({_n}台並び)"
                                            elif len(_ms) == 2:
                                                _tit = f"{_ms[0]}+{_ms[1]}({_n}台並び)"
                                            else:
                                                _tit = f"{_ms[0]}～{_ms[-1]}({_n}台並び)"
                                            _stat_p = {
                                                "total_diff":  int(_ds.sum()),
                                                "avg_diff":    int(round(_ds.mean())),
                                                "win_count":   int((_ds > 0).sum()),
                                                "total_count": _n,
                                            }
                                            _prev_list.append((_tit, _build_machine_img(_grp, _tit, _stat_p)))
                                        st.session_state[_prev_key]    = _prev_list
                                        st.session_state[_prev_rt_key] = ranges_text.strip()
                                        for _ci in range(len(_prev_list)):
                                            st.session_state[f"art_narabi_ck_{store}_{_ci}"] = True
                                    st.rerun()
                            narabi_ranges = _parsed_ranges
                            narabi_ok = uploaded is not None
                        else:
                            st.caption(f"📋 {len(_previews)}件のプレビュー　チェックした並びのみ生成されます")
                            for _ci, _item in enumerate(_previews):
                                _ck_key = f"art_narabi_ck_{store}_{_ci}"
                                if _item is None:
                                    st.warning(f"⚠️ 範囲 {_parsed_ranges[_ci]} がExcelに見つかりませんでした")
                                    continue
                                _tit, _img = _item
                                _col_ck, _col_img = st.columns([1, 12])
                                with _col_ck:
                                    if _ck_key not in st.session_state:
                                        st.session_state[_ck_key] = True
                                    st.checkbox("", key=_ck_key, label_visibility="collapsed")
                                with _col_img:
                                    st.image(_img, caption=_tit, use_container_width=True)
                            narabi_ranges = [
                                _parsed_ranges[_ci]
                                for _ci, _item in enumerate(_previews)
                                if _item is not None and st.session_state.get(f"art_narabi_ck_{store}_{_ci}", True)
                            ]
                            if narabi_ranges:
                                st.caption(f"✅ {len(narabi_ranges)}件を処理対象に設定")
                                narabi_ok = True
                            else:
                                st.warning("⚠️ 1件以上チェックしてください。")
                    else:
                        st.warning("範囲を正しく認識できませんでした。例: 409-413, 315-317, 508+424")
                except Exception:
                    st.warning("台番範囲の形式が正しくありません。例: 409-413, 315-317")
            else:
                st.info("台番範囲を入力してください。")

    # ── ④ 末尾画像オプション ─────────────────────────────────────────
    if "末尾画像" in STORES.get(store, []):
        st.markdown("### ④ 末尾画像")
        suebangai_enabled = st.checkbox("末尾画像も生成する", key="art_suebangai_enabled",
                                        on_change=_save_article_inputs, args=(store,))
        if suebangai_enabled:
            tail_input = st.text_input(
                "末尾（例: 5、ゾロ目は「ゾロ目」と入力）",
                value="",
                key="art_suebangai_tail_input",
                on_change=_save_article_inputs, args=(store,),
            )
            if tail_input.strip():
                _sc1, _sc2 = st.columns(2)
                with _sc1:
                    sue_zentai = st.button("全台", key="art_sue_zentai_btn",
                                           use_container_width=True, type="primary",
                                           disabled=(uploaded is None))
                with _sc2:
                    sue_yushu = st.button("優秀台", key="art_sue_yushu_btn",
                                          use_container_width=True, type="primary",
                                          disabled=(uploaded is None))

                if (sue_zentai or sue_yushu) and uploaded is not None:
                    _tail = tail_input.strip()
                    try:
                        _raw = pd.read_excel(uploaded)
                        uploaded.seek(0)
                        _df_s, _ = normalize_df(_raw)
                        nm_s, nm_norm_s = load_name_map()
                        if nm_s:
                            _df_s, _ = _apply_map(_df_s, nm_s, nm_norm_s)

                        if _tail == "ゾロ目":
                            _filtered = _df_s[_df_s["台番"].apply(
                                lambda b: (s := str(int(b))) and len(s) >= 2 and s[-2] == s[-1]
                            )].copy()
                            _base_label = "末尾ゾロ目の台"
                        elif _tail.isdigit() and len(_tail) in (1, 2):
                            _filtered = _df_s[_df_s["台番"].astype(str).str[-len(_tail):] == _tail].copy()
                            _base_label = f"末尾{_tail}番台"
                        else:
                            st.error("❌ 末尾を正しく入力してください（例: 5、ゾロ目）。")
                            _filtered = None
                            _base_label = ""

                        if _filtered is not None:
                            if _filtered.empty:
                                st.error(f"❌ {_base_label} の台が見つかりません。")
                            else:
                                if sue_yushu:
                                    _filtered = _filtered[_filtered["差枚"] > 0].copy()
                                    if _filtered.empty:
                                        st.error(f"❌ {_base_label}でプラスの台がありません。")
                                    else:
                                        _img_title = f"{_base_label}の優秀台"
                                        _stat = None
                                else:
                                    _img_title = _base_label
                                    _stat = {
                                        "total_diff":  int(_filtered["差枚"].sum()),
                                        "avg_diff":    int(round(_filtered["差枚"].mean())),
                                        "win_count":   int((_filtered["差枚"] > 0).sum()),
                                        "total_count": len(_filtered),
                                    }
                                if not _filtered.empty:
                                    with st.spinner("末尾画像を生成中..."):
                                        _img_s = _build_machine_img(_filtered, _img_title, _stat)
                                        st.image(_img_s, caption=_img_title, use_container_width=True)
                                        if _IS_CLOUD:
                                            _sue_buf = io.BytesIO()
                                            _img_s.convert("RGB").save(_sue_buf, format="JPEG", quality=85)
                                            st.download_button(
                                                "📥 ダウンロード",
                                                _sue_buf.getvalue(),
                                                f"{_make_safe_fn(_img_title)}.jpg",
                                                "image/jpeg",
                                                key=f"sue_dl_{_img_title[:10]}",
                                            )
                                        else:
                                            _stem_s   = os.path.splitext(uploaded.name)[0].replace("_20S", "")
                                            _out_dir  = os.path.join(_DESKTOP, _stem_s)
                                            os.makedirs(_out_dir, exist_ok=True)
                                            _save_path = os.path.join(_out_dir, f"{_make_safe_fn(_img_title)}.jpg")
                                            _save_jpeg(_img_s, _save_path)
                                            st.success(f"✅ `{_save_path}` に保存しました")
                    except Exception as _e:
                        st.error(f"❌ エラー: {_e}")
                        with st.expander("詳細"):
                            st.code(traceback.format_exc())
            else:
                st.info("末尾を入力してください。")

    # ── ⑤ プレビュー ────────────────────────────────────────────────
    st.markdown("### ⑤ プレビュー")
    if uploaded is not None:
        _art_aprev_key       = f"art_preview_imgs_{store}"
        _art_aprev_fname_key = f"art_preview_fname_{store}"
        _art_aprev_df_key    = f"art_preview_df_{store}"
        _art_aprev_di_key    = f"art_preview_diff_{store}"
        _art_aprev_ex_key    = f"art_preview_ex_{store}"
        _art_aprev_hr_key    = f"art_preview_hr_{store}"
        _art_aprev_zen_key   = f"art_preview_zen_{store}"
        _art_aprev_jug_ex_key   = f"art_preview_jug_ex_{store}"
        _art_aprev_jug_pool_key = f"art_preview_jug_pool_{store}"
        _art_aprev_narabi_key   = f"art_preview_narabi_{store}"
        if st.session_state.get(_art_aprev_fname_key) != uploaded.name:
            for _k in (_art_aprev_key, _art_aprev_df_key, _art_aprev_di_key, _art_aprev_ex_key,
                       _art_aprev_hr_key, _art_aprev_zen_key, _art_aprev_jug_ex_key,
                       _art_aprev_jug_pool_key, _art_aprev_narabi_key):
                st.session_state.pop(_k, None)
            st.session_state[_art_aprev_fname_key] = uploaded.name

        _art_auto_previews = st.session_state.get(_art_aprev_key)
        if _art_auto_previews is None:
            if st.button("🔍 プレビュー生成", key="art_preview_btn"):
                with st.spinner("画像を生成中（しばらくお待ちください）…"):
                    import tempfile as _atf
                    _art_xl_bytes = uploaded.getvalue()
                    uploaded.seek(0)
                    _art_pnb: set[int] = ranges_to_bans(narabi_ranges) if narabi_ok else set()
                    if kojin_enabled and kojin_narabi_ranges_text.strip():
                        try: _art_pnb |= ranges_to_bans(parse_ranges(kojin_narabi_ranges_text.strip()))
                        except Exception: pass
                    if kojin_enabled and kojin_narabi2_ranges_text.strip():
                        try: _art_pnb |= ranges_to_bans(parse_ranges(kojin_narabi2_ranges_text.strip()))
                        except Exception: pass
                    with _atf.TemporaryDirectory() as _atd:
                        _art_txl = os.path.join(_atd, uploaded.name)
                        with open(_art_txl, "wb") as _tf: _tf.write(_art_xl_bytes)
                        _art_prec: set[str] = set()
                        if kojin_enabled:
                            _art_prec |= {m.strip() for m in kojin_zentai_machines if m.strip()}
                            _art_prec |= {m.strip() for m in kojin_yushu_machines if m.strip()}
                        _art_stails: list[str] = []
                        if st.session_state.get("art_suebangai_enabled", False):
                            _at = st.session_state.get("art_suebangai_tail_input", "").strip()
                            if _at: _art_stails = [_at]
                        _art_pr = run_auto_pipeline(
                            _art_txl, _atd, store, _art_pnb,
                            lambda _m: None,
                            narabi_ranges=narabi_ranges if narabi_ok else None,
                            recommended_machines=_art_prec,
                            suebangai_tails=_art_stails,
                            article_mode=True,
                        )
                        _art_pil: list[tuple[str, "Image.Image"]] = []
                        _art_nb_map: dict[str, list[int]] = {}
                        if _art_pr["ok"]:
                            _art_fpm: dict[str, tuple[str, "Image.Image"]] = {}
                            for _fp in _art_pr["files"]:
                                if os.path.exists(_fp) and _fp.lower().endswith((".jpg", ".jpeg")):
                                    _bn = os.path.basename(_fp)
                                    _art_fpm[_bn] = (_bn, Image.open(_fp).copy())
                            _apdf  = _art_pr.get("df")
                            _apdi  = _art_pr.get("diff_raw")
                            # ① 全台系（avg_diff 降順・个別全台系を含む）
                            _azitems: list[tuple[int, str, object]] = []
                            for _it in _art_pr.get("zen_dai_list", []):
                                _azitems.append((_it.get("all_avg_diff", 0), "pipeline", _it))
                            if kojin_enabled and _apdf is not None and _apdi is not None:
                                for _km in kojin_zentai_machines:
                                    _km = _km.strip()
                                    if not _km: continue
                                    _kg = _apdf[_apdf["機種名"] == _km].copy().reset_index(drop=True)
                                    if _kg.empty: continue
                                    _kd = _apdi.loc[_apdf[_apdf["機種名"] == _km].index].reset_index(drop=True)
                                    _azitems.append((int(round(_kd.mean())), "kojin", (_km, _kg, _kd)))
                            for _av, _tp, _da in sorted(_azitems, key=lambda x: x[0], reverse=True):
                                if _tp == "pipeline":
                                    _fn = f"{_make_safe_fn(_da['name'])}.jpg"
                                    if _fn in _art_fpm: _art_pil.append(_art_fpm[_fn])
                                else:
                                    _km, _kg, _kd = _da
                                    _art_pil.append((f"{_km}.jpg", _build_article_machine_img(_kg, _km, _stat_from_diff(_kd))))
                            # ② 高配分（avg_diff 降順・个別優秀台を含む）
                            def _ahrk(x):
                                return x["all_avg_diff"] if "all_avg_diff" in x else (int(round(sum(x["diffs"])/len(x["diffs"]))) if x.get("diffs") else 0)
                            _ahitems: list[tuple[int, str, object]] = []
                            for _it in _art_pr.get("high_ratio_list", []):
                                _ahitems.append((_ahrk(_it), "pipeline", _it))
                            if kojin_enabled and _apdf is not None and _apdi is not None:
                                for _km in kojin_yushu_machines:
                                    _km = _km.strip()
                                    if not _km: continue
                                    _kga = _apdf[_apdf["機種名"] == _km]
                                    if _kga.empty: continue
                                    _kda = _apdi.loc[_kga.index]
                                    _kgp = _kojin_yushu_filter(_km, _kga, _kda, get_store_config(store)).reset_index(drop=True)
                                    if _kgp.empty: continue
                                    _ahitems.append((int(round(_kda.mean())), "kojin", (_km, _kgp)))
                            for _av, _tp, _da in sorted(_ahitems, key=lambda x: x[0], reverse=True):
                                if _tp == "pipeline":
                                    _fn = f"{_make_safe_fn(_da['name'])}_高配分.jpg"
                                    if _fn in _art_fpm: _art_pil.append(_art_fpm[_fn])
                                else:
                                    _km, _kgp = _da
                                    _kti = f"{_km}（優秀台）"
                                    _art_pil.append((f"{_kti}.jpg", _build_machine_img(_kgp, _kti, None)))
                            # ③ 並び画像
                            if narabi_ok and narabi_ranges and _apdf is not None:
                                _anbm = {int(row["台番"]): i for i, row in _apdf.iterrows()}
                                def _antit(nms, nn):
                                    if len(nms) == 1: return f"{nms[0]}({nn}台並び)"
                                    if len(nms) == 2: return f"{nms[0]}+{nms[1]}({nn}台並び)"
                                    return f"{nms[0]}～{nms[-1]}({nn}台並び)"
                                _anbinfos = []
                                for _bs2 in narabi_ranges:
                                    _ix = [_anbm[b] for b in _bs2 if b in _anbm]
                                    if not _ix: continue
                                    _ng = _apdf.loc[_ix].copy().reset_index(drop=True)
                                    _nms2 = list(dict.fromkeys(str(m) for m in _ng["機種名"]))
                                    _anbinfos.append((_ng, _nms2, _antit(_nms2, len(_ng))))
                                from collections import Counter as _ArtC
                                _adt = {t for t, c in _ArtC(i[2] for i in _anbinfos).items() if c > 1}
                                for _ng, _nms2, _nt in _anbinfos:
                                    _nds2 = _ng["差枚"]
                                    if _nt in _adt:
                                        _bs3, _be3 = int(_ng.iloc[0]["台番"]), int(_ng.iloc[-1]["台番"])
                                        _fnt = f"{_nt}（{_bs3}～{_be3}）"
                                    else:
                                        _fnt = _nt
                                    _art_nb_map[f"{_fnt}.jpg"] = [int(b) for b in _ng["台番"].tolist()]
                                    _nst = {"total_diff": int(_nds2.sum()), "avg_diff": int(round(_nds2.mean())),
                                            "win_count": int((_nds2 > 0).sum()), "total_count": len(_ng)}
                                    _art_pil.append((f"{_fnt}.jpg", _build_machine_img(_ng, _nt, _nst)))
                            if kojin_enabled and _apdf is not None and _apdi is not None:
                                for _rt, _rts in [(kojin_narabi_ranges_text, kojin_narabi_title),
                                                   (kojin_narabi2_ranges_text, kojin_narabi2_title)]:
                                    if _rt.strip():
                                        try:
                                            _rb = ranges_to_bans(parse_ranges(_rt.strip()))
                                            _rd = _apdf[_apdf["台番"].apply(lambda b: int(b) in _rb)].copy()
                                            _rdi = _apdi.loc[_rd.index]
                                            _rp = _rd.copy().reset_index(drop=True)
                                            if not _rp.empty:
                                                _base = _rts.strip() or f"{_rt.strip()}の優秀台"
                                                _art_pil.append((f"{_base}.jpg", _build_machine_img(_rp, _base, _stat_from_diff(_rdi))))
                                        except Exception: pass
                            # ④ ジャグラーシリーズ優秀台
                            if "ジャグラーシリーズ優秀台.jpg" in _art_fpm:
                                _art_pil.append(_art_fpm["ジャグラーシリーズ優秀台.jpg"])
                            # ⑤ その他の優秀台ピックアップ
                            if "その他の優秀台ピックアップ.jpg" in _art_fpm:
                                _art_pil.append(_art_fpm["その他の優秀台ピックアップ.jpg"])

                    # ── スランプグラフ合成（プレビュー）────────────────────────
                    _pv_api_key_sl = _get_pision_api_key()
                    if _pv_api_key_sl and _art_pr.get("ok") and _art_pil:
                        _pv_pr_df  = _art_pr.get("df")
                        _pv_pr_di  = _art_pr.get("diff_raw")
                        _pv_rd_sl  = _art_pr.get("date")
                        _pv_dt_key_sl = st.session_state.get(f"art_tb_date_{store}")
                        _pv_date_sl = (
                            _pv_rd_sl.strftime("%Y-%m-%d") if hasattr(_pv_rd_sl, "strftime") else str(_pv_rd_sl)
                        ) if _pv_rd_sl is not None else (
                            _pv_dt_key_sl.strftime("%Y-%m-%d") if hasattr(_pv_dt_key_sl, "strftime") else str(_pv_dt_key_sl or "")
                        )
                        # ban_map: ファイル名 → 台番リスト
                        _pv_bm_sl: dict[str, list[int]] = {}
                        for _zd_pv in _art_pr.get("zen_dai_list", []):
                            _pv_bm_sl[f"{_make_safe_fn(_zd_pv['name'])}.jpg"] = _zd_pv.get("bans", [])
                        for _hr_pv in _art_pr.get("high_ratio_list", []):
                            if _hr_pv.get("has_image", False):
                                _fn_hr_pv = f"{_make_safe_fn(_hr_pv['name'])}_高配分.jpg"
                                if _hr_pv.get("bans"):
                                    _pv_bm_sl[_fn_hr_pv] = _hr_pv["bans"]
                                elif _pv_pr_df is not None:
                                    _g_pv = _pv_pr_df[_pv_pr_df["機種名"] == _hr_pv["name"]]
                                    _pv_bm_sl[_fn_hr_pv] = [int(b) for b in _g_pv["台番"].tolist()]
                        if kojin_enabled and _pv_pr_df is not None and _pv_pr_di is not None:
                            for _km_pv2 in kojin_yushu_machines:
                                _km_pv2 = _km_pv2.strip()
                                if not _km_pv2:
                                    continue
                                _fn_ky_pv = f"{_make_safe_fn(_km_pv2)}（優秀台）.jpg"
                                if any(fn == _fn_ky_pv for fn, _ in _art_pil):
                                    _kga_pv = _pv_pr_df[_pv_pr_df["機種名"] == _km_pv2]
                                    if not _kga_pv.empty:
                                        _kda_pv = _pv_pr_di.loc[_kga_pv.index]
                                        _kgp_pv = _kojin_yushu_filter(_km_pv2, _kga_pv, _kda_pv, get_store_config(store)).reset_index(drop=True)
                                        if not _kgp_pv.empty:
                                            _pv_bm_sl[_fn_ky_pv] = [int(b) for b in _kgp_pv["台番"].tolist()]
                        _jpool_pv = _art_pr.get("jug_pool_df")
                        if _jpool_pv is not None and not _jpool_pv.empty:
                            _pv_bm_sl["ジャグラーシリーズ優秀台.jpg"] = [
                                int(str(b).split(".")[0]) for b in _jpool_pv["台番"].dropna()
                                if str(b).split(".")[0].lstrip("-").isdigit()
                            ]
                        _son_bns_pv = sorted({int(_e["ban"]) for _e in _art_pr.get("sonota_excellent_list", []) if "ban" in _e})
                        if _son_bns_pv:
                            _pv_bm_sl["その他の優秀台ピックアップ.jpg"] = _son_bns_pv
                        for _fn_nb_pv2, _bns_nb_pv2 in _art_nb_map.items():
                            _pv_bm_sl[_fn_nb_pv2] = _bns_nb_pv2
                        try:
                            _pv_rt_cached = st.session_state.get(f"_art_tb_rt_items_{store}")
                            _pv_rt_date   = st.session_state.get(f"_art_tb_rt_items_date_{store}", "")
                            _pv_pision_sl = None
                            if _pv_rt_cached and _pv_rt_date == _pv_date_sl:
                                _pv_pision_sl = _pv_rt_cached
                            else:
                                _pv_halls_sl = fetch_pision_halls(_pv_api_key_sl)
                                _pv_hall_id_sl = None
                                for _h_pv in _pv_halls_sl:
                                    _hn_pv = _h_pv.get("name") or _h_pv.get("displayName") or ""
                                    if store in _hn_pv and "エスパス" in _hn_pv:
                                        _pv_hall_id_sl = str(_h_pv.get("id") or _h_pv.get("hallId") or "")
                                        break
                                if _pv_hall_id_sl:
                                    _pv_pision_sl = fetch_pision_results(_pv_api_key_sl, _pv_hall_id_sl, _pv_date_sl)
                                    if _pv_pision_sl:
                                        _slump_apply_names(_pv_pision_sl)
                            if _pv_pision_sl:
                                _pv_by_uid   = {str(_it.get("unitId", "")): _it for _it in _pv_pision_sl}
                                _pv_tmpl_sl  = find_slump_template()
                                _pv_bgg_sl   = _find_slump_bg()
                                _pv_ban2mac: dict[str, str] = {}
                                _pv_ban2diff: dict[str, int] = {}
                                if _pv_pr_df is not None:
                                    for _idx_pv, _row_pv in _pv_pr_df.iterrows():
                                        _bs_pv = str(_row_pv.get("台番", "")).split(".")[0]
                                        if _bs_pv.lstrip("-").isdigit():
                                            _pv_ban2mac[_bs_pv] = str(_row_pv.get("機種名", ""))
                                            if _pv_pr_di is not None:
                                                try:
                                                    _pv_ban2diff[_bs_pv] = int(_pv_pr_di.loc[_idx_pv])
                                                except Exception:
                                                    pass
                                if _pv_tmpl_sl is not None:
                                    _merged_pil: list[tuple[str, "Image.Image"]] = []
                                    for (_fn_pv2, _img_pv2) in _art_pil:
                                        _bans_pv2 = _pv_bm_sl.get(_fn_pv2, [])
                                        if not _bans_pv2:
                                            _merged_pil.append((_fn_pv2, _img_pv2))
                                            continue
                                        _g_imgs_pv2: list["Image.Image"] = []
                                        _show_mn_pv2 = (_fn_pv2 in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg")
                                                        or _fn_pv2.startswith("末尾") or _fn_pv2.startswith("バラエティ"))
                                        _is_zentai_pv2 = (
                                            not _fn_pv2.endswith("_高配分.jpg") and
                                            not _fn_pv2.endswith("（優秀台）.jpg") and
                                            _fn_pv2 not in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg")
                                        )
                                        for _b_pv2 in _bans_pv2:
                                            _it_pv2 = _pv_by_uid.get(str(_b_pv2))
                                            if _it_pv2 is None or not _it_pv2.get("points"):
                                                continue
                                            _dn_pv2 = (_it_pv2.get("_convertedName")
                                                       or _it_pv2.get("displayName")
                                                       or _pv_ban2mac.get(str(_b_pv2), str(_b_pv2)))
                                            _sd_pv2 = not (_is_zentai_pv2 and _pv_ban2diff.get(str(_b_pv2), 0) < 0)
                                            try:
                                                _g_imgs_pv2.append(draw_slump_graph(
                                                    _pv_tmpl_sl, str(_b_pv2), _dn_pv2,
                                                    _it_pv2["points"], diff=_it_pv2.get("diff"),
                                                    machine_name=_dn_pv2 if _show_mn_pv2 else None,
                                                    show_diff=_sd_pv2,
                                                ))
                                            except Exception:
                                                pass
                                        if _g_imgs_pv2:
                                            _merged_pil.append((_fn_pv2, _attach_slump_to_table(_img_pv2, _g_imgs_pv2, _pv_bgg_sl)))
                                        else:
                                            _merged_pil.append((_fn_pv2, _img_pv2))
                                    _art_pil = _merged_pil
                        except Exception:
                            pass  # スランプ取得失敗時は表のみプレビュー

                    st.session_state[_art_aprev_key]       = _art_pil
                    st.session_state[_art_aprev_df_key]    = _art_pr.get("df")
                    st.session_state[_art_aprev_di_key]    = _art_pr.get("diff_raw")
                    st.session_state[_art_aprev_ex_key]    = _art_pr.get("sonota_excellent_list", [])
                    st.session_state[_art_aprev_hr_key]    = {
                        f"{_make_safe_fn(it['name'])}_高配分.jpg": it["name"]
                        for it in _art_pr.get("high_ratio_list", [])
                    }
                    st.session_state[_art_aprev_zen_key]   = {
                        f"{_make_safe_fn(it['name'])}.jpg": it["name"]
                        for it in _art_pr.get("zen_dai_list", [])
                    }
                    st.session_state[_art_aprev_jug_ex_key]   = _art_pr.get("jug_excellent_list", [])
                    st.session_state[_art_aprev_jug_pool_key] = _art_pr.get("jug_pool_df")
                    st.session_state[_art_aprev_narabi_key]   = _art_nb_map
                st.rerun()
        else:
            st.caption(f"📋 {len(_art_auto_previews)}枚の画像プレビュー　チェックした画像のみ生成されます")
            for _rs in range(0, len(_art_auto_previews), 3):
                _agc = st.columns(3)
                for _ci2, _ci in enumerate(range(_rs, min(_rs + 3, len(_art_auto_previews)))):
                    _pt, _pi = _art_auto_previews[_ci]
                    _ck = f"art_prev_ck_{store}_{_ci}"
                    if _ck not in st.session_state: st.session_state[_ck] = True
                    with _agc[_ci2]:
                        _sc, _si2 = st.columns([1, 10])
                        with _sc: st.checkbox("", key=_ck, label_visibility="collapsed")
                        with _si2: st.image(_pi, caption=_pt, use_container_width=True)
            _ab1, _ab2 = st.columns(2)
            with _ab1:
                if st.button("🔄 その他を更新", key="art_preview_update_btn", use_container_width=True):
                    _apdf2  = st.session_state.get(_art_aprev_df_key)
                    _apdi2  = st.session_state.get(_art_aprev_di_key)
                    _apex   = st.session_state.get(_art_aprev_ex_key, [])
                    _apje   = st.session_state.get(_art_aprev_jug_ex_key, [])
                    _apjp   = st.session_state.get(_art_aprev_jug_pool_key)
                    _aphr   = st.session_state.get(_art_aprev_hr_key, {})
                    _apzen  = st.session_state.get(_art_aprev_zen_key, {})
                    _apnb   = st.session_state.get(_art_aprev_narabi_key, {})
                    if _apdf2 is not None and _apdi2 is not None:
                        _ajss  = set(get_store_config(store)["juggler_series"])
                        _akset = {m.strip() for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip()}
                        _audfs:  list[pd.DataFrame] = []
                        _audis:  list[pd.Series]    = []
                        _ajdfs:  list[pd.DataFrame] = []
                        for _ci, (_pname, _) in enumerate(_art_auto_previews):
                            if not st.session_state.get(f"art_prev_ck_{store}_{_ci}", True):
                                _m2 = _aphr.get(_pname)
                                if _m2 and _m2 not in _akset:
                                    _mr = _apdf2[_apdf2["機種名"] == _m2]
                                    if not _mr.empty:
                                        _md = _apdi2.loc[_mr.index]
                                        _mm = _md >= 1000
                                        _mg = _mr[_mm.values].copy().reset_index(drop=True)
                                        _mgd = _md[_mm].reset_index(drop=True)
                                        if not _mg.empty:
                                            if _m2 in _ajss: _ajdfs.append(_mg)
                                            else: _audfs.append(_mg); _audis.append(_mgd)
                                _mz2 = _apzen.get(_pname)
                                if _mz2 and not _m2 and _mz2 not in _akset:
                                    _mr2 = _apdf2[_apdf2["機種名"] == _mz2]
                                    if not _mr2.empty:
                                        _md2 = _apdi2.loc[_mr2.index]
                                        _mm2 = _md2 >= 1000
                                        _mg2 = _mr2[_mm2.values].copy().reset_index(drop=True)
                                        _mgd2 = _md2[_mm2].reset_index(drop=True)
                                        if not _mg2.empty:
                                            if _mz2 in _ajss: _ajdfs.append(_mg2)
                                            else: _audfs.append(_mg2); _audis.append(_mgd2)
                                _nbb = _apnb.get(_pname)
                                if _nbb:
                                    _nbr = _apdf2[_apdf2["台番"].apply(lambda b: int(b) in set(_nbb))].copy()
                                    if not _nbr.empty:
                                        _nbd = _apdi2.loc[_nbr.index]
                                        _nbm2 = _nbd >= 1000
                                        _nbg = _nbr[_nbm2.values].copy().reset_index(drop=True)
                                        if not _nbg.empty:
                                            _nbgd = _nbd[_nbm2].reset_index(drop=True)
                                            _njm  = _nbg["機種名"].isin(_ajss)
                                            _nbo  = _nbg[~_njm.values].copy().reset_index(drop=True)
                                            _nbod = _nbgd[~_njm.values].reset_index(drop=True)
                                            if not _nbo.empty: _audfs.append(_nbo); _audis.append(_nbod)
                                _ASUF = "（優秀台）.jpg"
                                if _pname.endswith(_ASUF) and not _m2 and not _nbb:
                                    _kmy = _pname[:-len(_ASUF)]
                                    if _kmy in {m.strip() for m in kojin_yushu_machines if m.strip()}:
                                        _mry = _apdf2[_apdf2["機種名"] == _kmy]
                                        if not _mry.empty:
                                            _mdy = _apdi2.loc[_mry.index]
                                            _mmy = _mdy >= 1000
                                            _mgy = _mry[_mmy.values].copy().reset_index(drop=True)
                                            _mgdy = _mdy[_mmy].reset_index(drop=True)
                                            if not _mgy.empty:
                                                _hjea = any(_pn == "ジャグラーシリーズ優秀台.jpg" for _pn, _ in _art_auto_previews)
                                                if _kmy in _ajss and _hjea: _ajdfs.append(_mgy)
                                                else: _audfs.append(_mgy); _audis.append(_mgdy)
                        _anp = list(_art_auto_previews)
                        _aup = False
                        _upd_bm: dict[str, list[int]] = {}  # 更新画像の台番map（スランプ合成用）
                        def _bans_from_df(df: "pd.DataFrame") -> list[int]:
                            return [int(str(b).split(".")[0]) for b in df["台番"].dropna()
                                    if str(b).split(".")[0].lstrip("-").isdigit()]
                        if _audfs:
                            _aexb = {it["ban"] for it in _apex}
                            if _aexb:
                                _aexr = _apdf2[_apdf2["台番"].apply(lambda b: int(b) in _aexb)].copy().reset_index(drop=True)
                                _aexd = _apdi2.loc[_apdf2[_apdf2["台番"].apply(lambda b: int(b) in _aexb)].index].reset_index(drop=True)
                                _aadf = [_aexr] + _audfs; _aadi = [_aexd] + _audis
                            else:
                                _aadf = _audfs; _aadi = _audis
                            _asc = pd.concat(_aadf, ignore_index=True)
                            _asc = _asc.iloc[_asc["台番"].argsort()].reset_index(drop=True)
                            _asi = _build_machine_img(_asc, "その他の優秀台ピックアップ", None)
                            _upd_bm["その他の優秀台ピックアップ.jpg"] = _bans_from_df(_asc)
                            for _ci, (_pn2, _) in enumerate(_anp):
                                if _pn2 == "その他の優秀台ピックアップ.jpg":
                                    _anp[_ci] = (_pn2, _asi); break
                            else:
                                _anp.append(("その他の優秀台ピックアップ.jpg", _asi))
                                st.session_state[f"art_prev_ck_{store}_{len(_anp)-1}"] = True
                            _aup = True
                        if _ajdfs:
                            _ajbase = [_apjp.copy()] if _apjp is not None and not _apjp.empty else (
                                [_apdf2[_apdf2["台番"].apply(lambda b: int(b) in {it["ban"] for it in _apje})].copy().reset_index(drop=True)]
                                if _apje else []
                            )
                            _ajc = pd.concat(_ajbase + _ajdfs, ignore_index=True)
                            _ajc = _ajc.iloc[_ajc["台番"].argsort()].reset_index(drop=True)
                            if len(_ajc) <= 5:
                                # 5台以下 → overflowと同じ扱い: その他の優秀台ピックアップへ
                                _aov_exb = {it["ban"] for it in _apex}
                                _aov_rows = _apdf2[_apdf2["台番"].apply(lambda b: int(b) in _aov_exb)].copy().reset_index(drop=True) if _aov_exb else pd.DataFrame()
                                _aov_dfs = ([_aov_rows] if not _aov_rows.empty else []) + _ajdfs
                                _aov_son = pd.concat(_aov_dfs, ignore_index=True)
                                _aov_son = _aov_son.drop_duplicates(subset=["台番"])
                                _aov_son = _aov_son.iloc[_aov_son["台番"].argsort()].reset_index(drop=True)
                                _aov_img = _build_machine_img(_aov_son, "その他の優秀台ピックアップ", None)
                                _upd_bm["その他の優秀台ピックアップ.jpg"] = _bans_from_df(_aov_son)
                                for _ci2, (_pn2, _) in enumerate(_anp):
                                    if _pn2 == "その他の優秀台ピックアップ.jpg":
                                        _anp[_ci2] = (_pn2, _aov_img); break
                                else:
                                    _anp.append(("その他の優秀台ピックアップ.jpg", _aov_img))
                                    st.session_state[f"art_prev_ck_{store}_{len(_anp)-1}"] = True
                                _aup = True
                            else:
                                _ahkj = any(m.strip() in _ajss for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip())
                                _asto = _ahkj or any(
                                    st.session_state.get(f"art_prev_ck_{store}_{_si}", True) and
                                    (_aphr.get(_sp) in _ajss or _apzen.get(_sp) in _ajss)
                                    for _si, (_sp, _) in enumerate(_art_auto_previews)
                                )
                                _ajt = "その他のジャグラーシリーズの優秀台" if _asto else "ジャグラーシリーズの優秀台"
                                _aji = _build_machine_img(_ajc, _ajt, None)
                                _upd_bm["ジャグラーシリーズ優秀台.jpg"] = _bans_from_df(_ajc)
                                for _ji, (_jn, _) in enumerate(_anp):
                                    if _jn == "ジャグラーシリーズ優秀台.jpg":
                                        _anp[_ji] = (_jn, _aji); break
                                else:
                                    _anp.append(("ジャグラーシリーズ優秀台.jpg", _aji))
                                    st.session_state[f"art_prev_ck_{store}_{len(_anp)-1}"] = True
                                _aup = True
                        if _aup:
                            # スランプグラフ合成（その他を更新後）
                            _upd_api_sl = _get_pision_api_key()
                            if _upd_api_sl and _upd_bm:
                                try:
                                    _upd_rt_c  = st.session_state.get(f"_art_tb_rt_items_{store}")
                                    _upd_rt_d  = st.session_state.get(f"_art_tb_rt_items_date_{store}", "")
                                    _upd_dtk   = st.session_state.get(f"art_tb_date_{store}")
                                    _upd_datesl = (_upd_dtk.strftime("%Y-%m-%d") if hasattr(_upd_dtk, "strftime")
                                                   else str(_upd_dtk or ""))
                                    _upd_pision = None
                                    if _upd_rt_c and _upd_rt_d == _upd_datesl:
                                        _upd_pision = _upd_rt_c
                                    else:
                                        _upd_halls = fetch_pision_halls(_upd_api_sl)
                                        _upd_hid = None
                                        for _hh in _upd_halls:
                                            _hhn = _hh.get("name") or _hh.get("displayName") or ""
                                            if store in _hhn and "エスパス" in _hhn:
                                                _upd_hid = str(_hh.get("id") or _hh.get("hallId") or "")
                                                break
                                        if _upd_hid:
                                            _upd_pision = fetch_pision_results(_upd_api_sl, _upd_hid, _upd_datesl)
                                            if _upd_pision:
                                                _slump_apply_names(_upd_pision)
                                    if _upd_pision:
                                        _upd_by_uid  = {str(_it.get("unitId", "")): _it for _it in _upd_pision}
                                        _upd_tmpl    = find_slump_template()
                                        _upd_bgg     = _find_slump_bg()
                                        _upd_b2mac: dict[str, str] = {}
                                        _upd_b2diff: dict[str, int] = {}
                                        if _apdf2 is not None and _apdi2 is not None:
                                            for _idx_u, _row_u in _apdf2.iterrows():
                                                _bs_u = str(_row_u.get("台番", "")).split(".")[0]
                                                if _bs_u.lstrip("-").isdigit():
                                                    _upd_b2mac[_bs_u] = str(_row_u.get("機種名", ""))
                                                    try:
                                                        _upd_b2diff[_bs_u] = int(_apdi2.loc[_idx_u])
                                                    except Exception:
                                                        pass
                                        if _upd_tmpl is not None:
                                            _merged_anp: list[tuple[str, "Image.Image"]] = []
                                            for (_fn_u, _img_u) in _anp:
                                                _bans_u = _upd_bm.get(_fn_u, [])
                                                if not _bans_u:
                                                    _merged_anp.append((_fn_u, _img_u))
                                                    continue
                                                _g_imgs_u: list["Image.Image"] = []
                                                _show_mn_u = (_fn_u in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg"))
                                                for _b_u in _bans_u:
                                                    _it_u = _upd_by_uid.get(str(_b_u))
                                                    if _it_u is None or not _it_u.get("points"):
                                                        continue
                                                    _dn_u = (_it_u.get("_convertedName") or _it_u.get("displayName")
                                                             or _upd_b2mac.get(str(_b_u), str(_b_u)))
                                                    try:
                                                        _g_imgs_u.append(draw_slump_graph(
                                                            _upd_tmpl, str(_b_u), _dn_u,
                                                            _it_u["points"], diff=_it_u.get("diff"),
                                                            machine_name=_dn_u if _show_mn_u else None,
                                                            show_diff=True,
                                                        ))
                                                    except Exception:
                                                        pass
                                                if _g_imgs_u:
                                                    _merged_anp.append((_fn_u, _attach_slump_to_table(_img_u, _g_imgs_u, _upd_bgg)))
                                                else:
                                                    _merged_anp.append((_fn_u, _img_u))
                                            _anp = _merged_anp
                                except Exception:
                                    pass
                            st.session_state[_art_aprev_key] = _anp
                            st.rerun()
            with _ab2:
                if st.button("🔄 プレビューをクリア", key="art_preview_clear_btn", use_container_width=True):
                    for _k in (_art_aprev_key, _art_aprev_df_key, _art_aprev_di_key, _art_aprev_ex_key,
                               _art_aprev_hr_key, _art_aprev_zen_key, _art_aprev_jug_ex_key,
                               _art_aprev_jug_pool_key, _art_aprev_narabi_key):
                        st.session_state.pop(_k, None)
                    for _ci in range(len(_art_auto_previews)):
                        st.session_state.pop(f"art_prev_ck_{store}_{_ci}", None)
                    st.rerun()

    # ── ⑥ 実行ボタン ─────────────────────────────────────────────────
    st.markdown("### ⑥ 実行")
    run_clicked = st.button(
        "▶▶ 自動処理を開始",
        type="primary",
        use_container_width=True,
        disabled=(uploaded is None),
        key="art_run",
    )
    # Cloud のみ：ボタン直下にZIPダウンロード用スロットを確保
    _art_zip_slot = st.empty() if _IS_CLOUD else None

    if uploaded is None:
        st.info("⬆️ まずExcelをアップロードしてください。")
    elif run_clicked:
        _save_article_inputs(store)
        stem     = os.path.splitext(uploaded.name)[0]
        dir_stem = stem.replace("_20S", "")
        if _IS_CLOUD:
            _art_tmpdir = tempfile.mkdtemp()
            excel_path  = os.path.join(_art_tmpdir, uploaded.name)
            output_dir  = os.path.join(_art_tmpdir, dir_stem)
        else:
            _art_tmpdir = tempfile.mkdtemp()
            excel_path  = os.path.join(_art_tmpdir, uploaded.name)
            output_dir  = os.path.join(_DESKTOP, dir_stem)
        narabi_dir = os.path.join(output_dir, "並び画像")
        narabi_bans = ranges_to_bans(narabi_ranges) if narabi_ok else set()
        if kojin_enabled and kojin_narabi_ranges_text.strip():
            try:
                narabi_bans |= ranges_to_bans(parse_ranges(kojin_narabi_ranges_text.strip()))
            except Exception:
                pass
        if kojin_enabled and kojin_narabi2_ranges_text.strip():
            try:
                narabi_bans |= ranges_to_bans(parse_ranges(kojin_narabi2_ranges_text.strip()))
            except Exception:
                pass

        with open(excel_path, "wb") as f:
            f.write(uploaded.getvalue())
        os.makedirs(output_dir, exist_ok=True)

        log_lines: list[str] = []

        with st.status("自動処理を実行中…", expanded=True) as status_widget:

            def _log(msg: str) -> None:
                log_lines.append(msg)
                st.write(msg)

            _kojin_names: set[str] = {
                m.strip()
                for m in (kojin_zentai_machines + kojin_yushu_machines)
                if m.strip()
            }
            _sue_tails_art: list[str] = []
            if st.session_state.get("art_suebangai_enabled", False):
                _art_t = st.session_state.get("art_suebangai_tail_input", "").strip()
                if _art_t:
                    _sue_tails_art = [_art_t]
            result = run_auto_pipeline(
                excel_path, output_dir, store, narabi_bans, _log,
                narabi_ranges=narabi_ranges if narabi_ok else None,
                recommended_machines=_kojin_names,
                suebangai_tails=_sue_tails_art,
                article_mode=True,
            )

            # ── 並び画像（subprocess）────────────────────────────────
            narabi_result: dict | None = None
            if narabi_ok:
                st.write("⏳ 並び画像スクリプトを実行中…")
                os.makedirs(narabi_dir, exist_ok=True)
                ok_n, out_n, err_n = _patch_and_run_narabi(
                    STORE_NARABI_SCRIPT[store], excel_path, narabi_dir, narabi_ranges
                )
                narabi_result = {"ok": ok_n, "stdout": out_n, "stderr": err_n}
                st.write(f"{'✅' if ok_n else '❌'} 並び画像{'完了' if ok_n else 'エラー'}")
                # 記事用: 青タイトルバー＋赤ライン（6px）をトップから除去
                if ok_n and os.path.isdir(narabi_dir):
                    for _nf in os.listdir(narabi_dir):
                        if not _nf.lower().endswith((".jpg", ".jpeg", ".png")):
                            continue
                        _np = os.path.join(narabi_dir, _nf)
                        try:
                            _nim = Image.open(_np)
                            _narabi_bar_h = round(_nim.width * 73 / 950) + 6
                            _nim = _nim.crop((0, _narabi_bar_h, _nim.width, _nim.height))
                            _save_jpeg(_nim.convert("RGB"), _np)
                        except Exception:
                            pass

            # ── 個別画像生成 ─────────────────────────────────────────
            if kojin_enabled and result["ok"]:
                df_k   = result.get("df")
                diff_k = result.get("diff_raw")
                if df_k is not None and diff_k is not None:
                    for _km in kojin_zentai_machines:
                        _km = _km.strip()
                        if not _km:
                            continue
                        _kgrp = df_k[df_k["機種名"] == _km].copy().reset_index(drop=True)
                        if _kgrp.empty:
                            _log(f"  個別(全台)「{_km}」: 該当台なし")
                            continue
                        _kdr = diff_k.loc[df_k[df_k["機種名"] == _km].index].reset_index(drop=True)
                        _kimg = _build_article_machine_img(_kgrp, _km, _stat_from_diff(_kdr))
                        _kout = os.path.join(output_dir, f"{_make_safe_fn(_km)}.jpg")
                        _save_jpeg(_kimg, _kout)
                        result["files"].append(_kout)
                        result["zen_dai_list"].append({
                            "name":         _km,
                            "count":        int((_kdr > 0).sum()),
                            "total":        len(_kgrp),
                            "diffs":        sorted([int(d) for d in _kdr.tolist() if int(d) >= 1000], reverse=True),
                            "all_avg_diff": int(round(_kdr.mean())),
                        })
                        _log(f"  ✅ 個別(全台)「{_km}」({len(_kgrp)}台)")
                    for _km in kojin_yushu_machines:
                        _km = _km.strip()
                        if not _km:
                            continue
                        _kgrp_all = df_k[df_k["機種名"] == _km]
                        if _kgrp_all.empty:
                            _log(f"  個別(優秀台)「{_km}」: 該当台なし")
                            continue
                        _kdr_all  = diff_k.loc[_kgrp_all.index]
                        _kgrp_p   = _kojin_yushu_filter(_km, _kgrp_all, _kdr_all, get_store_config(store))
                        if _kgrp_p.empty:
                            _log(f"  個別(優秀台)「{_km}」: 条件を満たす台なし")
                            continue
                        _kdr_p2   = _kdr_all.loc[_kgrp_p.index]
                        _kgrp_p   = _kgrp_p.reset_index(drop=True)
                        _ktitle = "優秀台ピックアップ"
                        _kimg   = _build_machine_img(_kgrp_p, _ktitle, None)
                        _kout   = os.path.join(output_dir, f"{_make_safe_fn(_km)}（優秀台）.jpg")
                        _save_jpeg(_kimg, _kout)
                        result["files"].append(_kout)
                        result["high_ratio_list"].append({
                            "name":         _km,
                            "count":        len(_kgrp_p),
                            "total":        len(_kgrp_all),
                            "diffs":        sorted([int(d) for d in _kdr_p2.tolist() if int(d) >= 1000], reverse=True),
                            "all_avg_diff": int(round(_kdr_all.mean())),
                        })
                        _log(f"  ✅ 個別(優秀台)「{_km}」({len(_kgrp_p)}台)")

                    if kojin_narabi_ranges_text.strip():
                        try:
                            _rng_list = parse_ranges(kojin_narabi_ranges_text.strip())
                            _rng_bans = ranges_to_bans(_rng_list)
                            if _rng_bans:
                                _rng_df   = df_k[df_k["台番"].apply(lambda b: int(b) in _rng_bans)].copy()
                                _rng_diff = diff_k.loc[_rng_df.index]
                                _rng_p    = _rng_df.copy().reset_index(drop=True)
                                if not _rng_p.empty:
                                    _base = kojin_narabi_title.strip() if kojin_narabi_title.strip() else f"{kojin_narabi_ranges_text.strip()}の優秀台"
                                    _rng_stat  = _stat_from_diff(_rng_diff)
                                    _rng_img   = _build_machine_img(_rng_p, _base, _rng_stat)
                                    _rng_out   = os.path.join(output_dir, f"{_make_safe_fn(_base)}.jpg")
                                    _save_jpeg(_rng_img, _rng_out)
                                    result["files"].append(_rng_out)
                                    _log(f"  ✅ 台番範囲(優秀台・ピンクバーあり)「{_base}」({len(_rng_p)}台)")
                                else:
                                    _log(f"  台番範囲(優秀台): 台番 {sorted(_rng_bans)} に台なし")
                        except Exception:
                            _log(f"  ❌ 台番範囲優秀台エラー: {traceback.format_exc()}")

                    if kojin_narabi2_ranges_text.strip():
                        try:
                            _rng2_list = parse_ranges(kojin_narabi2_ranges_text.strip())
                            _rng2_bans = ranges_to_bans(_rng2_list)
                            if _rng2_bans:
                                _rng2_df   = df_k[df_k["台番"].apply(lambda b: int(b) in _rng2_bans)].copy()
                                _rng2_diff = diff_k.loc[_rng2_df.index]
                                _rng2_p    = _rng2_df.copy().reset_index(drop=True)
                                if not _rng2_p.empty:
                                    _base2 = kojin_narabi2_title.strip() if kojin_narabi2_title.strip() else f"{kojin_narabi2_ranges_text.strip()}の優秀台"
                                    _rng2_img = _build_machine_img(_rng2_p, _base2, None)
                                    _rng2_out = os.path.join(output_dir, f"{_make_safe_fn(_base2)}.jpg")
                                    _save_jpeg(_rng2_img, _rng2_out)
                                    result["files"].append(_rng2_out)
                                    _log(f"  ✅ 台番範囲(優秀台・ピンクバーなし)「{_base2}」({len(_rng2_p)}台)")
                                else:
                                    _log(f"  台番範囲(優秀台・ピンクバーなし): 台番 {sorted(_rng2_bans)} に台なし")
                        except Exception:
                            _log(f"  ❌ 台番範囲優秀台(ピンクバーなし)エラー: {traceback.format_exc()}")

            # ── プレビューでチェックを外した画像を削除・再生成 ────────────
            _art_aprev_imgs = st.session_state.get(f"art_preview_imgs_{store}")
            if _art_aprev_imgs and result["ok"]:
                _ardf   = result.get("df")
                _ardr   = result.get("diff_raw")
                _arjss  = set(get_store_config(store)["juggler_series"])
                _arks   = {m.strip() for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip()}
                _arhrm  = {f"{_make_safe_fn(it['name'])}_高配分.jpg": it["name"] for it in result.get("high_ratio_list", [])}
                _arzm   = {f"{_make_safe_fn(it['name'])}.jpg": it["name"] for it in result.get("zen_dai_list", [])}
                _arnbm  = st.session_state.get(f"art_preview_narabi_{store}", {})
                _arxdfs: list[pd.DataFrame] = []
                _arxdis: list[pd.Series]    = []
                _arjdfs: list[pd.DataFrame] = []
                for _ci, (_pname, _) in enumerate(_art_aprev_imgs):
                    if not st.session_state.get(f"art_prev_ck_{store}_{_ci}", True):
                        _dp = os.path.join(output_dir, _pname)
                        if os.path.exists(_dp):
                            os.remove(_dp)
                            _log(f"  🗑️ スキップ: {_pname}")
                        if _ardf is None or _ardr is None: continue
                        _arm = _arhrm.get(_pname)
                        if _arm and _arm not in _arks:
                            _armr = _ardf[_ardf["機種名"] == _arm]
                            if not _armr.empty:
                                _armd = _ardr.loc[_armr.index]
                                _armk = _armd >= 1000
                                _armg = _armr[_armk.values].copy().reset_index(drop=True)
                                _armgd = _armd[_armk].reset_index(drop=True)
                                if not _armg.empty:
                                    if _arm in _arjss: _arjdfs.append(_armg)
                                    else: _arxdfs.append(_armg); _arxdis.append(_armgd)
                        _arzv = _arzm.get(_pname)
                        if _arzv and not _arm and _arzv not in _arks:
                            _arzr = _ardf[_ardf["機種名"] == _arzv]
                            if not _arzr.empty:
                                _arzd = _ardr.loc[_arzr.index]
                                _arzk = _arzd >= 1000
                                _arzg = _arzr[_arzk.values].copy().reset_index(drop=True)
                                _arzgd = _arzd[_arzk].reset_index(drop=True)
                                if not _arzg.empty:
                                    if _arzv in _arjss: _arjdfs.append(_arzg)
                                    else: _arxdfs.append(_arzg); _arxdis.append(_arzgd)
                        _arnb = _arnbm.get(_pname)
                        if _arnb:
                            _arnbr = _ardf[_ardf["台番"].apply(lambda b: int(b) in set(_arnb))].copy()
                            if not _arnbr.empty:
                                _arnbd = _ardr.loc[_arnbr.index]
                                _arnbk = _arnbd >= 1000
                                _arnbg = _arnbr[_arnbk.values].copy().reset_index(drop=True)
                                if not _arnbg.empty:
                                    _arnbgd = _arnbd[_arnbk].reset_index(drop=True)
                                    _arnjm  = _arnbg["機種名"].isin(_arjss)
                                    _arno   = _arnbg[~_arnjm.values].copy()
                                    _arnod  = _arnbgd[~_arnjm.values]
                                    if not _arno.empty: _arxdfs.append(_arno); _arxdis.append(_arnod)
                        _ARSUF = "（優秀台）.jpg"
                        if _pname.endswith(_ARSUF) and not _arm and not _arnb:
                            _arky = _pname[:-len(_ARSUF)]
                            if _arky in {m.strip() for m in kojin_yushu_machines if m.strip()}:
                                _aryr = _ardf[_ardf["機種名"] == _arky]
                                if not _aryr.empty:
                                    _aryd = _ardr.loc[_aryr.index]
                                    _aryk = _aryd >= 1000
                                    _aryg = _aryr[_aryk.values].copy().reset_index(drop=True)
                                    _arygd = _aryd[_aryk].reset_index(drop=True)
                                    if not _aryg.empty:
                                        _arhjex = any(
                                            _pn2 == "ジャグラーシリーズ優秀台.jpg" and
                                            st.session_state.get(f"art_prev_ck_{store}_{_pni}", True)
                                            for _pni, (_pn2, _) in enumerate(_art_aprev_imgs)
                                        )
                                        if _arky in _arjss and _arhjex: _arjdfs.append(_aryg)
                                        else: _arxdfs.append(_aryg); _arxdis.append(_arygd)
                if _arxdfs and _ardf is not None and _ardr is not None:
                    _arsonp = os.path.join(output_dir, "その他の優秀台ピックアップ.jpg")
                    _arexb  = {it["ban"] for it in result.get("sonota_excellent_list", [])}
                    if _arexb:
                        _arexr = _ardf[_ardf["台番"].apply(lambda b: int(b) in _arexb)].copy().reset_index(drop=True)
                        _arexd = _ardr.loc[_ardf[_ardf["台番"].apply(lambda b: int(b) in _arexb)].index].reset_index(drop=True)
                        _aralld = [_arexr] + _arxdfs; _aralldi = [_arexd] + _arxdis
                    else:
                        _aralld = _arxdfs; _aralldi = _arxdis
                    _arsc = pd.concat(_aralld, ignore_index=True)
                    _arsc = _arsc.iloc[_arsc["台番"].argsort()].reset_index(drop=True)
                    _save_jpeg(_build_machine_img(_arsc, "その他の優秀台ピックアップ", None), _arsonp, target_kb=800)
                    _log(f"  ✅ その他の優秀台ピックアップ再生成: {len(_arsc)}台")
                if _arjdfs and _ardf is not None:
                    _arjp    = os.path.join(output_dir, "ジャグラーシリーズ優秀台.jpg")
                    _arjexb  = {it["ban"] for it in result.get("jug_excellent_list", [])}
                    _arjbase = [_ardf[_ardf["台番"].apply(lambda b: int(b) in _arjexb)].copy().reset_index(drop=True)] if _arjexb else []
                    _arjcomb = pd.concat(_arjbase + _arjdfs, ignore_index=True)
                    _arjcomb = _arjcomb.iloc[_arjcomb["台番"].argsort()].reset_index(drop=True)
                    if len(_arjcomb) <= 5:
                        # 5台以下 → overflowと同じ扱い: その他の優秀台ピックアップへ
                        _arsonp2 = os.path.join(output_dir, "その他の優秀台ピックアップ.jpg")
                        _arexb2  = {it["ban"] for it in result.get("sonota_excellent_list", [])}
                        _arexr2  = _ardf[_ardf["台番"].apply(lambda b: int(b) in _arexb2)].copy().reset_index(drop=True) if _arexb2 else pd.DataFrame()
                        _ardfs2  = ([_arexr2] if not _arexr2.empty else []) + _arjdfs
                        _arson2  = pd.concat(_ardfs2, ignore_index=True)
                        _arson2  = _arson2.drop_duplicates(subset=["台番"])
                        _arson2  = _arson2.iloc[_arson2["台番"].argsort()].reset_index(drop=True)
                        _save_jpeg(_build_machine_img(_arson2, "その他の優秀台ピックアップ", None), _arsonp2, target_kb=800)
                        _log(f"  ✅ ジャグラー{len(_arjcomb)}台→overflow: その他の優秀台ピックアップに追加({len(_arson2)}台)")
                    else:
                        _arjhkj  = any(m.strip() in _arjss for m in (kojin_zentai_machines + kojin_yushu_machines) if m.strip())
                        _arjt    = "その他のジャグラーシリーズの優秀台" if _arjhkj else "ジャグラーシリーズの優秀台"
                        _save_jpeg(_build_machine_img(_arjcomb, _arjt, None), _arjp, target_kb=800)
                        _log(f"  ✅ ジャグラーシリーズ優秀台再生成: {len(_arjcomb)}台")

            # ── スランプグラフ合成（記事用）────────────────────────────────
            _art_api_key_sl = _get_pision_api_key()
            if _art_api_key_sl and result.get("ok"):
                _art_rd_sl  = result.get("date")
                _art_dt_key_sl = st.session_state.get(f"art_tb_date_{store}")
                _art_date_sl = (
                    _art_rd_sl.strftime("%Y-%m-%d") if hasattr(_art_rd_sl, "strftime") else str(_art_rd_sl)
                ) if _art_rd_sl is not None else (
                    _art_dt_key_sl.strftime("%Y-%m-%d") if hasattr(_art_dt_key_sl, "strftime") else str(_art_dt_key_sl or "")
                )
                _art_df_sl = result.get("df")
                _art_dr_sl = result.get("diff_raw")

                # ban_map: 出力ファイル名 → 台番リスト
                _art_bm_sl: dict[str, list[int]] = {}
                for _zd_sl in result.get("zen_dai_list", []):
                    _art_bm_sl[f"{_make_safe_fn(_zd_sl['name'])}.jpg"] = _zd_sl.get("bans", [])
                for _hr_sl in result.get("high_ratio_list", []):
                    if _hr_sl.get("has_image", False):
                        _fn_hr_sl = f"{_make_safe_fn(_hr_sl['name'])}_高配分.jpg"
                        if _hr_sl.get("bans"):
                            _art_bm_sl[_fn_hr_sl] = _hr_sl["bans"]
                        elif _art_df_sl is not None:
                            _g_sl = _art_df_sl[_art_df_sl["機種名"] == _hr_sl["name"]]
                            _art_bm_sl[_fn_hr_sl] = [int(b) for b in _g_sl["台番"].tolist()]
                # kojin 優秀台 → {machine}（優秀台）.jpg
                if kojin_enabled and _art_df_sl is not None and _art_dr_sl is not None:
                    for _km_sl in kojin_yushu_machines:
                        _km_sl = _km_sl.strip()
                        if not _km_sl:
                            continue
                        _fn_ky_sl = f"{_make_safe_fn(_km_sl)}（優秀台）.jpg"
                        if os.path.exists(os.path.join(output_dir, _fn_ky_sl)):
                            _kga_sl = _art_df_sl[_art_df_sl["機種名"] == _km_sl]
                            if not _kga_sl.empty:
                                _kda_sl = _art_dr_sl.loc[_kga_sl.index]
                                _kgp_sl = _kojin_yushu_filter(_km_sl, _kga_sl, _kda_sl, get_store_config(store)).reset_index(drop=True)
                                if not _kgp_sl.empty:
                                    _art_bm_sl[_fn_ky_sl] = [int(b) for b in _kgp_sl["台番"].tolist()]
                _jpool_sl = result.get("jug_pool_df")
                if _jpool_sl is not None and not _jpool_sl.empty:
                    _art_bm_sl["ジャグラーシリーズ優秀台.jpg"] = [
                        int(str(b).split(".")[0]) for b in _jpool_sl["台番"].dropna()
                        if str(b).split(".")[0].lstrip("-").isdigit()
                    ]
                _son_bns_sl = sorted({int(_e["ban"]) for _e in result.get("sonota_excellent_list", []) if "ban" in _e})
                if _son_bns_sl:
                    _art_bm_sl["その他の優秀台ピックアップ.jpg"] = _son_bns_sl
                for _fn_nb_sl, _bns_nb_sl in st.session_state.get(f"art_preview_narabi_{store}", {}).items():
                    _art_bm_sl[_fn_nb_sl] = _bns_nb_sl

                _log(f"📡 スランプ: pisionデータ取得中（日付={_art_date_sl}）")
                try:
                    _art_rt_cached_sl = st.session_state.get(f"_art_tb_rt_items_{store}")
                    _art_rt_date_sl   = st.session_state.get(f"_art_tb_rt_items_date_{store}", "")
                    _art_pision_sl = None
                    if _art_rt_cached_sl and _art_rt_date_sl == _art_date_sl:
                        _art_pision_sl = _art_rt_cached_sl
                        _log(f"✅ スランプ: 速報キャッシュ使用（{len(_art_pision_sl)}台）")
                    else:
                        _art_halls_sl = fetch_pision_halls(_art_api_key_sl)
                        _art_hall_id_sl = None
                        for _h_sl in _art_halls_sl:
                            _hn_sl = _h_sl.get("name") or _h_sl.get("displayName") or ""
                            if store in _hn_sl and "エスパス" in _hn_sl:
                                _art_hall_id_sl = str(_h_sl.get("id") or _h_sl.get("hallId") or "")
                                break
                        if _art_hall_id_sl:
                            _art_pision_sl = fetch_pision_results(_art_api_key_sl, _art_hall_id_sl, _art_date_sl)
                            if _art_pision_sl:
                                _slump_apply_names(_art_pision_sl)
                                _log(f"✅ スランプ: {len(_art_pision_sl)}台分のデータ取得")
                            else:
                                _log(f"⚠️ スランプ: {_art_date_sl} の確定データなし（404/未公開）")
                        else:
                            _log(f"⚠️ スランプ: '{store}' のホールIDが見つかりません")
                    if _art_pision_sl:
                        _art_by_uid_sl   = {str(_it.get("unitId", "")): _it for _it in _art_pision_sl}
                        _art_tmpl_sl     = find_slump_template()
                        _art_bgg_sl      = _find_slump_bg()
                        _art_ban2mac_sl: dict[str, str] = {}
                        _art_ban2diff_sl: dict[str, int] = {}
                        if _art_df_sl is not None:
                            for _idx_sl, _row_sl in _art_df_sl.iterrows():
                                _bs_sl = str(_row_sl.get("台番", "")).split(".")[0]
                                if _bs_sl.lstrip("-").isdigit():
                                    _art_ban2mac_sl[_bs_sl] = str(_row_sl.get("機種名", ""))
                                    if _art_dr_sl is not None:
                                        try:
                                            _art_ban2diff_sl[_bs_sl] = int(_art_dr_sl.loc[_idx_sl])
                                        except Exception:
                                            pass
                        _art_slump_cnt = 0
                        if _art_tmpl_sl is None:
                            _log("⚠️ スランプ: テンプレート画像(base_3000_bk.png)が見つかりません")
                        else:
                            for _fp_sl in sorted(os.listdir(output_dir)):
                                if not _fp_sl.lower().endswith((".jpg", ".jpeg")):
                                    continue
                                _bans_sl = _art_bm_sl.get(_fp_sl, [])
                                if not _bans_sl:
                                    continue
                                _fpath_sl = os.path.join(output_dir, _fp_sl)
                                try:
                                    _t_img_sl = Image.open(_fpath_sl).convert("RGB")
                                except Exception:
                                    continue
                                _g_imgs_sl: list["Image.Image"] = []
                                _show_mn_sl = (_fp_sl in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg")
                                               or _fp_sl.startswith("末尾") or _fp_sl.startswith("バラエティ"))
                                _is_zentai_sl = (
                                    not _fp_sl.endswith("_高配分.jpg") and
                                    not _fp_sl.endswith("（優秀台）.jpg") and
                                    _fp_sl not in ("ジャグラーシリーズ優秀台.jpg", "その他の優秀台ピックアップ.jpg")
                                )
                                for _b_sl in _bans_sl:
                                    _it_sl = _art_by_uid_sl.get(str(_b_sl))
                                    if _it_sl is None or not _it_sl.get("points"):
                                        continue
                                    _dn_sl = (_it_sl.get("_convertedName")
                                              or _it_sl.get("displayName")
                                              or _art_ban2mac_sl.get(str(_b_sl), str(_b_sl)))
                                    _sd_sl = not (_is_zentai_sl and _art_ban2diff_sl.get(str(_b_sl), 0) < 0)
                                    try:
                                        _g_imgs_sl.append(draw_slump_graph(
                                            _art_tmpl_sl, str(_b_sl), _dn_sl,
                                            _it_sl["points"], diff=_it_sl.get("diff"),
                                            machine_name=_dn_sl if _show_mn_sl else None,
                                            show_diff=_sd_sl,
                                        ))
                                    except Exception:
                                        pass
                                if not _g_imgs_sl:
                                    continue
                                _combined_sl = _attach_slump_to_table(_t_img_sl, _g_imgs_sl, _art_bgg_sl)
                                _save_jpeg(_combined_sl, _fpath_sl)
                                _art_slump_cnt += 1
                        _log(f"✅ スランプ: {_art_slump_cnt}枚にスランプグラフを合成")
                except Exception as _sl_exc:
                    _log(f"❌ スランプグラフ合成エラー: {_sl_exc}")

            all_ok = result["ok"] and (narabi_result is None or narabi_result["ok"])
            if all_ok:
                status_widget.update(label="✅ 全処理完了！", state="complete", expanded=False)
            else:
                status_widget.update(label="⚠️ エラーあり", state="error", expanded=True)

        if not result["ok"]:
            st.markdown("### エラー詳細")
            st.error(result["error"])
        if narabi_result and not narabi_result["ok"]:
            st.markdown("### 並び画像エラー")
            if narabi_result["stderr"]:
                st.error(narabi_result["stderr"])

        st.markdown("### 生成されたファイル")
        if not _IS_CLOUD:
            st.info(f"📁 `{output_dir}`")
        if os.path.isdir(output_dir):
            imgs = sorted(
                f for f in os.listdir(output_dir)
                if f.lower().endswith((".png", ".jpg", ".jpeg"))
            )
            if imgs:
                for fname in imgs:
                    with st.expander(fname, expanded=False):
                        st.image(os.path.join(output_dir, fname), use_container_width=True)
            else:
                st.warning("画像ファイルが見つかりませんでした。")

        if narabi_ok and os.path.isdir(narabi_dir):
            narabi_imgs = sorted(
                f for f in os.listdir(narabi_dir)
                if f.lower().endswith((".png", ".jpg", ".jpeg"))
            )
            if narabi_imgs:
                st.info(f"📁 並び画像: `{narabi_dir}`")
                for fname in narabi_imgs:
                    with st.expander(f"[並び] {fname}", expanded=False):
                        st.image(os.path.join(narabi_dir, fname), use_container_width=True)

        if result["ok"]:
            st.markdown("---")
            report_text = generate_report_text(
                store_name=store,
                date=result.get("date"),
                zen_dai_list=result.get("zen_dai_list", []),
                high_ratio_list=result.get("high_ratio_list", []),
                nami_list=result.get("nami_list", []),
                excellent_list=result.get("excellent_list", []),
                diff_raw=result.get("diff_raw"),
                df=result.get("df"),
            )
            for _old, _new in STORE_RESULT_TRANSFORMS.get(store, []):
                report_text = report_text.replace(_old, _new)
            _date = result.get("date")
            if _date:
                _txt_name = f"{_date.month:02d}{_date.day:02d}_結果.txt"
            else:
                _txt_name = "結果.txt"
            _txt_path = os.path.join(output_dir, _txt_name)
            try:
                with open(_txt_path, "w", encoding="utf-8") as _f:
                    _f.write(report_text)
                if _IS_CLOUD:
                    st.caption(f"📄 {_txt_name} をZIPに含めます")
                else:
                    st.caption(f"📄 {_txt_name} を保存しました")
            except Exception as _e:
                st.warning(f"結果.txt の保存に失敗: {_e}")

            # ── ZIPデータをセッションに保存（ボタン直下スロットへ後で表示）──
            if os.path.isdir(output_dir):
                try:
                    _art_zip_data = _make_zip_bytes(output_dir)
                    if _IS_CLOUD:
                        st.session_state[f"_art_zip_data_{store}"] = _art_zip_data
                        st.session_state[f"_art_zip_stem_{store}"] = dir_stem
                    else:
                        st.download_button(
                            label="📥 画像・テキストをZIPでダウンロード",
                            data=_art_zip_data,
                            file_name=f"{dir_stem}.zip",
                            mime="application/zip",
                            key="art_zip_dl",
                            type="secondary",
                        )
                except Exception as _ze:
                    st.warning(f"ZIP生成に失敗: {_ze}")

            import html as _html
            _safe = _html.escape(report_text)
            _lines = report_text.count("\n") + 1
            _h = min(600, max(200, _lines * 20 + 110))
            st.iframe(f"""
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:8px;">
              <span style="font-size:1.4rem;font-weight:700;">結果報告</span>
              <button id="cb" onclick="
                var t=document.getElementById('rt');
                t.select();t.setSelectionRange(0,99999);
                document.execCommand('copy');
                this.textContent='✅ コピー済み';this.style.background='#4CAF50';
                var b=this;setTimeout(function(){{b.textContent='📋 コピー';b.style.background='#2F559E';}},2000);
              " style="padding:5px 14px;background:#2F559E;color:#fff;
                       border:none;border-radius:4px;cursor:pointer;font-size:14px;">
                📋 コピー
              </button>
            </div>
            <textarea id="rt" readonly
              style="width:100%;height:{_h - 70}px;font-family:monospace;font-size:13px;
                     border:1px solid #ccc;padding:8px;box-sizing:border-box;resize:vertical;"
            >{_safe}</textarea>
            """, height=_h)

    # ── ボタン直下スロットにZIPダウンロードボタンを表示（Cloud のみ）──
    if _IS_CLOUD and _art_zip_slot is not None and st.session_state.get(f"_art_zip_data_{store}"):
        with _art_zip_slot.container():
            st.success("✅ 処理が完了しました。ZIPをダウンロードしてください。")
            st.download_button(
                label="📥 画像・テキストをZIPでダウンロード",
                data=st.session_state[f"_art_zip_data_{store}"],
                file_name=f"{st.session_state.get(f'_art_zip_stem_{store}', 'output')}.zip",
                mime="application/zip",
                key="art_zip_dl",
                type="primary",
            )

    # ── 戻るボタン ─────────────────────────────────────────────────
    st.markdown("---")
    if st.button("← 戻る", key="art_back"):
        _navigate("image_type")


def _find_kisha_col(df: pd.DataFrame) -> pd.DataFrame | None:
    """「機種名」列を正規名で特定して返す。見つからなければ None。"""
    if "機種名" in df.columns:
        return df
    for alias in COLUMN_ALIASES.get("機種名", []):
        if alias in df.columns:
            return df.rename(columns={alias: "機種名"})
    return None


def _load_master_df() -> pd.DataFrame:
    """機種名変換.xlsx を「変換前」「変換後」の2列DataFrameとして読み込む"""
    raw = pd.read_excel(NAME_MAP_PATH, header=1, usecols=[1, 2])
    raw.columns = ["変換前（正式名）", "変換後（簡略名）"]
    raw = raw.dropna(how="all").reset_index(drop=True)
    raw = raw[
        raw["変換前（正式名）"].astype(str).str.strip().ne("nan") &
        raw["変換後（簡略名）"].astype(str).str.strip().ne("nan")
    ].reset_index(drop=True)
    return raw


def _save_master_df(df: pd.DataFrame) -> None:
    """編集済みDataFrameを機種名変換.xlsx に上書き保存する（行1・行2の構造を保持）"""
    wb = load_workbook(NAME_MAP_PATH)
    ws = wb.active
    # 3行目以降（データ行）を削除
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    # 新データを書き込む
    for _, row in df.iterrows():
        orig = str(row["変換前（正式名）"]).strip()
        conv = str(row["変換後（簡略名）"]).strip()
        if orig and conv and orig != "nan" and conv != "nan":
            ws.append([None, orig, conv])
    wb.save(NAME_MAP_PATH)
    load_name_map.clear()


def _generate_rote_result_text(
    df: pd.DataFrame,
    machine_inputs: list[str],
    date_obj,
    store_full: str,
    store: str = "",
) -> str:
    """ローテ結果テキストを生成して返す。"""
    import datetime as _dt
    import calendar as _cal
    dow = ["月", "火", "水", "木", "金", "土", "日"][date_obj.weekday()]

    if store == "新宿歌舞伎町":
        # 月内3期間（1～10 / 11～20 / 21～月末）で week_str と day_num を計算
        d = date_obj.day
        m = date_obj.month
        y = date_obj.year
        last_day = _cal.monthrange(y, m)[1]
        if d <= 10:
            period_start, period_end, day_num = 1, 10, d
        elif d <= 20:
            period_start, period_end, day_num = 11, 20, d - 10
        else:
            period_start, period_end, day_num = 21, last_day, d - 20
        week_str = f"{m}月{period_start}日～{m}月{period_end}日"
        # 機種名を「×」で結合してポスター行を生成
        _kw_names = [kw.strip() for kw in machine_inputs if (kw or "").strip()]
        poster_machines = "×".join(_kw_names) if _kw_names else ""
        lines = [
            f"{date_obj.month}/{date_obj.day}({dow})👨‍💻結果👨‍💻",
            "エスパス 新宿 歌舞伎 町",
            "",
            f"🏆{week_str}🏆",
            f"🏆{poster_machines}ポスター🏆",
            f"＼＼{day_num}日目結果／／",
        ]
    else:
        weekday = date_obj.weekday()          # 0=月
        day_num = weekday + 1
        monday  = date_obj - _dt.timedelta(days=weekday)
        sunday  = monday  + _dt.timedelta(days=6)
        week_str = f"{monday.month}月{monday.day}日～{sunday.month}月{sunday.day}日"
        lines = [
            f"{date_obj.month}/{date_obj.day}({dow})👨‍💻結果👨‍💻",
            store_full,
            "",
            f"🏆{week_str}オススメポスター🏆",
            f"＼＼{day_num}日目結果／／",
        ]

    _re, _te = ROTE_EMOJI_CONFIG.get(store, ("🌌", "🔥"))
    _be = ROTE_BAN_EMOJI_CONFIG.get(store, "💫")
    name_col = "機種名" if "機種名" in df.columns else None
    tiers = [
        (10000, None,  f"{_te}10,000枚超{_te}"),
        (5000,  10000, f"{_te}5,000枚超{_te}"),
        (3000,  5000,  f"{_te}3,000枚超{_te}"),
        (1000,  3000,  f"{_te}1,000枚超{_te}"),
    ]

    for kw in machine_inputs:
        kw = (kw or "").strip()
        if not kw or name_col is None:
            continue
        mask = df[name_col].astype(str).str.contains(kw, na=False)
        sub  = df[mask].copy()
        if sub.empty:
            continue
        sub["差枚"] = pd.to_numeric(sub["差枚"], errors="coerce").fillna(0).astype(int)
        sub["台番"] = pd.to_numeric(sub["台番"], errors="coerce").fillna(0).astype(int)
        sub = sub[sub["差枚"] >= 1000]
        if sub.empty:
            continue

        machine_name = sub[name_col].iloc[0]
        lines.append("")
        if store == "新宿歌舞伎町":
            lines.append(f"【{machine_name}】")
        else:
            lines.append(f"{_re}{machine_name}{_re}")

        for lo, hi, label in tiers:
            if hi is None:
                tier = sub[sub["差枚"] >= lo].sort_values(["差枚", "台番"], ascending=[False, True])
            else:
                tier = sub[(sub["差枚"] >= lo) & (sub["差枚"] < hi)].sort_values(["差枚", "台番"], ascending=[False, True])
            if tier.empty:
                continue
            lines.append(label)
            for _, row in tier.iterrows():
                lines.append(f"{_be}{int(row['台番'])}番台")
            lines.append("")

    return "\n".join(lines).rstrip()


def _generate_shibuyashinkan_result_texts(
    df: "pd.DataFrame",
    machine_inputs1: list[str],
    machine_inputs2: list[str],
    date_obj,
    store_full: str,
    weekly_items: list[str],
    machine_inputs3: list[str] | None = None,
    monthly_items: list[str] | None = None,
    weekly_items2: list[str] | None = None,
) -> tuple[str, str, str, str]:
    """渋谷新館専用：結果テキスト①②③④を生成して返す。
    weekly_items  = 週間オススメ表①の項目（テキスト②に使用）
    monthly_items = 月間オススメ表の項目（テキスト③に使用）
    weekly_items2 = 週間オススメ表②の項目（テキスト④に使用）
    """
    weekday = date_obj.weekday()
    dow     = ["月", "火", "水", "木", "金", "土", "日"][weekday]
    header  = f"{date_obj.month}/{date_obj.day}({dow})👨‍💻結果👨‍💻\n{store_full}"

    name_col = "機種名" if "機種名" in df.columns else None
    tiers = [
        (10000, None,  "📌10,000枚超📌"),
        (5000,  10000, "📌5,000枚超📌"),
        (3000,  5000,  "📌3,000枚超📌"),
        (1000,  3000,  "📌1,000枚超📌"),
    ]

    def _tier_block(inputs: list[str]) -> list[str]:
        if not name_col:
            return []
        combined = pd.DataFrame()
        for kw in inputs:
            kw = (kw or "").strip()
            if not kw:
                continue
            mask = df[name_col].astype(str).str.contains(kw, na=False)
            sub  = df[mask].copy()
            if sub.empty:
                continue
            sub["差枚"] = pd.to_numeric(sub["差枚"], errors="coerce").fillna(0).astype(int)
            sub["台番"] = pd.to_numeric(sub["台番"], errors="coerce").fillna(0).astype(int)
            combined = pd.concat([combined, sub], ignore_index=True)
        if combined.empty:
            return []
        combined = combined[combined["差枚"] >= 1000]
        if combined.empty:
            return []
        lines: list[str] = []
        for lo, hi, label in tiers:
            tier = (combined[combined["差枚"] >= lo] if hi is None
                    else combined[(combined["差枚"] >= lo) & (combined["差枚"] < hi)])
            tier = tier.sort_values(["差枚", "台番"], ascending=[False, True])
            if tier.empty:
                continue
            lines.append(label)
            for _, row in tier.iterrows():
                lines.append(f"💫{int(row['台番'])}番台")
            lines.append("")
        return lines

    # ── テキスト① ──────────────────────────────────────────────────
    name1   = next((n.strip() for n in machine_inputs1 if (n or "").strip()), "")
    lines1  = [header, ""]
    if name1:
        lines1.append(f"🔥{name1}🔥")
    lines1 += ["🔥週間オススメポスター🔥", ""]
    lines1 += _tier_block(machine_inputs1)
    text1   = "\n".join(lines1).rstrip()

    # ── テキスト② （週間オススメ） ────────────────────────────────────
    name2   = next((n.strip() for n in machine_inputs2 if (n or "").strip()), "")
    lines2  = [header, ""]
    if name2:
        lines2.append(f"👊{name2}👊")
    lines2 += ["👊週間オススメポスター👊", ""]
    lines2.append("✅毎日何かしらの仕掛けアリ!?")
    for item in (weekly_items or []):
        if (item or "").strip():
            lines2.append(f"📍{item.strip()}")
    lines2.append("")
    lines2 += _tier_block(machine_inputs2)
    text2   = "\n".join(lines2).rstrip()

    # ── テキスト③ （月間オススメ） ────────────────────────────────────
    _inputs3 = machine_inputs3 or []
    name3    = next((n.strip() for n in _inputs3 if (n or "").strip()), "")
    lines3   = [header, ""]
    if name3:
        lines3.append(f"🗼{name3}🗼")
    lines3 += ["🗼月間オススメポスター🗼", ""]
    lines3.append("✅毎日何かしらの仕掛けアリ!?")
    for item in (monthly_items or []):
        if (item or "").strip():
            lines3.append(f"📍{item.strip()}")
    lines3.append("")
    lines3 += _tier_block(_inputs3)
    text3   = "\n".join(lines3).rstrip()

    # ── テキスト④ （週間オススメ表②） ───────────────────────────────────
    lines4 = [header, ""]
    lines4 += ["🤡ジャグラーシリーズ🤡", "🚨週間オススメポスター🚨", ""]
    lines4.append("✅毎日何かしらの仕掛けアリ!?")
    for item in (weekly_items2 or []):
        _it = item.split("\n")[0].strip()
        if _it:
            lines4.append(f"📍{_it}")
    text4 = "\n".join(lines4).rstrip()

    return text1, text2, text3, text4


def _load_rote_machines(store: str) -> dict:
    if os.path.exists(_ROTE_SAVE_FILE):
        try:
            with open(_ROTE_SAVE_FILE, encoding="utf-8") as _f:
                return json.load(_f).get(store, {})
        except Exception:
            pass
    return {}


def _save_rote_machines(store: str, inputs1: list[str], inputs2: list[str],
                        inputs3: list[str] | None = None,
                        monthly_start: str | None = None) -> None:
    data: dict = {}
    if os.path.exists(_ROTE_SAVE_FILE):
        try:
            with open(_ROTE_SAVE_FILE, encoding="utf-8") as _f:
                data = json.load(_f)
        except Exception:
            pass
    _existing = data.get(store, {})
    _entry: dict = {"set1": inputs1, "set2": inputs2}
    if inputs3 is not None:
        _entry["set3"] = inputs3
    # monthly_start: 明示指定があれば更新、なければ既存値を保持
    _ms = monthly_start if monthly_start is not None else _existing.get("monthly_start")
    if _ms:
        _entry["monthly_start"] = _ms
    data[store] = _entry
    with open(_ROTE_SAVE_FILE, "w", encoding="utf-8") as _f:
        json.dump(data, _f, ensure_ascii=False, indent=2)


def _load_wrt_machines(store: str) -> dict:
    if os.path.exists(_WRT_SAVE_FILE):
        try:
            with open(_WRT_SAVE_FILE, encoding="utf-8") as _f:
                return json.load(_f).get(store, {})
        except Exception:
            pass
    return {}


def _save_wrt_machines(store: str, weekly: list[str], daily: dict) -> None:
    data: dict = {}
    if os.path.exists(_WRT_SAVE_FILE):
        try:
            with open(_WRT_SAVE_FILE, encoding="utf-8") as _f:
                data = json.load(_f)
        except Exception:
            pass
    data[store] = {"weekly": weekly, "daily": {str(k): v for k, v in daily.items()}}
    with open(_WRT_SAVE_FILE, "w", encoding="utf-8") as _f:
        json.dump(data, _f, ensure_ascii=False, indent=2)


_WEEKLY_N_ITEMS = 8  # 週間オススメ表の項目数

# 週間オススメ表②（t3）に表示するジャグラー機種リスト（追加時はここに1行足すだけ）
_T3_JUGGLER_MACHINES: list[str] = [
    "マイジャグV",
    "ネオアイム",
    "ファンキー2",
    "ゴージャグ3",
    "ハピジャグV",
    "ジャグラーガールズ",
    "ミスジャグ",
    "ウルトラミラジャグ",
]

# 「バーベルとらっぴ」「椅子に座るピエロ」項目用の選択肢（機種名ではなく結果表記）
_T3_SPECIAL_OPTS: list[str] = [
    "対象台が5/5でプラス差枚",
    "対象台が4/5でプラス差枚",
    "対象台が3/5でプラス差枚",
    "対象台が2/5でプラス差枚",
    "対象台が1/5でプラス差枚",
    "対象台が4/4でプラス差枚",
    "対象台が3/4でプラス差枚",
    "対象台が2/4でプラス差枚",
    "対象台が1/4でプラス差枚",
    "対象台が3/3でプラス差枚",
    "対象台が2/3でプラス差枚",
    "対象台が1/3でプラス差枚",
    "対象台が2/2でプラス差枚",
    "対象台が1/2でプラス差枚",
    "対象台が1/1でプラス差枚",
]
# 上記特殊選択肢を使う項目のキーワード（item テキストにいずれかが含まれれば適用）
_T3_SPECIAL_ITEM_KEYS: tuple[str, ...] = ("バーベルとらっぴ", "椅子に座るピエロ")


def _weekly_table_data(store: str, table_num: int = 1) -> dict:
    """weekly_items.json から store の t{table_num} ブロックを返す（旧形式も自動変換）。"""
    _tk = f"t{table_num}"
    if os.path.exists(_WEEKLY_SAVE_FILE):
        try:
            with open(_WEEKLY_SAVE_FILE, encoding="utf-8") as _f:
                _sd = json.load(_f).get(store, {})
            # 旧フォーマット（items/title が直下にある）→ t1 扱い
            if isinstance(_sd, list):
                _sd = {"t1": {"items": _sd}}
            elif isinstance(_sd, dict) and "items" in _sd and "t1" not in _sd:
                _sd = {"t1": _sd}
            return _sd.get(_tk, {})
        except Exception:
            pass
    return {}


def _load_weekly_items(store: str, table_num: int = 1) -> list[str]:
    _v = _weekly_table_data(store, table_num).get("items", [])
    return (_v + [""] * _WEEKLY_N_ITEMS)[:_WEEKLY_N_ITEMS]


def _load_weekly_title(store: str, table_num: int = 1) -> str:
    return _weekly_table_data(store, table_num).get("title", "週間オススメ")


def _load_weekly_checks(store: str, table_num: int = 1) -> list[list[bool]]:
    _raw  = _weekly_table_data(store, table_num).get("checks", [])
    _rows = [(_r + [False] * 7)[:7] for _r in _raw]
    return (_rows + [[False] * 7] * _WEEKLY_N_ITEMS)[:_WEEKLY_N_ITEMS]


def _load_weekly_start_date(store: str, table_num: int = 1) -> str:
    return _weekly_table_data(store, table_num).get("start_date", "")


def _load_weekly_machine(store: str, table_num: int = 1) -> str:
    return _weekly_table_data(store, table_num).get("machine_name", "")

def _load_weekly_blank_days(store: str, table_num: int = 1) -> list[bool]:
    return _weekly_table_data(store, table_num).get("blank_days", [False] * 7)

def _load_weekly_blank_date_checks(store: str, table_num: int = 1) -> dict:
    return _weekly_table_data(store, table_num).get("blank_date_checks", {})


def _load_t3_cell_machines(store: str) -> dict:
    """t3（週間オススメ表②）の cell_machines: {"i,j": ["m1", ...]} を返す。"""
    return _weekly_table_data(store, 3).get("cell_machines", {})


def _load_weekly_date_checks(store: str, table_num: int = 2) -> dict:
    """date_checks dict: {date_iso: [bool×N_ITEMS]} を返す（月間オススメ表用）。
    date_checks が空の場合、checks + start_date から変換してフォールバック。"""
    import datetime as _dt_wdc
    _tdata = _weekly_table_data(store, table_num)
    _dc = _tdata.get("date_checks", {})
    if not _dc:
        _checks = _tdata.get("checks", [])
        _start_str = _tdata.get("start_date", "")
        if _checks and _start_str:
            try:
                _start = _dt_wdc.date.fromisoformat(_start_str)
                _dc = {}
                for _cj in range(7):
                    _d = _start + _dt_wdc.timedelta(days=_cj)
                    _dc[_d.isoformat()] = [
                        (_checks[_ci][_cj] if _ci < len(_checks) and _cj < len(_checks[_ci]) else False)
                        for _ci in range(_WEEKLY_N_ITEMS)
                    ]
            except Exception:
                pass
    return _dc


def _save_weekly_items(
    store: str, items: list[str], title: str | None = None,
    checks: list[list[bool]] | None = None, start_date: str | None = None,
    table_num: int = 1, machine_name: str | None = None,
    date_checks: dict | None = None,
    cell_machines: dict | None = None,
    blank_days: list[bool] | None = None,
    blank_date_checks: dict | None = None,
    monthly_start_date: str | None = None,
) -> None:
    _data: dict = {}
    if os.path.exists(_WEEKLY_SAVE_FILE):
        try:
            with open(_WEEKLY_SAVE_FILE, encoding="utf-8") as _f:
                _data = json.load(_f)
        except Exception:
            pass
    _sd = _data.get(store, {})
    # 旧フォーマット変換
    if isinstance(_sd, list):
        _sd = {"t1": {"items": _sd}}
    elif isinstance(_sd, dict) and "items" in _sd and "t1" not in _sd:
        _sd = {"t1": _sd}
    _tk  = f"t{table_num}"
    _cur = _sd.get(_tk, {})
    _cur["items"] = items
    if title is not None:
        _cur["title"] = title
    if checks is not None:
        _cur["checks"] = checks
    if start_date is not None:
        _cur["start_date"] = start_date
    if machine_name is not None:
        _cur["machine_name"] = machine_name
    if date_checks is not None:
        _cur["date_checks"] = date_checks
    if cell_machines is not None:
        _cur["cell_machines"] = cell_machines
    if blank_days is not None:
        _cur["blank_days"] = blank_days
    if blank_date_checks is not None:
        _cur["blank_date_checks"] = blank_date_checks
    if monthly_start_date is not None:
        _cur["monthly_start"] = monthly_start_date
    _sd[_tk]    = _cur
    _data[store] = _sd
    _weekly_json_str = json.dumps(_data, ensure_ascii=False, indent=2)
    with open(_WEEKLY_SAVE_FILE, "w", encoding="utf-8") as _f:
        _f.write(_weekly_json_str)
    if _IS_CLOUD:
        _ok, _msg = _github_push_file(_weekly_json_str)
        _log = ("✅ " if _ok else "❌ ") + _msg
        st.session_state["_github_sync_log"] = _log
        try:
            st.toast(_log, icon="✅" if _ok else "❌")
        except Exception:
            pass


def _weekly_table_html_image(
    items: list[str],
    date_labels: list[str],
    cell_machines,
    title: str = "週間オススメ",
) -> "Image.Image":
    """週間オススメ表② (cell_machinesモード) を PIL ImageDraw で生成する。
    Playwright不使用・Cloud/ローカル共通版。"""
    from PIL import ImageFont as _IFont
    import io as _io

    # フォントパス（MochiyPopOne + NotoSansJP フォールバック）
    _fp_m = os.path.join(_FONTS_DIR, "MochiyPopOne-Regular.ttf")
    _fp_n = os.path.join(_FONTS_DIR, "NotoSansJP-Regular.ttf")
    if not os.path.exists(_fp_m):
        _fp_m = r"C:\Users\23-3\AppData\Local\Microsoft\Windows\Fonts\MochiyPopOne-Regular.ttf"

    def _lf_m(sz):
        try:    return _IFont.truetype(_fp_m, sz)
        except: return _IFont.load_default()

    def _lf_n(sz):
        if os.path.exists(_fp_n):
            try:    return _IFont.truetype(_fp_n, sz)
            except: pass
        return _lf_m(sz)

    # レイアウト定数（CSS px × deviceScaleFactor=2 の物理ピクセル）
    SC          = 2
    _TH         = 44 * SC    # タイトルバー高さ (88px)
    _HH         = 38 * SC    # ヘッダー行高さ  (76px)
    _RH         = 40 * SC    # 機種スロット1つの高さ (80px)
    _IW_MIN     = 190 * SC   # アイテム列最小幅 (380px)
    _DW         = 130 * SC   # 日付列幅       (260px)
    _ITEM_PAD_X = 12 * SC    # アイテム列左右パディング

    # フォントサイズ（CSS px × SC）
    _TITLE_SZ   = 22 * SC   # 44px
    _HDR_SZ     = 12 * SC   # 24px
    _ITEM_SZ    = 15 * SC   # 30px
    _MACHINE_SZ = 10 * SC   # 20px

    # 色
    C_TITLE_BG = (0, 0, 0)
    C_TITLE_FG = (255, 255, 255)
    C_HDR_BG   = (208, 208, 208)
    C_ITEM_BG  = (247, 235, 203)
    C_CELL_YEL = (255, 255, 0)
    C_CELL_WHT = (255, 255, 255)
    C_BORDER   = (0, 0, 0)

    f_title   = _lf_m(_TITLE_SZ)
    f_hdr     = _lf_m(_HDR_SZ)
    f_item    = _lf_m(_ITEM_SZ)     # MochiyPopOne（アイテム本体）
    f_item_fb = _lf_n(_ITEM_SZ)     # NotoSansJP（◎等の記号フォールバック）
    f_machine = _lf_m(_MACHINE_SZ)

    # ◎○等 MochiyPopOneで文字化けする記号 → NotoSansJPで描画
    _SYM_CHARS = set("◎○●◯△▲▽▼□■◇◆★☆")

    def _item_font_for(char):
        return f_item_fb if char in _SYM_CHARS else f_item

    active = [(i, it) for i, it in enumerate(items) if it.strip()]
    n_days = len(date_labels)

    # ── アイテム列幅を実テキスト幅から動的決定 ──────────────────────
    _tmp_img  = Image.new("RGB", (10, 10))
    _tmp_draw = ImageDraw.Draw(_tmp_img)

    def _item_tw(text):
        """アイテムテキストの実幅（文字ごとフォント切り替え）"""
        total = 0
        for ch in text:
            fn = _item_font_for(ch)
            try:
                bb = _tmp_draw.textbbox((0, 0), ch, font=fn)
                total += bb[2] - bb[0]
            except Exception:
                total += _ITEM_SZ // 2
        return total

    _IW = _IW_MIN
    for _, item in active:
        for ln in [l.strip() for l in item.split("\n") if l.strip()]:
            _IW = max(_IW, _item_tw(ln) + 2 * _ITEM_PAD_X)

    # 行ごとの最大機種数 → 行高さ決定
    _row_max = []
    for _, (i, _) in enumerate(active):
        _n = max(
            (len(cell_machines[i][j]) for j in range(n_days)
             if i < len(cell_machines) and j < len(cell_machines[i])),
            default=1,
        )
        _row_max.append(max(1, _n))

    _TW      = _IW + _DW * n_days
    _total_h = _TH + _HH + sum(rm * _RH for rm in _row_max)

    img  = Image.new("RGB", (_TW, _total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    def _tsize(text, font):
        try:
            bb = draw.textbbox((0, 0), text, font=font)
            return bb[2] - bb[0], bb[3] - bb[1], bb
        except Exception:
            sz = getattr(font, "size", 10)
            return len(text) * sz // 2, sz, (0, 0, len(text) * sz // 2, sz)

    def _fill_rect(x, y, w, h, fill):
        draw.rectangle([x, y, x + w - 1, y + h - 1], fill=fill, outline=C_BORDER, width=1)

    def _centered(text, x, y, w, h, font, fg=(0, 0, 0)):
        tw, th, bb = _tsize(text, font)
        tx = x + (w - tw) // 2 - bb[0]
        ty = y + (h - th) // 2 - bb[1]
        draw.text((tx, ty), text, fill=fg, font=font)

    def _item_centered(text, x, y, w, h):
        """アイテムテキストを中央揃え（◎等は文字ごとにフォント切り替え）
        全文字で共通の y_origin（PILのdraw.textに渡すy）を使うことで
        文字ごとのbb[1]差による垂直ズレ（、が上に飛ぶ等）を防ぐ。
        """
        chars   = []
        total_tw = 0
        min_bb1  = 0    # テキスト全体の最上部（ascender）
        max_bb3  = 0    # テキスト全体の最下部（descender）

        for ch in text:
            fn = _item_font_for(ch)
            try:
                bb = draw.textbbox((0, 0), ch, font=fn)
            except Exception:
                bb = (0, 0, _ITEM_SZ // 2, _ITEM_SZ)
            cw = bb[2] - bb[0]
            chars.append((ch, fn, bb, cw))
            total_tw += cw
            min_bb1 = min(min_bb1, bb[1])
            max_bb3 = max(max_bb3, bb[3])

        text_h  = max_bb3 - min_bb1               # 文字列の実高さ
        text_top = y + (h - text_h) // 2          # セル内垂直中央の上端
        # PILの draw.text y は「テキスト原点（カーソルy）」
        # text_top = y_origin + min_bb1  →  y_origin = text_top - min_bb1
        y_origin = text_top - min_bb1

        cx = x + (w - total_tw) // 2
        for ch, fn, bb, cw in chars:
            # x方向のみbb[0]補正。y方向は全文字共通のy_originを使う
            draw.text((cx - bb[0], y_origin), ch, fill=(0, 0, 0), font=fn)
            cx += cw

    # ── タイトルバー ──
    _fill_rect(0, 0, _TW, _TH, C_TITLE_BG)
    _centered(title, 0, 0, _TW, _TH, f_title, C_TITLE_FG)

    # ── ヘッダー行 ──
    _fill_rect(0, _TH, _IW, _HH, C_HDR_BG)
    for j, dl in enumerate(date_labels):
        xj = _IW + j * _DW
        _fill_rect(xj, _TH, _DW, _HH, C_HDR_BG)
        lines = [l for l in dl.split("\n") if l]
        if len(lines) <= 1:
            _centered(dl, xj, _TH, _DW, _HH, f_hdr)
        else:
            lh = _HH // len(lines)
            for li, ln in enumerate(lines):
                _centered(ln, xj, _TH + li * lh, _DW, lh, f_hdr)

    # ── データ行 ──
    _y = _TH + _HH
    _SP = "対象台が"
    for ri, (i, item) in enumerate(active):
        row_h = _row_max[ri] * _RH

        # アイテム列（◎等を含む混在テキスト対応）
        _fill_rect(0, _y, _IW, row_h, C_ITEM_BG)
        item_lines = [l.strip() for l in item.split("\n") if l.strip()]
        if len(item_lines) == 1:
            _item_centered(item.strip(), 0, _y, _IW, row_h)
        else:
            lh = row_h // len(item_lines)
            for li, ln in enumerate(item_lines):
                _item_centered(ln, 0, _y + li * lh, _IW, lh)

        # 日付列
        for j in range(n_days):
            ms = (cell_machines[i][j]
                  if i < len(cell_machines) and j < len(cell_machines[i]) else [])
            xj = _IW + j * _DW
            if ms:
                slot_h_base = row_h // len(ms)
                for mi, m in enumerate(ms):
                    sy = _y + mi * slot_h_base
                    sh = slot_h_base if mi < len(ms) - 1 else row_h - mi * slot_h_base
                    _fill_rect(xj, sy, _DW, sh, C_CELL_YEL)
                    if mi > 0:
                        draw.line([(xj, sy), (xj + _DW - 1, sy)], fill=C_BORDER, width=1)
                    if m.startswith(_SP):
                        _centered(_SP,          xj, sy,           _DW, sh // 2,       f_machine)
                        _centered(m[len(_SP):], xj, sy + sh // 2, _DW, sh - sh // 2,  f_machine)
                    else:
                        _centered(m, xj, sy, _DW, sh, f_machine)
            else:
                _fill_rect(xj, _y, _DW, row_h, C_CELL_WHT)

        _y += row_h

    return img


def _draw_weekly_table_image(
    items: list[str], date_labels: list[str], checks: list[list[bool]],
    title: str = "週間オススメ",
    cell_machines=None,
) -> "Image.Image":
    """週間オススメ表をPIL画像で生成する。"""
    # cell_machinesモードはHTML+CSS+Playwright実装に委譲
    if cell_machines is not None:
        return _weekly_table_html_image(items, date_labels, cell_machines, title)
    SC          = 2
    TITLE_H     = int(44 * SC)
    HDR_H       = int(38 * SC)
    ROW_H       = int(40 * SC)
    DAY_W       = int(105 * SC) if cell_machines is not None else int(80 * SC)
    FONT_SZ     = int(15 * SC)
    MACHINE_SZ  = int(11 * SC)
    TEXT_LINE_H = FONT_SZ + int(4 * SC)   # 行間詰め用ライン高（cell_machinesモードのみ）
    TITLE_SZ    = int(22 * SC)
    SM_SZ       = int(12 * SC)
    PAD         = int(10 * SC)

    C_BK = "#000000"
    C_WH = "#FFFFFF"
    C_BG = "#F7EBCB"
    C_YL = "#FFFF00"
    C_GY = "#D0D0D0"

    # ◎○など記号だけ Meiryo、それ以外は MochiyPopOne（1文字単位で切り替え）
    _SYM_CHARS = set("◎○●◯△▲▽▼□■◇◆")
    _sym_cache: dict[int, "ImageFont.ImageFont"] = {}
    def _sym_font(size: int) -> "ImageFont.ImageFont":
        if size not in _sym_cache:
            # ◎○記号フォント: MochiyPopOneは○グリフ欠如のため除外、NotoSansJP → Windows フォールバック
            for _p, _i in [
                (os.path.join(_FONTS_DIR, "NotoSansJP-Regular.ttf"),   None),
                (r"C:\Windows\Fonts\YuGoth-M.ttc",  0),
                (r"C:\Windows\Fonts\YuGothB.ttc",   0),
                (r"C:\Windows\Fonts\yugothm.ttf",    None),
                (r"C:\Windows\Fonts\meiryo.ttc",     0),
                (r"C:\Windows\Fonts\msgothic.ttc",   0),
            ]:
                if not os.path.exists(_p):
                    continue
                try:
                    _sym_cache[size] = (ImageFont.truetype(_p, size) if _i is None
                                        else ImageFont.truetype(_p, size, index=_i))
                    break
                except Exception:
                    continue
            else:
                _sym_cache[size] = load_font(size)
        return _sym_cache[size]

    def _char_font(c: str, size: int) -> "ImageFont.ImageFont":
        return _sym_font(size) if c in _SYM_CHARS else load_font(size)

    # 測定用テンポラリ draw（ITEM_W 計算に使用）
    _td = ImageDraw.Draw(Image.new("RGB", (1, 1)))

    def _measure(text: str, size: int) -> int:
        return sum(_td.textbbox((0, 0), c, font=_char_font(c, size))[2]
                   - _td.textbbox((0, 0), c, font=_char_font(c, size))[0]
                   for c in text)

    active = [(i, it) for i, it in enumerate(items) if it.strip()]
    n_days = len(date_labels)

    # 項目列幅：改行がある場合は行ごとに計測して最大幅を使用
    max_item_w = max(
        (_measure(line, FONT_SZ) for _, it in active for line in it.split('\n') if line),
        default=0,
    )
    ITEM_W = int(max_item_w * 1.15) + PAD * 2  # 15% buffer for measurement underestimate

    # 行ごとの高さ（cell_machines使用時は機種数も考慮）
    if cell_machines is not None:
        _row_heights = []
        _row_nmach   = []   # 行ごとの最大機種数（センタリング計算に使用）
        for _, (i, it) in enumerate(active):
            _nel = [l for l in it.split('\n') if l.strip()]
            n_text = max(1, len(_nel))
            n_mach = max(
                (len(cell_machines[i][j]) for j in range(n_days)
                 if i < len(cell_machines) and j < len(cell_machines[i])),
                default=0,
            )
            n_mach_eff = max(1, n_mach)
            h_text = n_text * TEXT_LINE_H
            h_mach = n_mach_eff * ROW_H
            _row_heights.append(max(h_text, h_mach))
            _row_nmach.append(n_mach_eff)
    else:
        _row_heights = [ROW_H * max(1, len(it.split('\n'))) for _, it in active]
        _row_nmach   = [max(1, len(it.split('\n'))) for _, it in active]

    W = ITEM_W + DAY_W * n_days
    H = TITLE_H + HDR_H + sum(_row_heights)
    img = Image.new("RGB", (W, H), C_WH)
    d   = ImageDraw.Draw(img)

    def _draw_mixed(text: str, cx: int, cell_y: int, cell_h: int, size: int, fg: str, sym_stroke: int = 0) -> None:
        """1文字ずつフォントを切り替え、全文字共通のy_origin（ベースライン）で縦中央揃えして描画する。
        文字ごとにbb[1]で補正すると、、や◎などbb[1]が異なる文字でベースラインがズレるため、
        全文字のbb[1]/bb[3]を先に収集して共通y_originを算出する。"""
        if not text:
            return
        # 1st pass: 全文字のbb・幅を収集しテキスト全体の高さを計算
        chars_info = []
        min_bb1, max_bb3 = None, None
        for c in text:
            f  = _char_font(c, size)
            bb = d.textbbox((0, 0), c, font=f)
            tw = bb[2] - bb[0]
            chars_info.append((c, f, bb, tw))
            min_bb1 = bb[1] if min_bb1 is None else min(min_bb1, bb[1])
            max_bb3 = bb[3] if max_bb3 is None else max(max_bb3, bb[3])
        # 全体の高さでセル内垂直中央のy_origin（PILのdraw.text y引数）を算出
        text_h   = max_bb3 - min_bb1
        text_top = cell_y + (cell_h - text_h) // 2
        y_origin = text_top - min_bb1
        # 2nd pass: 全文字を共通y_originで描画
        for c, f, bb, tw in chars_info:
            if sym_stroke > 0 and c in _SYM_CHARS:
                d.text((cx - bb[0], y_origin), c, fill=fg, font=f,
                       stroke_width=sym_stroke, stroke_fill=fg)
            else:
                d.text((cx - bb[0], y_origin), c, fill=fg, font=f)
            cx += tw

    def _cell(x, y, w, h, bg, text="", fg=C_BK, size=None, sym_stroke: int = 0):
        d.rectangle([x, y, x + w - 1, y + h - 1], fill=bg, outline=C_BK)
        if not text:
            return
        sz      = size or FONT_SZ
        total_w = _measure(text, sz)
        tx      = x + (w - total_w) // 2
        _draw_mixed(text, tx, y, h, sz, fg, sym_stroke=sym_stroke)

    _cell(0, 0, W, TITLE_H, C_BK, title, C_WH, size=TITLE_SZ)
    cy = TITLE_H
    _cell(0, cy, ITEM_W, HDR_H, C_GY, "", C_BK)
    for j, dl in enumerate(date_labels):
        _cell(ITEM_W + j * DAY_W, cy, DAY_W, HDR_H, C_GY, dl, C_BK, size=SM_SZ)
    cy += HDR_H
    for _ri, (i, item) in enumerate(active):
        _rh    = _row_heights[_ri]
        _lines = item.split('\n')
        _nel   = [l for l in _lines if l.strip()]  # non-empty lines
        # 項目セル
        d.rectangle([0, cy, ITEM_W - 1, cy + _rh - 1], fill=C_BG, outline=C_BK)
        if cell_machines is not None:
            # 縦中央・行間詰め描画
            _tbh = len(_nel) * TEXT_LINE_H
            _ty0 = cy + (_rh - _tbh) // 2
            for _li, _line in enumerate(_nel):
                _lw = _measure(_line, FONT_SZ)
                _tx = (ITEM_W - _lw) // 2
                _draw_mixed(_line, _tx, _ty0 + _li * TEXT_LINE_H, TEXT_LINE_H, FONT_SZ, C_BK)
        else:
            for _li, _line in enumerate(_lines):
                if _line.strip():
                    _lw = _measure(_line, FONT_SZ)
                    _tx = (ITEM_W - _lw) // 2
                    _draw_mixed(_line, _tx, cy + _li * ROW_H, ROW_H, FONT_SZ, C_BK)
        if cell_machines is not None:
            # 機種名モード：縦中央揃えで描画
            # _row_nmach[_ri] = このタイトル行の最大機種数（全日付列の最大値）
            # 機種数が少ない列は上下に等しいパディングを入れて中央揃えにする
            _n_mach_max = _row_nmach[_ri]
            for j in range(n_days):
                _ms = (cell_machines[i][j]
                       if i < len(cell_machines) and j < len(cell_machines[i]) else [])
                _bg = C_YL if _ms else C_WH
                _x0 = ITEM_W + j * DAY_W
                _x1 = _x0 + DAY_W - 1
                d.rectangle([_x0, cy, _x1, cy + _rh - 1], fill=_bg, outline=C_BK)
                # 上余白 = (最大機種数 - この列の機種数) × ROW_H ÷ 2
                _n_ms       = len(_ms)
                _ms_top_pad = (_n_mach_max - _n_ms) * ROW_H // 2
                _ms_start   = cy + _ms_top_pad
                for _mi, _mname in enumerate(_ms):
                    _sub_y = _ms_start + _mi * ROW_H
                    if _mi > 0:
                        d.line([_x0 + 1, _sub_y, _x1 - 1, _sub_y], fill=C_BK, width=1)
                    _mw  = _measure(_mname, MACHINE_SZ)
                    _mtx = _x0 + (DAY_W - _mw) // 2
                    _draw_mixed(_mname, _mtx, _sub_y, ROW_H, MACHINE_SZ, C_BK)
        else:
            # チェックセルは行高全体に対して縦中央
            for j in range(n_days):
                ck = checks[i][j] if i < len(checks) and j < len(checks[i]) else False
                _cell(ITEM_W + j * DAY_W, cy, DAY_W, _rh, C_YL if ck else C_WH, "○" if ck else "", size=int(FONT_SZ * 1.3), sym_stroke=2)
        cy += _rh
    d.rectangle([0, 0, W - 1, H - 1], outline=C_BK, width=2)
    return img


def show_weekly_table_section(store: str, table_num: int = 1, excel_date=None) -> None:
    """週間オススメ表セクション（渋谷新館専用）。table_num=1 or 2 で2つ並べて使う。
    excel_date: t2 のとき渡すと過去7日間を自動表示（日付ピッカー非表示）。"""
    import datetime as _dt, io as _io

    _tn = table_num  # 短縮エイリアス
    _tname        = "週間オススメ表①" if _tn == 1 else ("週間オススメ表②" if _tn == 3 else ("月間オススメ表①" if _tn == 2 else ("月間オススメ表②" if _tn == 4 else "月間オススメ表③")))
    _default_title = "週間オススメ"   if _tn in (1, 3) else "月間オススメ"

    st.markdown("---")
    st.markdown(f"### 📅 {_tname}")

    # ── Cloud: GITHUB_TOKEN確認 & 同期ログ表示 ──
    if _IS_CLOUD and table_num in (1, 2, 3, 4, 5):
        _tok_ok, _tok_msg = _check_github_token()
        if _tok_ok:
            st.success(f"🔑 {_tok_msg}", icon="✅")
        else:
            st.error(f"🔑 {_tok_msg}", icon="❌")
        _sync_log = st.session_state.get("_github_sync_log")
        if _sync_log:
            if _sync_log.startswith("✅"):
                st.info(_sync_log)
            else:
                st.warning(_sync_log)

    # ── ① 機種名入力
    st.markdown("**① 機種名を入力**")
    _mname_key = f"weekly_machine_{store}_t{_tn}"
    if _mname_key not in st.session_state:
        st.session_state[_mname_key] = st.session_state.get(f"_weekly_init_machine_{store}_t{_tn}", "")

    def _on_machine_change():
        _save_weekly_items(
            store,
            [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
            title=st.session_state.get(f"weekly_title_{store}_t{_tn}", _default_title),
            machine_name=st.session_state.get(_mname_key, ""),
            table_num=_tn,
        )
        # _weekly_init_machine_ も更新（ナビゲーション離脱→復帰時にwidgetキーが消え古い値に戻るのを防ぐ）
        st.session_state[f"_weekly_init_machine_{store}_t{_tn}"] = st.session_state.get(_mname_key, "")

    st.text_input("機種名", value=st.session_state[_mname_key], key=_mname_key,
                  on_change=_on_machine_change)

    # ── ② タイトル入力
    st.markdown("**② タイトルを入力**")
    _title_key = f"weekly_title_{store}_t{_tn}"
    if _title_key not in st.session_state:
        st.session_state[_title_key] = st.session_state.get(f"_weekly_init_title_{store}_t{_tn}", _default_title)

    def _on_title_change():
        _save_weekly_items(
            store,
            [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
            title=st.session_state.get(_title_key, _default_title),
            table_num=_tn,
        )

    _title = st.text_input("画像タイトル", value=st.session_state[_title_key], key=_title_key,
                           on_change=_on_title_change)

    # ── ③ 項目入力（8個・2行4列）
    st.markdown(f"**③ 項目を入力（最大{_WEEKLY_N_ITEMS}つ）**")

    def _on_item_change(_idx: int):
        _save_weekly_items(
            store,
            [st.session_state.get(f"weekly_item_{store}_t{_tn}_{j}", "") for j in range(_WEEKLY_N_ITEMS)],
            title=st.session_state.get(f"weekly_title_{store}_t{_tn}", "週間オススメ"),
            table_num=_tn,
        )

    _icols_a = st.columns(4)
    _icols_b = st.columns(4)
    for _i, _col in enumerate(list(_icols_a) + list(_icols_b)):
        with _col:
            _init = st.session_state.get(f"_weekly_init_{store}_t{_tn}_{_i}", "")
            _key  = f"weekly_item_{store}_t{_tn}_{_i}"
            if _key not in st.session_state:
                st.session_state[_key] = _init
            if _tn == 3:
                st.text_area(f"項目{_i + 1}", value=st.session_state[_key], key=_key,
                             height=80, on_change=_on_item_change, args=(_i,))
            else:
                st.text_input(f"項目{_i + 1}", value=st.session_state[_key], key=_key,
                              on_change=_on_item_change, args=(_i,))

    _items = [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_i}", "") for _i in range(_WEEKLY_N_ITEMS)]

    # ── ③/④ 開始日
    _dow_names = ["月", "火", "水", "木", "金", "土", "日"]
    # 上野本館の月間オススメ表はExcel未アップロード時もdate_checks（日付キー）で読む
    _use_excel_date = _tn in (2, 4, 5) and (excel_date is not None or store == "上野本館")

    if _use_excel_date:
        st.markdown("**④ 表示期間**")
        if store == "上野本館":
            _ref = excel_date if excel_date is not None else _dt.date.today()
            # monthly_start（リセット日）が6日以内なら初日固定増列、7日超でスライド
            _ms_raw = _weekly_table_data(store, _tn).get("monthly_start")
            _ms_dt = None
            if _ms_raw:
                try:
                    _ms_dt = _dt.date.fromisoformat(_ms_raw)
                except Exception:
                    pass
            if _ms_dt and _ms_dt <= _ref and (_ref - _ms_dt).days <= 6:
                _start = _ms_dt
                _end   = _ref
            else:
                _start = _ref - _dt.timedelta(days=6)
                _end   = _ref
            _cap_suffix = "（Excelの日付より自動設定）" if excel_date is not None else "（今日の日付より自動設定）"
        else:
            # 渋谷新館：Excelの日付が最終日、過去7日間
            _start = excel_date - _dt.timedelta(days=6)
            _end   = excel_date
            _cap_suffix = "（Excelの日付より自動設定）"
        st.caption(
            f"📅 {_start.month}/{_start.day}({_dow_names[_start.weekday()]}) ～ "
            f"{_end.month}/{_end.day}({_dow_names[_end.weekday()]})"
            + _cap_suffix
        )
        if store == "上野本館":
            def _do_monthly_reset(_tn=_tn):
                _today = _dt.date.today()
                _mit = [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_mi}", "") for _mi in range(_WEEKLY_N_ITEMS)]
                _save_weekly_items(store, _mit, date_checks={}, blank_days=[False] * 7,
                                   monthly_start_date=_today.isoformat(), table_num=_tn)
                for _ci in range(_WEEKLY_N_ITEMS):
                    for _cj in range(7):
                        st.session_state.pop(f"weekly_ck_{store}_t{_tn}_{_ci}_{_cj}", None)
                for _cj in range(7):
                    st.session_state.pop(f"weekly_blank_{store}_t{_tn}_{_cj}", None)
                for _sk in [k for k in list(st.session_state.keys()) if k.startswith(f"_monthly_slide_{store}_t{_tn}_")]:
                    del st.session_state[_sk]
                st.session_state.pop(f"_weekly_prev_start_{store}_t{_tn}", None)
                st.session_state.pop(f"weekly_items_loaded_{store}", None)
            with st.expander("⚠️ リセット（誤操作注意）"):
                st.caption("この月間表のチェック・空欄設定をすべてクリアし、今日を初日として再スタートします。")
                st.button(f"🔄 月間表を今日からリセット", key=f"monthly_reset_{store}_t{_tn}",
                          on_click=_do_monthly_reset)
    else:
        st.markdown("**④ 開始日を選択**")
        _date_key = f"weekly_start_{store}_t{_tn}"
        if _date_key not in st.session_state:
            _saved_start_iso = st.session_state.get(f"_weekly_init_start_{store}_t{_tn}", "")
            if _saved_start_iso:
                try:
                    st.session_state[_date_key] = _dt.date.fromisoformat(_saved_start_iso)
                except Exception:
                    _td = _dt.date.today()
                    st.session_state[_date_key] = _td - _dt.timedelta(days=_td.weekday())
            else:
                _td = _dt.date.today()
                st.session_state[_date_key] = _td - _dt.timedelta(days=_td.weekday())

        def _on_date_change():
            _new_date_str = str(st.session_state.get(_date_key, ""))
            st.session_state[f"_weekly_init_start_{store}_t{_tn}"] = _new_date_str
            _save_weekly_items(
                store,
                [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
                title=st.session_state.get(f"weekly_title_{store}_t{_tn}", _default_title),
                start_date=_new_date_str,
                table_num=_tn,
            )

        _start = st.date_input("開始日", value=st.session_state[_date_key], key=_date_key, on_change=_on_date_change)

    _n_ui_dates = 7
    if _use_excel_date:
        try:
            _n_ui_dates = min(int((_end - _start).days) + 1, 7)
        except Exception:
            pass
    _dates   = [_start + _dt.timedelta(days=_j) for _j in range(_n_ui_dates)]
    _dlabels = [f"{_d.month}/{_d.day}({_dow_names[_d.weekday()]})" for _d in _dates]

    # 開始日変更検知
    _current_start_str = str(_start)
    _prev_start_key    = f"_weekly_prev_start_{store}_t{_tn}"
    _prev_start        = st.session_state.get(_prev_start_key)
    if _prev_start is not None and _prev_start != _current_start_str:
        # ウィジェットキーを削除して次回描画時に再初期化させる
        for _ci in range(_WEEKLY_N_ITEMS):
            for _cj in range(7):
                _ck_wk = f"weekly_ck_{store}_t{_tn}_{_ci}_{_cj}"
                if _ck_wk in st.session_state:
                    del st.session_state[_ck_wk]
                if _tn == 3:
                    _ms_wk = f"t3_ms_{store}_{_ci}_{_cj}"
                    if _ms_wk in st.session_state:
                        del st.session_state[_ms_wk]
        # blank widget キーを削除して次回描画時に日付キーで再初期化させる
        if _use_excel_date:
            for _cj in range(7):
                st.session_state.pop(f"weekly_blank_{store}_t{_tn}_{_cj}", None)
        if not _use_excel_date:
            if _tn == 3:
                # t3（週間オススメ表②）: 開始日変更で機種選択をリセット
                _save_weekly_items(
                    store,
                    [st.session_state.get(f"weekly_item_{store}_t3_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
                    table_num=3, cell_machines={},
                )
            else:
                # t1（週間）: 開始日変更でチェックをリセット
                for _ci in range(_WEEKLY_N_ITEMS):
                    for _cj in range(7):
                        st.session_state[f"_weekly_init_ck_{store}_t{_tn}_{_ci}_{_cj}"] = False
                _save_weekly_items(
                    store,
                    [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
                    title=st.session_state.get(f"weekly_title_{store}_t{_tn}", _default_title),
                    checks=[[False] * 7] * _WEEKLY_N_ITEMS,
                    start_date=_current_start_str,
                    table_num=_tn,
                )
    st.session_state[_prev_start_key] = _current_start_str

    # ── ④ 表入力
    _active = [it for it in _items if it.strip()]
    if not _active:
        st.info("項目を入力してください。")
        return

    st.markdown("**⑤ 表入力**")
    _ratio = [2] + [1] * len(_dlabels)
    _hdr = st.columns(_ratio)
    with _hdr[0]:
        st.markdown("**機種**")
    for _j, _dl in enumerate(_dlabels):
        with _hdr[_j + 1]:
            st.markdown(f"**{_dl}**")

    if _tn == 3:
        # ── t3（週間オススメ表②）: multiselect で機種名選択 ────────────
        _cands3 = _T3_JUGGLER_MACHINES
        _cm_dict3 = _load_t3_cell_machines(store)

        # ── 台番ルックアップ ─────────────────────────────────────────────
        _ban_map3 = st.session_state.get(f"ban_map_{store}", {})
        if _ban_map3:
            _bc1, _bc2 = st.columns([4, 3])
            with _bc1:
                st.text_input(
                    "台番で機種を調べる",
                    key=f"t3_ban_inp_{store}",
                    placeholder="例: 2028-2029, 2130-2132, 2157-2159",
                )
            _ban_val3 = st.session_state.get(f"t3_ban_inp_{store}", "").strip()
            with _bc2:
                if _ban_val3:
                    _found3: list[str] = []
                    _seen3: set[str] = set()
                    for _p3 in re.split(r"[,、\s　]+", _ban_val3):
                        _p3 = _p3.strip()
                        if not _p3:
                            continue
                        _mr3 = re.match(r"(\d+)[-–~～](\d+)$", _p3)
                        if _mr3:
                            for _b3 in range(int(_mr3.group(1)), int(_mr3.group(2)) + 1):
                                _mn3 = _ban_map3.get(_b3)
                                if _mn3 and _mn3 not in _seen3:
                                    _seen3.add(_mn3)
                                    _found3.append(_mn3)
                        else:
                            _bs3 = re.match(r"(\d+)$", _p3)
                            if _bs3:
                                _mn3 = _ban_map3.get(int(_bs3.group(1)))
                                if _mn3 and _mn3 not in _seen3:
                                    _seen3.add(_mn3)
                                    _found3.append(_mn3)
                    st.markdown("　")
                    if _found3:
                        st.success("→ " + " / ".join(_found3))
                    else:
                        st.warning("該当なし")

            # ── 台番でプラスを調べる ─────────────────────────────────────
            _diff_map3 = st.session_state.get(f"diff_map_{store}", {})
            if _diff_map3:
                _pc1, _pc2 = st.columns([4, 3])
                with _pc1:
                    st.text_input(
                        "台番でプラスを調べる",
                        key=f"t3_plus_inp_{store}",
                        placeholder="例: 2216.2195.2221.2183.2244",
                    )
                _plus_val3 = st.session_state.get(f"t3_plus_inp_{store}", "").strip()
                with _pc2:
                    if _plus_val3:
                        _plus_total = 0
                        _plus_count = 0
                        _plus_lines: list[str] = []
                        for _pp in re.split(r"[.,、,\s　]+", _plus_val3):
                            _pp = _pp.strip()
                            if not _pp or not re.match(r"^\d+$", _pp):
                                continue
                            _pb = int(_pp)
                            if _pb not in _diff_map3:
                                continue
                            _plus_total += 1
                            _pd = _diff_map3[_pb]
                            if _pd > 0:
                                _plus_count += 1
                                _plus_lines.append(f"{_pb}番台 +{_pd:,}枚")
                        st.markdown("　")
                        if _plus_total == 0:
                            st.warning("該当台番なし")
                        else:
                            st.info(f"**{_plus_count}/{_plus_total}台プラス**")
                            if _plus_lines:
                                st.markdown("  \n".join(_plus_lines))

        def _on_ms_save():
            _new_cm: dict = {}
            for _ci2 in range(_WEEKLY_N_ITEMS):
                for _cj2 in range(7):
                    _sel = st.session_state.get(f"t3_ms_{store}_{_ci2}_{_cj2}", [])
                    if _sel:
                        _new_cm[f"{_ci2},{_cj2}"] = list(_sel)
            _bdays3 = [st.session_state.get(f"weekly_blank_{store}_t{_tn}_{_j2}", False) for _j2 in range(7)]
            _save_weekly_items(
                store,
                [st.session_state.get(f"weekly_item_{store}_t3_{_j2}", "") for _j2 in range(_WEEKLY_N_ITEMS)],
                table_num=3, cell_machines=_new_cm, blank_days=_bdays3,
            )

        for _i3, _item3 in enumerate(_items):
            if not _item3.strip():
                continue
            _row3 = st.columns(_ratio)
            with _row3[0]:
                st.markdown(_item3.replace('\n', '  \n'))
            _is_special3 = any(k in _item3 for k in _T3_SPECIAL_ITEM_KEYS)
            for _j3 in range(7):
                with _row3[_j3 + 1]:
                    _ms_key = f"t3_ms_{store}_{_i3}_{_j3}"
                    _saved_sel3 = _cm_dict3.get(f"{_i3},{_j3}", [])
                    if _is_special3:
                        _opts3 = list(_T3_SPECIAL_OPTS) + [m for m in _saved_sel3 if m not in set(_T3_SPECIAL_OPTS)]
                        _max_sel3 = 1
                    else:
                        _opts3 = list(_cands3) + [m for m in _saved_sel3 if m not in set(_cands3)]
                        _max_sel3 = 9
                    if _ms_key not in st.session_state:
                        st.session_state[_ms_key] = _saved_sel3
                    st.multiselect(
                        "", options=_opts3,
                        key=_ms_key, label_visibility="collapsed",
                        max_selections=_max_sel3,
                        on_change=_on_ms_save,
                    )

        # ── 空欄にする行（t3）
        _blank_row3 = st.columns(_ratio)
        with _blank_row3[0]:
            st.markdown("**（空欄にする）**")
        for _j3b in range(7):
            with _blank_row3[_j3b + 1]:
                _bk_key3 = f"weekly_blank_{store}_t{_tn}_{_j3b}"
                if _bk_key3 not in st.session_state:
                    st.session_state[_bk_key3] = st.session_state.get(f"_weekly_init_blank_{store}_t{_tn}_{_j3b}", False)
                st.checkbox("", key=_bk_key3, label_visibility="collapsed", on_change=_on_ms_save)

        # ── t3 PNG出力
        st.markdown("---")
        if st.button(f"💾 {_tname}をPNGで保存", key=f"weekly_png_btn_{store}_t{_tn}"):
            _cm3 = _load_t3_cell_machines(store)
            _bdays3s = [st.session_state.get(f"weekly_blank_{store}_t{_tn}_{_j3s}", False) for _j3s in range(7)]
            _cm3_arr = [[
                [] if _bdays3s[_wj3] else [m for m in _cm3.get(f"{_wi3},{_wj3}", []) if m]
                for _wj3 in range(7)
            ] for _wi3 in range(_WEEKLY_N_ITEMS)]
            _out_labels3 = _dlabels
            _out_cm3 = _cm3_arr
            _wimg3 = _add_margin(_draw_weekly_table_image(
                _items, _out_labels3, [], title=_title or "週間オススメ", cell_machines=_out_cm3,
            ))
            _buf3 = _io.BytesIO()
            _wimg3.save(_buf3, format="PNG", dpi=(300, 300))
            _buf3.seek(0)
            st.download_button(
                label=f"⬇️ {_tname} PNG ダウンロード",
                data=_buf3, file_name=f"{_tname}.png", mime="image/png",
                key=f"weekly_dl_btn_{store}_t{_tn}",
            )

    else:
        # ── t1/t2: チェックボックス ──────────────────────────────────────
        def _on_ck_change():
            _bdays = [st.session_state.get(f"weekly_blank_{store}_t{_tn}_{_j2}", False) for _j2 in range(len(_dates))]
            if _use_excel_date:
                _dc = {
                    _dates[_cj2].isoformat(): [
                        st.session_state.get(f"weekly_ck_{store}_t{_tn}_{_ci2}_{_cj2}", False)
                        for _ci2 in range(_WEEKLY_N_ITEMS)
                    ]
                    for _cj2 in range(len(_dates))
                }
                _blank_dc = {_dates[_cj2].isoformat(): _bdays[_cj2] for _cj2 in range(len(_dates))}
                _save_weekly_items(
                    store,
                    [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
                    title=st.session_state.get(f"weekly_title_{store}_t{_tn}", _default_title),
                    date_checks=_dc, table_num=_tn, blank_date_checks=_blank_dc,
                )
            else:
                _save_weekly_items(
                    store,
                    [st.session_state.get(f"weekly_item_{store}_t{_tn}_{_j}", "") for _j in range(_WEEKLY_N_ITEMS)],
                    title=st.session_state.get(f"weekly_title_{store}_t{_tn}", _default_title),
                    checks=[[st.session_state.get(f"weekly_ck_{store}_t{_tn}_{_ci}_{_cj}", False) for _cj in range(7)] for _ci in range(_WEEKLY_N_ITEMS)],
                    start_date=_current_start_str, table_num=_tn, blank_days=_bdays,
                )

        _saved_start_str = st.session_state.get(f"_weekly_init_start_{store}_t{_tn}", "")
        _date_checks_data = _load_weekly_date_checks(store, _tn) if _use_excel_date else {}
        _checks: list[list[bool]] = []
        for _i, _item in enumerate(_items):
            if not _item.strip():
                _checks.append([False] * len(_dates))
                continue
            _row = st.columns(_ratio)
            with _row[0]:
                st.markdown(_item)
            _row_ck = []
            for _j in range(len(_dates)):
                with _row[_j + 1]:
                    _ck_key = f"weekly_ck_{store}_t{_tn}_{_i}_{_j}"
                    if _ck_key not in st.session_state:
                        if _use_excel_date:
                            _date_iso = _dates[_j].isoformat()
                            _dc_row = _date_checks_data.get(_date_iso, [False] * _WEEKLY_N_ITEMS)
                            _iv = _dc_row[_i] if _i < len(_dc_row) else False
                        else:
                            _iv = st.session_state.get(f"_weekly_init_ck_{store}_t{_tn}_{_i}_{_j}", False)
                            if _saved_start_str != _current_start_str:
                                _iv = False
                        st.session_state[_ck_key] = _iv
                    _row_ck.append(st.checkbox("", key=_ck_key, label_visibility="collapsed", on_change=_on_ck_change))
            _checks.append(_row_ck)

        # ── 空欄にする行（t1/t2）
        _blank_row = st.columns(_ratio)
        with _blank_row[0]:
            st.markdown("**（空欄にする）**")
        _blank_days = []
        _bdc_loaded = _load_weekly_blank_date_checks(store, _tn) if _use_excel_date else {}
        for _j in range(len(_dates)):
            with _blank_row[_j + 1]:
                _bk_key = f"weekly_blank_{store}_t{_tn}_{_j}"
                if _bk_key not in st.session_state:
                    if _use_excel_date:
                        st.session_state[_bk_key] = _bdc_loaded.get(_dates[_j].isoformat(), False)
                    else:
                        st.session_state[_bk_key] = st.session_state.get(f"_weekly_init_blank_{store}_t{_tn}_{_j}", False)
                _blank_days.append(st.checkbox("", key=_bk_key, label_visibility="collapsed", on_change=_on_ck_change))

        # ── t1/t2 PNG出力
        st.markdown("---")
        if st.button(f"💾 {_tname}をPNGで保存", key=f"weekly_png_btn_{store}_t{_tn}"):
            _bdays_save = [st.session_state.get(f"weekly_blank_{store}_t{_tn}_{_j}", False) for _j in range(len(_dates))]
            _checks_out = [
                [False if _bdays_save[_cj2] else _checks[_ci2][_cj2] for _cj2 in range(len(_dates))]
                for _ci2 in range(len(_checks))
            ]
            _last_col = -1
            for _cj in range(len(_dates)):
                if _bdays_save[_cj]:
                    _last_col = max(_last_col, _cj)
                else:
                    for _ci in range(len(_items)):
                        if _checks_out[_ci][_cj]:
                            _last_col = max(_last_col, _cj)
                            break
            if _last_col >= 0:
                _out_labels = _dlabels[: _last_col + 1]
                _out_checks = [_row[: _last_col + 1] for _row in _checks_out]
            else:
                # 全チェックなし → 初日のみ印なしで表示
                _out_labels = _dlabels[:1]
                _out_checks = [[False] for _ in _checks]
            _wimg = _add_margin(_draw_weekly_table_image(_items, _out_labels, _out_checks, title=_title or "週間オススメ"))
            _buf  = _io.BytesIO()
            _wimg.save(_buf, format="PNG", dpi=(300, 300))
            _buf.seek(0)
            st.download_button(
                label=f"⬇️ {_tname} PNG ダウンロード",
                data=_buf, file_name=f"{_tname}.png", mime="image/png",
                key=f"weekly_dl_btn_{store}_t{_tn}",
            )


def show_rote_page() -> None:
    """Excel一括処理(ローテ)ページ"""
    import io as _io
    store = st.session_state.selected_store

    st.markdown(f"## 【{store}】ローテ用")
    st.markdown("---")

    # 保存済み機種名をウィジェットキーとは別のキーに保持
    # （Streamlitはwidgetが描画されないrunでwidgetキーのsession_stateを削除するため）
    _rote_loaded_key = f"rote_machines_loaded_{store}"
    if not st.session_state.get(_rote_loaded_key):
        _saved = _load_rote_machines(store)
        for _i in range(6):
            st.session_state[f"_rote_init_{store}_1_{_i}"] = _saved["set1"][_i] if _i < len(_saved.get("set1", [])) else ""
            st.session_state[f"_rote_init_{store}_2_{_i}"] = _saved["set2"][_i] if _i < len(_saved.get("set2", [])) else ""
            st.session_state[f"_rote_init_{store}_3_{_i}"] = _saved["set3"][_i] if _i < len(_saved.get("set3", [])) else ""
        st.session_state[_rote_loaded_key] = True

    # 週間オススメ表の項目復元（渋谷新館・上野本館・非ウィジェットキーパターン）
    _wt_tn_list = (1, 2, 3) if store == "渋谷新館" else ((2, 4, 5) if store == "上野本館" else ())
    if _wt_tn_list:
        _wt_loaded_key = f"weekly_items_loaded_{store}"
        if not st.session_state.get(_wt_loaded_key):
            for _tn in _wt_tn_list:
                for _i, _v in enumerate(_load_weekly_items(store, _tn)):
                    st.session_state[f"_weekly_init_{store}_t{_tn}_{_i}"] = _v
                st.session_state[f"_weekly_init_title_{store}_t{_tn}"] = _load_weekly_title(store, _tn)
                _saved_cks = _load_weekly_checks(store, _tn)
                for _ci in range(_WEEKLY_N_ITEMS):
                    for _cj in range(7):
                        st.session_state[f"_weekly_init_ck_{store}_t{_tn}_{_ci}_{_cj}"] = (
                            _saved_cks[_ci][_cj] if _ci < len(_saved_cks) and _cj < len(_saved_cks[_ci]) else False
                        )
                st.session_state[f"_weekly_init_start_{store}_t{_tn}"] = _load_weekly_start_date(store, _tn)
                st.session_state[f"_weekly_init_machine_{store}_t{_tn}"] = _load_weekly_machine(store, _tn)
                _saved_blanks = _load_weekly_blank_days(store, _tn)
                for _cj in range(7):
                    st.session_state[f"_weekly_init_blank_{store}_t{_tn}_{_cj}"] = (
                        _saved_blanks[_cj] if _cj < len(_saved_blanks) else False
                    )
            st.session_state[_wt_loaded_key] = True
        else:
            # ロード済みでも JSON の monthly_start がセッション記録と異なれば blank_days を再読み込み
            for _tn in _wt_tn_list:
                _json_ms  = _weekly_table_data(store, _tn).get("monthly_start")
                _sess_ms  = st.session_state.get(f"_weekly_prev_start_{store}_t{_tn}")
                if _json_ms and _sess_ms and _json_ms != _sess_ms:
                    _fresh_blanks = _load_weekly_blank_days(store, _tn)
                    for _cj in range(7):
                        st.session_state[f"_weekly_init_blank_{store}_t{_tn}_{_cj}"] = (
                            _fresh_blanks[_cj] if _cj < len(_fresh_blanks) else False
                        )
                        st.session_state.pop(f"weekly_blank_{store}_t{_tn}_{_cj}", None)

    # ── ⓪ pision.io から日付データを自動取得 ──────────────────────────
    _rote_tb_uploaded = None
    if store in STORES:
        st.markdown(f"### 📈 日付からデータを自動取得（{store}）")
        _rote_api_key = _get_pision_api_key()
        if not _rote_api_key:
            st.caption("⚠️ PISION_API_KEY が未設定のため利用できません。手動でアップロードしてください。")
        else:
            _rote_mode = st.radio(
                "データ種別",
                ["確定データ", "速報データ（当日・営業中）"],
                horizontal=True,
                key=f"rote_tb_mode_{store}",
                help="確定データ＝前日まで（X-Api-Key）。速報データ＝当日の営業中データ（realtimeログインが必要）。",
            )
            _rote_is_rt = _rote_mode.startswith("速報")
            _rote_rt_ok = True
            if _rote_is_rt:
                _rote_rt_user, _rote_rt_pass = _get_pision_rt_credentials()
                if not _rote_rt_user or not _rote_rt_pass:
                    st.error("❌ 速報データには realtime のログイン情報が必要です。"
                             ".env に PISION_RT_USER / PISION_RT_PASS を設定してください。")
                    _rote_rt_ok = False

            if _rote_is_rt:
                import datetime as _dt_r
                _jst_r = _dt_r.timezone(_dt_r.timedelta(hours=9))
                _now_r = _dt_r.datetime.now(_jst_r)
                _rt_def_r = _now_r.date() if _now_r.hour >= 7 else _now_r.date() - _dt_r.timedelta(days=1)
                _rote_prev_mode_key = f"_rote_tb_prev_mode_{store}"
                _rote_cnt_key = f"_rote_tb_cnt_{store}"
                if st.session_state.get(_rote_prev_mode_key) != "rt":
                    st.session_state[_rote_cnt_key] = st.session_state.get(_rote_cnt_key, 0) + 1
                st.session_state[_rote_prev_mode_key] = "rt"
                _rote_cnt = st.session_state.get(_rote_cnt_key, 0)
                _rote_date = st.date_input(
                    "日付を選択（速報は営業日に合わせて選択・日付変更では自動取得しません）",
                    value=_rt_def_r,
                    key=f"rote_tb_date_rt_{store}_{_rote_cnt}",
                )
            else:
                st.session_state[f"_rote_tb_prev_mode_{store}"] = "fix"
                _rote_date = st.date_input(
                    "日付を選択",
                    value=datetime.date.today() - datetime.timedelta(days=1),
                    key=f"rote_tb_date_{store}",
                )
            _rote_date_str = _rote_date.strftime("%Y-%m-%d")

            _rote_mode_tag          = "rt" if _rote_is_rt else "fix"
            _rote_seen_key          = f"_rote_tb_seen_{_rote_mode_tag}_{store}"
            _rote_bytes_key         = f"_rote_tb_file_bytes_{_rote_mode_tag}_{store}"
            _rote_name_key          = f"_rote_tb_file_name_{_rote_mode_tag}_{store}"
            _rote_count_key         = f"_rote_tb_count_{_rote_mode_tag}_{store}"
            _rote_fetched_key       = f"_rote_tb_fetched_{_rote_mode_tag}_{store}"
            _rote_collecting_key    = f"_rote_tb_collecting_rt_{store}"
            _rote_rt_items_key      = f"_rote_tb_rt_items_{store}"
            _rote_rt_items_date_key = f"_rote_tb_rt_items_date_{store}"
            _rote_baseline_key      = f"_rote_tb_baseline_artid_{store}"

            _rote_is_collecting = _rote_is_rt and st.session_state.get(_rote_collecting_key) == _rote_date_str

            # ── ボタン描画 ─────────────────────────────────────────────
            _rote_do_rt_check    = False
            _rote_do_rt_existing = False
            if _rote_is_collecting:
                _btn_r1, _btn_r2 = st.columns(2)
                with _btn_r1:
                    _rote_refetch = st.button("⏳ 収集中...", key=f"rote_tb_refetch_{store}",
                                              disabled=True, use_container_width=True)
                with _btn_r2:
                    _rote_do_rt_check = st.button("🔍 今すぐ確認", key=f"rote_tb_rt_check_{store}",
                                                  use_container_width=True)
            elif _rote_is_rt and _rote_rt_ok:
                _btn_r1, _btn_r2 = st.columns(2)
                with _btn_r1:
                    _rote_refetch = st.button("⚡ 速報を取得", key=f"rote_tb_refetch_{store}",
                                              use_container_width=True, type="primary")
                with _btn_r2:
                    _rote_do_rt_existing = st.button("📂 既存のデータを取得", key=f"rote_tb_rt_existing_{store}",
                                                     use_container_width=True)
            else:
                _rote_refetch = st.button("🔄 取得", key=f"rote_tb_refetch_{store}")

            _rote_seen = st.session_state.get(_rote_seen_key)
            _rote_date_changed = (not _rote_is_rt) and _rote_seen is not None and _rote_seen != _rote_date_str
            st.session_state[_rote_seen_key] = _rote_date_str

            def _rote_save_rt_items(items: list) -> None:
                _slump_apply_names(items)
                _rows_r = [
                    {"台番": it.get("unitId"),
                     "機種名": it.get("_machineName") or it.get("_convertedName") or it.get("displayName") or "",
                     "差枚": it.get("diff", 0), "ゲーム数": it.get("games", 0),
                     "BB": it.get("bb", 0), "RB": it.get("rb", 0), "AT": it.get("art", 0)}
                    for it in items
                ]
                _df_r = pd.DataFrame(_rows_r)
                _buf_r = io.BytesIO()
                _df_r.to_excel(_buf_r, index=False)
                st.session_state[_rote_bytes_key]         = _buf_r.getvalue()
                st.session_state[_rote_name_key]          = f"{_rote_date.strftime('%Y%m%d')}_{store}_20S.xlsx"
                st.session_state[_rote_count_key]         = len(_df_r)
                st.session_state[_rote_fetched_key]       = _rote_date_str
                st.session_state[_rote_rt_items_key]      = items
                st.session_state[_rote_rt_items_date_key] = _rote_date_str
                st.session_state.pop(_rote_collecting_key, None)

            def _rote_is_new_artid(poll_result: dict) -> bool:
                _baseline = st.session_state.get(_rote_baseline_key)
                _new_id   = poll_result.get("article_id")
                if _baseline is None:
                    return True
                try:
                    return int(_new_id) > int(_baseline)
                except (TypeError, ValueError):
                    return False

            # 今すぐ確認
            if _rote_do_rt_check:
                with st.spinner("収集状況を確認中..."):
                    _chk_r = fetch_pision_realtime(store, _rote_date_str, trigger=False)
                if _chk_r["ok"] and _rote_is_new_artid(_chk_r):
                    _rote_save_rt_items(_chk_r["items"])
                    st.rerun()
                elif _chk_r["ok"]:
                    st.info("⏳ 収集はまだ完了していません（前回と同じスナップショット）。自動確認に戻ります。")
                    st.rerun()
                else:
                    st.rerun()

            # 自動ポーリング
            if _rote_is_collecting and not _rote_do_rt_check:
                import time as _time_r
                _ap_ph_r = st.empty()
                _ap_ph_r.info("⏳ 速報データを収集中... 30秒後に自動で確認します。")
                _time_r.sleep(30)
                _ap_ph_r.empty()
                with st.spinner("収集状況を自動確認中..."):
                    _poll_r = fetch_pision_realtime(store, _rote_date_str, trigger=False)
                if _poll_r["ok"] and _rote_is_new_artid(_poll_r):
                    _rote_save_rt_items(_poll_r["items"])
                    st.rerun()
                elif _poll_r.get("running") or (_poll_r["ok"] and not _rote_is_new_artid(_poll_r)):
                    st.rerun()
                else:
                    st.warning("⚠️ 収集が完了しましたがデータが取得できませんでした。「⚡ 速報を取得」をもう一度押してください。")
                    st.session_state.pop(_rote_collecting_key, None)
                    st.rerun()

            # 既存データ取得
            if _rote_do_rt_existing:
                with st.spinner("既存の速報データを確認中..."):
                    _exist_r = fetch_pision_realtime(store, _rote_date_str, trigger=False)
                if _exist_r["ok"]:
                    _rote_save_rt_items(_exist_r["items"])
                    st.rerun()
                else:
                    st.warning("⚠️ 既存の速報データが見つかりませんでした。「⚡ 速報を取得」で新しい収集を開始してください。")

            # 新規取得（確定 or 速報ボタン or 日付変更）
            if (_rote_date_changed or _rote_refetch) and _rote_rt_ok:
                with st.spinner(f"{_rote_date_str} のデータを取得中..."):
                    _rote_fetched_data = None
                    if _rote_is_rt:
                        _rt_r = fetch_pision_realtime(store, _rote_date_str)
                        if not _rt_r["ok"]:
                            st.error(f"❌ {_rt_r['error']}")
                            if _rt_r.get("collect_started"):
                                st.session_state[_rote_collecting_key] = _rote_date_str
                                st.session_state[_rote_baseline_key]   = None
                        else:
                            st.session_state[_rote_collecting_key] = _rote_date_str
                            st.session_state[_rote_baseline_key]   = _rt_r.get("article_id")
                    else:
                        st.session_state.pop(_rote_collecting_key, None)
                        st.session_state.pop(_rote_rt_items_key, None)
                        st.session_state.pop(_rote_rt_items_date_key, None)
                        try:
                            _rote_halls = fetch_pision_halls(_rote_api_key)
                        except Exception as _e_r:
                            st.error(f"❌ ホール一覧取得失敗: {_e_r}")
                            _rote_halls = []
                        _rote_hall_id = None
                        for _h_r in _rote_halls:
                            _hn_r = _h_r.get("name") or _h_r.get("displayName") or ""
                            if store in _hn_r and "エスパス" in _hn_r:
                                _rote_hall_id = str(_h_r.get("id") or _h_r.get("hallId") or "")
                                break
                        if _rote_hall_id is not None:
                            try:
                                _rote_fetched_data = fetch_pision_results(_rote_api_key, _rote_hall_id, _rote_date_str)
                            except Exception as _e_r:
                                st.error(f"❌ データ取得失敗: {_e_r}")
                    if not _rote_fetched_data:
                        st.session_state[_rote_bytes_key] = None
                    else:
                        _rote_rows2 = [
                            {"台番": _it.get("unitId"),
                             "機種名": _it.get("_machineName") or _it.get("_convertedName") or _it.get("displayName") or "",
                             "差枚": _it.get("diff", 0), "ゲーム数": _it.get("games", 0),
                             "BB": _it.get("bb", 0), "RB": _it.get("rb", 0), "AT": _it.get("art", 0)}
                            for _it in _rote_fetched_data
                        ]
                        _rote_df2  = pd.DataFrame(_rote_rows2)
                        _rote_buf2 = io.BytesIO()
                        _rote_df2.to_excel(_rote_buf2, index=False)
                        st.session_state[_rote_bytes_key] = _rote_buf2.getvalue()
                        st.session_state[_rote_name_key]  = f"{_rote_date.strftime('%Y%m%d')}_{store}_20S.xlsx"
                        st.session_state[_rote_count_key] = len(_rote_df2)
                    st.session_state[_rote_fetched_key] = _rote_date_str
                    st.rerun()

            if st.session_state.get(_rote_fetched_key) == _rote_date_str:
                _rote_tb_data = st.session_state.get(_rote_bytes_key)
                if _rote_tb_data:
                    _rote_tb_uploaded = io.BytesIO(_rote_tb_data)
                    _rote_tb_uploaded.name = st.session_state.get(
                        _rote_name_key, f"{_rote_date.strftime('%Y%m%d')}_{store}_20S.xlsx")
                    _rote_label = "速報" if _rote_is_rt else "確定"
                    st.success(f"✅ {_rote_date_str} の{_rote_label}データ"
                               f"（{st.session_state.get(_rote_count_key, '?')}台）を取得しました。")
                    # 機種別データ表示（機種名入力①②に絞り込み・サマリー非表示）
                    try:
                        _rv_df = pd.read_excel(io.BytesIO(_rote_tb_data))
                        _rv_df, _ = normalize_df(_rv_df)
                        _rv_df = apply_name_conversion(_rv_df)
                        # 入力①②で指定された機種名をセッションから収集
                        _rv_filter: set[str] = set()
                        for _rfi in range(6):
                            for _rset in ("1", "2"):
                                _rv_m = (
                                    st.session_state.get(f"rote{_rset}_mname_{_rfi}", "")
                                    or st.session_state.get(f"_rote_init_{store}_{_rset}_{_rfi}", "")
                                ).strip()
                                if _rv_m:
                                    _rv_filter.add(_rv_m)
                        if _rv_filter and "機種名" in _rv_df.columns:
                            _rv_df = _rv_df[_rv_df["機種名"].isin(_rv_filter)]
                        _rv_title = (f"{_rote_date.year}/{_rote_date.month}/{_rote_date.day}"
                                     f" エスパス{store}")
                        # summary=None でサマリーボックスを非表示にして機種テーブルのみ描画
                        if {"機種名", "差枚"} <= set(_rv_df.columns):
                            _rv_g = _rv_df.groupby("機種名", sort=False)
                            _rv_agg = pd.DataFrame({
                                "機種名":   list(_rv_g.groups.keys()),
                                "台数":     _rv_g["差枚"].size().values,
                                "勝台数":   _rv_g["差枚"].apply(lambda s: int((s > 0).sum())).values,
                                "総差枚":   _rv_g["差枚"].sum().astype(int).values,
                                "平均差枚": _rv_g["差枚"].mean().round().astype(int).values,
                                "平均G数":  (_rv_g["ゲーム数"].mean().round().astype(int).values
                                             if "ゲーム数" in _rv_df.columns else [0]*len(_rv_g)),
                            })
                            _rv_multi  = _rv_agg[_rv_agg["台数"] >= 2].sort_values("平均差枚", ascending=False)
                            _rv_single = _rv_agg[_rv_agg["台数"] == 1]
                            _rv_rows = [
                                (r["機種名"], int(r["台数"]), int(r["勝台数"]),
                                 int(r["総差枚"]), int(r["平均差枚"]), int(r["平均G数"]))
                                for _, r in _rv_multi.iterrows()
                            ]
                            if not _rv_single.empty:
                                _rvvn = int(_rv_single["台数"].sum())
                                _rvvw = int(_rv_single["勝台数"].sum())
                                _rvvd = int(_rv_single["総差枚"].sum())
                                _rvva = int(round(_rvvd / _rvvn)) if _rvvn else 0
                                _rvvg = int(round((_rv_single["平均G数"] * _rv_single["台数"]).sum() / _rvvn)) if _rvvn else 0
                                _rv_rows.append(("バラエティ", _rvvn, _rvvw, _rvvd, _rvva, _rvvg))
                            _rv_ucols = [c for c in ["台番", "機種名", "差枚", "BB", "RB", "AT", "ゲーム数"] if c in _rv_df.columns]
                            _rv_units = _rv_df[_rv_ucols].copy() if _rv_ucols else None
                            _rv_snames = set(_rv_single["機種名"].tolist()) if not _rv_single.empty else None
                            _rv_h = max(300, min(700, len(_rv_rows) * 42 + 180))
                            components.html(
                                _build_pision_interactive_html(_rv_title, None, _rv_rows, _rv_units, _rv_snames),
                                height=_rv_h, scrolling=True,
                            )
                    except Exception:
                        pass
                elif not _rote_is_collecting:
                    st.info(f"📭 {_rote_date_str} のデータを取得できませんでした"
                            "（404 / 未公開 / 店休日の可能性があります）。")

    st.markdown("---")

    uploaded = st.file_uploader(
        "Excelファイルをアップロード",
        type=["xlsx", "xls"],
        key="rote_excel_upload",
    )
    if uploaded is None and _rote_tb_uploaded is not None:
        uploaded = _rote_tb_uploaded
        st.caption(f"📈 自動取得データを使用中: `{uploaded.name}`（手動でアップロードすると優先されます）")

    # 台番→機種名マップをキャッシュ（週間オススメ表②の台番ルックアップ用）
    if uploaded is not None:
        _ban_src_key = f"ban_map_src_{store}"
        if st.session_state.get(_ban_src_key) != uploaded.name:
            try:
                _df_ban, _ = normalize_df(pd.read_excel(_io.BytesIO(uploaded.getvalue())))
                _nm_ban, _nm_norm_ban = load_name_map()
                if _nm_ban:
                    _df_ban, _ = _apply_map(_df_ban, _nm_ban, _nm_norm_ban)
                if "台番" in _df_ban.columns and "機種名" in _df_ban.columns:
                    st.session_state[f"ban_map_{store}"] = {
                        int(row["台番"]): row["機種名"]
                        for _, row in _df_ban.iterrows()
                        if pd.notna(row.get("台番")) and pd.notna(row.get("機種名"))
                    }
                    # 差枚マップ（台番でプラスを調べる用）
                    if "差枚" in _df_ban.columns:
                        st.session_state[f"diff_map_{store}"] = {
                            int(row["台番"]): int(row["差枚"])
                            for _, row in _df_ban.iterrows()
                            if pd.notna(row.get("台番")) and pd.notna(row.get("差枚"))
                        }
                else:
                    st.session_state[f"ban_map_{store}"] = {}
                    st.session_state[f"diff_map_{store}"] = {}
                st.session_state[_ban_src_key] = uploaded.name
            except Exception:
                pass

    if uploaded is None and store not in ("渋谷新館", "上野本館", "溝の口新館", "溝の口本館", "西武新宿", "新大久保"):
        st.info("Excelをアップロードすると機種名入力欄が表示されます。")
        st.markdown("---")
        if st.button("← 画像種類選択に戻る", key="rote_back_top"):
            _navigate("image_type")
        return

    _ban_map_vals = list(st.session_state.get(f"ban_map_{store}", {}).values())
    _rote_candidates = sorted(set(_ban_map_vals)) if _ban_map_vals else load_machine_candidates()

    # 機種名変更時に即座にJSONへ保存するコールバック
    def _on_rote_name_change():
        _m1 = [st.session_state.get(f"rote1_mname_{_i}", "") for _i in range(6)]
        _m2 = [st.session_state.get(f"rote2_mname_{_i}", "") for _i in range(6)]
        _save_rote_machines(store, _m1, _m2)
        # _rote_init_ も更新（ナビゲーション離脱→復帰時にwidgetキーが消え古い値に戻るのを防ぐ）
        for _i in range(6):
            st.session_state[f"_rote_init_{store}_1_{_i}"] = _m1[_i]
            st.session_state[f"_rote_init_{store}_2_{_i}"] = _m2[_i]

    # ── ① セット ────────────────────────────────────────────────────
    if store == "上野本館":
        # 上野本館：ローテ①の機種名は月間オススメ表①の機種名欄から取得
        _uo_m1 = st.session_state.get(f"weekly_machine_{store}_t2", "").strip()
        machine_inputs1: list[str] = [_uo_m1] if _uo_m1 else []
    else:
        st.markdown("**機種名を入力①（部分一致・最大6機種・入力順に表示）**")
        _r1a = st.columns(3)
        _r1b = st.columns(3)
        for _i, _col in enumerate(list(_r1a) + list(_r1b)):
            with _col:
                render_machine_autocomplete_input(f"機種名 {_i + 1}", f"rote1_mname_{_i}", _rote_candidates,
                                                  default=st.session_state.get(f"_rote_init_{store}_1_{_i}", ""),
                                                  on_change=_on_rote_name_change)
        machine_inputs1 = [st.session_state.get(f"rote1_mname_{_i}", "") for _i in range(6)]

    st.markdown("---")

    # ── ② セット ────────────────────────────────────────────────────
    if store == "渋谷新館":
        # 渋谷新館：ローテ②の機種名は週間オススメ表①の① 機種名欄から取得
        _mw1 = st.session_state.get(f"weekly_machine_{store}_t1", "").strip()
        machine_inputs2: list[str] = [_mw1] if _mw1 else []
    elif store == "上野本館":
        # 上野本館：ローテ②の機種名は月間オススメ表②の機種名欄から取得
        _uo_m2 = st.session_state.get(f"weekly_machine_{store}_t4", "").strip()
        machine_inputs2 = [_uo_m2] if _uo_m2 else []
    else:
        st.markdown("**機種名を入力②（部分一致・最大6機種・入力順に表示）**")
        _r2a = st.columns(3)
        _r2b = st.columns(3)
        for _i, _col in enumerate(list(_r2a) + list(_r2b)):
            with _col:
                render_machine_autocomplete_input(f"機種名 {_i + 1} ", f"rote2_mname_{_i}", _rote_candidates,
                                                  default=st.session_state.get(f"_rote_init_{store}_2_{_i}", ""),
                                                  on_change=_on_rote_name_change)
        machine_inputs2 = [st.session_state.get(f"rote2_mname_{_i}", "") for _i in range(6)]

    # 渋谷新館・上野本館のみ週間/月間オススメ表セクションを表示
    if store in ("渋谷新館", "上野本館"):
        import datetime as _dt_early, re as _re_early
        _rd_early = None
        if uploaded is not None:
            _m_early = _re_early.search(r"(\d{4})(\d{2})(\d{2})", uploaded.name)
            if _m_early:
                _rd_early = _dt_early.date(int(_m_early.group(1)), int(_m_early.group(2)), int(_m_early.group(3)))
        if store == "渋谷新館":
            show_weekly_table_section(store, table_num=1)
            show_weekly_table_section(store, table_num=3)
        show_weekly_table_section(store, table_num=2, excel_date=_rd_early)
        if store == "上野本館":
            with st.expander("📅 月間オススメ表②", expanded=False):
                show_weekly_table_section(store, table_num=4, excel_date=_rd_early)
            with st.expander("📅 月間オススメ表③", expanded=False):
                show_weekly_table_section(store, table_num=5, excel_date=_rd_early)

    if store == "渋谷新館":
        _m3 = st.session_state.get(f"weekly_machine_{store}_t2", "").strip()
        machine_inputs3: list[str] = [_m3] if _m3 else []
    elif store == "上野本館":
        _uo_m5_v = st.session_state.get(f"weekly_machine_{store}_t5", "").strip()
        machine_inputs3 = [_uo_m5_v] if _uo_m5_v else []
    else:
        machine_inputs3 = []

    st.markdown("---")
    # ファイル名（例: 20260427_店名_20S.xlsx）から日付・フォルダ情報を生成
    import datetime as _dt
    _rote_date_label = ""
    _rd = None
    _rote_out_dir  = None
    _rote_store_full = store
    if uploaded is not None:
        _m = re.search(r"(\d{4})(\d{2})(\d{2})", uploaded.name)
        if _m:
            _rd  = _dt.date(int(_m.group(1)), int(_m.group(2)), int(_m.group(3)))
            _dow = ["月", "火", "水", "木", "金", "土", "日"][_rd.weekday()]
            _rote_date_label = f"{_rd.month}/{_rd.day}({_dow})"
        _rote_stem    = os.path.splitext(uploaded.name)[0].replace("_20S", "")
        _rote_out_dir = tempfile.mkdtemp() if _IS_CLOUD else os.path.join(_DESKTOP, _rote_stem)
        _m_store      = re.match(r'^\d{8}_(.*)', _rote_stem)
        _rote_store_full = _m_store.group(1) if _m_store else store
        # ファイル名の店名に「エスパス」が付かない場合は補完（Cloud/ローカルで表記統一）
        if _rote_store_full and not _rote_store_full.startswith("エスパス"):
            _rote_store_full = f"エスパス{_rote_store_full}"
    else:
        st.info("Excelをアップロードすると画像を生成できます。")

    _rote_gen_clicked = st.button("🎰 画像を生成する", type="primary", use_container_width=True, key="rote_gen_btn",
                                   disabled=uploaded is None)
    # Cloud のみ：ボタン直下にZIPダウンロード用スロットを確保
    _zip_slot = st.empty() if _IS_CLOUD else None

    if _rote_gen_clicked:
        names1 = [n for n in machine_inputs1 if n.strip()]
        names2 = [n for n in machine_inputs2 if n.strip()]
        names3 = [n for n in machine_inputs3 if n.strip()]
        _uo_m2 = st.session_state.get(f"weekly_machine_{store}_t2", "").strip()
        _uo_m4 = st.session_state.get(f"weekly_machine_{store}_t4", "").strip()
        _uo_m5 = st.session_state.get(f"weekly_machine_{store}_t5", "").strip()
        _has_any = bool(names1 or names2 or names3 or (store == "上野本館" and (_uo_m2 or _uo_m4 or _uo_m5)))
        if uploaded is None or _rote_out_dir is None:
            st.warning("Excelをアップロードしてください。")
        elif not _has_any:
            st.warning("①か②いずれかに機種名を1つ以上入力してください。")
        else:
            st.session_state[f"rote_gen_saved_{store}"] = _rote_out_dir
            if not _IS_CLOUD:
                st.success(f"✅ `{_rote_out_dir}` に保存しました")
            # 上野本館の月間オススメ表スタート日を決定（初回=当日、以降は保存済み初日を継続）
            _uo_monthly_start = None
            if store == "上野本館" and _rd is not None:
                _saved_ms = _load_rote_machines(store).get("monthly_start")
                if _saved_ms:
                    try:
                        _saved_ms_dt = datetime.date.fromisoformat(_saved_ms)
                        if _saved_ms_dt <= _rd:
                            _diff_ms = (_rd - _saved_ms_dt).days
                            if _diff_ms <= 7:
                                _uo_monthly_start = _saved_ms_dt          # 8日以内：初日固定
                            else:
                                _uo_monthly_start = _rd - datetime.timedelta(days=6)  # 9日目以降：末尾7日スライド
                    except Exception:
                        pass
                if _uo_monthly_start is None:
                    _uo_monthly_start = _rd
            _save_rote_machines(store, machine_inputs1, machine_inputs2,
                                inputs3=machine_inputs3 if machine_inputs3 else None,
                                monthly_start=_uo_monthly_start.isoformat() if _uo_monthly_start else None)
            os.makedirs(_rote_out_dir, exist_ok=True)

            with st.spinner("データ読み込み中…"):
                df_raw = pd.read_excel(uploaded)
                df, _  = normalize_df(df_raw)
                nm, nm_norm = load_name_map()
                if nm:
                    df, _ = _apply_map(df, nm, nm_norm)
                if "差枚" in df.columns:
                    df["差枚"] = df["差枚"].apply(_pipeline_calc_d)

            # ── 画像生成 ─────────────────────────────────────────────
            with st.spinner("画像生成中…"):
                img1 = _add_margin(generate_rote_image(df, machine_inputs1, date_label=_rote_date_label, store=store)) if names1 else None
                img2 = _add_margin(generate_rote_image(df, machine_inputs2, date_label=_rote_date_label, store=store)) if names2 else None
                img3 = _add_margin(generate_rote_image(df, machine_inputs3, date_label=_rote_date_label, store=store)) if names3 else None
                # ランキング画像は新宿歌舞伎町のみ・①②それぞれ1枚
                if store == "新宿歌舞伎町":
                    ranking_img1 = generate_ranking_image(df, machine_inputs1, date_label=_rote_date_label, store=store) if names1 else None
                    ranking_img2 = generate_ranking_image(df, machine_inputs2, date_label=_rote_date_label, store=store) if names2 else None
                else:
                    ranking_img1 = None
                    ranking_img2 = None

            # ── フォルダへ保存 ────────────────────────────────────────
            if store == "新宿歌舞伎町":
                _r1_mac = next((m.strip() for m in machine_inputs1 if m.strip()), "")
                _r2_mac = next((m.strip() for m in machine_inputs2 if m.strip()), "")
                _r3_mac = next((m.strip() for m in machine_inputs3 if m.strip()), "")
            elif store in ("上野本館", "渋谷新館"):
                _r1_mac = next((m.strip() for m in machine_inputs1 if m.strip()), "")
                _r2_mac = machine_inputs2[0].strip() if machine_inputs2 else ""
                _r3_mac = machine_inputs3[0].strip() if machine_inputs3 else ""
            else:
                _r1_mac = ""
                _r2_mac = ""
                _r3_mac = ""
            if img1:
                img1.save(os.path.join(_rote_out_dir, f"{_r1_mac}ローテ.png" if _r1_mac else "ローテ①.png"), format="PNG", dpi=(300, 300))
            if img2:
                img2.save(os.path.join(_rote_out_dir, f"{_r2_mac}ローテ.png" if _r2_mac else "ローテ②.png"), format="PNG", dpi=(300, 300))
            if img3:
                img3.save(os.path.join(_rote_out_dir, f"{_r3_mac}ローテ.png" if _r3_mac else "ローテ③.png"), format="PNG", dpi=(300, 300))
            if ranking_img1:
                ranking_img1.save(os.path.join(_rote_out_dir, f"ranking_{_r1_mac}ローテ.png" if _r1_mac else "ranking_ローテ①.png"), format="PNG", dpi=(300, 300))
            if ranking_img2:
                ranking_img2.save(os.path.join(_rote_out_dir, f"ranking_{_r2_mac}ローテ.png" if _r2_mac else "ranking_ローテ②.png"), format="PNG", dpi=(300, 300))

            # ── 週間/月間オススメ表の保存（渋谷新館・上野本館）──────────────
            if store in ("渋谷新館", "上野本館"):
                _wt_save_list = (1, 2, 3) if store == "渋谷新館" else (2, 4, 5)
                for _wtn in _wt_save_list:
                    # 月間オススメ表②（t4）・③（t5）は機種名・項目がすべて空ならスキップ
                    if _wtn in (4, 5):
                        _tN_mac = st.session_state.get(f"weekly_machine_{store}_t{_wtn}", "").strip()
                        _tN_its = [st.session_state.get(f"weekly_item_{store}_t{_wtn}_{_wi}", "").strip() for _wi in range(_WEEKLY_N_ITEMS)]
                        if not _tN_mac and not any(_tN_its):
                            continue
                    _wt_items  = [st.session_state.get(f"weekly_item_{store}_t{_wtn}_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    _wt_title  = st.session_state.get(f"weekly_title_{store}_t{_wtn}", "週間オススメ")
                    # 月間オススメ表はExcel日付から自動設定
                    # Excel日付が最終日、過去7日間（上野本館・渋谷新館共通）
                    if _wtn in (2, 4, 5):
                        _rd_ref = _rd if _rd is not None else datetime.date.today()
                        _ms_gen_raw = _weekly_table_data(store, _wtn).get("monthly_start")
                        _ms_gen_dt = None
                        if _ms_gen_raw:
                            try:
                                _ms_gen_dt = datetime.date.fromisoformat(_ms_gen_raw)
                            except Exception:
                                pass
                        if _ms_gen_dt and _ms_gen_dt <= _rd_ref and (_rd_ref - _ms_gen_dt).days <= 6:
                            _wt_start  = _ms_gen_dt
                            _wt_n_cols = (_rd_ref - _ms_gen_dt).days + 1
                        else:
                            _wt_start  = _rd_ref - datetime.timedelta(days=6)
                            _wt_n_cols = 7
                    else:
                        _wt_start  = st.session_state.get(f"weekly_start_{store}_t{_wtn}")
                        _wt_n_cols = 7
                    if _wt_start:
                        _wt_dates   = [_wt_start + datetime.timedelta(days=_wj) for _wj in range(_wt_n_cols)]
                        _wt_dow     = ["月", "火", "水", "木", "金", "土", "日"]
                        _wt_dlabels = [f"{_d.month}/{_d.day}({_wt_dow[_d.weekday()]})" for _d in _wt_dates]
                        _wt_default_title = "週間オススメ" if _wtn in (1, 3) else "月間オススメ"
                        if _wtn == 3:
                            # 週間オススメ表②: 機種名モード
                            _cm3r = _load_t3_cell_machines(store)
                            _wt_blank3 = _load_weekly_blank_days(store, 3)
                            _cm3_arr = [[
                                [] if _wt_blank3[_wj] else [m for m in _cm3r.get(f"{_wi},{_wj}", []) if m]
                                for _wj in range(7)
                            ] for _wi in range(_WEEKLY_N_ITEMS)]
                            _wt_last_col = -1
                            for _wi in range(_WEEKLY_N_ITEMS):
                                for _wj in range(7):
                                    if _cm3_arr[_wi][_wj] or _wt_blank3[_wj]:
                                        _wt_last_col = max(_wt_last_col, _wj)
                            if _wt_last_col >= 0:
                                _wt_dlabels = _wt_dlabels[:_wt_last_col + 1]
                                _wt_cm3 = [[_cm3_arr[_wi][_wj] for _wj in range(_wt_last_col + 1)] for _wi in range(_WEEKLY_N_ITEMS)]
                            else:
                                _wt_cm3 = _cm3_arr
                            _wt_img = _add_margin(_draw_weekly_table_image(
                                _wt_items, _wt_dlabels, [],
                                title=_wt_title or _wt_default_title,
                                cell_machines=_wt_cm3,
                            ))
                        else:
                            _wt_checks  = [
                                [st.session_state.get(f"weekly_ck_{store}_t{_wtn}_{_wi}_{_wj}", False) for _wj in range(_wt_n_cols)]
                                for _wi in range(_WEEKLY_N_ITEMS)
                            ]
                            if store == "上野本館" and _wtn in (2, 4, 5):
                                _wt_blank_dc = _load_weekly_blank_date_checks(store, _wtn)
                                _wt_blank = [_wt_blank_dc.get((_wt_start + datetime.timedelta(days=_wj)).isoformat(), False) for _wj in range(_wt_n_cols)]
                            else:
                                _wt_blank = _load_weekly_blank_days(store, _wtn)
                            _wt_checks = [
                                [False if _wt_blank[_wj] else _wt_checks[_wi][_wj] for _wj in range(_wt_n_cols)]
                                for _wi in range(_WEEKLY_N_ITEMS)
                            ]
                            # 上野本館の月間表は全列を常に表示（日付が増えていく仕様）
                            _uo_monthly_no_trim = (store == "上野本館" and _wtn in (2, 4, 5))
                            if not _uo_monthly_no_trim:
                                _wt_last_col = -1
                                for _wj in range(_wt_n_cols):
                                    if _wt_blank[_wj]:
                                        _wt_last_col = max(_wt_last_col, _wj)
                                    else:
                                        for _wi in range(_WEEKLY_N_ITEMS):
                                            if _wt_checks[_wi][_wj]:
                                                _wt_last_col = max(_wt_last_col, _wj)
                                                break
                                if _wt_last_col >= 0:
                                    _wt_dlabels = _wt_dlabels[:_wt_last_col + 1]
                                    _wt_checks  = [_r[:_wt_last_col + 1] for _r in _wt_checks]
                            _wt_img = _add_margin(_draw_weekly_table_image(
                                _wt_items, _wt_dlabels, _wt_checks,
                                title=_wt_title or _wt_default_title,
                            ))
                        if _wtn == 1:
                            _wt1_mac = st.session_state.get(f"weekly_machine_{store}_t1", "").strip()
                            _fname = f"{_wt1_mac}表.png" if _wt1_mac else "週間オススメ表①.png"
                        elif _wtn == 3:
                            _wt3_mac = st.session_state.get(f"weekly_machine_{store}_t3", "").strip()
                            _fname = f"{_wt3_mac}表.png" if _wt3_mac else "週間オススメ表②.png"
                        elif _wtn == 4:
                            _wt4_mac = st.session_state.get(f"weekly_machine_{store}_t4", "").strip()
                            _fname = f"{_wt4_mac}表.png" if _wt4_mac else "月間オススメ表②.png"
                        elif _wtn == 5:
                            _wt5_mac = st.session_state.get(f"weekly_machine_{store}_t5", "").strip()
                            _fname = f"{_wt5_mac}表.png" if _wt5_mac else "月間オススメ表③.png"
                        elif store == "上野本館":
                            _wt2_mac = st.session_state.get(f"weekly_machine_{store}_t2", "").strip()
                            _fname = f"{_wt2_mac}表.png" if _wt2_mac else "月間オススメ表①.png"
                        else:
                            _wt2_mac = st.session_state.get(f"weekly_machine_{store}_t2", "").strip()
                            _fname = f"{_wt2_mac}表.png" if _wt2_mac else "月間オススメ表.png"
                        _wt_img.save(os.path.join(_rote_out_dir, _fname), format="PNG", dpi=(300, 300))

            # ── 結果テキスト生成・保存 ────────────────────────────────
            _rote_result  = ""
            _rote_result2 = ""
            _rote_result3 = ""
            _rote_result4 = ""
            if _rd is not None:
                if store == "渋谷新館":
                    _wt_items1 = [st.session_state.get(f"weekly_item_{store}_t1_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    _wt_items2 = [st.session_state.get(f"weekly_item_{store}_t2_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    _wt_items3 = [st.session_state.get(f"weekly_item_{store}_t3_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    _rote_result, _rote_result2, _rote_result3, _rote_result4 = _generate_shibuyashinkan_result_texts(
                        df, machine_inputs1, machine_inputs2, _rd, _rote_store_full,
                        weekly_items=_wt_items1,
                        machine_inputs3=machine_inputs3,
                        monthly_items=_wt_items2,
                        weekly_items2=_wt_items3,
                    )
                    _sh_r1_fn = f"{_r1_mac}結果.txt" if _r1_mac else "ローテ①結果.txt"
                    _sh_r2_fn = f"{_r2_mac}結果.txt" if _r2_mac else "ローテ②結果.txt"
                    _sh_r3_fn = f"{_r3_mac}結果.txt" if _r3_mac else "ローテ③結果.txt"
                    with open(os.path.join(_rote_out_dir, _sh_r1_fn), "w", encoding="utf-8") as _f:
                        _f.write(_rote_result)
                    with open(os.path.join(_rote_out_dir, _sh_r2_fn), "w", encoding="utf-8") as _f:
                        _f.write(_rote_result2)
                    with open(os.path.join(_rote_out_dir, _sh_r3_fn), "w", encoding="utf-8") as _f:
                        _f.write(_rote_result3)
                    _sh_r4_mac = st.session_state.get(f"weekly_machine_{store}_t3", "").strip()
                    _sh_r4_fn = f"{_sh_r4_mac}結果.txt" if _sh_r4_mac else "週間オススメ②結果.txt"
                    with open(os.path.join(_rote_out_dir, _sh_r4_fn), "w", encoding="utf-8") as _f:
                        _f.write(_rote_result4)
                elif store == "上野本館":
                    _re_uo, _te_uo = ROTE_EMOJI_CONFIG.get(store, ("🌌", "🔥"))
                    _dow_uo = ["月", "火", "水", "木", "金", "土", "日"][_rd.weekday()]
                    _header_uo = f"{_rd.month}/{_rd.day}({_dow_uo})👨‍💻結果👨‍💻\n{_rote_store_full}"
                    _nc_uo = "機種名" if "機種名" in df.columns else None
                    _tiers_uo = [
                        (10000, None,  f"{_te_uo}10,000枚超{_te_uo}"),
                        (5000,  10000, f"{_te_uo}5,000枚超{_te_uo}"),
                        (3000,  5000,  f"{_te_uo}3,000枚超{_te_uo}"),
                        (1000,  3000,  f"{_te_uo}1,000枚超{_te_uo}"),
                    ]
                    def _uo_monthly_text(machine_input, monthly_items, re_emoji=None, poster_text=None):
                        _re = re_emoji if re_emoji else _re_uo
                        _pt = poster_text if poster_text else "月間オススメポスター"
                        _nm = (machine_input or "").strip()
                        _lm = [_header_uo, ""]
                        if _nm:
                            _lm.append(f"{_re}{_nm}{_re}")
                        _lm += [f"{_re}{_pt}{_re}", ""]
                        _lm.append("✅毎日何かしらの仕掛けアリ!?")
                        for _it in (monthly_items or []):
                            if (_it or "").strip():
                                _lm.append(f"📍{_it.strip()}")
                        _lm.append("")
                        if _nm and _nc_uo:
                            _sm = df[df[_nc_uo].astype(str).str.contains(_nm, na=False)].copy()
                            if not _sm.empty:
                                _sm["差枚"] = pd.to_numeric(_sm["差枚"], errors="coerce").fillna(0).astype(int)
                                _sm["台番"] = pd.to_numeric(_sm["台番"], errors="coerce").fillna(0).astype(int)
                                _sm = _sm[_sm["差枚"] >= 1000]
                                for _lo, _hi, _lb in _tiers_uo:
                                    _tr = (_sm[_sm["差枚"] >= _lo] if _hi is None
                                           else _sm[(_sm["差枚"] >= _lo) & (_sm["差枚"] < _hi)])
                                    _tr = _tr.sort_values(["差枚", "台番"], ascending=[False, True])
                                    if _tr.empty:
                                        continue
                                    _lm.append(_lb)
                                    for _, _row in _tr.iterrows():
                                        _lm.append(f"💎{int(_row['台番'])}番台")
                                    _lm.append("")
                        return "\n".join(_lm).rstrip()
                    # 月間オススメ表結果.txtと重複するためローテ①②結果.txtは生成しない
                    # 月間オススメ表①結果.txt（🗼）
                    _uo_m2 = st.session_state.get(f"weekly_machine_{store}_t2", "")
                    _uo_items2 = [st.session_state.get(f"weekly_item_{store}_t2_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    _uo_t2_fn = f"{_uo_m2.strip()}結果.txt" if _uo_m2.strip() else "月間オススメ表①結果.txt"
                    with open(os.path.join(_rote_out_dir, _uo_t2_fn), "w", encoding="utf-8") as _f:
                        _f.write(_uo_monthly_text(_uo_m2, _uo_items2, re_emoji="🤡", poster_text="毎日オススメポスター"))
                    # 月間オススメ表②結果.txt（🚂）：機種名・項目いずれかが入力済みの場合のみ生成
                    _uo_m4 = st.session_state.get(f"weekly_machine_{store}_t4", "")
                    _uo_items4 = [st.session_state.get(f"weekly_item_{store}_t4_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    if _uo_m4.strip() or any(i.strip() for i in _uo_items4):
                        _uo_t4_fn = f"{_uo_m4.strip()}結果.txt" if _uo_m4.strip() else "月間オススメ表②結果.txt"
                        with open(os.path.join(_rote_out_dir, _uo_t4_fn), "w", encoding="utf-8") as _f:
                            _f.write(_uo_monthly_text(_uo_m4, _uo_items4, re_emoji="🚂"))
                    # 月間オススメ表③結果.txt（👊）：機種名・項目いずれかが入力済みの場合のみ生成
                    _uo_m5_txt = st.session_state.get(f"weekly_machine_{store}_t5", "")
                    _uo_items5 = [st.session_state.get(f"weekly_item_{store}_t5_{_wi}", "") for _wi in range(_WEEKLY_N_ITEMS)]
                    if _uo_m5_txt.strip() or any(i.strip() for i in _uo_items5):
                        _uo_t5_fn = f"{_uo_m5_txt.strip()}結果.txt" if _uo_m5_txt.strip() else "月間オススメ表③結果.txt"
                        with open(os.path.join(_rote_out_dir, _uo_t5_fn), "w", encoding="utf-8") as _f:
                            _f.write(_uo_monthly_text(_uo_m5_txt, _uo_items5, re_emoji="👊"))
                else:
                    all_inputs   = machine_inputs1 + machine_inputs2
                    _rote_result = _generate_rote_result_text(df, all_inputs, _rd, _rote_store_full, store=store)
                    _txt_path    = os.path.join(_rote_out_dir, "結果テキスト.txt")
                    with open(_txt_path, "w", encoding="utf-8") as _f:
                        _f.write(_rote_result)

            # ── 横並びプレビュー ─────────────────────────────────────
            _col1, _col2 = st.columns(2)
            if img1:
                with _col1:
                    st.markdown("#### ①")
                    st.image(img1, use_container_width=True)
                    buf1 = _io.BytesIO()
                    img1.save(buf1, format="PNG", dpi=(300, 300))
                    buf1.seek(0)
                    st.download_button(
                        label="⬇️ ①PNG をダウンロード",
                        data=buf1,
                        file_name="ローテ①.png",
                        mime="image/png",
                        key="rote_dl_btn1",
                    )
            if img2:
                with _col2:
                    st.markdown("#### ②")
                    st.image(img2, use_container_width=True)
                    buf2 = _io.BytesIO()
                    img2.save(buf2, format="PNG", dpi=(300, 300))
                    buf2.seek(0)
                    st.download_button(
                        label="⬇️ ②PNG をダウンロード",
                        data=buf2,
                        file_name="ローテ②.png",
                        mime="image/png",
                        key="rote_dl_btn2",
                    )
            if img3:
                _col3, _ = st.columns(2)
                with _col3:
                    st.markdown("#### ③")
                    st.image(img3, use_container_width=True)
                    buf3 = _io.BytesIO()
                    img3.save(buf3, format="PNG", dpi=(300, 300))
                    buf3.seek(0)
                    st.download_button(
                        label="⬇️ ③PNG をダウンロード",
                        data=buf3,
                        file_name="ローテ③.png",
                        mime="image/png",
                        key="rote_dl_btn3",
                    )

            # ── 結果テキスト表示 ──────────────────────────────────────
            if _rote_result:
                st.markdown("---")
                if store == "渋谷新館":
                    st.markdown("### ローテ①結果")
                    st.text_area("", value=_rote_result, height=300, key="rote_result_area")
                    if _rote_result2:
                        st.markdown("### ローテ②結果")
                        st.text_area("", value=_rote_result2, height=400, key="rote_result_area2")
                    if _rote_result3:
                        st.markdown("### ローテ③結果")
                        st.text_area("", value=_rote_result3, height=400, key="rote_result_area3")
                    if _rote_result4:
                        st.markdown("### 週間オススメ②結果")
                        st.text_area("", value=_rote_result4, height=300, key="rote_result_area4")
                else:
                    st.markdown("### 結果テキスト")
                    st.text_area("", value=_rote_result, height=400, key="rote_result_area")

            # ── ZIP データをセッションに保存（ボタン直下スロットへ後で表示）──
            if _IS_CLOUD and _rote_out_dir and os.path.isdir(_rote_out_dir):
                try:
                    st.session_state[f"_rote_zip_data_{store}"] = _make_zip_bytes(_rote_out_dir)
                    st.session_state[f"_rote_zip_stem_{store}"] = _rote_stem
                except Exception as _rze:
                    st.warning(f"ZIP生成に失敗: {_rze}")

            # ── ローカルのみ: 設定ファイルを自動push ──
            if not _IS_CLOUD:
                _push_ok, _push_msg = _git_auto_push("ローテ画像生成")
                if _push_ok:
                    st.info(f"🔄 {_push_msg}")
                else:
                    st.warning(f"⚠️ {_push_msg}")

    if not st.session_state.get("rote_gen_btn", False):
        _rote_saved = st.session_state.get(f"rote_gen_saved_{store}")
        if _rote_saved and not _IS_CLOUD:
            st.success(f"✅ `{_rote_saved}` に保存しました")

    # ── ボタン直下スロットにZIPダウンロードボタンを表示（Cloud のみ）──
    if _IS_CLOUD and _zip_slot is not None and st.session_state.get(f"_rote_zip_data_{store}"):
        with _zip_slot.container():
            st.success("✅ 処理が完了しました。ZIPをダウンロードしてください。")
            st.download_button(
                label="📥 画像・テキストをZIPでダウンロード",
                data=st.session_state[f"_rote_zip_data_{store}"],
                file_name=f"{st.session_state.get(f'_rote_zip_stem_{store}', 'rote')}.zip",
                mime="application/zip",
                key="rote_zip_dl",
                type="primary",
            )

    st.markdown("---")
    if st.button("← 画像種類選択に戻る", key="rote_back_bottom"):
        _navigate("image_type")


# =============================================================================
# ■ 週間結果テキスト抽出ページ
# =============================================================================

def _wrt_build_machine_block(df: pd.DataFrame, machine: str, machine_prefix: str = "") -> "str | None":
    """機種名部分一致 + 差枚>=1000 の台番ブロックを生成。該当なしは None。
    全角/半角・スペース差を _normalize_key で吸収してから照合する。"""
    _nq = _normalize_key(machine).lower()
    _norm_col = df["機種名"].astype(str).apply(lambda x: _normalize_key(x).lower())
    mask = _norm_col.str.contains(_nq, na=False, regex=False) & (df["差枚"] >= 1000)
    rows = df[mask].sort_values("差枚", ascending=False)
    if rows.empty:
        return None
    lines = [f"{machine_prefix}{machine}"]
    for _, row in rows.iterrows():
        try:
            ban = str(int(float(row["台番"])))
        except Exception:
            ban = str(row["台番"])
        lines.append(f"【{ban}番台】{fmt_diff(int(row['差枚']))}")
    return "\n".join(lines)


def _generate_weekly_result_text(
    date_labels: list,
    uploaded_files: dict,
    weekly_machines: list,
    daily_machines: dict,
    store_name: str = "",
) -> str:
    """週間結果テキストを生成して返す"""
    import io as _wrt_io
    import re as _wrt_re

    def _fmt_date_label(label: str) -> str:
        m = _wrt_re.match(r'^(\d+/\d+)（(.+?)）$', label)
        if m:
            return f"🗓️{m.group(1)}({m.group(2)})🗓️"
        return label

    def _extract_date(label: str) -> str:
        m = _wrt_re.match(r'^(\d+/\d+)（(.+?)）$', label)
        if m:
            return f"{m.group(1)}({m.group(2)})"
        m2 = _wrt_re.match(r'^(\d+/\d+)', label)
        return m2.group(1) if m2 else label

    seen_w: set = set()
    uniq_weekly: list = []
    for m in weekly_machines:
        if m not in seen_w:
            seen_w.add(m)
            uniq_weekly.append(m)

    day_sections: list = []

    for di in sorted(uploaded_files.keys()):
        f = uploaded_files[di]
        date_label = date_labels[di]

        try:
            df_raw = pd.read_excel(_wrt_io.BytesIO(f.getvalue()))
            df, _ = normalize_df(df_raw)
            _nm, _nm_norm = load_name_map()
            if _nm:
                df, _ = _apply_map(df, _nm, _nm_norm)
        except Exception as e:
            day_sections.append(f"{_fmt_date_label(date_label)}\n\n❌ Excel読み込みエラー: {e}")
            continue

        missing_cols = [c for c in ("台番", "機種名", "差枚") if c not in df.columns]
        if missing_cols:
            day_sections.append(
                f"{_fmt_date_label(date_label)}\n\n❌ 必須列が見つかりません: {missing_cols}"
                f"\n実際の列: {list(df.columns)}"
            )
            continue

        try:
            df["差枚"] = pd.to_numeric(df["差枚"], errors="coerce").fillna(0).astype(int)
        except Exception as e:
            day_sections.append(f"{_fmt_date_label(date_label)}\n\n❌ 差枚列を数値化できません: {e}")
            continue

        df["機種名"] = df["機種名"].astype(str)

        weekly_blocks: list = []
        for m in uniq_weekly:
            blk = _wrt_build_machine_block(df, m, machine_prefix="💖")
            if blk:
                weekly_blocks.append(blk)

        _day_raw = [m.strip() for m in daily_machines.get(di, []) if m.strip()]
        seen_d: set = set()
        uniq_daily: list = []
        for m in _day_raw:
            if m not in seen_d:
                seen_d.add(m)
                uniq_daily.append(m)

        daily_blocks: list = []
        for m in uniq_daily:
            blk = _wrt_build_machine_block(df, m, machine_prefix="🏅")
            if blk:
                daily_blocks.append(blk)

        lines: list = [_fmt_date_label(date_label)]
        lines.append("👑週間オススメ機種")
        if weekly_blocks:
            lines.append("\n\n".join(weekly_blocks))
        else:
            lines.append("（該当台なし）")

        if uniq_daily:
            lines.append("")
            lines.append("🎁毎日オススメ機種")
            if daily_blocks:
                lines.append("\n\n".join(daily_blocks))
            else:
                lines.append("（該当台なし）")

        day_sections.append("\n".join(lines))

    if not day_sections:
        return "（Excelがアップロードされていません）"

    body = "\n\n────────────────────\n\n".join(day_sections)

    if store_name and len(date_labels) >= 7:
        date_from = _extract_date(date_labels[0])
        date_to = _extract_date(date_labels[6])
        header = f"{date_from}～{date_to}\nエスパス{store_name}\n\n🏆週間&毎日オススメ機種の結果🏆"
        return f"{header}\n\n{body}"

    return body


def show_weekly_result_text_page() -> None:
    """週間結果テキスト抽出ページ"""
    store = st.session_state.selected_store
    _WRT_DAYS = ["月", "火", "水", "木", "金", "土", "日"]
    _N_DAILY = 15
    _wrt_cands = load_machine_candidates()

    # ページ初回表示時にJSONから init キーへ復元（ローテと同じパターン）
    # ウィジェットキーを直接上書きせず default= 経由で渡すことで
    # 同一セッション内の入力値を誤って消さない
    _wrt_loaded_key = f"wrt_machines_loaded_{store}"
    if not st.session_state.get(_wrt_loaded_key):
        _saved = _load_wrt_machines(store)
        for _wi in range(3):
            st.session_state[f"_wrt_init_{store}_weekly_{_wi}"] = (
                _saved.get("weekly", [])[_wi]
                if _wi < len(_saved.get("weekly", [])) else ""
            )
        for _di in range(7):
            for _idx in range(15):
                st.session_state[f"_wrt_init_{store}_daily_{_di}_{_idx}"] = (
                    _saved.get("daily", {}).get(str(_di), [])[_idx]
                    if _idx < len(_saved.get("daily", {}).get(str(_di), [])) else ""
                )
        st.session_state[_wrt_loaded_key] = True

    st.markdown(f"## 【{store}】　1週間の結果テキスト")
    st.markdown("---")

    # ① 対象週設定
    st.markdown("### ① 対象週設定")
    _today = datetime.date.today()
    _default_mon = _today - datetime.timedelta(days=_today.weekday())
    _wstart_key = f"wrt_week_start_{store}"
    week_start = st.date_input(
        "週の開始日（月曜日）",
        value=st.session_state.get(_wstart_key, _default_mon),
        key=f"wrt_date_input_{store}",
    )
    st.session_state[_wstart_key] = week_start

    dates = [week_start + datetime.timedelta(days=i) for i in range(7)]
    date_labels = [f"{d.month}/{d.day}（{_WRT_DAYS[i]}）" for i, d in enumerate(dates)]
    st.info(f"対象週: **{date_labels[0]}** 〜 **{date_labels[6]}**")
    st.markdown("---")

    # ② 週間オススメ機種（全曜日共通・上部固定）
    st.markdown("### ② 週間オススメ機種（全曜日共通）")
    _wcols = st.columns(3)
    for _wi, _wc in enumerate(_wcols):
        with _wc:
            _wkey = f"wrt_weekly_{store}_{_wi}"
            render_machine_autocomplete_input(
                f"週間オススメ {_wi + 1}", _wkey, _wrt_cands,
                default=st.session_state.get(f"_wrt_init_{store}_weekly_{_wi}", ""),
            )
    weekly_machines: list = [st.session_state.get(f"wrt_weekly_{store}_{_wi}", "") for _wi in range(3)]
    st.markdown("---")

    # ③ 曜日別 expander（Excel + 毎日オススメ機種）
    st.markdown("### ③ 曜日別設定（Excelアップロード・毎日オススメ機種）")

    # 一括アップロード
    import re as _wrt_re_bulk
    _bulk_key = f"wrt_bulk_{store}"
    _bulk_uploaded = st.file_uploader(
        "📂 一括アップロード（ファイル名にYYYYMMDDが含まれていれば自動で曜日に割り当て）",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key=_bulk_key,
    )
    _bulk_assigned: dict = {}
    _bulk_unmatched: list = []
    for _bf in (_bulk_uploaded or []):
        _bm = _wrt_re_bulk.search(r'(\d{8})', _bf.name)
        if _bm:
            try:
                _ds = _bm.group(1)
                _bd = datetime.date(int(_ds[:4]), int(_ds[4:6]), int(_ds[6:8]))
                for _di, _date in enumerate(dates):
                    if _bd == _date:
                        _bulk_assigned[_di] = _bf
                        break
                else:
                    _bulk_unmatched.append(_bf.name)
            except Exception:
                _bulk_unmatched.append(_bf.name)
        else:
            _bulk_unmatched.append(_bf.name)
    if _bulk_assigned:
        st.success("✅ 自動割り当て: " + "　".join(
            f"{_WRT_DAYS[_di]}曜: {_bf.name}" for _di, _bf in sorted(_bulk_assigned.items())
        ))
    if _bulk_unmatched:
        st.warning("⚠️ 日付不明（手動でアップロードしてください）: " + "　".join(_bulk_unmatched))

    uploaded_files: dict = {}
    daily_machines: dict = {}

    for di in range(7):
        _upload_key = f"wrt_excel_{store}_{di}"
        _f_individual = st.session_state.get(_upload_key)
        _is_up = (_f_individual is not None) or (di in _bulk_assigned)
        _status = "✅アップロード済み" if _is_up else "未アップロード"
        with st.expander(f"**{date_labels[di]}**　{_status}"):
            _f = st.file_uploader(
                f"{_WRT_DAYS[di]}曜Excel",
                type=["xlsx", "xls"],
                key=_upload_key,
            )
            _effective_f = _f if _f is not None else _bulk_assigned.get(di)
            if _effective_f is not None:
                uploaded_files[di] = _effective_f
                if _f is not None:
                    st.success(f"✅ {_f.name}")
                else:
                    st.info(f"📂 一括アップロード: {_effective_f.name}")

            st.markdown(f"**{_WRT_DAYS[di]}曜オススメ機種**（15機種まで）")
            for _ri in range(5):
                _row_cols = st.columns(3)
                for _ci, _rc in enumerate(_row_cols):
                    _idx = _ri * 3 + _ci
                    with _rc:
                        _dkey = f"wrt_daily_{store}_{di}_{_idx}"
                        render_machine_autocomplete_input(
                            f"オススメ{_idx + 1}", _dkey, _wrt_cands,
                            default=st.session_state.get(f"_wrt_init_{store}_daily_{di}_{_idx}", ""),
                        )
            daily_machines[di] = [st.session_state.get(f"wrt_daily_{store}_{di}_{_idx}", "") for _idx in range(_N_DAILY)]

    st.markdown("---")

    # 実行ボタン
    if st.button("📝 結果テキスト生成", type="primary", use_container_width=True, key="wrt_gen_btn"):
        if not uploaded_files:
            st.error("❌ Excelが1つもアップロードされていません。")
        else:
            _valid_weekly = [m.strip() for m in weekly_machines if m.strip()]
            _has_daily = any(
                any(m.strip() for m in daily_machines.get(di, []))
                for di in uploaded_files
            )
            if not _valid_weekly and not _has_daily:
                st.warning("⚠️ 週間オススメ機種と毎日オススメ機種がすべて空欄です。機種名を1つ以上入力してください。")
            else:
                _save_wrt_machines(store, weekly_machines, daily_machines)
                _result = _generate_weekly_result_text(
                    date_labels, uploaded_files, _valid_weekly, daily_machines,
                    store_name=store,
                )
                st.session_state[f"wrt_result_{store}"] = _result
                _wrt_fname = f"週間結果テキスト_{date_labels[0].replace('/', '-').replace('（', '(').replace('）', ')')}〜{date_labels[6].replace('/', '-').replace('（', '(').replace('）', ')')}.txt"
                if _IS_CLOUD:
                    st.download_button(
                        label="📥 ZIPをダウンロードしてください",
                        data=_result.encode("utf-8"),
                        file_name=_wrt_fname,
                        mime="text/plain",
                        key="wrt_dl",
                        type="primary",
                    )
                else:
                    _wrt_save_path = os.path.join(_DESKTOP, _wrt_fname)
                    with open(_wrt_save_path, "w", encoding="utf-8") as _f:
                        _f.write(_result)
                    st.success(f"✅ デスクトップに保存しました: {_wrt_fname}")

    _result_val = st.session_state.get(f"wrt_result_{store}", "")
    if _result_val:
        st.markdown("---")
        st.markdown("### 抽出結果")
        st.text_area("", value=_result_val, height=600, key="wrt_output_area")

    st.markdown("---")
    if st.button("← 画像種類選択に戻る", key="wrt_back"):
        _navigate("image_type")


def show_name_conversion_page() -> None:
    """機種名変換ページ（タブ構成）"""
    st.markdown("## 機種名変換")
    st.markdown("---")

    tab_convert, tab_master, tab_pision = st.tabs(["🔄 変換実行", "📋 マスタ管理", "🔍 pisionチェック"])

    # ================================================================
    # TAB1: マスタ管理
    # ================================================================
    with tab_master:
        st.markdown("### 変換マスタ一覧（機種名変換.xlsx）")

        if not os.path.exists(NAME_MAP_PATH):
            st.error(f"❌ マスタファイルが見つかりません: {NAME_MAP_PATH}")
        else:
            try:
                master_df = _load_master_df()
                st.caption(f"現在の登録件数: **{len(master_df):,} 件**　（行を直接編集・追加できます）")

                edited_master = st.data_editor(
                    master_df,
                    key="nc_master_editor",
                    use_container_width=True,
                    num_rows="dynamic",
                    hide_index=True,
                    column_config={
                        "変換前（正式名）": st.column_config.TextColumn(
                            "変換前（正式名）", width="large"
                        ),
                        "変換後（簡略名）": st.column_config.TextColumn(
                            "変換後（簡略名）", width="large"
                        ),
                    },
                )

                if st.button("💾 マスタを保存", type="primary", key="nc_save_master", use_container_width=True):
                    try:
                        _save_master_df(edited_master)
                        st.success(f"✅ 保存しました（{len(edited_master):,} 件）")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ 保存に失敗しました: {e}")
                        with st.expander("詳細"):
                            st.code(traceback.format_exc())

            except Exception as e:
                st.error(f"❌ マスタの読み込みに失敗しました: {e}")

    # ================================================================
    # TAB2: 変換実行
    # ================================================================
    with tab_convert:
        # ── ① 変換元ファイル ─────────────────────────────────────────
        st.markdown("### ① 変換元ファイルを選択")
        src_file = st.file_uploader(
            "変換したいファイル（.xlsx / .xls / .csv）",
            type=["xlsx", "xls", "csv"],
            key="nc_src",
        )

        df_src: pd.DataFrame | None = None
        raw_src: pd.DataFrame | None = None
        if src_file:
            try:
                raw = _read_uploaded_df(src_file)
                raw_src = raw
                df_src = _find_kisha_col(raw)
                if df_src is None:
                    st.error(f"❌ 「機種名」列が見つかりません。実際の列名: {list(raw.columns)}")
                else:
                    st.success(f"✅ {len(df_src):,} 行を読み込みました")
            except Exception as e:
                st.error(f"❌ ファイルの読み込みに失敗しました: {e}")

        # ── ② 未登録機種の表示 ─────────────────────────────────────
        name_map, name_map_norm = load_name_map()

        if df_src is not None:
            # 全機種名列（正式名・データサイト表記など）からすべての機種名を収集
            _kisha_cands = ["機種名"] + COLUMN_ALIASES.get("機種名", [])
            _seen: dict[str, str] = {}  # 機種名 -> 出典列名
            _raw_for_check = raw_src if raw_src is not None else df_src
            for _col in _kisha_cands:
                if _col in _raw_for_check.columns:
                    for _m in _raw_for_check[_col].dropna().astype(str).str.strip().unique():
                        if _m and _m != "nan" and _m not in _seen:
                            _seen[_m] = _col
            unregistered_pairs = [
                (_m, _col) for _m, _col in _seen.items()
                if _m not in name_map and _normalize_key(_m) not in name_map_norm
            ]
            unregistered = [_m for _m, _ in unregistered_pairs]

            if unregistered:
                st.markdown("### ② 未登録機種をマスタに追加")
                st.warning(
                    f"⚠️ 変換マスタに未登録の機種が **{len(unregistered)} 件** あります。"
                    "　変換後の名前を入力して「マスタに追加」を押してください。"
                )
                edit_df = pd.DataFrame({
                    "機種名（元）":   [_m for _m, _ in unregistered_pairs],
                    "出典列":        [_col for _, _col in unregistered_pairs],
                    "変換後の機種名": [""] * len(unregistered_pairs),
                })
                edited = st.data_editor(
                    edit_df,
                    key="nc_unreg_edit",
                    use_container_width=True,
                    disabled=["機種名（元）", "出典列"],
                    hide_index=True,
                )
                new_entries = [
                    (str(row["機種名（元）"]).strip(), str(row["変換後の機種名"]).strip())
                    for _, row in edited.iterrows()
                    if str(row["変換後の機種名"]).strip()
                ]
                if new_entries:
                    if st.button(
                        f"📝 {len(new_entries)} 件をマスタに追加",
                        key="nc_add_master",
                        type="secondary",
                    ):
                        try:
                            wb = load_workbook(NAME_MAP_PATH)
                            ws = wb.active
                            for orig, conv in new_entries:
                                ws.append([None, orig, conv])
                            wb.save(NAME_MAP_PATH)
                            load_name_map.clear()
                            st.success(f"✅ {len(new_entries)} 件を追加しました。")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ 追加に失敗しました: {e}")
                else:
                    st.info("「変換後の機種名」列に入力すると追加ボタンが表示されます。")
            else:
                st.markdown("### ② 未登録機種")
                st.success("✅ すべての機種が変換マスタに登録済みです。")

        # ── ③ 出力先フォルダ ───────────────────────────────────────
        st.markdown("### ③ 出力先フォルダを指定")
        out_folder = st.text_input(
            "出力先フォルダパス",
            value=r"C:\Users\23-3\Desktop\画像作成",
            key="nc_out_folder",
        )

        # ── ④ 変換実行 ────────────────────────────────────────────
        st.markdown("### ④ 変換実行")
        run_clicked = st.button(
            "▶ 変換実行",
            type="primary",
            key="nc_run",
            use_container_width=True,
            disabled=(df_src is None),
        )

        if df_src is None:
            st.info("⬆️ まず変換元Excelを選択してください。")
        elif run_clicked:
            if not out_folder.strip():
                st.error("❌ 出力先フォルダを指定してください。")
            elif not os.path.isdir(out_folder):
                st.error(f"❌ 出力先フォルダが存在しません: {out_folder}")
            else:
                with st.spinner("変換中..."):
                    try:
                        cur_map, cur_norm = load_name_map()
                        if not cur_map:
                            st.error(f"❌ 変換マスタの読み込みに失敗しました: {NAME_MAP_PATH}")
                        else:
                            df_converted, conv_count = _apply_map(df_src, cur_map, cur_norm)
                            base_name = os.path.splitext(src_file.name)[0]
                            out_name  = f"{base_name}_機種名変換済.xlsx"
                            st.success(f"✅ 変換完了！{conv_count:,} 件を変換しました。")
                            if not _IS_CLOUD:
                                out_path = os.path.join(out_folder, out_name)
                                df_converted.to_excel(out_path, index=False)
                                st.info(f"📁 保存先: {out_path}")
                            buf = io.BytesIO()
                            df_converted.to_excel(buf, index=False)
                            st.download_button(
                                label="💾 変換済みExcelをダウンロード",
                                data=buf.getvalue(),
                                file_name=out_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="nc_download",
                            )
                            with st.expander("📋 変換後プレビュー（先頭5行）"):
                                st.dataframe(df_converted.head(), use_container_width=True)
                    except Exception as e:
                        st.error(f"❌ エラーが発生しました: {e}")
                        with st.expander("詳細（開発者向け）"):
                            st.code(traceback.format_exc())

    # ================================================================
    # TAB3: pisionチェック
    # ================================================================
    with tab_pision:
        st.markdown("### 🔍 pisionデータ取得 → 変換マスタ照合")
        st.caption("pisionから各店舗のデータを取得し、変換マスタへの登録状況を確認します。")

        _nc_api_key = _get_pision_api_key()
        if not _nc_api_key:
            st.error("❌ PISION_API_KEY が未設定のため利用できません。")
        else:
            _nc_col1, _nc_col2 = st.columns(2)
            with _nc_col1:
                _nc_store = st.selectbox("店舗を選択", list(STORES.keys()), key="nc_pision_store")
            with _nc_col2:
                _nc_mode = st.radio(
                    "データ種別",
                    ["確定データ", "速報データ（当日）"],
                    horizontal=True,
                    key="nc_pision_mode",
                )
            _nc_is_rt = _nc_mode.startswith("速報")
            if _nc_is_rt:
                _nc_rt_user, _nc_rt_pass = _get_pision_rt_credentials()
                if not _nc_rt_user or not _nc_rt_pass:
                    st.error("❌ 速報データには PISION_RT_USER / PISION_RT_PASS が必要です。")
                    _nc_is_rt = False

            # 確定→速報に切り替わった瞬間はキーを削除してdefault値（当日）を使わせる
            _nc_today = datetime.date.today()
            if _nc_is_rt and st.session_state.get("_nc_pision_prev_mode") != "rt":
                st.session_state.pop("nc_pision_date", None)
            st.session_state["_nc_pision_prev_mode"] = "rt" if _nc_is_rt else "fix"
            _nc_date = st.date_input(
                "日付",
                value=_nc_today if _nc_is_rt else _nc_today - datetime.timedelta(days=1),
                key="nc_pision_date",
            )
            _nc_date_str = _nc_date.strftime("%Y-%m-%d")

            _nc_items_key      = f"_nc_pision_items_{_nc_store}_{_nc_date_str}_{'rt' if _nc_is_rt else 'fix'}"
            _nc_collecting_key = f"_nc_pision_collecting_{_nc_store}"
            _nc_baseline_key   = f"_nc_pision_baseline_{_nc_store}"
            _nc_is_collecting  = _nc_is_rt and st.session_state.get(_nc_collecting_key) == _nc_date_str

            def _nc_is_new_artid(poll_result: dict) -> bool:
                _baseline = st.session_state.get(_nc_baseline_key)
                _new_id   = poll_result.get("article_id")
                if _baseline is None:
                    return True
                try:
                    return int(_new_id) > int(_baseline)
                except (TypeError, ValueError):
                    return False

            # ── ボタン描画 ──────────────────────────────────────────────
            _nc_do_rt_check   = False
            _nc_fetch_clicked = False
            _nc_do_existing   = False
            if _nc_is_collecting:
                _nc_bc1, _nc_bc2 = st.columns(2)
                with _nc_bc1:
                    st.button("⏳ 収集中...", key="nc_pision_fetch_col", disabled=True, use_container_width=True)
                with _nc_bc2:
                    _nc_do_rt_check = st.button("🔍 今すぐ確認", key="nc_pision_rt_check",
                                                 use_container_width=True, help="30秒待たずに今すぐ確認します。")
            elif _nc_is_rt and _nc_is_rt:
                _nc_bc1, _nc_bc2 = st.columns(2)
                with _nc_bc1:
                    _nc_fetch_clicked = st.button("📥 データを取得", key="nc_pision_fetch",
                                                   type="primary", use_container_width=True)
                with _nc_bc2:
                    _nc_do_existing = st.button("📂 既存のデータを取得", key="nc_pision_existing",
                                                 use_container_width=True,
                                                 help="新しい収集を開始せず、過去の直近データを読み込みます。")
            else:
                _nc_fetch_clicked = st.button("📥 データを取得", key="nc_pision_fetch", type="primary")

            # ── 取得ボタン押下時 ─────────────────────────────────────────
            if _nc_fetch_clicked:
                with st.spinner(f"{_nc_store} / {_nc_date_str} を取得中..."):
                    _nc_fetched = None
                    if _nc_is_rt:
                        _nc_rt = fetch_pision_realtime(_nc_store, _nc_date_str)
                        if _nc_rt["ok"] and _nc_is_new_artid(_nc_rt):
                            _nc_fetched = _nc_rt["items"]
                        elif _nc_rt.get("collect_started") and not _nc_rt["ok"]:
                            st.session_state[_nc_collecting_key] = _nc_date_str
                            st.session_state[_nc_baseline_key]   = None
                            st.rerun()
                        else:
                            if _nc_rt.get("error"):
                                st.error(f"❌ {_nc_rt['error']}")
                    else:
                        try:
                            _nc_halls = fetch_pision_halls(_nc_api_key)
                        except Exception as _e:
                            st.error(f"❌ ホール一覧取得失敗: {_e}")
                            _nc_halls = []
                        _nc_hall_id = None
                        for _h in _nc_halls:
                            _hname = _h.get("name") or _h.get("displayName") or ""
                            if _nc_store in _hname and "エスパス" in _hname:
                                _nc_hall_id = str(_h.get("id") or _h.get("hallId") or "")
                                break
                        if _nc_hall_id:
                            try:
                                _nc_fetched = fetch_pision_results(_nc_api_key, _nc_hall_id, _nc_date_str)
                            except Exception as _e:
                                st.error(f"❌ データ取得失敗: {_e}")
                        else:
                            st.error(f"❌ {_nc_store} のホールIDが見つかりません。")
                if _nc_fetched:
                    st.session_state[_nc_items_key] = _nc_fetched
                    st.session_state.pop(_nc_collecting_key, None)
                    st.success(f"✅ {len(_nc_fetched)} 台分のデータを取得しました。")
                    st.rerun()
                elif _nc_fetched is not None:
                    st.info("📭 データが空です（404 / 未公開 / 店休日の可能性があります）。")

            # ── 既存データ取得（新規収集なし）────────────────────────────
            if _nc_do_existing:
                with st.spinner("既存の速報データを確認中（新しい収集は開始しません）..."):
                    _nc_exist = fetch_pision_realtime(_nc_store, _nc_date_str, trigger=False)
                if _nc_exist["ok"]:
                    st.session_state[_nc_items_key] = _nc_exist["items"]
                    st.session_state.pop(_nc_collecting_key, None)
                    st.rerun()
                else:
                    st.warning("⚠️ 既存の速報データが見つかりませんでした。「📥 データを取得」で新しい収集を開始してください。")

            # ── 手動確認（「今すぐ確認」ボタン）───────────────────────────
            if _nc_do_rt_check:
                with st.spinner("収集状況を確認中..."):
                    _nc_chk = fetch_pision_realtime(_nc_store, _nc_date_str, trigger=False)
                if _nc_chk["ok"] and _nc_is_new_artid(_nc_chk):
                    st.session_state[_nc_items_key] = _nc_chk["items"]
                    st.session_state.pop(_nc_collecting_key, None)
                    st.rerun()
                elif _nc_chk["ok"]:
                    st.info("⏳ まだ完了していません。自動確認に戻ります。")
                    st.rerun()
                else:
                    st.rerun()

            # ── 自動ポーリング（収集中・手動確認なし）─────────────────────
            if _nc_is_collecting and not _nc_do_rt_check:
                import time as _nc_time
                _nc_ph = st.empty()
                _nc_ph.info("⏳ 速報データを収集中... 30秒後に自動で確認します。")
                _nc_time.sleep(30)
                _nc_ph.empty()
                with st.spinner("収集状況を自動確認中..."):
                    _nc_auto = fetch_pision_realtime(_nc_store, _nc_date_str, trigger=False)
                if _nc_auto["ok"] and _nc_is_new_artid(_nc_auto):
                    st.session_state[_nc_items_key] = _nc_auto["items"]
                    st.session_state.pop(_nc_collecting_key, None)
                    st.rerun()
                elif _nc_auto.get("running") or (_nc_auto["ok"] and not _nc_is_new_artid(_nc_auto)):
                    st.rerun()  # まだ収集中 → 30秒ループ継続
                else:
                    st.warning("⚠️ 収集が完了しましたがデータが取得できませんでした。「📥 データを取得」をもう一度押してください。")
                    st.session_state.pop(_nc_collecting_key, None)
                    st.rerun()

            _nc_items = st.session_state.get(_nc_items_key)
            if _nc_items:
                st.markdown("---")
                _nc_nm, _nc_nm_norm = load_name_map()

                def _nc_conv(raw: str) -> str:
                    raw = str(raw).strip()
                    if raw in _nc_nm:
                        return _nc_nm[raw]
                    k = _normalize_key(raw)
                    if k in _nc_nm_norm:
                        return _nc_nm_norm[k]
                    return ""

                # 機種名の重複排除（displayName / modelName 両方チェック）
                _nc_seen: dict[str, str] = {}  # raw_name → "displayName" or "modelName"
                for _it in _nc_items:
                    for _field, _label in [("displayName", "short"), ("modelName", "model")]:
                        _raw = str(_it.get(_field) or "").strip()
                        if _raw and _raw != "nan" and _raw not in _nc_seen:
                            _nc_seen[_raw] = _label

                _nc_rows = []
                for _raw, _src in sorted(_nc_seen.items()):
                    _conv = _nc_conv(_raw)
                    _nc_rows.append({
                        "pision名":    _raw,
                        "種別":        _src,
                        "変換後":      _conv,
                        "登録済み":    "✅" if _conv else "❌ 未登録",
                    })

                _nc_df = pd.DataFrame(_nc_rows)
                _unreg = _nc_df[_nc_df["登録済み"] == "❌ 未登録"]
                _reg   = _nc_df[_nc_df["登録済み"] == "✅"]

                st.markdown(f"**登録済み: {len(_reg)} 件　／　未登録: {len(_unreg)} 件**")

                if not _unreg.empty:
                    st.warning(f"⚠️ {len(_unreg)} 件が変換マスタに未登録です。")
                    _nc_edit = st.data_editor(
                        _unreg[["pision名", "種別"]].assign(**{"変換後（入力）": ""}),
                        key="nc_pision_unreg_edit",
                        use_container_width=True,
                        disabled=["pision名", "種別"],
                        hide_index=True,
                    )
                    _nc_new = [
                        (str(r["pision名"]).strip(), str(r["変換後（入力）"]).strip())
                        for _, r in _nc_edit.iterrows()
                        if str(r["変換後（入力）"]).strip()
                    ]
                    if _nc_new:
                        if st.button(f"📝 {len(_nc_new)} 件をマスタに追加", key="nc_pision_add", type="secondary"):
                            try:
                                from openpyxl import load_workbook as _lw
                                _wb = _lw(NAME_MAP_PATH)
                                _ws = _wb.active
                                for _orig, _cv in _nc_new:
                                    _ws.append([None, _orig, _cv])
                                _wb.save(NAME_MAP_PATH)
                                load_name_map.clear()
                                st.success(f"✅ {len(_nc_new)} 件を追加しました。")
                                st.rerun()
                            except Exception as _e:
                                st.error(f"❌ 追加に失敗しました: {_e}")
                    else:
                        st.info("「変換後（入力）」列に変換後の名前を入力すると追加ボタンが表示されます。")

                with st.expander(f"✅ 登録済み一覧（{len(_reg)} 件）"):
                    st.dataframe(_reg[["pision名", "変換後"]], use_container_width=True, hide_index=True)


# =============================================================================
# ■ ⑩スランプグラフ生成
# =============================================================================

_PISION_BASE_URL = "https://www.pision.io"

# 表示対象ホール（部分一致）。将来の追加はここにキーワードを1行足す。
_PISION_ALLOWED_HALLS: list[str] = [
    "新宿歌舞伎町",
    "西武新宿",
    "新大久保",
    "高田馬場",
    "上野本館",
    "上野新館",
    "渋谷新館",
    "赤坂見附",
    "新小岩",
    "溝の口本館",
    "溝の口新館",
    "稲毛",
    "秋葉原",
]


def get_secret_value(key: str, default=None):
    """st.secrets → os.getenv の優先順でシークレット値を取得する。
    ローカルは .env（起動時ロード済み）、Cloud は st.secrets を使う。"""
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return os.getenv(key, default)


def _get_pision_api_key() -> "str | None":
    """APIキーを st.secrets → 環境変数 の優先順で取得する。"""
    return get_secret_value("PISION_API_KEY")


def find_slump_template() -> "object | None":
    """テンプレート画像 base_3000_bk.png を優先順で検索して Path を返す。"""
    from pathlib import Path
    candidates = [
        Path(os.path.abspath(__file__)).parent / "base_3000_bk.png",
        Path(os.path.abspath(__file__)).parent / "images" / "base_3000_bk.png",
        Path(r"C:\Users\23-3\Desktop\画像作成\base_3000_bk.png"),
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _pision_request(api_key: str, path: str) -> "dict | list | None":
    """pision.io API へ GET リクエストを送る。404 は None を返す。"""
    import urllib.request
    import urllib.error
    url = f"{_PISION_BASE_URL}{path}"
    req = urllib.request.Request(url, headers={"X-Api-Key": api_key})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise


def fetch_pision_halls(api_key: str) -> list:
    """ホール一覧を取得する。"""
    data = _pision_request(api_key, "/api/v2/halls")
    if data is None:
        return []
    if isinstance(data, list):
        return data
    return data.get("halls", data.get("data", []))


def fetch_pision_results(api_key: str, hall_id: str, date: str) -> "list | None":
    """台別データを取得する。404 / 未公開 / 店休日は None を返す。"""
    import time
    data = _pision_request(api_key, f"/api/v2/halls/{hall_id}/results/{date}")
    time.sleep(1)
    if data is None:
        return None
    if isinstance(data, list):
        return data
    return data.get("details", data.get("data", []))


# ── 速報データ（realtime ログインアプリ）────────────────────────────────
# 確定データ API（X-Api-Key）とは別系統。/login のセッション Cookie を使う。
# realtime 用 hallId は X-Api-Key 側の hall id とは無関係（/realtime の店舗 select から取得）。

def _get_pision_rt_credentials() -> "tuple[str | None, str | None]":
    """realtime ログイン情報を st.secrets → 環境変数 の優先順で取得する。"""
    return get_secret_value("PISION_RT_USER"), get_secret_value("PISION_RT_PASS")


def _pision_rt_csrf(session, path: str) -> "str | None":
    """指定パスの HTML から hidden input[name=csrfToken] の値を取り出す。"""
    from bs4 import BeautifulSoup
    r = session.get(f"{_PISION_BASE_URL}{path}", timeout=20)
    soup = BeautifulSoup(r.text, "html.parser")
    tok = soup.find("input", {"name": "csrfToken"})
    return (tok.get("value") if tok else None), r


def _pision_rt_login(session) -> "str | None":
    """realtime にログインする。成功で None、失敗でエラーメッセージ文字列を返す。"""
    user, pw = _get_pision_rt_credentials()
    if not user or not pw:
        return ("ログイン情報が未設定です。.env に PISION_RT_USER / PISION_RT_PASS を設定してください。")
    csrf, _ = _pision_rt_csrf(session, "/login")
    if not csrf:
        return "CSRFトークンの取得に失敗しました（ログインページの構造が変わった可能性があります）。"
    session.post(
        f"{_PISION_BASE_URL}/login",
        data={"csrfToken": csrf, "userName": user, "password": pw},
        timeout=20, allow_redirects=True,
    )
    # ログイン確認：/realtime が /login にリダイレクトされなければ成功
    chk = session.get(f"{_PISION_BASE_URL}/realtime", timeout=20, allow_redirects=False)
    loc = chk.headers.get("Location", "")
    if chk.status_code in (301, 302, 303) and "/login" in loc:
        return "ログインに失敗しました（ID／パスワードを確認してください）。"
    return None


def _pision_rt_halls(session) -> "tuple[dict, str | None]":
    """/realtime/create の店舗 select から {店舗名: realtime用hallId} と作成用 csrfToken を取得する。
    realtime用hallId は X-Api-Key 側の hall id とは別に、この select の値を正とする。"""
    from bs4 import BeautifulSoup
    r = session.get(f"{_PISION_BASE_URL}/realtime/create", timeout=20)
    soup = BeautifulSoup(r.text, "html.parser")
    sel = soup.find("select", {"name": "hallId"}) or soup.find("select")
    halls: dict = {}
    if sel:
        for opt in sel.find_all("option"):
            val = (opt.get("value") or "").strip()
            name = opt.get_text(strip=True)
            if val and name:
                halls[name] = val
    tok = soup.find("input", {"name": "csrfToken"})
    return halls, (tok.get("value") if tok else None)


def _match_rt_hall(halls: dict, hall_name: str) -> "str | None":
    """店舗名（例: エスパス新小岩）を realtime の店舗 select 候補に突き合わせて hallId を返す。"""
    if hall_name in halls:
        return halls[hall_name]
    # 「エスパス」有無やスペースを無視した緩い一致
    def norm(s: str) -> str:
        return _normalize_key(s).replace("エスパス", "")
    target = norm(hall_name)
    for name, hid in halls.items():
        if norm(name) == target:
            return hid
    for name, hid in halls.items():
        if target and (target in norm(name) or norm(name) in target):
            return hid
    return None


def _pision_rt_find_article_id(html: str, hall_id: str, date_str: str) -> "tuple[int | None, str | None, list]":
    """/realtime 一覧 HTML から対象ホール・対象日の記事 id を探す。
    各行の「再実行」リンク (/realtime/create?targetDate=...&hallId=...) で行の (hallId, 日付) を判定し、
    同じ行の /articles/{id} とペアにする。同一条件で複数あれば最大id（最新の収集）を返す。
    行内の最後の日時（収集完了時刻）も取り出す。
    戻り値: (記事id or None, 収集時刻 or None, 候補リスト[(id, hallId, date)])。"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    cands: list = []   # (article_id, hallId, date)
    times: dict = {}   # article_id -> 収集完了時刻文字列
    for tr in soup.find_all("tr"):
        art = hid = td = None
        for a in tr.find_all("a", href=True):
            m1 = re.search(r"/articles/(\d+)", a["href"])
            if m1:
                art = int(m1.group(1))
            m2 = re.search(r"/realtime/create\?targetDate=([\d\-]+)&hallId=(\d+)", a["href"])
            if m2:
                td, hid = m2.group(1), m2.group(2)
        if art and hid and td:
            cands.append((art, hid, td))
            # 行内の日時（YYYY/MM/DD HH:MM:SS）の最後＝収集完了時刻
            ts = re.findall(r"\d{4}/\d{2}/\d{2}\s+\d{2}:\d{2}(?::\d{2})?", tr.get_text(" ", strip=True))
            if ts:
                times[art] = ts[-1]
    cands.sort(key=lambda x: x[0], reverse=True)  # id 降順（新しい順）

    # 店舗 hallId と日付が厳密一致するものだけを採用（他店舗の同日記事は拾わない）
    match = [c for c in cands if str(c[1]) == str(hall_id) and c[2] == date_str]
    if match:
        aid = match[0][0]
        return aid, times.get(aid), cands
    return None, None, cands


def _pision_rt_is_running(html: str, hall_id: str, date_str: str) -> bool:
    """対象ホール・対象日の収集が「実行中」の行が一覧にあるかを判定する。
    実行中の行は「再実行」リンク(hallId,date)を持つが /articles/{id} リンクが無く、
    『実行中』の文字を含む。"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    for tr in soup.find_all("tr"):
        has_target = has_article = False
        for a in tr.find_all("a", href=True):
            if re.search(r"/articles/(\d+)", a["href"]):
                has_article = True
            m = re.search(r"/realtime/create\?targetDate=([\d\-]+)&hallId=(\d+)", a["href"])
            if m and m.group(1) == date_str and m.group(2) == str(hall_id):
                has_target = True
        if has_target and not has_article and "実行中" in tr.get_text(" ", strip=True):
            return True
    return False


def _rt_int(text: str) -> int:
    """「-1,800」「+2,900」「±0」などを整数に変換する。"""
    t = (text or "").replace(",", "").replace("±", "").strip()
    m = re.match(r"^[+\-]?\d+", t)
    return int(m.group(0)) if m else 0


def _rt_count(text: str) -> int:
    """「43(1/37)」「0」などBB/RB/ARTセルから回数だけを取り出す。"""
    m = re.match(r"^\s*(\d+)", text or "")
    return int(m.group(1)) if m else 0


def _parse_pision_rt_detail(html: str) -> "tuple[list, int]":
    """/articles/{id}/detail HTML から速報データを抽出する。
    差枚・G数・BB・RB・ART は台別テーブル（ヘッダー: 台番/差枚/前日差枚/BB/RB/合算/ART/G数）の
    実値を使い、points（時系列 {x,y}）は div.graph の data-points から台番で紐づける。
    points の x は間引き値で実G数ではないため、グラフ描画専用に使う。
    戻り値: (items, グラフ用pointsが無い台数)。items は確定APIの details と同形:
    {unitId, displayName, points, diff, games, bb, rb, art}。"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")

    # ① div.graph から {台番: (points, 機種短縮名)}
    gmap: dict = {}
    for g in soup.select("div.graph"):
        uid = g.get("data-unit-id")
        if not uid:
            continue
        raw = g.get("data-points")
        pts = None
        if raw:
            try:
                # BeautifulSoup は属性値の HTML エスケープを復元済み。JSON をパース。
                pts = json.loads(raw)
            except (ValueError, TypeError):
                pts = None
        gmap[str(uid)] = {
            "points":     pts if (pts and len(pts) >= 2) else None,
            "name":       g.get("data-short-name") or g.get("data-model-name") or "",
            "model_name": g.get("data-model-name") or g.get("data-short-name") or "",
        }

    # ② 台別テーブルから実データを集める（台番をキーに）
    HEAD = ["台番", "差枚", "前日差枚", "BB", "RB", "合算", "ART", "G数"]
    items: list = []
    skipped = 0
    seen: set = set()
    for t in soup.find_all("table"):
        ths = [th.get_text(strip=True) for th in t.find_all("th")]
        if ths[:8] != HEAD:
            continue
        for tr in t.find_all("tr"):
            cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) < 8 or not re.match(r"^\d+$", cells[0]):
                continue  # ヘッダー行・平均行・空行を除外
            uid = cells[0]
            if uid in seen:
                continue
            seen.add(uid)
            g = gmap.get(uid, {})
            pts = g.get("points")
            if not pts:
                skipped += 1
            items.append({
                "unitId":      uid,
                "displayName": g.get("name", ""),        # data-short-name（グラフラベル用）
                "modelName":   g.get("model_name", ""),  # data-model-name（機種名変換用）
                "points":      pts,
                "diff":        _rt_int(cells[1]),
                "bb":          _rt_count(cells[3]),
                "rb":          _rt_count(cells[4]),
                "art":         _rt_count(cells[6]),
                "games":       _rt_int(cells[7]),
            })
    return items, skipped


def fetch_pision_realtime(hall_name: str, date_str: str, trigger: bool = True) -> dict:
    """速報データを取得する。realtime にログイン → （trigger時のみ）/realtime/create で最新収集を開始 →
    一覧から直近の完了済み記事idを特定 → /articles/{id}/detail をパースして items を返す。
    収集は数分かかるため、表示するのは「今すぐ読める最新の完了スナップショット」。
    trigger=False なら収集を起こさず（実行中にせず）に最新完了分を読むだけ（完了確認用）。
    戻り値: {ok, items, skipped, article_id, snapshot_time, collect_started, running, error, debug}。"""
    import requests
    result = {"ok": False, "items": [], "skipped": 0, "article_id": None,
              "snapshot_time": None, "collect_started": False, "running": False,
              "error": None, "debug": {}}
    try:
        session = requests.Session()
        session.headers.update({"User-Agent": "Mozilla/5.0 (slump-graph-app)"})

        err = _pision_rt_login(session)
        if err:
            result["error"] = err
            return result

        halls, csrf = _pision_rt_halls(session)
        result["debug"]["hall_options"] = list(halls.keys())
        if not halls:
            result["error"] = "realtime の店舗一覧（select）の取得に失敗しました。"
            return result

        hall_id = _match_rt_hall(halls, hall_name)
        if not hall_id:
            result["error"] = (f"realtime の店舗一覧に「{hall_name}」が見つかりません。"
                               f"候補: {', '.join(halls.keys())}")
            return result
        result["debug"]["rt_hall_id"] = hall_id

        # 最新収集を開始（POST /realtime/create）→ realtime 上でその店舗が「実行中」になる。
        # 収集完了まで数分かかるため、ここでは待たずに直近の完了済み記事を読む。
        if trigger and csrf:
            try:
                session.post(
                    f"{_PISION_BASE_URL}/realtime/create",
                    data={"csrfToken": csrf, "hallId": hall_id, "targetDate": date_str},
                    timeout=60, allow_redirects=True,
                )
                result["collect_started"] = True
            except Exception:
                pass  # 収集開始に失敗しても、既存記事の読み取りは続行する

        # 一覧から対象ホール・対象日の直近完了記事idを特定（新しい順・数ページ走査）
        article_id = snap_time = None
        last_cands: list = []
        first_page_html = None
        for _page in range(1, 5):
            list_html = session.get(
                f"{_PISION_BASE_URL}/realtime?page={_page}", timeout=20).text
            if _page == 1:
                first_page_html = list_html
            aid, ts, cands = _pision_rt_find_article_id(list_html, hall_id, date_str)
            if cands:
                last_cands = cands
            if aid:
                article_id, snap_time = aid, ts
                break
            if not cands:        # これ以上記事が無い
                break
        # この店舗・日付の収集が「実行中」かどうか（ページ1で判定）
        result["running"] = _pision_rt_is_running(first_page_html or "", hall_id, date_str)
        result["debug"]["article_candidates"] = last_cands[:15]
        if article_id is None:
            result["error"] = (
                f"対象店舗・対象日（{date_str}）の完了済み速報記事が realtime 一覧に見つかりませんでした。"
                + ("収集を開始したので、数分後にもう一度『速報データを取得』を押してください。"
                   if result["collect_started"] else
                   "realtime 側で『新規』収集を実行してから再度お試しください。"))
            return result
        result["article_id"]   = article_id
        result["snapshot_time"] = snap_time

        # 詳細ページを取得してパース
        dr = session.get(f"{_PISION_BASE_URL}/articles/{article_id}/detail", timeout=30)
        if dr.status_code != 200 or "/login" in (dr.url or ""):
            result["error"] = f"詳細ページの取得に失敗しました（status={dr.status_code}）。"
            return result
        items, skipped = _parse_pision_rt_detail(dr.text)
        result["skipped"] = skipped
        if not items:
            result["error"] = ("台別データが見つかりませんでした。"
                               "pointsが無いためスランプグラフは生成できません。")
            result["debug"]["detail_len"] = len(dr.text)
            return result
        result["items"] = items
        result["ok"] = True
        return result
    except Exception as e:
        result["error"] = f"速報データ取得中にエラー: {e}"
        return result


def draw_slump_graph(
    template_path,
    unit_id: str,
    display_name: str,
    points: list,
    diff: "int | None" = None,
    machine_name: "str | None" = None,
    show_diff: bool = True,
) -> "Image.Image":
    """スランプグラフを template に描画して PIL Image を返す。"""
    SCALE     = 3
    X_START   = 24
    X_END     = 364
    Y_ZERO    = 290
    PX_1000   = 47
    DARK_Y1   = 462  # 黒地グラフ枠の下端（これより下の区間は描画しない）
    LINE_RGB  = (255, 0, 0, 255)
    LINE_W    = 15  # 3xキャンバス上の線幅（1x換算で約5px）

    base = Image.open(str(template_path)).convert("RGBA")
    w, h = base.size

    # 高解像度キャンバスでアンチエイリアスを強化
    big      = base.resize((w * SCALE, h * SCALE), Image.NEAREST)

    if len(points) >= 2:
        max_x = max(p["x"] for p in points)
        if max_x > 0:
            x_range = X_END - X_START
            y_limit = DARK_Y1 * SCALE
            raw = [
                (
                    (X_START + (p["x"] / max_x) * x_range) * SCALE,
                    (Y_ZERO - (p["y"] / 1000) * PX_1000) * SCALE,
                )
                for p in points
            ]
            # グラフ上部（ヘッダー領域）まで線がはみ出してもクリップしない（cats.jpg準拠）。
            # 黒地下端(y_limit)より下にはみ出す区間は描画しない。下端を跨ぐ区間は境界でクリップし、
            # 黒地内へ戻る場合のみ境界点から線を再開する。
            strokes = []
            cur = []
            for i, (x, y) in enumerate(raw):
                inside = y <= y_limit
                if i == 0:
                    if inside:
                        cur = [(x, y)]
                    continue
                px, py = raw[i - 1]
                prev_inside = py <= y_limit
                if inside and prev_inside:
                    cur.append((x, y))
                elif inside and not prev_inside:
                    t = (y_limit - py) / (y - py)
                    cur = [(px + (x - px) * t, y_limit), (x, y)]
                elif not inside and prev_inside:
                    t = (y_limit - py) / (y - py)
                    cur.append((px + (x - px) * t, y_limit))
                    strokes.append(cur)
                    cur = []
            if cur:
                strokes.append(cur)

            d = ImageDraw.Draw(big)
            for stroke in strokes:
                if len(stroke) >= 2:
                    d.line(stroke, fill=LINE_RGB, width=LINE_W, joint="curve")

    # 1x にスケールダウン（ライン AA）
    result = big.resize((w, h), Image.LANCZOS).convert("RGB")
    draw   = ImageDraw.Draw(result)

    # ヘッダーテキスト（純白・1回だけ描画）
    font_name = load_font(24)
    font_uid  = load_font(24)

    def _center_xy(text, font, box_y0, box_h):
        bb = font.getbbox(text)
        tw = bb[2] - bb[0]
        th = bb[3] - bb[1]
        x = (w - tw) // 2 - bb[0]
        y = box_y0 + (box_h - th) // 2 - bb[1]
        return x, y

    name_x, name_y = _center_xy(display_name, font_name, 10, 41)
    draw.text((name_x, name_y), display_name, fill=(255, 255, 255), font=font_name)

    uid_text = f"{unit_id}番台"
    uid_x, uid_y = _center_xy(uid_text, font_uid, 57, 41)
    draw.text((uid_x, uid_y), uid_text, fill=(255, 255, 255), font=font_uid)

    # 差枚テキスト（黄色・中央寄せ）
    if points and show_diff:
        font_diff  = load_font(42)
        _raw = diff if diff is not None else points[-1]["y"]
        diff_text  = _fmt_diff(_pipeline_calc_d(_raw))
        bb = font_diff.getbbox(diff_text)
        diff_x = (w - (bb[2] - bb[0])) // 2 - bb[0]
        diff_y = (h - 18) - bb[3]
        draw.text((diff_x, diff_y), diff_text, fill=(255, 255, 0), font=font_diff)

        # 機種名テキスト（差枚数の直上・黄色・縁取り）
        if machine_name:
            _mn_sz = 34
            _mn_font = load_font(_mn_sz)
            _mn_bb = _mn_font.getbbox(machine_name)
            _mn_w  = _mn_bb[2] - _mn_bb[0]
            while _mn_w > w - 8 and _mn_sz > 14:
                _mn_sz -= 2
                _mn_font = load_font(_mn_sz)
                _mn_bb = _mn_font.getbbox(machine_name)
                _mn_w  = _mn_bb[2] - _mn_bb[0]
            _mn_h  = _mn_bb[3] - _mn_bb[1]
            _mn_x  = (w - _mn_w) // 2 - _mn_bb[0]
            _mn_y  = diff_y - _mn_h - 4 - _mn_bb[1]
            for _ox, _oy in ((-1,-1),(1,-1),(-1,1),(1,1),(0,-1),(0,1),(-1,0),(1,0)):
                draw.text((_mn_x + _ox, _mn_y + _oy), machine_name, fill=(0, 0, 0), font=_mn_font)
            draw.text((_mn_x, _mn_y), machine_name, fill=(255, 255, 0), font=_mn_font)

    return result


def _make_slump_zip(images: list) -> bytes:
    """(ファイル名, PIL Image) リストを ZIP バイト列に変換する。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, img in images:
            ibuf = io.BytesIO()
            img.save(ibuf, format="PNG")
            zf.writestr(fname, ibuf.getvalue())
    return buf.getvalue()


def _safe_filename(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name


def _slump_apply_names(fetched: list) -> list:
    """取得データ（確定/速報共通）に機種名変換を適用して _convertedName を付与し、
    points ありの機種名一覧（multiselect 用・五十音ソート）を返す。fetched は in-place で更新。"""
    _nm, _nm_norm = load_name_map()
    def _conv(raw: str) -> str:
        raw = str(raw).strip()
        if raw in _nm:
            return _nm[raw]
        k = _normalize_key(raw)
        if k in _nm_norm:
            return _nm_norm[k]
        return raw
    for _it in fetched:
        _raw = _it.get("displayName", "")
        _converted = _conv(_raw)
        # data-short-name で変換できなかった場合は data-model-name を試す
        if _converted == _raw:
            _model = _it.get("modelName", "")
            if _model and _model != _raw:
                _cm = _conv(_model)
                if _cm != _model:
                    _converted = _cm
        _it["_convertedName"] = _converted
        # Excel の機種名列用：modelName が変換マスタにあればそれを優先
        _model2 = _it.get("modelName", "")
        _cm2 = _conv(_model2) if _model2 else _raw
        _it["_machineName"] = _cm2 if _cm2 != _model2 else _converted
    return sorted({
        _it["_convertedName"]
        for _it in fetched
        if _it.get("points") and _it.get("_convertedName")
    })


def _build_inagawa_composite(
    table_imgs: "list[tuple[str, Image.Image]]",
    graph_imgs: "list[Image.Image]",
    title: str,
) -> "Image.Image":
    """稲毛専用: テーブルJPG群＋スランプグラフを1枚の白背景PNGに合成する。"""
    MARGIN     = 20
    TITLE_H    = 70
    SEC_GAP    = 20
    CANVAS_W   = 1600
    BG         = (255, 255, 255)

    # テーブル画像を CANVAS_W 幅にスケール
    scaled_tables = []
    for _fn, _img in table_imgs:
        _w, _h = _img.size
        _scale = CANVAS_W / _w if _w > 0 else 1.0
        scaled_tables.append(_img.resize((CANVAS_W, max(1, int(_h * _scale))), Image.LANCZOS))

    # グラフグリッドのレイアウト決定
    n_graphs = len(graph_imgs)
    if n_graphs == 0:
        graph_cols = 1
    elif n_graphs <= 3:
        graph_cols = n_graphs
    else:
        graph_cols = 3

    scaled_graphs = []
    g_row_h = 0
    if graph_imgs:
        g_cell_w = CANVAS_W // graph_cols
        for _g in graph_imgs:
            _gw, _gh = _g.size
            _gscale = g_cell_w / _gw if _gw > 0 else 1.0
            scaled_graphs.append(_g.resize((g_cell_w, max(1, int(_gh * _gscale))), Image.LANCZOS))
        g_row_h = max(g.size[1] for g in scaled_graphs)

    graph_rows = math.ceil(n_graphs / graph_cols) if n_graphs > 0 else 0
    total_table_h = sum(img.size[1] for img in scaled_tables) + SEC_GAP * max(0, len(scaled_tables) - 1)
    total_graph_h = graph_rows * g_row_h + SEC_GAP * max(0, graph_rows - 1) if graph_rows > 0 else 0

    canvas_h = TITLE_H + MARGIN + total_table_h
    if total_graph_h > 0:
        canvas_h += SEC_GAP * 2 + total_graph_h
    canvas_h += MARGIN

    canvas = Image.new("RGB", (CANVAS_W, canvas_h), BG)
    draw   = ImageDraw.Draw(canvas)

    # タイトル描画
    _fnt = load_font(36)
    _bb  = _fnt.getbbox(title)
    _tx  = (CANVAS_W - (_bb[2] - _bb[0])) // 2 - _bb[0]
    _ty  = (TITLE_H - (_bb[3] - _bb[1])) // 2 - _bb[1]
    draw.text((_tx, _ty), title, fill=(0, 0, 0), font=_fnt)

    # テーブル画像を縦に配置
    _y = TITLE_H + MARGIN
    for _timg in scaled_tables:
        canvas.paste(_timg, (0, _y))
        _y += _timg.size[1] + SEC_GAP

    # スランプグラフをグリッドに配置
    if scaled_graphs:
        _y += SEC_GAP  # テーブルとグラフの間に余白
        for _i, _gimg in enumerate(scaled_graphs):
            _row = _i // graph_cols
            _col = _i % graph_cols
            _gx  = _col * (CANVAS_W // graph_cols)
            _gy  = _y + _row * (g_row_h + SEC_GAP)
            canvas.paste(_gimg, (_gx, _gy))

    return canvas


def _attach_slump_to_table_side(
    table_img: "Image.Image",
    graph_imgs: "list[Image.Image]",
    bg_path=None,
) -> "Image.Image":
    """表画像（左）＋スランプグラフ4列（右）の横レイアウト合成（16台以上用）。"""
    COLS     = 4
    PAD      = 12
    GAP      = 8
    SIDE_GAP = 24  # 表とグラフエリアの間隔

    tw, th = table_img.size
    if not graph_imgs:
        return table_img.copy()

    n    = len(graph_imgs)
    rows = math.ceil(n / COLS)

    # グラフエリア幅を表と同じ幅に固定（縦レイアウトと同幅）
    graph_area_w = tw
    cell_w = max(1, (graph_area_w - PAD * 2 - GAP * (COLS - 1)) // COLS)

    scaled: "list[Image.Image]" = []
    for g in graph_imgs:
        gw, gh = g.size
        new_h = max(1, round(gh * cell_w / gw)) if gw > 0 else gh
        scaled.append(g.resize((cell_w, new_h), Image.LANCZOS))

    row_h        = max(g.size[1] for g in scaled) if scaled else 0
    graph_area_h = PAD + rows * row_h + max(0, rows - 1) * GAP + PAD
    total_h      = max(th, graph_area_h)

    # 表をキャンバス高さに合わせて縦横比維持で拡大
    if total_h > th:
        new_tw       = max(1, round(tw * total_h / th))
        table_scaled = table_img.resize((new_tw, total_h), Image.LANCZOS)
    else:
        new_tw       = tw
        table_scaled = table_img

    total_w  = new_tw + SIDE_GAP + graph_area_w

    canvas = Image.new("RGB", (total_w, total_h), (255, 255, 255))
    canvas.paste(table_scaled, (0, 0))

    graph_x0 = new_tw + SIDE_GAP
    if bg_path is not None:
        try:
            _bg = Image.open(str(bg_path)).convert("RGB").resize((graph_area_w, total_h), Image.LANCZOS)
            canvas.paste(_bg, (graph_x0, 0))
        except Exception:
            pass

    for i, g in enumerate(scaled):
        row = i // COLS
        col = i % COLS
        x = graph_x0 + PAD + col * (cell_w + GAP)
        y = PAD + row * (row_h + GAP)
        canvas.paste(g, (x, y))

    return canvas


def _build_slump_title_img(
    title: str,
    graph_imgs: "list[Image.Image]",
    bg_path=None,
) -> "Image.Image | None":
    """青タイトルバー＋スランプグラフのみの画像（表なし・秋葉原スランプ付き結果ポスト用）。"""
    if not graph_imgs:
        return None

    COLS   = 3
    PAD    = 12
    GAP    = 8
    LINE_H = 6

    gw0, gh0 = graph_imgs[0].size
    cell_w   = gw0
    total_w  = PAD * 2 + GAP * (COLS - 1) + cell_w * COLS

    # 青タイトルバー（横幅比例・_build_machine_imgと同じ描画ロジック）
    BAR_H   = round(total_w * 73 / 950)
    FONT_SZ = round(BAR_H * 40 / 73)
    font    = load_font(FONT_SZ)
    SUB     = "（優秀台）"
    GAP_TITLE = -22
    _dummy  = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    while FONT_SZ > 12:
        if title.endswith(SUB):
            _mt = title[:-len(SUB)]
            _b1 = _dummy.textbbox((0, 0), _mt,  font=font)
            _b2 = _dummy.textbbox((0, 0), SUB,   font=font)
            _tw = (_b1[2]-_b1[0]) + GAP_TITLE + (_b2[2]-_b2[0])
        else:
            _bb = _dummy.textbbox((0, 0), title, font=font)
            _tw = _bb[2] - _bb[0]
        if _tw <= total_w - 20:
            break
        FONT_SZ -= 2
        font = load_font(FONT_SZ)

    rows         = math.ceil(len(graph_imgs) / COLS)
    graph_area_h = PAD + rows * gh0 + max(0, rows - 1) * GAP + PAD
    total_h      = BAR_H + LINE_H + graph_area_h

    canvas = Image.new("RGB", (total_w, total_h), (255, 255, 255))

    bar = Image.new("RGBA", (total_w, BAR_H), (38, 76, 161, 255))
    bd  = ImageDraw.Draw(bar)
    if title.endswith(SUB):
        main_text = title[:-len(SUB)].replace('･', '・')
        b1 = bd.textbbox((0, 0), main_text, font=font)
        b2 = bd.textbbox((0, 0), SUB,       font=font)
        w1, w2   = b1[2]-b1[0], b2[2]-b2[0]
        _tw2     = w1 + GAP_TITLE + w2
        x1 = (total_w - _tw2) // 2 - b1[0]
        x2 = x1 + w1 + GAP_TITLE - b2[0]
        ty = (BAR_H - (b1[3]-b1[1])) // 2 - b1[1]
        bd.text((x1, ty), main_text, fill=(255, 255, 255, 255), font=font)
        bd.text((x2, ty), SUB,       fill=(255, 255, 255, 255), font=font)
    else:
        disp_title = title.replace('･', '・')
        _bb = bd.textbbox((0, 0), disp_title, font=font)
        bd.text(
            ((total_w - (_bb[2]-_bb[0])) // 2 - _bb[0],
             (BAR_H  - (_bb[3]-_bb[1])) // 2 - _bb[1]),
            disp_title, fill=(255, 255, 255, 255), font=font,
        )
    canvas.paste(bar.convert("RGB"), (0, 0))
    canvas.paste(Image.new("RGB", (total_w, LINE_H), (204, 0, 0)), (0, BAR_H))

    y0 = BAR_H + LINE_H
    if bg_path is not None:
        try:
            _bg = Image.open(str(bg_path)).convert("RGB").resize((total_w, graph_area_h), Image.LANCZOS)
            canvas.paste(_bg, (0, y0))
        except Exception:
            pass

    for i, g in enumerate(graph_imgs):
        row = i // COLS
        col = i % COLS
        canvas.paste(g, (PAD + col * (cell_w + GAP), y0 + PAD + row * (gh0 + GAP)))

    return canvas


def _find_slump_bg() -> "object | None":
    """bbb.jpg（グラフエリア背景）を探してPathを返す。"""
    from pathlib import Path as _P
    candidates = [
        _P(r"C:\Users\23-3\Desktop\bbb.jpg"),
        _P(os.path.abspath(__file__)).parent / "bbb.jpg",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _attach_slump_to_table(
    table_img: "Image.Image",
    graph_imgs: "list[Image.Image]",
    bg_path=None,
) -> "Image.Image":
    """表画像の下にスランプグラフを3列で並べて合成する（稲毛スランプ付き専用）。

    * グラフは表幅に収まるよう3列グリッドにスケール（縦横比維持）。
    * 最終行が3未満の場合は中央寄せ。
    * bg_path が指定されている場合、グラフエリアの背景に貼り付ける。
    """
    COLS = 3
    PAD  = 12  # 外周余白 (px)
    GAP  = 8   # グラフ間隙間 (px)

    tw, th = table_img.size

    if not graph_imgs:
        return table_img.copy()

    n    = len(graph_imgs)
    # 3列基準でセル幅を決定（1列・2列でも同幅セルを使い中央寄せ）
    cell_w = max(1, (tw - PAD * 2 - GAP * (COLS - 1)) // COLS)

    scaled: "list[Image.Image]" = []
    for g in graph_imgs:
        gw, gh = g.size
        new_h = max(1, round(gh * cell_w / gw)) if gw > 0 else gh
        scaled.append(g.resize((cell_w, new_h), Image.LANCZOS))

    row_h = max(g.size[1] for g in scaled) if scaled else 0
    rows  = math.ceil(n / COLS)
    graph_area_h = PAD + rows * row_h + max(0, rows - 1) * GAP + PAD

    canvas = Image.new("RGB", (tw, th + graph_area_h), (255, 255, 255))
    canvas.paste(table_img, (0, 0))

    if bg_path is not None:
        try:
            _bg = Image.open(str(bg_path)).convert("RGB").resize((tw, graph_area_h), Image.LANCZOS)
            canvas.paste(_bg, (0, th))
        except Exception:
            pass

    for i, g in enumerate(scaled):
        row = i // COLS
        col = i % COLS
        x_off = PAD
        x = x_off + col * (cell_w + GAP)
        y = th + PAD + row * (row_h + GAP)
        canvas.paste(g, (x, y))

    return canvas


def show_slump_graph_page() -> None:
    st.header("📈 スランプグラフ生成")

    # ── APIキー確認 ──────────────────────────────────────────────────
    api_key = _get_pision_api_key()
    if not api_key:
        st.error("❌ APIキーが未設定です。.env に `PISION_API_KEY` を設定してください。")
        st.code("PISION_API_KEY=your-api-key-here", language="text")
        return

    # ── テンプレート確認 ─────────────────────────────────────────────
    template_path = find_slump_template()
    if template_path is None:
        st.warning("⚠️ `base_3000_bk.png` が見つかりません。画像作成フォルダまたはアプリと同じフォルダに配置してください。")
    else:
        st.caption(f"✅ テンプレート: `{template_path}`")

    # ── データ種別（確定 / 速報）─────────────────────────────────────
    _mode = st.radio(
        "データ種別",
        ["確定データ", "速報データ（当日・営業中）"],
        horizontal=True,
        key="slump_mode",
        help="確定データ＝前日まで（X-Api-Key）。速報データ＝当日の営業中データ（realtimeログインが必要）。",
    )
    is_realtime = _mode.startswith("速報")
    if is_realtime:
        _rt_user, _rt_pass = _get_pision_rt_credentials()
        if not _rt_user or not _rt_pass:
            st.error("❌ 速報データには realtime のログイン情報が必要です。"
                     ".env に以下を追記してください。")
            st.code("PISION_RT_USER=ログインID\nPISION_RT_PASS=ログインパスワード", language="text")
            return
        st.caption("🟢 速報モード：取得時に realtime で最新収集を開始します"
                   "（その店舗が一時的に『実行中』になります）。収集完了まで数分かかるため、"
                   "表示されるのは『今すぐ読める直近の完了スナップショット』です。"
                   "完了後にもう一度取得すると、今開始した収集の結果になります。")

    # ── ホール一覧（セッションキャッシュ）───────────────────────────
    _hall_key = "_slump_halls"
    if _hall_key not in st.session_state:
        with st.spinner("ホール一覧を取得中..."):
            try:
                halls = fetch_pision_halls(api_key)
                st.session_state[_hall_key] = halls
            except Exception as e:
                st.error(f"❌ ホール一覧取得失敗: {e}")
                return

    halls = st.session_state.get(_hall_key, [])
    halls_visible = [
        h for h in halls
        if not h.get("secret", False)
        and "エスパス" in (h.get("name") or h.get("displayName") or "")
        and any(kw in (h.get("name") or h.get("displayName") or "") for kw in _PISION_ALLOWED_HALLS)
    ]
    if not halls_visible:
        st.error("❌ 利用可能なホールが見つかりません。")
        if st.button("🔄 再取得"):
            st.session_state.pop(_hall_key, None)
            st.rerun()
        return

    def _hname(h):
        return h.get("name") or h.get("displayName") or str(h.get("id", ""))

    def _hid(h):
        return str(h.get("id") or h.get("hallId") or "")

    hall_names = [_hname(h) for h in halls_visible]
    hall_ids   = [_hid(h)   for h in halls_visible]

    # ── 入力フォーム ─────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        sel_idx = st.selectbox(
            "店舗を選択",
            options=range(len(hall_names)),
            format_func=lambda i: hall_names[i],
            key="slump_hall_idx",
        )
        sel_hall_id = hall_ids[sel_idx]
    with col2:
        sel_date = st.date_input("日付を選択", key="slump_date")
        date_str = sel_date.strftime("%Y-%m-%d")

    # ── Step1: 機種一覧を取得（ホール・日付・モードごとにキャッシュ）──────
    _mode_tag  = "rt" if is_realtime else "fix"
    _data_key  = f"_slump_data_{_mode_tag}_{sel_hall_id}_{date_str}"
    _names_key = f"_slump_names_{_mode_tag}_{sel_hall_id}_{date_str}"
    _artid_key = f"_slump_artid_{_mode_tag}_{sel_hall_id}_{date_str}"
    _btn_label = "🟢 速報データを取得" if is_realtime else "🔍 機種一覧を取得"

    if _data_key not in st.session_state:
        if st.button(_btn_label, use_container_width=True, key="slump_fetch"):
            _fetched = None
            if is_realtime:
                # ── 速報データ（realtime ログイン・最新収集を開始して直近完了分を読む）──
                with st.spinner(f"{date_str} の速報データを取得中...（ログイン→最新収集を開始→直近完了分を取得）"):
                    _rt = fetch_pision_realtime(hall_names[sel_idx], date_str)
                if not _rt["ok"]:
                    st.error(f"❌ {_rt['error']}")
                    if _rt.get("collect_started"):
                        st.info("🔄 realtime で最新収集を開始しました（実行中）。完了まで数分かかります。")
                    if _rt.get("debug"):
                        with st.expander("🔧 デバッグ情報"):
                            st.json(_rt["debug"])
                    return
                _fetched = _rt["items"]
            else:
                # ── 確定データ（X-Api-Key）─────────────────────────────
                with st.spinner(f"{date_str} のデータを取得中..."):
                    try:
                        _fetched = fetch_pision_results(api_key, sel_hall_id, date_str)
                    except Exception as e:
                        st.error(f"❌ データ取得失敗: {e}")
                        return
                if _fetched is None:
                    st.info("📭 データを取得できませんでした（404 / 未公開 / 店休日の可能性があります）。")
                    return
            if not _fetched:
                st.info("📭 この日付のデータがありません。")
                return
            # 機種名変換マスタを適用して _convertedName を付与
            st.session_state[_data_key]  = _fetched
            st.session_state[_names_key] = _slump_apply_names(_fetched)
            if is_realtime:
                st.session_state[_artid_key] = _rt.get("article_id")
                st.session_state[f"{_artid_key}_time"] = _rt.get("snapshot_time")
            st.rerun()
        return  # 取得前はここで終了

    # ── Step2: フィルタ選択・グラフ生成 ──────────────────────────────
    cached_details = st.session_state[_data_key]
    cached_names   = st.session_state.get(_names_key, [])

    col3, col4 = st.columns(2)
    with col3:
        selected_machines = st.multiselect(
            "機種名を選択（空欄=全台・入力で絞り込み）",
            options=cached_names,
            key="slump_machine_ms",
        )
    with col4:
        unit_filter = st.text_input("台番号検索（空欄可）", key="slump_unit")

    st.caption(f"📋 取得済み: {len(cached_details)} 台 / points あり: {len(cached_names)} 機種")

    if is_realtime:
        # 速報モード：現在表示中のスナップショット時刻と、収集完了の確認ボタン
        _cur_artid = st.session_state.get(_artid_key)
        _snap = st.session_state.get(f"{_artid_key}_time")
        if _snap:
            st.info(f"🕒 表示中データの収集時刻: **{_snap}**　"
                    "（取得時に最新収集を開始済み。完了したら下の『収集の完了を確認』で最新化できます）")
        _c_chk, _c_re = st.columns(2)
        with _c_chk:
            if st.button("🔄 収集の完了を確認", use_container_width=True, key="slump_rt_check",
                         help="先ほど開始した収集が完了したか確認します（収集は再実行しません）。"):
                with st.spinner("収集状況を確認中..."):
                    _chk = fetch_pision_realtime(hall_names[sel_idx], date_str, trigger=False)
                if not _chk["ok"]:
                    st.error(f"❌ {_chk['error']}")
                else:
                    _new_id = _chk.get("article_id")
                    if _cur_artid is not None and _new_id and int(_new_id) > int(_cur_artid):
                        # 新しい完了スナップショットに更新（収集はトリガーしない）
                        _items = _chk["items"]
                        st.session_state[_data_key]  = _items
                        st.session_state[_names_key] = _slump_apply_names(_items)
                        st.session_state[_artid_key] = _new_id
                        st.session_state[f"{_artid_key}_time"] = _chk.get("snapshot_time")
                        st.success(f"✅ 収集完了！最新データ（収集時刻 {_chk.get('snapshot_time')}）に更新しました。")
                        st.rerun()
                    elif _chk.get("running"):
                        st.info("⏳ まだ実行中です。完了まで数分かかります。少し待ってからもう一度押してください。")
                    else:
                        st.info(f"ℹ️ 新しい完了データはまだありません（最新の完了は {_chk.get('snapshot_time')}）。"
                                "少し待ってからもう一度押してください。")
        with _c_re:
            if st.button("🟢 最新を再収集して取得", use_container_width=True, key="slump_refetch",
                         help="あらためて realtime で収集を開始します（一時的に実行中になります）。"):
                st.session_state.pop(_data_key, None)
                st.session_state.pop(_names_key, None)
                st.session_state.pop(_artid_key, None)
                st.session_state.pop(f"{_artid_key}_time", None)
                st.rerun()
    else:
        if st.button("🔄 再取得", key="slump_refetch"):
            st.session_state.pop(_data_key, None)
            st.session_state.pop(_names_key, None)
            st.rerun()

    with st.expander("🔍 取得済み全機種一覧（機種が見つからない場合に確認）"):
        _all_names_debug = sorted({
            (
                _it.get("_convertedName") or _it.get("displayName", ""),
                _it.get("displayName", ""),
                "✅" if _it.get("points") else "❌ pointsなし",
            )
            for _it in cached_details
            if _it.get("displayName")
        }, key=lambda x: x[0])
        if _all_names_debug:
            _debug_df = pd.DataFrame(_all_names_debug, columns=["表示名（変換後）", "API名（displayName）", "グラフ"])
            st.dataframe(_debug_df, use_container_width=True, hide_index=True)
        else:
            st.info("データなし")

    # ── 📋 取得データをpision風に表示（結果ポスト用と同じビュー）──────────
    _slump_rows = [
        {
            "台番":     _it.get("unitId"),
            "機種名":   _it.get("_convertedName") or _it.get("displayName", ""),
            "差枚":     _it.get("diff", 0),
            "ゲーム数": _it.get("games", 0),
            "BB":       _it.get("bb", 0),
            "RB":       _it.get("rb", 0),
            "AT":       _it.get("art", 0),
        }
        for _it in cached_details
    ]
    _slump_vt_df, _ = normalize_df(pd.DataFrame(_slump_rows))
    _m_sl = re.match(r"(\d{4})-(\d{2})-(\d{2})", date_str)
    _sl_title = (f"{int(_m_sl.group(1))}/{int(_m_sl.group(2))}/{int(_m_sl.group(3))} {hall_names[sel_idx]}"
                 if _m_sl else hall_names[sel_idx])
    render_pision_data_view(_slump_vt_df, _sl_title, sel_key="slump_detail_machine")

    st.divider()

    if st.button("グラフ生成", type="primary", use_container_width=True, key="slump_run"):
        if template_path is None:
            st.error("❌ `base_3000_bk.png` が見つかりません。テンプレート画像を配置してください。")
            return

        # フィルタ
        filtered = []
        skipped_no_points = 0
        for item in cached_details:
            if not item.get("points"):
                skipped_no_points += 1
                continue
            name = item.get("_convertedName") or item.get("displayName", "")
            uid  = str(item.get("unitId", ""))
            if selected_machines and name not in selected_machines:
                continue
            if unit_filter and unit_filter != uid:
                continue
            filtered.append(item)

        if skipped_no_points:
            st.caption(f"ℹ️ points なしの台 {skipped_no_points} 件をスキップしました。")

        if not filtered:
            st.warning("⚠️ 条件に一致する台（points あり）が見つかりませんでした。")
            return

        # グラフ生成
        pb     = st.progress(0, text="グラフ生成中...")
        images = []
        errors = []
        for i, item in enumerate(filtered):
            try:
                uid  = str(item.get("unitId", ""))
                name = item.get("_convertedName") or item.get("displayName", "")
                pts  = item["points"]
                img  = draw_slump_graph(template_path, uid, name, pts, diff=item.get("diff"))
                fname = _safe_filename(f"{uid}_{name}.png")
                images.append((fname, img))
            except Exception as e:
                errors.append(f"台{item.get('unitId','?')}: {e}")
            pb.progress((i + 1) / len(filtered), text=f"生成中 {i + 1}/{len(filtered)}")
        pb.empty()

        if errors:
            with st.expander(f"⚠️ 生成エラー {len(errors)} 件"):
                for err in errors:
                    st.text(err)

        if not images:
            st.error("❌ グラフを1件も生成できませんでした。")
            return

        st.success(f"✅ {len(images)} 台分のグラフを生成しました")

        # サムネイル表示（4列）
        cols_per_row = 4
        for row_start in range(0, len(images), cols_per_row):
            cols = st.columns(cols_per_row)
            for ci, idx in enumerate(range(row_start, min(row_start + cols_per_row, len(images)))):
                fname, img = images[idx]
                with cols[ci]:
                    st.image(img, caption=fname, use_container_width=True)

        # ZIP ダウンロード
        with st.spinner("ZIP 作成中..."):
            zip_bytes = _make_slump_zip(images)

        st.download_button(
            label="📦 ZIP ダウンロード",
            data=zip_bytes,
            file_name=_safe_filename(f"slump_{hall_names[sel_idx]}_{date_str}.zip"),
            mime="application/zip",
            use_container_width=True,
            key="slump_zip_dl",
        )


def show_machine_image_page() -> None:
    """簡略名と機種画像（パネル/液晶）の紐づけ確認・プレビュー。"""
    st.header("🖼️ 機種画像紐づけ")
    st.caption(
        "簡略名と機種画像（パネル/液晶）の紐づけを確認します。"
        "画像は `assets/machine_images/`、マスタは `masters/machine_image_master.xlsx`。"
    )

    master = load_machine_image_master()
    name_map, _ = load_name_map()
    used_short_names = sorted(set(name_map.values()))

    # フォルダ/マスタの状態表示
    if not os.path.isdir(_MACHINE_IMAGES_DIR):
        st.warning(f"画像フォルダが見つかりません: `{_rel_machine_path(_MACHINE_IMAGES_DIR)}`")
    if not master:
        st.warning("画像マスタが未読込です: `masters/machine_image_master.xlsx`")

    # ── 紐づけ一覧 ────────────────────────────────────────────────
    st.subheader("紐づけ一覧")
    rows = []
    for sn, gid in sorted(master.items()):
        info = get_machine_images(sn)
        panel_ok = bool(info and info["panel"])
        n_screens = len(info["screens"]) if info else 0
        status = "OK" if (panel_ok or n_screens > 0) else "要確認"
        rows.append({
            "簡略名": sn,
            "画像ID": gid,
            "パネル": "あり" if panel_ok else "なし",
            "液晶枚数": f"{n_screens}枚",
            "状態": status,
        })
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info("マスタに登録がありません。")

    # ── プレビュー ────────────────────────────────────────────────
    st.subheader("プレビュー")
    if master:
        sel = st.selectbox("簡略名を選択", options=sorted(master.keys()), key="mi_preview_sel")
        info = get_machine_images(sel)
        if info:
            st.write(f"画像グループID: `{info['image_id']}`")
            _c1, _c2 = st.columns([1, 2])
            with _c1:
                st.markdown("**パネル画像**")
                if info["panel"]:
                    st.image(os.path.join(BASE_DIR, info["panel"]),
                             caption=os.path.basename(info["panel"]), use_container_width=True)
                else:
                    st.info("パネル画像なし")
            with _c2:
                st.markdown(f"**液晶画像（{len(info['screens'])}枚）**")
                if info["screens"]:
                    _cols = st.columns(min(3, len(info["screens"])))
                    for _i, _sp in enumerate(info["screens"]):
                        with _cols[_i % len(_cols)]:
                            st.image(os.path.join(BASE_DIR, _sp),
                                     caption=os.path.basename(_sp), use_container_width=True)
                else:
                    st.info("液晶画像なし")
        else:
            st.info("この簡略名には画像グループIDが紐づいていません。")

    # ── 新規紐づけ追加 / 更新フォーム ──────────────────────────────
    st.markdown("---")
    st.subheader("新規紐づけを追加")
    st.caption(
        "新機種の 簡略名 と 画像グループID を入力して紐づけます。"
        "画像グループIDは `○○_panel.png` の **`○○` の部分だけ**を入力すればOK。"
        "パネル `{ID}_panel.png` と液晶 `{ID}_01.png` 以降を自動で探します。"
    )

    # 直近保存の結果＆GitHub同期パネル（rerunをまたいで表示）
    _saved = st.session_state.get("mi_last_saved")
    if _saved:
        st.success(f"保存しました: 「{_saved[0]}」→ 画像グループID `{_saved[1]}`（{_saved[2]}）")
        _sc1, _sc2 = st.columns([1, 3])
        with _sc1:
            if st.button("GitHubへ同期", key="mi_sync_github"):
                _ok, _msg = _sync_machine_image_master()
                (st.success if _ok else st.error)(_msg)
        with _sc2:
            if _IS_CLOUD:
                st.caption("⚠️ Cloudでは再起動時に消える可能性があるため、最終的にGitHubへ反映が必要です。")
            else:
                st.caption("ローカルに保存済み。Cloudにも反映するには「GitHubへ同期」を押してください。")
        if st.button("閉じる", key="mi_saved_dismiss"):
            st.session_state.pop("mi_last_saved", None)
            st.rerun()

    _f1, _f2 = st.columns(2)
    with _f1:
        _in_sn = st.text_input("簡略名", key="mi_add_sn", placeholder="例: 北斗")
    with _f2:
        _in_gid = st.text_input(
            "画像グループID（_panel.png の前の部分）", key="mi_add_gid", placeholder="例: hokuto"
        )
    _sn, _gid = _in_sn.strip(), _in_gid.strip()

    if _gid:
        _panel = _find_panel_image(_gid)
        _screens = _find_screen_images(_gid)
        _exists_gid = master.get(_sn) if _sn else None

        # 状態サマリー
        _state = "登録可能" if _panel else "画像未配置（登録は可能）"
        st.markdown(
            f"**簡略名:** {_sn or '（未入力）'}　／　**画像グループID:** `{_gid}`　／　"
            f"**パネル:** {'あり' if _panel else 'なし'}　／　"
            f"**液晶画像:** {len(_screens)}枚　／　**状態:** {_state}"
        )
        if not _panel:
            st.warning("パネル画像が見つかりません。後から画像を追加する場合はこのまま登録できます。")

        # プレビュー
        _pc1, _pc2 = st.columns([1, 2])
        with _pc1:
            st.markdown("**パネル画像**")
            if _panel:
                st.image(os.path.join(BASE_DIR, _panel),
                         caption=os.path.basename(_panel), use_container_width=True)
            else:
                st.info("なし")
        with _pc2:
            st.markdown(f"**液晶画像（{len(_screens)}枚）**")
            if _screens:
                _cols = st.columns(min(3, len(_screens)))
                for _i, _sp in enumerate(_screens):
                    with _cols[_i % len(_cols)]:
                        st.image(os.path.join(BASE_DIR, _sp),
                                 caption=os.path.basename(_sp), use_container_width=True)
            else:
                st.info("なし")

        # 重複時は更新 / それ以外は新規追加
        if _exists_gid is not None:
            st.info(f"「{_sn}」は既に登録済みです（現在: `{_exists_gid}`）。")
            _mode = st.radio(
                "処理を選択", options=["既存の画像グループIDを更新する", "何もしない"],
                key="mi_add_mode", horizontal=True,
            )
            if st.button("この紐づけを更新", key="mi_add_update", type="primary"):
                if not _sn or not _gid:
                    st.error("簡略名と画像グループIDを入力してください。")
                elif _mode != "既存の画像グループIDを更新する":
                    st.warning("更新するには「既存の画像グループIDを更新する」を選択してください。")
                else:
                    _act, _g = save_machine_image_mapping(_sn, _gid, overwrite=True)
                    st.session_state["mi_last_saved"] = (_sn, _g, "更新")
                    st.rerun()
        else:
            if st.button("この紐づけを追加", key="mi_add_new", type="primary"):
                if not _sn or not _gid:
                    st.error("簡略名と画像グループIDを入力してください。")
                else:
                    _act, _g = save_machine_image_mapping(_sn, _gid, overwrite=False)
                    _label = "新規追加" if _act == "added" else "既存"
                    st.session_state["mi_last_saved"] = (_sn, _g, _label)
                    st.rerun()
    else:
        st.caption("画像グループIDを入力すると、プレビューと登録ボタンが表示されます。")

    # 未登録の簡略名を算出（自動紐づけ・下部の一覧で使用）
    _master_norm = {_normalize_key(k) for k in master}
    unregistered = [s for s in used_short_names
                    if s not in master and _normalize_key(s) not in _master_norm]

    # ── 未登録機種の画像候補を自動作成 ────────────────────────────
    st.markdown("---")
    st.subheader("未登録機種の自動紐づけ")
    st.caption(
        "未登録の簡略名と `*_panel.png` のファイル名を比較し、画像グループIDを予測します。"
        "類似度 0.85以上=自動採用候補 / 0.65〜0.85=要確認 / 0.65未満=候補なし。"
    )

    _AUTO_TH, _CHECK_TH = 0.65, 0.85
    if st.button("未登録機種の画像候補を自動作成", key="mi_gen_candidates"):
        group_ids = list_image_group_ids()
        cands = []
        for sn in unregistered:
            gid, score, reason = predict_image_id(sn, group_ids)
            if gid is None or score < _AUTO_TH:
                status = "候補なし"
                gid_out = ""
            elif score >= _CHECK_TH:
                status = "自動採用候補"
                gid_out = gid
            else:
                status = "要確認"
                gid_out = gid
            cands.append({
                "簡略名": sn,
                "推定画像グループID": gid_out,
                "類似度": score,
                "判定": reason,
                "状態": status,
            })
        st.session_state["mi_candidates"] = cands
        # 生成のたびにウィジェットキーを一新（古いチェック状態を残さない）
        st.session_state["mi_cand_gen"] = st.session_state.get("mi_cand_gen", 0) + 1

    _cands = st.session_state.get("mi_candidates")
    _gen = st.session_state.get("mi_cand_gen", 0)
    if _cands is not None:
        if not _cands:
            st.success("未登録の簡略名はありません。")
        else:
            st.markdown(f"**候補 {len(_cands)}件**（採用にチェック→下部のボタンで保存）")
            # ヘッダー行
            _h = st.columns([2, 2, 1, 1.2, 1, 2])
            for _c, _t in zip(_h, ["簡略名", "画像グループID(手修正可)", "類似度", "状態", "採用", "パネル"]):
                _c.markdown(f"**{_t}**")
            for _i, _cd in enumerate(_cands):
                _c1, _c2, _c3, _c4, _c5, _c6 = st.columns([2, 2, 1, 1.2, 1, 2])
                _c1.write(_cd["簡略名"])
                # 入力欄は空。推定IDはプレースホルダー（薄いヒント）に表示するだけ
                _gid = _c2.text_input(
                    "gid", value="",
                    key=f"mi_cand_gid_{_gen}_{_i}", label_visibility="collapsed",
                    placeholder=(_cd["推定画像グループID"] or "画像グループID"),
                )
                _c3.write(f"{_cd['類似度']:.2f}")
                _c4.write(_cd["状態"])
                # チェックは常にデフォルトOFF（自動でONにしない＝勝手に登録しない）
                _c5.checkbox(
                    "採用", value=False,
                    key=f"mi_cand_use_{_gen}_{_i}", label_visibility="collapsed",
                )
                # パネルプレビュー（入力値優先・空欄なら推定IDで表示）
                _eff_gid = _gid.strip() or str(_cd["推定画像グループID"]).strip()
                _panel = _find_panel_image(_eff_gid) if _eff_gid else None
                with _c6:
                    if _panel:
                        st.image(os.path.join(BASE_DIR, _panel), use_container_width=True)
                    else:
                        st.caption("画像なし")

            if st.button("採用してマスタに追加", key="mi_save_candidates", type="primary"):
                _rows = []
                _adopted_names = set()
                for _i, _cd in enumerate(_cands):
                    if st.session_state.get(f"mi_cand_use_{_gen}_{_i}"):
                        # 手入力があればそれを、空欄なら推定ID（プレースホルダー）を採用
                        _gid = str(st.session_state.get(f"mi_cand_gid_{_gen}_{_i}", "")).strip()
                        if not _gid:
                            _gid = str(_cd["推定画像グループID"]).strip()
                        if _gid:
                            _rows.append((_cd["簡略名"], _gid))
                            _adopted_names.add(_cd["簡略名"])
                if not _rows:
                    st.warning("採用チェックされた候補がありません。")
                else:
                    _added = append_machine_image_master(_rows)
                    # 採用しなかった候補は「要確認」として残す
                    _remaining = []
                    for _cd in _cands:
                        if _cd["簡略名"] not in _adopted_names:
                            _nc = dict(_cd)
                            _nc["状態"] = "要確認"
                            _remaining.append(_nc)
                    st.session_state["mi_candidates"] = _remaining
                    st.session_state["mi_cand_gen"] = _gen + 1
                    st.success(f"{_added}件をマスタに追加しました。残り{len(_remaining)}件は要確認です。")
                    if not _IS_CLOUD:
                        st.info("反映するには masters/machine_image_master.xlsx をcommit/pushしてください。")
                    st.rerun()

    # ── 未登録の簡略名 ────────────────────────────────────────────
    st.markdown("---")
    st.subheader("未登録の簡略名")
    if unregistered:
        st.warning(f"{len(unregistered)}件が画像マスタ未登録です（機種名変換の変換後の名前が基準）。")
        st.dataframe(pd.DataFrame({"簡略名": unregistered}), use_container_width=True, hide_index=True)
    else:
        st.success("未登録の簡略名はありません。")

    st.markdown("---")
    if st.button("← 画像生成トップへ戻る", key="mi_back"):
        _navigate("store")


# =============================================================================
# ■ ⑨メイン
# =============================================================================

def init_session_state() -> None:
    """セッション状態の初期化（初回アクセス時のみ実行される）"""
    defaults = {
        "page":                "store",
        "selected_store":      "",
        "selected_image_type": "",
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def main() -> None:
    # ページ設定（ブラウザタブのタイトルとレイアウト）
    st.set_page_config(
        page_title="ギルド画像生成",
        page_icon="🎰",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    init_session_state()
    _sync_from_query_params()

    # ローカルのみ: 起動時に1回だけgit pullしてCloudの変更を取り込む
    if not _IS_CLOUD and not st.session_state.get("_git_pulled"):
        st.session_state["_git_pulled"] = True
        _pull_ok, _pull_msg = _git_auto_pull()
        if not _pull_ok:
            st.toast(f"⚠️ git pull失敗: {_pull_msg}", icon="⚠️")

    # ブラウザの戻る/進むボタンを検知してページをリロードさせる
    # components.html は iframe 内なので window.parent を使う
    # st.iframe は height=0 非対応のためレイアウトが崩れる→こちらを維持（警告はロガーで抑制済み）
    components.html(
        """
        <script>
        (function() {
            var p = window.parent;
            if (p._popstateAttached) return;
            p._popstateAttached = true;
            p.addEventListener('popstate', function() {
                p.location.reload();
            });
        })();
        </script>
        """,
        height=0,
    )

    # 入力欄のブラウザ履歴オートコンプリートを無効化
    components.html(
        """
        <script>
        (function() {
            var p = window.parent;
            if (p._autocompleteDisabled) return;
            p._autocompleteDisabled = true;
            function disableAutocomplete() {
                p.document.querySelectorAll('input[type="text"], input:not([type])').forEach(function(el) {
                    el.setAttribute('autocomplete', 'off');
                });
            }
            disableAutocomplete();
            new p.MutationObserver(disableAutocomplete).observe(
                p.document.body, { childList: true, subtree: true }
            );
        })();
        </script>
        """,
        height=0,
    )

    # ── サイドバー ────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## 🎰 データ画像生成")
        st.markdown("---")

        # カテゴリボタン
        if st.button("🖼️ 画像生成", use_container_width=True, key="nav_image"):
            _navigate("store")
        if st.button("📈 スランプグラフ生成", use_container_width=True, key="nav_slump"):
            _navigate("slump_graph")
        if st.button("🔄 機種名変換", use_container_width=True, key="nav_nc"):
            _navigate("name_conversion")
        if st.button("🖼️ 機種画像紐づけ", use_container_width=True, key="nav_mi"):
            _navigate("machine_image")

        st.markdown("---")

        # 現在地をパンくずで表示
        page = st.session_state.page
        if page == "store":
            st.markdown("📍 **店舗選択**")
        elif page == "image_type":
            st.markdown(f"📍 **{st.session_state.selected_store}**")
            st.markdown("　→ 画像種類選択")
        elif page == "work":
            st.markdown(f"📍 **{st.session_state.selected_store}**")
            st.markdown(f"　→ **{st.session_state.selected_image_type}**")
        elif page == "auto":
            st.markdown(f"📍 **{st.session_state.selected_store}**")
            st.markdown("　→ **⚡ 結果ポスト用**")
        elif page == "auto_slump":
            st.markdown(f"📍 **{st.session_state.selected_store}**")
            st.markdown("　→ **📊 スランプ付き結果ポスト用**")
        elif page == "rote":
            st.markdown(f"📍 **{st.session_state.selected_store}**")
            st.markdown("　→ **📋 ローテ用**")
        elif page == "weekly_result_text":
            st.markdown(f"📍 **{st.session_state.selected_store}**")
            st.markdown("　→ **📅 1週間の結果テキスト**")
        elif page == "name_conversion":
            st.markdown("📍 **機種名変換**")
        elif page == "slump_graph":
            st.markdown("📍 **スランプグラフ生成**")
        elif page == "machine_image":
            st.markdown("📍 **機種画像紐づけ**")

    # ── アプリタイトル ────────────────────────────────────────────
    st.title("ギルド画像生成")

    # ── ページルーティング ────────────────────────────────────────
    if st.session_state.page == "store":
        show_store_page()
    elif st.session_state.page == "image_type":
        show_image_type_page()
    elif st.session_state.page == "work":
        show_work_page()
    elif st.session_state.page == "auto":
        show_auto_page()
    elif st.session_state.page == "auto_slump":
        show_auto_page(with_slump=True)
    elif st.session_state.page == "rote":
        show_rote_page()
    elif st.session_state.page == "weekly_result_text":
        show_weekly_result_text_page()
    elif st.session_state.page == "auto_article":
        show_auto_article_page()
    elif st.session_state.page == "name_conversion":
        show_name_conversion_page()
    elif st.session_state.page == "slump_graph":
        show_slump_graph_page()
    elif st.session_state.page == "machine_image":
        show_machine_image_page()
    else:
        # 不正な状態はトップに戻す
        st.session_state.page = "store"
        st.rerun()


if __name__ == "__main__":
    main()
